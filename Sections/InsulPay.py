# import packages
import os
import shutil
from openpyxl import load_workbook
import pandas as pd
import subprocess
import datetime
import tkinter as tk
from tkinter import messagebox, filedialog, ttk

import customtkinter as ctk

print('Importing Modules')


class Employee:
    def __init__(self, first, last, vacationdays, passwd):
        self.first = first
        self.last = last
        self.vacationdays = vacationdays
        self.fullname = first + ' ' + last
        self.pwd = passwd
        self.username = first[0] + last


employee0 = Employee('Admin', 'Account', 0, '')
employee1 = Employee("Justin", "Bausinger", 5, '')
employee2 = Employee("Cody", "Billings", 0, '')
employee3 = Employee("Jeremy", "Denton", 5, '')
employee4 = Employee("Glen", "Hayes", 10, '')
employee5 = Employee("Brian", "Maikranz", 5, '')
employee6 = Employee("Mike", "Mato", 5, '')
employee7 = Employee("Kenny", "Ray", 10, '')
employee8 = Employee("Edgar", "Sanchez", 5, '')
employee9 = Employee("Chris", "Waldridge", 10, '')
employee10 = Employee("New", "Employee", 0, '')

employee0.username = ''
employee0.pwd = ""


class Payrate:
    def __init__(self, name, rate):
        self.name = name
        self.rate = float(rate)


# update pay rates here!!!!
batt1 = Payrate("Batt >10ft", 0.09)
batt2 = Payrate("Batt <10ft", 0.10)
soffit = Payrate("Soffit - New", 0.55)
existing_soffit = Payrate("Soffit - Existing", 2.50)
bibs_full = Payrate("Bibs - Full Install", 0.20)
bibs_hung = Payrate("Bibs - Netting", 0.07)
bibs_tacked = Payrate("Bibs - Air Tack", 0.07)
bibs_blown = Payrate("Bibs - Blow", 0.06)
r19_blown = Payrate("R-19 Blow", 0.08)
r30_blown = Payrate("R-30 Blow", 0.09)
r38_blown = Payrate("R-38/40 Blow", 0.10)
r49_blown = Payrate("R-49/50 Blow", 0.11)
cell_blown = Payrate("Cellulose - Attic", 0.10)
closed_under34 = Payrate('Closed Cell 3/4"', 0.15)
closed_under1 = Payrate('Closed Cell > 1"', 0.22)
closed_under2 = Payrate('Closed Cell >= 2"', 0.22)
closed_3 = Payrate('Closed Cell <= 3"', 0.33)
open_cell4 = Payrate('Open Cell >= 4"', 0.15)
open_cell6 = Payrate('Open Cell >= 6" ', 0.20)
open_cell8 = Payrate('Open Cell <= 8"', 0.22)
vacation_day = Payrate("Vacation Day Pay", 130.0)

ctk.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
ctk.set_default_color_theme("dark-blue")  # Themes: "blue" (standard), "green", "dark-blue"


# Menu Sections
# noinspection PyGlobalUndefined
class Calculator(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Calculator")
        width = 500
        height = 500
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)

        def button_press(num):

            # noinspection PyGlobalUndefined
            global equation_text

            self.equation_text = self.equation_text + str(num)

            self.equation_label.set(self.equation_text)

        def equals():

            global equation_text

            try:

                total = str(eval(self.equation_text))

                self.equation_label.set(total)

                self.equation_text = total

            except SyntaxError:

                self.equation_label.set("syntax error")

                self.equation_text = ""

            except ZeroDivisionError:

                self.equation_label.set("arithmetic error")

                self.equation_text = ""

        def clear():

            global equation_text

            self.equation_label.set("")

            self.equation_text = ""

        self.equation_text = ""

        self.equation_label = tk.StringVar()

        Label = ctk.CTkLabel(self, width=300, height=75,
                             font=ctk.CTkFont(size=20, weight="normal"), textvariable=self.equation_label)
        Label.pack()

        frame = ctk.CTkFrame(self)
        frame.pack()

        button1 = ctk.CTkButton(frame,
                                text='1',
                                height=75,
                                width=75,
                                font=ctk.CTkFont(size=17, weight="normal"),
                                command=lambda: button_press(1))
        button1.grid(row=0, column=0)

        button2 = ctk.CTkButton(frame,
                                text='2',
                                height=75,
                                width=75,
                                font=ctk.CTkFont(size=17, weight="normal"),
                                command=lambda: button_press(2))
        button2.grid(row=0, column=1)

        button3 = ctk.CTkButton(frame,
                                text='3',
                                height=75,
                                width=75,
                                font=ctk.CTkFont(size=17, weight="normal"),
                                command=lambda: button_press(3))
        button3.grid(row=0, column=2)

        button4 = ctk.CTkButton(frame,
                                text='4',
                                height=75,
                                width=75,
                                font=ctk.CTkFont(size=17, weight="normal"),
                                command=lambda: button_press(4))
        button4.grid(row=1, column=0)

        button5 = ctk.CTkButton(frame,
                                text='5',
                                height=75,
                                width=75,
                                font=ctk.CTkFont(size=17, weight="normal"),
                                command=lambda: button_press(5))
        button5.grid(row=1, column=1)

        button6 = ctk.CTkButton(frame,
                                text='6',
                                height=75,
                                width=75,
                                font=ctk.CTkFont(size=17, weight="normal"),
                                command=lambda: button_press(6))
        button6.grid(row=1, column=2)

        button7 = ctk.CTkButton(frame,
                                text='7',
                                height=75,
                                width=75,
                                font=ctk.CTkFont(size=17, weight="normal"),
                                command=lambda: button_press(7))
        button7.grid(row=2, column=0)

        button8 = ctk.CTkButton(frame,
                                text='8',
                                height=75,
                                width=75,
                                font=ctk.CTkFont(size=17, weight="normal"),
                                command=lambda: button_press(8))
        button8.grid(row=2, column=1)

        button9 = ctk.CTkButton(frame,
                                text='9',
                                height=75,
                                width=75,
                                font=ctk.CTkFont(size=17, weight="normal"),
                                command=lambda: button_press(9))
        button9.grid(row=2, column=2)

        button0 = ctk.CTkButton(frame,
                                text='0',
                                height=75,
                                width=75,
                                font=ctk.CTkFont(size=17, weight="normal"),
                                command=lambda: button_press(0))
        button0.grid(row=3, column=0)

        plus = ctk.CTkButton(frame,
                             text='+',
                             height=75,
                             width=75,
                             font=ctk.CTkFont(size=17, weight="normal"),
                             command=lambda: button_press('+'))
        plus.grid(row=0, column=3)

        minus = ctk.CTkButton(frame,
                              text='-',
                              height=75,
                              width=75,
                              font=ctk.CTkFont(size=17, weight="normal"),
                              command=lambda: button_press('-'))
        minus.grid(row=1, column=3)

        multiply = ctk.CTkButton(frame,
                                 text='*',
                                 height=75,
                                 width=75,
                                 font=ctk.CTkFont(size=17, weight="normal"),
                                 command=lambda: button_press('*'))
        multiply.grid(row=2, column=3)

        divide = ctk.CTkButton(frame,
                               text='/',
                               height=75,
                               width=75,
                               font=ctk.CTkFont(size=17, weight="normal"),
                               command=lambda: button_press('/'))
        divide.grid(row=3, column=3)

        equal = ctk.CTkButton(frame,
                              text='=',
                              height=75,
                              width=75,
                              font=ctk.CTkFont(size=17, weight="normal"),
                              command=equals)
        equal.grid(row=3, column=2)

        decimal = ctk.CTkButton(frame,
                                text='.',
                                height=75,
                                width=75,
                                font=ctk.CTkFont(size=17, weight="normal"),
                                command=lambda: button_press('.'))
        decimal.grid(row=3, column=1)

        clear = ctk.CTkButton(self,
                              text='clear',
                              height=75,
                              width=75,
                              font=ctk.CTkFont(size=17, weight="normal"),
                              command=clear)
        clear.pack()


# Paysheet Sections
class BattPaySheet(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("800x600")
        self.title("Batt Pay Sheet")
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)

        self.count = 0

        self.emp1Checkbox_var = tk.IntVar(value=0)
        self.emp2Checkbox_var = tk.IntVar(value=0)
        self.emp3Checkbox_var = tk.IntVar(value=0)
        self.emp4Checkbox_var = tk.IntVar(value=0)
        self.emp5Checkbox_var = tk.IntVar(value=0)
        self.emp6Checkbox_var = tk.IntVar(value=0)
        self.emp7Checkbox_var = tk.IntVar(value=0)
        self.emp8Checkbox_var = tk.IntVar(value=0)
        self.emp9Checkbox_var = tk.IntVar(value=0)
        self.emp10Checkbox_var = tk.IntVar(value=0)

        emp1Checkbox = ctk.CTkCheckBox(self,
                                       text=employee1.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp1Checkbox_command,
                                       variable=self.emp1Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp1Checkbox.place(x=5, y=30)

        emp2Checkbox = ctk.CTkCheckBox(self,
                                       text=employee2.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp2Checkbox_command,
                                       variable=self.emp2Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp2Checkbox.place(x=5, y=65)

        emp3Checkbox = ctk.CTkCheckBox(self,
                                       text=employee3.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp3Checkbox_command,
                                       variable=self.emp3Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp3Checkbox.place(x=5, y=100)

        emp4Checkbox = ctk.CTkCheckBox(self,
                                       text=employee4.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp4Checkbox_command,
                                       variable=self.emp4Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp4Checkbox.place(x=5, y=135)

        emp5Checkbox = ctk.CTkCheckBox(self,
                                       text=employee5.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp5Checkbox_command,
                                       variable=self.emp5Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp5Checkbox.place(x=5, y=170)

        emp6Checkbox = ctk.CTkCheckBox(self,
                                       text=employee6.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp6Checkbox_command,
                                       variable=self.emp6Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp6Checkbox.place(x=5, y=205)

        emp7Checkbox = ctk.CTkCheckBox(self,
                                       text=employee7.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp7Checkbox_command,
                                       variable=self.emp7Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp7Checkbox.place(x=5, y=240)

        emp8Checkbox = ctk.CTkCheckBox(self,
                                       text=employee8.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp8Checkbox_command,
                                       variable=self.emp8Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp8Checkbox.place(x=5, y=275)

        emp9Checkbox = ctk.CTkCheckBox(self,
                                       text=employee9.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp9Checkbox_command,
                                       variable=self.emp9Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp9Checkbox.place(x=5, y=310)

        emp10Checkbox = ctk.CTkCheckBox(self,
                                        text=employee10.fullname,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.emp10Checkbox_command,
                                        variable=self.emp10Checkbox_var,
                                        onvalue=1,
                                        offvalue=0)
        emp10Checkbox.place(x=5, y=345)

        jobNameLabel = ctk.CTkLabel(self,
                                    text="Job Name",
                                    width=100,
                                    height=25,
                                    fg_color="transparent",
                                    font=ctk.CTkFont(size=18, weight="normal"))
        jobNameLabel.place(x=180, y=30)

        self.jobName = ctk.CTkEntry(self,
                                    width=150,
                                    height=30,
                                    font=ctk.CTkFont(size=18, weight="normal"))

        self.jobName.place(x=300, y=30)

        battUnder10Label = ctk.CTkLabel(self,
                                        text="Batt > 10ft",
                                        width=100,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))
        battUnder10Label.place(x=180, y=65)

        self.battUnder10 = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.battUnder10.place(x=300, y=65)
        self.battUnder10.insert(0, '0')

        battOver10Label = ctk.CTkLabel(self,
                                       text="Batt < 10ft",
                                       width=100,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        battOver10Label.place(x=180, y=100)

        self.battOver10 = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))
        self.battOver10.place(x=300, y=100)
        self.battOver10.insert(0, '0')

        newSoffitLabel = ctk.CTkLabel(self,
                                      text="Soffit-New",
                                      width=100,
                                      height=30,
                                      fg_color="transparent",
                                      font=ctk.CTkFont(size=18, weight="normal"))

        newSoffitLabel.place(x=180, y=135)

        self.newSoffit = ctk.CTkEntry(self,
                                      width=100,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"))

        self.newSoffit.place(x=300, y=135)
        self.newSoffit.insert(0, '0')

        caulkFoamLabel = ctk.CTkLabel(self,
                                      text="Caulk & Foam",
                                      width=100,
                                      height=30,
                                      fg_color="transparent",
                                      font=ctk.CTkFont(size=18, weight="normal"))

        caulkFoamLabel.place(x=180, y=170)

        self.caulkFoam = ctk.CTkEntry(self,
                                      width=100,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"))

        self.caulkFoam.place(x=300, y=170)
        self.caulkFoam.insert(0, '0')

        bonusLabel = ctk.CTkLabel(self,
                                  text="Bonus",
                                  width=100,
                                  height=30,
                                  fg_color="transparent",
                                  font=ctk.CTkFont(size=18, weight="normal"))

        bonusLabel.place(x=180, y=205)

        self.bonusAmount = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.bonusAmount.place(x=300, y=205)
        self.bonusAmount.insert(0, '0')

        celluloseLabel = ctk.CTkLabel(self,
                                      text="Cellulose",
                                      width=100,
                                      height=30,
                                      fg_color="transparent",
                                      font=ctk.CTkFont(size=18, weight="normal"))

        celluloseLabel.place(x=180, y=240)

        self.celluloseAmount = ctk.CTkEntry(self,
                                            width=100,
                                            height=30,
                                            font=ctk.CTkFont(size=18, weight="normal"))

        self.celluloseAmount.place(x=300, y=240)
        self.celluloseAmount.insert(0, '0')

        otherAmountLabel = ctk.CTkLabel(self,
                                        text="Other Amount",
                                        width=100,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        otherAmountLabel.place(x=180, y=275)

        self.otherAmount = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.otherAmount.place(x=300, y=275)
        self.otherAmount.insert(0, '0')

        bibsTitleLabel = ctk.CTkLabel(self,
                                      text="BIBS",
                                      width=200,
                                      height=30,
                                      fg_color="transparent",
                                      font=ctk.CTkFont(size=18, weight="normal"))

        bibsTitleLabel.place(x=450, y=30)

        bibsFullInstallLabel = ctk.CTkLabel(self,
                                            text="Full Install",
                                            width=100,
                                            height=30,
                                            fg_color="transparent",
                                            font=ctk.CTkFont(size=18, weight="normal"))

        bibsFullInstallLabel.place(x=450, y=65)

        self.bibsFullInstall = ctk.CTkEntry(self,
                                            width=100,
                                            height=30,
                                            font=ctk.CTkFont(size=18, weight="normal"))

        self.bibsFullInstall.place(x=565, y=65)
        self.bibsFullInstall.insert(0, '0')

        bibsHangInstallLabel = ctk.CTkLabel(self,
                                            text="Hang Netting",
                                            width=100,
                                            height=30,
                                            fg_color="transparent",
                                            font=ctk.CTkFont(size=18, weight="normal"))

        bibsHangInstallLabel.place(x=450, y=100)

        self.bibsHangInstall = ctk.CTkEntry(self,
                                            width=100,
                                            height=30,
                                            font=ctk.CTkFont(size=18, weight="normal"))

        self.bibsHangInstall.place(x=565, y=100)
        self.bibsHangInstall.insert(0, '0')

        bibsTackInstallLabel = ctk.CTkLabel(self,
                                            text="Tack Netting",
                                            width=100,
                                            height=30,
                                            fg_color="transparent",
                                            font=ctk.CTkFont(size=18, weight="normal"))

        bibsTackInstallLabel.place(x=450, y=135)

        self.bibsTackInstall = ctk.CTkEntry(self,
                                            width=100,
                                            height=30,
                                            font=ctk.CTkFont(size=18, weight="normal"))

        self.bibsTackInstall.place(x=565, y=135)
        self.bibsTackInstall.insert(0, '0')

        bibsBlowInstallLabel = ctk.CTkLabel(self,
                                            text="Walls Blown",
                                            width=100,
                                            height=30,
                                            fg_color="transparent",
                                            font=ctk.CTkFont(size=18, weight="normal"))

        bibsBlowInstallLabel.place(x=450, y=170)

        self.bibsBlowInstall = ctk.CTkEntry(self,
                                            width=100,
                                            height=30,
                                            font=ctk.CTkFont(size=18, weight="normal"))

        self.bibsBlowInstall.place(x=565, y=170)
        self.bibsBlowInstall.insert(0, '0')

        jobTotalPayLabel = ctk.CTkLabel(self,
                                        text="Job Total Pay",
                                        width=100,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        jobTotalPayLabel.place(x=180, y=350)

        self.jobTotalPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.jobTotalPay.place(x=300, y=350)
        self.jobTotalPay.configure(state='disabled')

        numWorkersLabel = ctk.CTkLabel(self,
                                       text="Total Workers",
                                       width=100,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        numWorkersLabel.place(x=180, y=385)

        self.numWorkers = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.numWorkers.place(x=300, y=385)
        self.numWorkers.configure(state='disabled')

        totalPayPerLabel = ctk.CTkLabel(self,
                                        text="Total Pay Per Person",
                                        width=150,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        totalPayPerLabel.place(x=450, y=350)

        self.totalPayPer = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.totalPayPer.place(x=625, y=350)
        self.totalPayPer.configure(state='disabled')

        calculateButton = ctk.CTkButton(self,
                                        text="Calculate",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=17),
                                        command=self.calculateButton_command)

        calculateButton.place(x=10, y=400)

        resetFormButton = ctk.CTkButton(self,
                                        text="Reset Form",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=17),
                                        command=self.resetFormButton_command)

        resetFormButton.place(x=10, y=435)

        saveFileButton = ctk.CTkButton(self,
                                       text="Save File",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=17),
                                       command=self.saveFileButton_command)

        saveFileButton.place(x=10, y=470)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=17),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=10, y=505)

    def calculateButton_command(self):
        self.numWorkers.configure(state='normal')
        self.jobTotalPay.configure(state='normal')
        self.totalPayPer.configure(state='normal')
        self.numWorkers.insert(0, str(self.count))

        nameFile = str(self.jobName.get())
        battPayGet1 = self.battUnder10.get()
        battPayGet2 = self.battOver10.get()
        battPayGet3 = self.newSoffit.get()
        battPayGet4 = self.caulkFoam.get()
        battPayGet5 = self.bonusAmount.get()
        battPayGet6 = self.celluloseAmount.get()
        battPayGet7 = self.otherAmount.get()
        battPayGet8 = self.bibsFullInstall.get()
        battPayGet9 = self.bibsHangInstall.get()
        battPayGet10 = self.bibsTackInstall.get()
        battPayGet11 = self.bibsBlowInstall.get()

        battPayCalc1 = float(battPayGet1) * batt1.rate
        battPayCalc2 = float(battPayGet2) * batt2.rate
        battPayCalc3 = float(battPayGet3) * soffit.rate
        battPayCalc4 = float(battPayGet4)
        battPayCalc5 = float(battPayGet5)
        battPayCalc6 = float(battPayGet6)
        battPayCalc7 = float(battPayGet7)
        battPayCalc8 = float(battPayGet8) * bibs_full.rate
        battPayCalc9 = float(battPayGet9) * bibs_hung.rate
        battPayCalc10 = float(battPayGet10) * bibs_tacked.rate
        battPayCalc11 = float(battPayGet11) * bibs_blown.rate

        battTotalPayCalc = battPayCalc1 + battPayCalc2 + battPayCalc3 + battPayCalc4 + battPayCalc5 + battPayCalc6 + \
                           battPayCalc7 + battPayCalc8 + battPayCalc9 + battPayCalc10 + battPayCalc11

        battPaySplit = float(battTotalPayCalc) / float(self.numWorkers.get())

        self.jobTotalPay.insert(0, '$' + str(round(battTotalPayCalc, 2)))
        self.totalPayPer.insert(0, '$' + str(round(battPaySplit, 2)))

        print("")
        print("Job Name: " + nameFile)
        print("")
        print("Total Batt Work <10ft = " + str(battPayGet1) + " sqft")
        print(str(battPayGet1) + " x " + str(batt1.rate) + "= $" + str(battPayCalc1))
        print("")
        print("Total Batt Work >10ft = " + str(battPayGet2) + " sqft")
        print(str(battPayGet2) + " x " + str(batt2.rate) + "= $" + str(battPayCalc2))
        print("")
        print("Total Number Of Soffits = " + str(battPayGet3))
        print(str(battPayGet3) + " x " + str(soffit.rate) + "= $" + str(battPayCalc3))
        print("")
        print("Total Amount Of Caulk & Foam = $" + str(battPayGet4))
        print("")
        print("Total Bonus Amount = $" + str(battPayGet5))
        print("")
        print("Total Cellulose Amount = $" + str(battPayGet6))
        print("")
        print("Total Other Amount = $" + str(battPayGet7))
        print("")
        print("Total Full BIBS Install = " + str(battPayGet8) + " sqft")
        print(str(battPayGet8) + " x " + str(bibs_full.rate) + "= $" + str(battPayCalc8))
        print("")
        print("Total BIBS Mesh Hung = " + str(battPayGet9) + " sqft")
        print(str(battPayGet9) + " x " + str(bibs_hung.rate) + "= $" + str(battPayCalc9))
        print("")
        print("Total BIBS Mesh Tacked = " + str(battPayGet10) + " sqft")
        print(str(battPayGet10) + " x " + str(bibs_tacked.rate) + "= $" + str(battPayCalc10))
        print("")
        print("Total BIBS Mesh Blown = " + str(battPayGet11) + " sqft")
        print(str(battPayGet11) + " x " + str(bibs_blown.rate) + "= $" + str(battPayCalc11))
        print("")
        print("Total Job Pay: $" + str(battTotalPayCalc))
        print("")
        print("Total Number Of Employee's: " + str(self.numWorkers.get()))
        print("")
        print("Total Pay Per Person: $" + str(battPaySplit))
        print("")

    def resetFormButton_command(self):
        self.count = 0
        self.jobName.delete(0, "end")
        self.numWorkers.delete(0, 'end')
        self.battOver10.delete(0, 'end')
        self.battOver10.insert(0, '0')
        self.battUnder10.delete(0, 'end')
        self.battUnder10.insert(0, '0')
        self.newSoffit.delete(0, 'end')
        self.newSoffit.insert(0, '0')
        self.caulkFoam.delete(0, 'end')
        self.caulkFoam.insert(0, '0')
        self.bonusAmount.delete(0, 'end')
        self.bonusAmount.insert(0, '0')
        self.otherAmount.delete(0, 'end')
        self.otherAmount.insert(0, '0')
        self.bibsBlowInstall.delete(0, 'end')
        self.bibsBlowInstall.insert(0, '0')
        self.bibsTackInstall.delete(0, 'end')
        self.bibsTackInstall.insert(0, '0')
        self.bibsHangInstall.delete(0, 'end')
        self.bibsHangInstall.insert(0, '0')
        self.bibsFullInstall.delete(0, 'end')
        self.bibsFullInstall.insert(0, '0')
        self.celluloseAmount.delete(0, 'end')
        self.celluloseAmount.insert(0, '0')
        self.jobTotalPay.delete(0, 'end')
        self.totalPayPer.delete(0, 'end')

        if self.emp1Checkbox_var.get() == 1:
            self.emp1Checkbox_var.set(0)

        if self.emp2Checkbox_var.get() == 1:
            self.emp2Checkbox_var.set(0)

        if self.emp3Checkbox_var.get() == 1:
            self.emp3Checkbox_var.set(0)

        if self.emp4Checkbox_var.get() == 1:
            self.emp4Checkbox_var.set(0)

        if self.emp5Checkbox_var.get() == 1:
            self.emp5Checkbox_var.set(0)

        if self.emp6Checkbox_var.get() == 1:
            self.emp6Checkbox_var.set(0)

        if self.emp7Checkbox_var.get() == 1:
            self.emp7Checkbox_var.set(0)

        if self.emp8Checkbox_var.get() == 1:
            self.emp8Checkbox_var.set(0)

        if self.emp9Checkbox_var.get() == 1:
            self.emp9Checkbox_var.set(0)

        if self.emp10Checkbox_var.get() == 1:
            self.emp10Checkbox_var.set(0)

        self.numWorkers.configure(state='disabled')
        self.jobTotalPay.configure(state='disabled')
        self.totalPayPer.configure(state='disabled')

    def saveFileButton_command(self):
        namefile = self.jobName.get()
        filename = str(namefile + '.xlsx')

        if not os.path.exists(
                "D:/Insulpro/temp/paysheets/batt pay sheet/" + filename):
            print('Saving File...')

            self.savePaySheet()
            self.saveWorkRecord()
            self.saveEmployeeRecord()
            self.savePaytreeBatt()
            self.savePayrollSpreadsheet()

            self.resetFormButton_command()
            self.closeWindowButton_command()

            messagebox.showinfo(title='Save', message='File Save Successful')
            print("File was saved successfully")

        else:
            messagebox.showerror(title='File Name Already Exist',
                                 message='Please change job name to something else!')

    def closeWindowButton_command(self):
        self.destroy()

    def savePaySheet(self):
        print('Saving Batt Pay Sheet')

        battPayGet1 = self.battUnder10.get()
        battPayGet2 = self.battOver10.get()
        battPayGet3 = self.newSoffit.get()
        battPayGet4 = self.caulkFoam.get()
        battPayGet5 = self.bonusAmount.get()
        battPayGet6 = self.celluloseAmount.get()
        battPayGet7 = self.otherAmount.get()
        battPayGet8 = self.bibsFullInstall.get()
        battPayGet9 = self.bibsHangInstall.get()
        battPayGet10 = self.bibsTackInstall.get()
        battPayGet11 = self.bibsBlowInstall.get()
        battPayCalc3 = float(battPayGet3) * soffit.rate

        splitPay1 = round(float(battPayGet1) / float(self.numWorkers.get()), 0)
        splitPay2 = round(float(battPayGet2) / float(self.numWorkers.get()), 0)
        splitPay3 = round(float(battPayCalc3) / float(self.numWorkers.get()), 2)
        splitPay4 = round(float(battPayGet4) / float(self.numWorkers.get()), 2)
        splitPay5 = round(float(battPayGet5) / float(self.numWorkers.get()), 2)
        splitPay6 = round(float(battPayGet6) / float(self.numWorkers.get()), 2)
        splitPay7 = round(float(battPayGet7) / float(self.numWorkers.get()), 2)
        splitPay8 = (float(battPayGet8) / float(self.numWorkers.get())) * bibs_full.rate
        splitPay9 = (float(battPayGet9) / float(self.numWorkers.get())) * bibs_hung.rate
        splitPay10 = (float(battPayGet10) / float(self.numWorkers.get())) * bibs_tacked.rate
        splitPay11 = (float(battPayGet11) / float(self.numWorkers.get())) * bibs_blown.rate
        splitPay12 = round(float(splitPay8) + float(splitPay9) + float(splitPay10) + float(splitPay11), 2)

        source = "D:/Insulpro/system/spreadsheet/Batt Pay Sheet.xlsx"

        destination = ("D:/Insulpro/temp/paysheets/batt pay sheet/Batt Pay "
                       "Sheet.xlsx")

        shutil.copy(source, destination)

        nameFile = self.jobName.get()

        filename = str(nameFile + '.xlsx')

        os.rename("D:/Insulpro/temp/paysheets/batt pay sheet/Batt Pay Sheet.xlsx",
                  "D:/Insulpro/temp/paysheets/batt pay sheet/" + filename)

        workbook = load_workbook(
            "D:/Insulpro/temp/paysheets/batt pay sheet/" + filename)

        sheet = workbook.active

        jobNameSheet = str(nameFile)

        sheet["A2"] = jobNameSheet
        sheet['A5'] = employee1.fullname
        sheet['A6'] = employee2.fullname
        sheet['A7'] = employee3.fullname
        sheet['A8'] = employee4.fullname
        sheet['A9'] = employee5.fullname
        sheet['A10'] = employee6.fullname
        sheet['A11'] = employee7.fullname
        sheet['A12'] = employee8.fullname
        sheet['A13'] = employee9.fullname
        sheet['A14'] = employee10.fullname

        if self.emp1Checkbox_var.get() == 1:
            sheet['B5'] = splitPay1
            sheet['C5'] = splitPay2
            sheet['D5'] = splitPay3
            sheet['E5'] = splitPay4
            sheet['F5'] = splitPay5
            sheet['G5'] = splitPay6
            sheet['H5'] = splitPay7
            sheet['I5'] = splitPay12

        if self.emp2Checkbox_var.get() == 1:
            sheet['B6'] = splitPay1
            sheet['C6'] = splitPay2
            sheet['D6'] = splitPay3
            sheet['E6'] = splitPay4
            sheet['F6'] = splitPay5
            sheet['G6'] = splitPay6
            sheet['H6'] = splitPay7
            sheet['I6'] = splitPay12

        if self.emp3Checkbox_var.get() == 1:
            sheet['B7'] = splitPay1
            sheet['C7'] = splitPay2
            sheet['D7'] = splitPay3
            sheet['E7'] = splitPay4
            sheet['F7'] = splitPay5
            sheet['G7'] = splitPay6
            sheet['H7'] = splitPay7
            sheet['I7'] = splitPay12

        if self.emp4Checkbox_var.get() == 1:
            sheet['B8'] = splitPay1
            sheet['C8'] = splitPay2
            sheet['D8'] = splitPay3
            sheet['E8'] = splitPay4
            sheet['F8'] = splitPay5
            sheet['G8'] = splitPay6
            sheet['H8'] = splitPay7
            sheet['I8'] = splitPay12

        if self.emp5Checkbox_var.get() == 1:
            sheet['B9'] = splitPay1
            sheet['C9'] = splitPay2
            sheet['D9'] = splitPay3
            sheet['E9'] = splitPay4
            sheet['F9'] = splitPay5
            sheet['G9'] = splitPay6
            sheet['H9'] = splitPay7
            sheet['I9'] = splitPay12

        if self.emp6Checkbox_var.get() == 1:
            sheet['B10'] = splitPay1
            sheet['C10'] = splitPay2
            sheet['D10'] = splitPay3
            sheet['E10'] = splitPay4
            sheet['F10'] = splitPay5
            sheet['G10'] = splitPay6
            sheet['H10'] = splitPay7
            sheet['I10'] = splitPay12

        if self.emp7Checkbox_var.get() == 1:
            sheet['B11'] = splitPay1
            sheet['C11'] = splitPay2
            sheet['D11'] = splitPay3
            sheet['E11'] = splitPay4
            sheet['F11'] = splitPay5
            sheet['G11'] = splitPay6
            sheet['H11'] = splitPay7
            sheet['I11'] = splitPay12

        if self.emp8Checkbox_var.get() == 1:
            sheet['B12'] = splitPay1
            sheet['C12'] = splitPay2
            sheet['D12'] = splitPay3
            sheet['E12'] = splitPay4
            sheet['F12'] = splitPay5
            sheet['G12'] = splitPay6
            sheet['H12'] = splitPay7
            sheet['I12'] = splitPay12

        if self.emp9Checkbox_var.get() == 1:
            sheet['B13'] = splitPay1
            sheet['C13'] = splitPay2
            sheet['D13'] = splitPay3
            sheet['E13'] = splitPay4
            sheet['F13'] = splitPay5
            sheet['G13'] = splitPay6
            sheet['H13'] = splitPay7
            sheet['I13'] = splitPay12

        if self.emp10Checkbox_var.get() == 1:
            sheet['B14'] = splitPay1
            sheet['C14'] = splitPay2
            sheet['D14'] = splitPay3
            sheet['E14'] = splitPay4
            sheet['F14'] = splitPay5
            sheet['G14'] = splitPay6
            sheet['H14'] = splitPay7
            sheet['I14'] = splitPay12

        workbook.save("D:/Insulpro/temp/paysheets/batt pay sheet/" + filename)

        workbook.close()

    def saveWorkRecord(self):
        print('Saving Batt Work Record')
        nameFile = self.jobName.get()
        battPayGet1 = self.battUnder10.get()
        battPayGet2 = self.battOver10.get()
        battPayGet3 = self.newSoffit.get()
        battPayGet4 = self.caulkFoam.get()
        battPayGet5 = self.bonusAmount.get()
        battPayGet6 = self.celluloseAmount.get()
        battPayGet7 = self.otherAmount.get()
        battPayGet8 = self.bibsFullInstall.get()
        battPayGet9 = self.bibsHangInstall.get()
        battPayGet10 = self.bibsTackInstall.get()
        battPayGet11 = self.bibsBlowInstall.get()

        splitPay8 = float(battPayGet8) * bibs_full.rate
        splitPay9 = float(battPayGet9) * bibs_hung.rate
        splitPay10 = float(battPayGet10) * bibs_tacked.rate
        splitPay11 = float(battPayGet11) * bibs_blown.rate
        splitPay12 = float(splitPay8) + float(splitPay9) + float(splitPay10) + float(splitPay11)

        workbook = load_workbook("D:/Insulpro/payroll records/Work Record.xlsx")

        sheet = workbook['Batt Work Record']

        sheet.append([nameFile, battPayGet1, battPayGet2, battPayGet3, battPayGet4, battPayGet5, battPayGet6,
                      battPayGet7, splitPay12])

        workbook.save("D:/Insulpro/payroll records/Work Record.xlsx")

        workbook.close()

    def savePaytreeBatt(self):
        empFile1 = str(employee1.fullname) + "/batt record.xlsx"
        empFile2 = str(employee2.fullname) + "/batt record.xlsx"
        empFile3 = str(employee3.fullname) + "/batt record.xlsx"
        empFile4 = str(employee4.fullname) + "/batt record.xlsx"
        empFile5 = str(employee5.fullname) + "/batt record.xlsx"
        empFile6 = str(employee6.fullname) + "/batt record.xlsx"
        empFile7 = str(employee7.fullname) + "/batt record.xlsx"
        empFile8 = str(employee8.fullname) + "/batt record.xlsx"
        empFile9 = str(employee9.fullname) + "/batt record.xlsx"
        empFile10 = str(employee10.fullname) + "/batt record.xlsx"

        namefile = self.jobName.get()
        battPayGet1 = self.battUnder10.get()
        battPayGet2 = self.battOver10.get()
        battPayGet3 = self.newSoffit.get()
        battPayGet4 = self.caulkFoam.get()
        battPayGet5 = self.bonusAmount.get()
        battPayGet6 = self.celluloseAmount.get()
        battPayGet7 = self.otherAmount.get()
        battPayGet8 = self.bibsFullInstall.get()
        battPayGet9 = self.bibsHangInstall.get()
        battPayGet10 = self.bibsTackInstall.get()
        battPayGet11 = self.bibsBlowInstall.get()

        battPayCalc1 = float(battPayGet1) * batt1.rate
        battPayCalc2 = float(battPayGet2) * batt2.rate
        battPayCalc3 = float(battPayGet3) * soffit.rate
        battPayCalc4 = float(battPayGet4)
        battPayCalc5 = float(battPayGet5)
        battPayCalc6 = float(battPayGet6)
        battPayCalc7 = float(battPayGet7)
        battPayCalc8 = float(battPayGet8) * bibs_full.rate
        battPayCalc9 = float(battPayGet9) * bibs_hung.rate
        battPayCalc10 = float(battPayGet10) * bibs_tacked.rate
        battPayCalc11 = float(battPayGet11) * bibs_blown.rate
        battPayCalc12 = float(battPayCalc8) + float(battPayCalc9) + float(battPayCalc10) + float(battPayCalc11)

        splitPay1 = float(battPayCalc1) / float(self.numWorkers.get())
        splitPay2 = float(battPayCalc2) / float(self.numWorkers.get())
        splitPay3 = float(battPayCalc3) / float(self.numWorkers.get())
        splitPay4 = float(battPayCalc4) / float(self.numWorkers.get())
        splitPay5 = float(battPayCalc5) / float(self.numWorkers.get())
        splitPay6 = float(battPayCalc6) / float(self.numWorkers.get())
        splitPay7 = float(battPayCalc7) / float(self.numWorkers.get())
        splitPay8 = float(battPayCalc12) / float(self.numWorkers.get())

        battPaySplit = splitPay1 + splitPay2 + splitPay3 + splitPay4 + splitPay5 + splitPay6 + splitPay7 + splitPay8

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile1)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, battPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile1)

            workbook.close()

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile2)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, battPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile2)

            workbook.close()

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile3)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, battPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile3)

            workbook.close()

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile4)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, battPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile4)

            workbook.close()

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile5)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, battPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile5)

            workbook.close()

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile6)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, battPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile6)

            workbook.close()

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile7)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, battPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile7)

            workbook.close()

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile8)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, battPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile8)

            workbook.close()

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile9)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, battPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile9)

            workbook.close()

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile10)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, battPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile10)

            workbook.close()

    def saveEmployeeRecord(self):
        print('Saving Employee Work Records')
        empFile1 = str(employee1.fullname + '.xlsx')
        empFile2 = str(employee2.fullname + '.xlsx')
        empFile3 = str(employee3.fullname + '.xlsx')
        empFile4 = str(employee4.fullname + '.xlsx')
        empFile5 = str(employee5.fullname + '.xlsx')
        empFile6 = str(employee6.fullname + '.xlsx')
        empFile7 = str(employee7.fullname + '.xlsx')
        empFile8 = str(employee8.fullname + '.xlsx')
        empFile9 = str(employee9.fullname + '.xlsx')
        empFile10 = str(employee10.fullname + '.xlsx')

        namefile = self.jobName.get()
        battPayGet1 = self.battUnder10.get()
        battPayGet2 = self.battOver10.get()
        battPayGet3 = self.newSoffit.get()
        battPayGet4 = self.caulkFoam.get()
        battPayGet5 = self.bonusAmount.get()
        battPayGet6 = self.celluloseAmount.get()
        battPayGet7 = self.otherAmount.get()
        battPayGet8 = self.bibsFullInstall.get()
        battPayGet9 = self.bibsHangInstall.get()
        battPayGet10 = self.bibsTackInstall.get()
        battPayGet11 = self.bibsBlowInstall.get()

        battPayCalc1 = float(battPayGet1) * batt1.rate
        battPayCalc2 = float(battPayGet2) * batt2.rate
        battPayCalc3 = float(battPayGet3) * soffit.rate
        battPayCalc4 = float(battPayGet4)
        battPayCalc5 = float(battPayGet5)
        battPayCalc6 = float(battPayGet6)
        battPayCalc7 = float(battPayGet7)
        battPayCalc8 = float(battPayGet8) * bibs_full.rate
        battPayCalc9 = float(battPayGet9) * bibs_hung.rate
        battPayCalc10 = float(battPayGet10) * bibs_tacked.rate
        battPayCalc11 = float(battPayGet11) * bibs_blown.rate
        battPayCalc12 = float(battPayCalc8) + float(battPayCalc9) + float(battPayCalc10) + float(battPayCalc11)

        splitPay1 = float(battPayCalc1) / float(self.numWorkers.get())
        splitPay2 = float(battPayCalc2) / float(self.numWorkers.get())
        splitPay3 = float(battPayCalc3) / float(self.numWorkers.get())
        splitPay4 = float(battPayCalc4) / float(self.numWorkers.get())
        splitPay5 = float(battPayCalc5) / float(self.numWorkers.get())
        splitPay6 = float(battPayCalc6) / float(self.numWorkers.get())
        splitPay7 = float(battPayCalc7) / float(self.numWorkers.get())
        splitPay8 = float(battPayCalc12) / float(self.numWorkers.get())

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile1)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile1)

            workbook.close()

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile2)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile2)

            workbook.close()

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile3)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile3)

            workbook.close()

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile4)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile4)

            workbook.close()

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile5)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile5)

            workbook.close()

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile6)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile6)

            workbook.close()

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile7)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile7)

            workbook.close()

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile8)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile8)

            workbook.close()

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile9)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile9)

            workbook.close()

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile10)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile10)

            workbook.close()

    def savePayrollSpreadsheet(self):
        print('Saving Payroll Spreadsheet')
        namefile = self.jobName.get()
        battPayGet1 = self.battUnder10.get()
        battPayGet2 = self.battOver10.get()
        battPayGet3 = self.newSoffit.get()
        battPayGet4 = self.caulkFoam.get()
        battPayGet5 = self.bonusAmount.get()
        battPayGet6 = self.celluloseAmount.get()
        battPayGet7 = self.otherAmount.get()
        battPayGet8 = self.bibsFullInstall.get()
        battPayGet9 = self.bibsHangInstall.get()
        battPayGet10 = self.bibsTackInstall.get()
        battPayGet11 = self.bibsBlowInstall.get()
        battPayCalc3 = float(battPayGet3) * soffit.rate

        splitPay1 = round(float(battPayGet1) / float(self.numWorkers.get()), 0)
        splitPay2 = round(float(battPayGet2) / float(self.numWorkers.get()), 0)
        splitPay3 = round(float(battPayCalc3) / float(self.numWorkers.get()), 2)
        splitPay4 = round(float(battPayGet4) / float(self.numWorkers.get()), 2)
        splitPay5 = round(float(battPayGet5) / float(self.numWorkers.get()), 2)
        splitPay6 = round(float(battPayGet6) / float(self.numWorkers.get()), 2)
        splitPay7 = round(float(battPayGet7) / float(self.numWorkers.get()), 2)
        splitPay8 = (float(battPayGet8) / float(self.numWorkers.get())) * bibs_full.rate
        splitPay9 = (float(battPayGet9) / float(self.numWorkers.get())) * bibs_hung.rate
        splitPay10 = (float(battPayGet10) / float(self.numWorkers.get())) * bibs_tacked.rate
        splitPay11 = (float(battPayGet11) / float(self.numWorkers.get())) * bibs_blown.rate
        splitPay12 = round(float(splitPay8) + float(splitPay9) + float(splitPay10) + float(splitPay11), 2)

        workbook = load_workbook("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        if self.emp1Checkbox_var.get() == 1:
            sheet = workbook['Emp1Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay12, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp2Checkbox_var.get() == 1:
            sheet = workbook['Emp2Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay12, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp3Checkbox_var.get() == 1:
            sheet = workbook['Emp3Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay12, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp4Checkbox_var.get() == 1:
            sheet = workbook['Emp4Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay12, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp5Checkbox_var.get() == 1:
            sheet = workbook['Emp5Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay12, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp6Checkbox_var.get() == 1:
            sheet = workbook['Emp6Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay12, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp7Checkbox_var.get() == 1:
            sheet = workbook['Emp7Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay12, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp8Checkbox_var.get() == 1:
            sheet = workbook['Emp8Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay12, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp9Checkbox_var.get() == 1:
            sheet = workbook['Emp9Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay12, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp10Checkbox_var.get() == 1:
            sheet = workbook['Emp10Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay12, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        workbook.save("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        workbook.close()

    def emp1Checkbox_command(self):
        print(employee1.fullname + " selected")
        self.emp1Checkbox_var.set(1)
        self.count += 1

    def emp2Checkbox_command(self):
        print(employee2.fullname + " selected")
        self.emp2Checkbox_var.set(1)
        self.count += 1

    def emp3Checkbox_command(self):
        print(employee3.fullname + " selected")
        self.emp3Checkbox_var.set(1)
        self.count += 1

    def emp4Checkbox_command(self):
        print(employee4.fullname + " selected")
        self.emp4Checkbox_var.set(1)
        self.count += 1

    def emp5Checkbox_command(self):
        print(employee5.fullname + " selected")
        self.emp5Checkbox_var.set(1)
        self.count += 1

    def emp6Checkbox_command(self):
        print(employee6.fullname + " selected")
        self.emp6Checkbox_var.set(1)
        self.count += 1

    def emp7Checkbox_command(self):
        print(employee7.fullname + " selected")
        self.emp7Checkbox_var.set(1)
        self.count += 1

    def emp8Checkbox_command(self):
        print(employee8.fullname + " selected")
        self.emp8Checkbox_var.set(1)
        self.count += 1

    def emp9Checkbox_command(self):
        print(employee9.fullname + " selected")
        self.emp9Checkbox_var.set(1)
        self.count += 1

    def emp10Checkbox_command(self):
        print(employee10.fullname + " selected")
        self.emp10Checkbox_var.set(1)
        self.count += 1


class AtticPaySheet(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("800x600")
        self.title("Attic Pay Sheet")
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)

        self.count = 0

        self.emp1Checkbox_var = tk.IntVar(value=0)
        self.emp2Checkbox_var = tk.IntVar(value=0)
        self.emp3Checkbox_var = tk.IntVar(value=0)
        self.emp4Checkbox_var = tk.IntVar(value=0)
        self.emp5Checkbox_var = tk.IntVar(value=0)
        self.emp6Checkbox_var = tk.IntVar(value=0)
        self.emp7Checkbox_var = tk.IntVar(value=0)
        self.emp8Checkbox_var = tk.IntVar(value=0)
        self.emp9Checkbox_var = tk.IntVar(value=0)
        self.emp10Checkbox_var = tk.IntVar(value=0)

        emp1Checkbox = ctk.CTkCheckBox(self,
                                       text=employee1.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp1Checkbox_command,
                                       variable=self.emp1Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp1Checkbox.place(x=5, y=30)

        emp2Checkbox = ctk.CTkCheckBox(self,
                                       text=employee2.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp2Checkbox_command,
                                       variable=self.emp2Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp2Checkbox.place(x=5, y=65)

        emp3Checkbox = ctk.CTkCheckBox(self,
                                       text=employee3.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp3Checkbox_command,
                                       variable=self.emp3Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp3Checkbox.place(x=5, y=100)

        emp4Checkbox = ctk.CTkCheckBox(self,
                                       text=employee4.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp4Checkbox_command,
                                       variable=self.emp4Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp4Checkbox.place(x=5, y=135)

        emp5Checkbox = ctk.CTkCheckBox(self,
                                       text=employee5.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp5Checkbox_command,
                                       variable=self.emp5Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp5Checkbox.place(x=5, y=170)

        emp6Checkbox = ctk.CTkCheckBox(self,
                                       text=employee6.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp6Checkbox_command,
                                       variable=self.emp6Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp6Checkbox.place(x=5, y=205)

        emp7Checkbox = ctk.CTkCheckBox(self,
                                       text=employee7.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp7Checkbox_command,
                                       variable=self.emp7Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp7Checkbox.place(x=5, y=240)

        emp8Checkbox = ctk.CTkCheckBox(self,
                                       text=employee8.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp8Checkbox_command,
                                       variable=self.emp8Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp8Checkbox.place(x=5, y=275)

        emp9Checkbox = ctk.CTkCheckBox(self,
                                       text=employee9.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp9Checkbox_command,
                                       variable=self.emp9Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp9Checkbox.place(x=5, y=310)

        emp10Checkbox = ctk.CTkCheckBox(self,
                                        text=employee10.fullname,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.emp10Checkbox_command,
                                        variable=self.emp10Checkbox_var,
                                        onvalue=1,
                                        offvalue=0)
        emp10Checkbox.place(x=5, y=345)

        jobNameLabel = ctk.CTkLabel(self,
                                    text="Job Name",
                                    width=100,
                                    height=30,
                                    fg_color="transparent",
                                    font=ctk.CTkFont(size=18, weight="normal"))

        jobNameLabel.place(x=180, y=30)

        self.jobName = ctk.CTkEntry(self,
                                    width=150,
                                    height=30,
                                    font=ctk.CTkFont(size=18, weight="normal"))

        self.jobName.place(x=300, y=30)

        r19InstalledLabel = ctk.CTkLabel(self,
                                         text="R19 Blow",
                                         width=100,
                                         height=30,
                                         fg_color="transparent",
                                         font=ctk.CTkFont(size=18, weight="normal"))

        r19InstalledLabel.place(x=180, y=65)

        self.r19Installed = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.r19Installed.place(x=300, y=65, )
        self.r19Installed.insert(0, '0')

        r30InstalledLabel = ctk.CTkLabel(self,
                                         text="R30 Blow",
                                         width=100,
                                         height=30,
                                         fg_color="transparent",
                                         font=ctk.CTkFont(size=18, weight="normal"))

        r30InstalledLabel.place(x=180, y=100)

        self.r30Installed = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.r30Installed.place(x=300, y=100)
        self.r30Installed.insert(0, '0')

        r38InstalledLabel = ctk.CTkLabel(self,
                                         text="R38/40 Blow",
                                         width=100,
                                         height=30,
                                         fg_color="transparent",
                                         font=ctk.CTkFont(size=18, weight="normal"))

        r38InstalledLabel.place(x=180, y=135)

        self.r38Installed = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.r38Installed.place(x=300, y=135)
        self.r38Installed.insert(0, '0')

        r49InstalledLabel = ctk.CTkLabel(self,
                                         text="R49/50 Blow",
                                         width=100,
                                         height=30,
                                         fg_color="transparent",
                                         font=ctk.CTkFont(size=18, weight="normal"))

        r49InstalledLabel.place(x=180, y=170)

        self.r49Installed = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.r49Installed.place(x=300, y=170)
        self.r49Installed.insert(0, '0')

        celluloseInstalledLabel = ctk.CTkLabel(self,
                                               text="Cellulose Blow",
                                               width=100,
                                               height=30,
                                               fg_color="transparent",
                                               font=ctk.CTkFont(size=18, weight="normal"))

        celluloseInstalledLabel.place(x=180, y=205)

        self.celluloseInstalled = ctk.CTkEntry(self,
                                               width=100,
                                               height=30,
                                               font=ctk.CTkFont(size=18, weight="normal"))

        self.celluloseInstalled.place(x=300, y=205)
        self.celluloseInstalled.insert(0, '0')

        soffitInstalledLabel = ctk.CTkLabel(self,
                                            text="Soffits-Existing",
                                            width=100,
                                            height=30,
                                            fg_color="transparent",
                                            font=ctk.CTkFont(size=18, weight="normal"))

        soffitInstalledLabel.place(x=180, y=240)

        self.soffitInstalled = ctk.CTkEntry(self,
                                            width=100,
                                            height=30,
                                            font=ctk.CTkFont(size=18, weight="normal"))

        self.soffitInstalled.place(x=300, y=240)
        self.soffitInstalled.insert(0, '0')

        airSealLabel = ctk.CTkLabel(self,
                                    text="Air Seal",
                                    width=100,
                                    height=30,
                                    fg_color="transparent",
                                    font=ctk.CTkFont(size=18, weight="normal"))

        airSealLabel.place(x=180, y=275)

        self.airSeal = ctk.CTkEntry(self,
                                    width=100,
                                    height=30,
                                    font=ctk.CTkFont(size=18, weight="normal"))

        self.airSeal.place(x=300, y=275)
        self.airSeal.insert(0, '0')

        bonusAmountLabel = ctk.CTkLabel(self,
                                        text="Bonus",
                                        width=100,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        bonusAmountLabel.place(x=180, y=310)

        self.bonusAmount = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.bonusAmount.place(x=300, y=310)
        self.bonusAmount.insert(0, '0')

        otherAmountLabel = ctk.CTkLabel(self,
                                        text="Other Amount",
                                        width=100,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        otherAmountLabel.place(x=180, y=345)

        self.otherAmount = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.otherAmount.place(x=300, y=345)
        self.otherAmount.insert(0, '0')

        jobTotalPayLabel = ctk.CTkLabel(self,
                                        text="Job Total Pay",
                                        width=100,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        jobTotalPayLabel.place(x=180, y=450)

        self.jobTotalPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.jobTotalPay.place(x=300, y=450)
        self.jobTotalPay.configure(state='disabled')

        numWorkersLabel = ctk.CTkLabel(self,
                                       text="Total Workers",
                                       width=100,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        numWorkersLabel.place(x=180, y=485)

        self.numWorkers = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.numWorkers.place(x=300, y=485)
        self.numWorkers.configure(state='disabled')

        totalPayPerLabel = ctk.CTkLabel(self,
                                        text="Total Pay Per Person",
                                        width=150,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        totalPayPerLabel.place(x=430, y=450)

        self.totalPayPer = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.totalPayPer.place(x=605, y=450)
        self.totalPayPer.configure(state='disabled')

        calculateButton = ctk.CTkButton(self,
                                        text="Calculate",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.calculateButton_command)

        calculateButton.place(x=5, y=430)

        resetFormButton = ctk.CTkButton(self,
                                        text="Reset Form",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.resetFormButton_command)

        resetFormButton.place(x=5, y=465)

        saveFileButton = ctk.CTkButton(self,
                                       text="Save File",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.saveFileButton_command)

        saveFileButton.place(x=5, y=500)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=535)

    def calculateButton_command(self):
        print("")
        print('Calculating Attic Pay Sheet...')
        self.numWorkers.configure(state='normal')
        self.jobTotalPay.configure(state='normal')
        self.totalPayPer.configure(state='normal')
        self.numWorkers.insert(0, str(self.count))
        nameFile = str(self.jobName.get())
        atticPayGet1 = self.r19Installed.get()
        atticPayGet2 = self.r30Installed.get()
        atticPayGet3 = self.r38Installed.get()
        atticPayGet4 = self.r49Installed.get()
        atticPayGet5 = self.celluloseInstalled.get()
        atticPayGet6 = self.soffitInstalled.get()
        atticPayGet7 = self.bonusAmount.get()
        atticPayGet8 = self.otherAmount.get()
        atticPayGet9 = self.airSeal.get()

        atticPayCalc1 = float(atticPayGet1) * r19_blown.rate
        atticPayCalc2 = float(atticPayGet2) * r30_blown.rate
        atticPayCalc3 = float(atticPayGet3) * r38_blown.rate
        atticPayCalc4 = float(atticPayGet4) * r49_blown.rate
        atticPayCalc5 = float(atticPayGet5) * cell_blown.rate
        atticPayCalc6 = float(atticPayGet6) * existing_soffit.rate
        atticPayCalc7 = float(atticPayGet7)
        atticPayCalc8 = float(atticPayGet8)
        atticPayCalc9 = float(atticPayGet9)

        atticTotalPayCalc = round(atticPayCalc1 + atticPayCalc2 + atticPayCalc3 + atticPayCalc4 +
                                  atticPayCalc5 + atticPayCalc6 + atticPayCalc7 + atticPayCalc8 + atticPayCalc9, 2)

        atticPaySplit = round(float(atticTotalPayCalc) / float(self.numWorkers.get()), 2)

        self.jobTotalPay.insert(0, '$' + str(round(atticTotalPayCalc, 2)))
        self.totalPayPer.insert(0, '$' + str(round(atticPaySplit, 2)))

        print("")
        print("Job Name: " + nameFile)
        print("")
        print("Total R19 Installed: " + str(atticPayGet1) + " sqft")
        print(str(atticPayGet1) + " x " + str(r19_blown.rate) + " = $" + str(atticPayCalc1))
        print("")
        print("Total R30 Installed: " + str(atticPayGet2) + " sqft")
        print(str(atticPayGet2) + " x " + str(r30_blown.rate) + " = $" + str(atticPayCalc2))
        print("")
        print("Total R38-R44 Installed: " + str(atticPayGet3) + " sqft")
        print(str(atticPayGet3) + " x " + str(r38_blown.rate) + " = $" + str(atticPayCalc3))
        print("")
        print("Total R49+ Installed: " + str(atticPayGet4) + " sqft")
        print(str(atticPayGet4) + " x " + str(r49_blown.rate) + " = $" + str(atticPayCalc4))
        print("")
        print("Total Cellulose Installed: " + str(atticPayGet5) + " sqft")
        print(str(atticPayGet5) + " x " + str(cell_blown.rate) + " = $" + str(atticPayCalc5))
        print("")
        print("Total Soffits Installed: " + str(atticPayGet6))
        print(str(atticPayGet6) + " x " + str(existing_soffit.rate) + " = $" + str(atticPayCalc6))
        print("")
        print("Total Bonus Amount = $" + str(atticPayGet7))
        print("")
        print("Total Other Amount = $" + str(atticPayGet8))
        print("")
        print("Total Air Seal Amount = $" + str(atticPayGet9))
        print("")
        print("Total Job Pay: $" + str(atticTotalPayCalc))
        print("")
        print("Total Number Of Employee's: " + str(self.numWorkers.get()))
        print("")
        print("Total Pay Per Person: $" + str(atticPaySplit))
        print("")

    def resetFormButton_command(self):
        print('Resetting Form')
        self.count = 0
        self.jobName.delete(0, "end")
        self.numWorkers.delete(0, 'end')
        self.r19Installed.delete(0, 'end')
        self.r19Installed.insert(0, "0")
        self.r30Installed.delete(0, 'end')
        self.r30Installed.insert(0, "0")
        self.r38Installed.delete(0, 'end')
        self.r38Installed.insert(0, "0")
        self.r49Installed.delete(0, 'end')
        self.r49Installed.insert(0, "0")
        self.celluloseInstalled.delete(0, 'end')
        self.celluloseInstalled.insert(0, "0")
        self.bonusAmount.delete(0, 'end')
        self.bonusAmount.insert(0, "0")
        self.otherAmount.delete(0, 'end')
        self.otherAmount.insert(0, "0")
        self.soffitInstalled.delete(0, 'end')
        self.soffitInstalled.insert(0, "0")
        self.airSeal.delete(0, 'end')
        self.airSeal.insert(0, "0")
        self.jobTotalPay.delete(0, 'end')
        self.totalPayPer.delete(0, 'end')

        if self.emp1Checkbox_var.get() == 1:
            self.emp1Checkbox_var.set(0)

        if self.emp2Checkbox_var.get() == 1:
            self.emp2Checkbox_var.set(0)

        if self.emp3Checkbox_var.get() == 1:
            self.emp3Checkbox_var.set(0)

        if self.emp4Checkbox_var.get() == 1:
            self.emp4Checkbox_var.set(0)

        if self.emp5Checkbox_var.get() == 1:
            self.emp5Checkbox_var.set(0)

        if self.emp6Checkbox_var.get() == 1:
            self.emp6Checkbox_var.set(0)

        if self.emp7Checkbox_var.get() == 1:
            self.emp7Checkbox_var.set(0)

        if self.emp8Checkbox_var.get() == 1:
            self.emp8Checkbox_var.set(0)

        if self.emp9Checkbox_var.get() == 1:
            self.emp9Checkbox_var.set(0)

        if self.emp10Checkbox_var.get() == 1:
            self.emp10Checkbox_var.set(0)

        self.numWorkers.configure(state='disabled')
        self.jobTotalPay.configure(state='disabled')
        self.totalPayPer.configure(state='disabled')

    def saveFileButton_command(self):
        namefile = self.jobName.get()
        filename = str(namefile + ".xlsx")

        if not os.path.exists(
                "D:/Insulpro/temp/paysheets/attic pay sheet/" + filename):
            print('Saving File...')

            self.savePaySheet()
            self.saveWorkRecord()
            self.saveEmployeeRecord()
            self.savePaytreeAttic()
            self.savePayrollSpreadsheet()
            self.resetFormButton_command()

            self.closeWindowButton_command()

            messagebox.showinfo(title='Save', message='File Save Successful')

        else:
            messagebox.showerror(title='File Name Already Exist',
                                 message='Please change job name to something else!')

    def closeWindowButton_command(self):
        self.destroy()

    def savePaySheet(self):
        print('Saving Pay Sheet')
        atticPayGet1 = self.r19Installed.get()
        atticPayGet2 = self.r30Installed.get()
        atticPayGet3 = self.r38Installed.get()
        atticPayGet4 = self.r49Installed.get()
        atticPayGet5 = self.celluloseInstalled.get()
        atticPayGet6 = self.soffitInstalled.get()
        atticPayGet7 = self.bonusAmount.get()
        atticPayGet8 = self.otherAmount.get()
        atticPayGet9 = self.airSeal.get()
        atticPayCalc6 = float(atticPayGet6) * existing_soffit.rate

        splitPay1 = round(float(atticPayGet1) / float(self.numWorkers.get()), 0)
        splitPay2 = round(float(atticPayGet2) / float(self.numWorkers.get()), 0)
        splitPay3 = round(float(atticPayGet3) / float(self.numWorkers.get()), 0)
        splitPay4 = round(float(atticPayGet4) / float(self.numWorkers.get()), 0)
        splitPay5 = round(float(atticPayGet5) / float(self.numWorkers.get()), 0)
        splitPay6 = round(float(atticPayCalc6) / float(self.numWorkers.get()), 2)
        splitPay7 = round(float(atticPayGet7) / float(self.numWorkers.get()), 2)
        splitPay8 = round(float(atticPayGet8) / float(self.numWorkers.get()), 2)
        splitPay9 = round(float(atticPayGet9) / float(self.numWorkers.get()), 2)

        source = "D:/Insulpro/system/spreadsheet/Attic Pay Sheet.xlsx"

        destination = ("D:/Insulpro/temp/paysheets/attic pay sheet/Attic Pay "
                       "Sheet.xlsx")

        shutil.copy(source, destination)

        nameFile = self.jobName.get()

        filename = str(nameFile + '.xlsx')

        os.rename("D:/Insulpro/temp/paysheets/attic pay sheet/Attic Pay Sheet.xlsx",
                  "D:/Insulpro/temp/paysheets/attic pay sheet/" + filename)

        workbook = load_workbook(
            "D:/Insulpro/temp/paysheets/attic pay sheet/" + filename)

        sheet = workbook.active

        jobNameSheet = str(nameFile)

        sheet["A2"] = jobNameSheet
        sheet['A5'] = employee1.fullname
        sheet['A6'] = employee2.fullname
        sheet['A7'] = employee3.fullname
        sheet['A8'] = employee4.fullname
        sheet['A9'] = employee5.fullname
        sheet['A10'] = employee6.fullname
        sheet['A11'] = employee7.fullname
        sheet['A12'] = employee8.fullname
        sheet['A13'] = employee9.fullname
        sheet['A14'] = employee10.fullname

        if self.emp1Checkbox_var.get() == 1:
            sheet['B5'] = splitPay1
            sheet['C5'] = splitPay2
            sheet['D5'] = splitPay3
            sheet['E5'] = splitPay4
            sheet['F5'] = splitPay5
            sheet['G5'] = splitPay6
            sheet['H5'] = splitPay7
            sheet['I5'] = splitPay8
            sheet['J5'] = splitPay9

        if self.emp2Checkbox_var.get() == 1:
            sheet['B6'] = splitPay1
            sheet['C6'] = splitPay2
            sheet['D6'] = splitPay3
            sheet['E6'] = splitPay4
            sheet['F6'] = splitPay5
            sheet['G6'] = splitPay6
            sheet['H6'] = splitPay7
            sheet['I6'] = splitPay8
            sheet['J6'] = splitPay9

        if self.emp3Checkbox_var.get() == 1:
            sheet['B7'] = splitPay1
            sheet['C7'] = splitPay2
            sheet['D7'] = splitPay3
            sheet['E7'] = splitPay4
            sheet['F7'] = splitPay5
            sheet['G7'] = splitPay6
            sheet['H7'] = splitPay7
            sheet['I7'] = splitPay8
            sheet['J7'] = splitPay9

        if self.emp4Checkbox_var.get() == 1:
            sheet['B8'] = splitPay1
            sheet['C8'] = splitPay2
            sheet['D8'] = splitPay3
            sheet['E8'] = splitPay4
            sheet['F8'] = splitPay5
            sheet['G8'] = splitPay6
            sheet['H8'] = splitPay7
            sheet['I8'] = splitPay8
            sheet['J8'] = splitPay9

        if self.emp5Checkbox_var.get() == 1:
            sheet['B9'] = splitPay1
            sheet['C9'] = splitPay2
            sheet['D9'] = splitPay3
            sheet['E9'] = splitPay4
            sheet['F9'] = splitPay5
            sheet['G9'] = splitPay6
            sheet['H9'] = splitPay7
            sheet['I9'] = splitPay8
            sheet['J9'] = splitPay9

        if self.emp6Checkbox_var.get() == 1:
            sheet['B10'] = splitPay1
            sheet['C10'] = splitPay2
            sheet['D10'] = splitPay3
            sheet['E10'] = splitPay4
            sheet['F10'] = splitPay5
            sheet['G10'] = splitPay6
            sheet['H10'] = splitPay7
            sheet['I10'] = splitPay8
            sheet['J10'] = splitPay9

        if self.emp7Checkbox_var.get() == 1:
            sheet['B11'] = splitPay1
            sheet['C11'] = splitPay2
            sheet['D11'] = splitPay3
            sheet['E11'] = splitPay4
            sheet['F11'] = splitPay5
            sheet['G11'] = splitPay6
            sheet['H11'] = splitPay7
            sheet['I11'] = splitPay8
            sheet['J11'] = splitPay9

        if self.emp8Checkbox_var.get() == 1:
            sheet['B12'] = splitPay1
            sheet['C12'] = splitPay2
            sheet['D12'] = splitPay3
            sheet['E12'] = splitPay4
            sheet['F12'] = splitPay5
            sheet['G12'] = splitPay6
            sheet['H12'] = splitPay7
            sheet['I12'] = splitPay8
            sheet['J12'] = splitPay9

        if self.emp9Checkbox_var.get() == 1:
            sheet['B13'] = splitPay1
            sheet['C13'] = splitPay2
            sheet['D13'] = splitPay3
            sheet['E13'] = splitPay4
            sheet['F13'] = splitPay5
            sheet['G13'] = splitPay6
            sheet['H13'] = splitPay7
            sheet['I13'] = splitPay8
            sheet['J13'] = splitPay9

        if self.emp10Checkbox_var.get() == 1:
            sheet['B14'] = splitPay1
            sheet['C14'] = splitPay2
            sheet['D14'] = splitPay3
            sheet['E14'] = splitPay4
            sheet['F14'] = splitPay5
            sheet['G14'] = splitPay6
            sheet['H14'] = splitPay7
            sheet['I14'] = splitPay8
            sheet['J14'] = splitPay9

        workbook.save("D:/Insulpro/temp/paysheets/attic pay sheet/" + filename)

    def saveWorkRecord(self):
        print('Saving Work Record')
        nameFile = self.jobName.get()
        atticPayGet1 = self.r19Installed.get()
        atticPayGet2 = self.r30Installed.get()
        atticPayGet3 = self.r38Installed.get()
        atticPayGet4 = self.r49Installed.get()
        atticPayGet5 = self.celluloseInstalled.get()
        atticPayGet6 = self.soffitInstalled.get()
        atticPayGet7 = self.bonusAmount.get()
        atticPayGet8 = self.otherAmount.get()
        atticPayGet9 = self.airSeal.get()

        workbook = load_workbook("D:/Insulpro/payroll records/Work Record.xlsx")

        sheet = workbook['Attic Work Record']

        sheet.append([nameFile, atticPayGet1, atticPayGet2, atticPayGet3, atticPayGet4, atticPayGet5, atticPayGet6,
                      atticPayGet7, atticPayGet8, atticPayGet9])

        workbook.save("D:/Insulpro/payroll records/Work Record.xlsx")

    def savePaytreeAttic(self):
        empFile1 = str(employee1.fullname) + "/attic record.xlsx"
        empFile2 = str(employee2.fullname) + "/attic record.xlsx"
        empFile3 = str(employee3.fullname) + "/attic record.xlsx"
        empFile4 = str(employee4.fullname) + "/attic record.xlsx"
        empFile5 = str(employee5.fullname) + "/attic record.xlsx"
        empFile6 = str(employee6.fullname) + "/attic record.xlsx"
        empFile7 = str(employee7.fullname) + "/attic record.xlsx"
        empFile8 = str(employee8.fullname) + "/attic record.xlsx"
        empFile9 = str(employee9.fullname) + "/attic record.xlsx"
        empFile10 = str(employee10.fullname) + "/attic record.xlsx"

        namefile = self.jobName.get()
        atticPayGet1 = self.r19Installed.get()
        atticPayGet2 = self.r30Installed.get()
        atticPayGet3 = self.r38Installed.get()
        atticPayGet4 = self.r49Installed.get()
        atticPayGet5 = self.celluloseInstalled.get()
        atticPayGet6 = self.soffitInstalled.get()
        atticPayGet7 = self.bonusAmount.get()
        atticPayGet8 = self.otherAmount.get()
        atticPayGet9 = self.airSeal.get()

        atticPayCalc1 = float(atticPayGet1) * r19_blown.rate
        atticPayCalc2 = float(atticPayGet2) * r30_blown.rate
        atticPayCalc3 = float(atticPayGet3) * r38_blown.rate
        atticPayCalc4 = float(atticPayGet4) * r49_blown.rate
        atticPayCalc5 = float(atticPayGet5) * cell_blown.rate
        atticPayCalc6 = float(atticPayGet6) * existing_soffit.rate

        splitPay1 = float(atticPayCalc1) / float(self.numWorkers.get())
        splitPay2 = float(atticPayCalc2) / float(self.numWorkers.get())
        splitPay3 = float(atticPayCalc3) / float(self.numWorkers.get())
        splitPay4 = float(atticPayCalc4) / float(self.numWorkers.get())
        splitPay5 = float(atticPayCalc5) / float(self.numWorkers.get())
        splitPay6 = float(atticPayCalc6) / float(self.numWorkers.get())
        splitPay7 = float(atticPayGet7) / float(self.numWorkers.get())
        splitPay8 = float(atticPayGet8) / float(self.numWorkers.get())
        splitPay9 = float(atticPayGet9) / float(self.numWorkers.get())

        atticPaySplit = splitPay1 + splitPay2 + splitPay3 + splitPay4 + splitPay5 + splitPay6 + splitPay7 + splitPay8 \
                        + splitPay9

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile1)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, atticPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee paytree/" + empFile1)

            workbook.close()

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile2)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, atticPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee paytree/" + empFile2)

            workbook.close()

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile3)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, atticPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee paytree/" + empFile3)

            workbook.close()

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile4)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, atticPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee paytree/" + empFile4)

            workbook.close()

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile5)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, atticPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee paytree/" + empFile5)

            workbook.close()

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile6)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, atticPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee paytree/" + empFile6)

            workbook.close()

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile7)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, atticPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee paytree/" + empFile7)

            workbook.close()

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile8)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, atticPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee paytree/" + empFile8)

            workbook.close()

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile9)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, atticPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee paytree/" + empFile9)

            workbook.close()

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile10)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, atticPaySplit, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee paytree/" + empFile10)

            workbook.close()

    def saveEmployeeRecord(self):
        print('Saving Employee Records')
        empFile1 = str(employee1.fullname + '.xlsx')
        empFile2 = str(employee2.fullname + '.xlsx')
        empFile3 = str(employee3.fullname + '.xlsx')
        empFile4 = str(employee4.fullname + '.xlsx')
        empFile5 = str(employee5.fullname + '.xlsx')
        empFile6 = str(employee6.fullname + '.xlsx')
        empFile7 = str(employee7.fullname + '.xlsx')
        empFile8 = str(employee8.fullname + '.xlsx')
        empFile9 = str(employee9.fullname + '.xlsx')
        empFile10 = str(employee10.fullname + '.xlsx')

        namefile = self.jobName.get()
        atticPayGet1 = self.r19Installed.get()
        atticPayGet2 = self.r30Installed.get()
        atticPayGet3 = self.r38Installed.get()
        atticPayGet4 = self.r49Installed.get()
        atticPayGet5 = self.celluloseInstalled.get()
        atticPayGet6 = self.soffitInstalled.get()
        atticPayGet7 = self.bonusAmount.get()
        atticPayGet8 = self.otherAmount.get()
        atticPayGet9 = self.airSeal.get()

        atticPayCalc1 = float(atticPayGet1) * r19_blown.rate
        atticPayCalc2 = float(atticPayGet2) * r30_blown.rate
        atticPayCalc3 = float(atticPayGet3) * r38_blown.rate
        atticPayCalc4 = float(atticPayGet4) * r49_blown.rate
        atticPayCalc5 = float(atticPayGet5) * cell_blown.rate
        atticPayCalc6 = float(atticPayGet6) * existing_soffit.rate

        splitPay1 = float(atticPayCalc1) / float(self.numWorkers.get())
        splitPay2 = float(atticPayCalc2) / float(self.numWorkers.get())
        splitPay3 = float(atticPayCalc3) / float(self.numWorkers.get())
        splitPay4 = float(atticPayCalc4) / float(self.numWorkers.get())
        splitPay5 = float(atticPayCalc5) / float(self.numWorkers.get())
        splitPay6 = float(atticPayCalc6) / float(self.numWorkers.get())
        splitPay7 = float(atticPayGet7) / float(self.numWorkers.get())
        splitPay8 = float(atticPayGet8) / float(self.numWorkers.get())
        splitPay9 = float(atticPayGet9) / float(self.numWorkers.get())

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile1)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee work records/" + empFile1)

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile2)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee work records/" + empFile2)

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile3)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee work records/" + empFile3)

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile4)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee work records/" + empFile4)

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile5)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee work records/" + empFile5)

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile6)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee work records/" + empFile6)

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile7)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee work records/" + empFile7)

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile8)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee work records/" + empFile8)

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile9)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee work records/" + empFile9)

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile10)

            sheet = workbook['Attic Work Record']

            sheet.append([namefile, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9])

            workbook.save("D:/Insulpro/employee work records/" + empFile10)

    def savePayrollSpreadsheet(self):
        print('Saving Payroll Spreadsheet')
        namefile = self.jobName.get()
        atticPayGet1 = self.r19Installed.get()
        atticPayGet2 = self.r30Installed.get()
        atticPayGet3 = self.r38Installed.get()
        atticPayGet4 = self.r49Installed.get()
        atticPayGet5 = self.celluloseInstalled.get()
        atticPayGet6 = self.soffitInstalled.get()
        atticPayGet7 = self.bonusAmount.get()
        atticPayGet8 = self.otherAmount.get()
        atticPayGet9 = self.airSeal.get()
        atticPayCalc6 = float(atticPayGet6) * existing_soffit.rate

        splitPay1 = round(float(atticPayGet1) / float(self.numWorkers.get()), 0)
        splitPay2 = round(float(atticPayGet2) / float(self.numWorkers.get()), 0)
        splitPay3 = round(float(atticPayGet3) / float(self.numWorkers.get()), 0)
        splitPay4 = round(float(atticPayGet4) / float(self.numWorkers.get()), 0)
        splitPay5 = round(float(atticPayGet5) / float(self.numWorkers.get()), 0)
        splitPay6 = round(float(atticPayCalc6) / float(self.numWorkers.get()), 2)
        splitPay7 = round(float(atticPayGet7) / float(self.numWorkers.get()), 2)
        splitPay8 = round(float(atticPayGet8) / float(self.numWorkers.get()), 2)
        splitPay9 = round(float(atticPayGet9) / float(self.numWorkers.get()), 2)

        workbook = load_workbook("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        if self.emp1Checkbox_var.get() == 1:
            sheet = workbook['Emp1Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp2Checkbox_var.get() == 1:
            sheet = workbook['Emp2Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp3Checkbox_var.get() == 1:
            sheet = workbook['Emp3Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp4Checkbox_var.get() == 1:
            sheet = workbook['Emp4Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp5Checkbox_var.get() == 1:
            sheet = workbook['Emp5Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp6Checkbox_var.get() == 1:
            sheet = workbook['Emp6Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp7Checkbox_var.get() == 1:
            sheet = workbook['Emp7Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp8Checkbox_var.get() == 1:
            sheet = workbook['Emp8Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp9Checkbox_var.get() == 1:
            sheet = workbook['Emp9Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp10Checkbox_var.get() == 1:
            sheet = workbook['Emp10Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, splitPay9,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        workbook.save("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        workbook.close()

    def emp1Checkbox_command(self):
        print(employee1.fullname + " selected")
        self.emp1Checkbox_var.set(1)
        self.count += 1

    def emp2Checkbox_command(self):
        print(employee2.fullname + " selected")
        self.emp2Checkbox_var.set(1)
        self.count += 1

    def emp3Checkbox_command(self):
        print(employee3.fullname + " selected")
        self.emp3Checkbox_var.set(1)
        self.count += 1

    def emp4Checkbox_command(self):
        print(employee4.fullname + " selected")
        self.emp4Checkbox_var.set(1)
        self.count += 1

    def emp5Checkbox_command(self):
        print(employee5.fullname + " selected")
        self.emp5Checkbox_var.set(1)
        self.count += 1

    def emp6Checkbox_command(self):
        print(employee6.fullname + " selected")
        self.emp6Checkbox_var.set(1)
        self.count += 1

    def emp7Checkbox_command(self):
        print(employee7.fullname + " selected")
        self.emp7Checkbox_var.set(1)
        self.count += 1

    def emp8Checkbox_command(self):
        print(employee8.fullname + " selected")
        self.emp8Checkbox_var.set(1)
        self.count += 1

    def emp9Checkbox_command(self):
        print(employee9.fullname + " selected")
        self.emp9Checkbox_var.set(1)
        self.count += 1

    def emp10Checkbox_command(self):
        print(employee10.fullname + " selected")
        self.emp10Checkbox_var.set(1)
        self.count += 1


class SprayFoamPaySheet(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("800x600")
        self.title("Foam Pay Sheet")
        # setting window size
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)

        self.count = 0

        self.emp1Checkbox_var = tk.IntVar(value=0)
        self.emp2Checkbox_var = tk.IntVar(value=0)
        self.emp3Checkbox_var = tk.IntVar(value=0)
        self.emp4Checkbox_var = tk.IntVar(value=0)
        self.emp5Checkbox_var = tk.IntVar(value=0)
        self.emp6Checkbox_var = tk.IntVar(value=0)
        self.emp7Checkbox_var = tk.IntVar(value=0)
        self.emp8Checkbox_var = tk.IntVar(value=0)
        self.emp9Checkbox_var = tk.IntVar(value=0)
        self.emp10Checkbox_var = tk.IntVar(value=0)

        emp1Checkbox = ctk.CTkCheckBox(self,
                                       text=employee1.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp1Checkbox_command,
                                       variable=self.emp1Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp1Checkbox.place(x=5, y=30)

        emp2Checkbox = ctk.CTkCheckBox(self,
                                       text=employee2.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp2Checkbox_command,
                                       variable=self.emp2Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp2Checkbox.place(x=5, y=65)

        emp3Checkbox = ctk.CTkCheckBox(self,
                                       text=employee3.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp3Checkbox_command,
                                       variable=self.emp3Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp3Checkbox.place(x=5, y=100)

        emp4Checkbox = ctk.CTkCheckBox(self,
                                       text=employee4.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp4Checkbox_command,
                                       variable=self.emp4Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp4Checkbox.place(x=5, y=135)

        emp5Checkbox = ctk.CTkCheckBox(self,
                                       text=employee5.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp5Checkbox_command,
                                       variable=self.emp5Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp5Checkbox.place(x=5, y=170)

        emp6Checkbox = ctk.CTkCheckBox(self,
                                       text=employee6.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp6Checkbox_command,
                                       variable=self.emp6Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp6Checkbox.place(x=5, y=205)

        emp7Checkbox = ctk.CTkCheckBox(self,
                                       text=employee7.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp7Checkbox_command,
                                       variable=self.emp7Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp7Checkbox.place(x=5, y=240)

        emp8Checkbox = ctk.CTkCheckBox(self,
                                       text=employee8.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp8Checkbox_command,
                                       variable=self.emp8Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp8Checkbox.place(x=5, y=275)

        emp9Checkbox = ctk.CTkCheckBox(self,
                                       text=employee9.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp9Checkbox_command,
                                       variable=self.emp9Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp9Checkbox.place(x=5, y=310)

        emp10Checkbox = ctk.CTkCheckBox(self,
                                        text=employee10.fullname,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.emp10Checkbox_command,
                                        variable=self.emp10Checkbox_var,
                                        onvalue=1,
                                        offvalue=0)
        emp10Checkbox.place(x=5, y=345)

        jobNameLabel = ctk.CTkLabel(self,
                                    text="Job Name",
                                    width=100,
                                    height=30,
                                    fg_color="transparent",
                                    font=ctk.CTkFont(size=18, weight="normal"))

        jobNameLabel.place(x=180, y=30)

        self.jobName = ctk.CTkEntry(self,
                                    width=150,
                                    height=30,
                                    font=ctk.CTkFont(size=18, weight="normal"))

        self.jobName.place(x=300, y=30)

        closed34Label = ctk.CTkLabel(self,
                                     text="Closed 3/4 ",
                                     width=100,
                                     height=30,
                                     fg_color="transparent",
                                     font=ctk.CTkFont(size=18, weight="normal"))

        closed34Label.place(x=180, y=65)

        self.closed34 = ctk.CTkEntry(self,
                                     width=100,
                                     height=30,
                                     font=ctk.CTkFont(size=18, weight="normal"))

        self.closed34.place(x=300, y=65)
        self.closed34.insert(0, "0")

        closed1Label = ctk.CTkLabel(self,
                                    text="Closed 1",
                                    width=100,
                                    height=30,
                                    fg_color="transparent",
                                    font=ctk.CTkFont(size=18, weight="normal"))

        closed1Label.place(x=180, y=100)

        self.closed1 = ctk.CTkEntry(self,
                                    width=100,
                                    height=30,
                                    font=ctk.CTkFont(size=18, weight="normal"))

        self.closed1.place(x=300, y=100)
        self.closed1.insert(0, "0")

        closed2Label = ctk.CTkLabel(self,
                                    text="Closed 2",
                                    width=100,
                                    height=30,
                                    fg_color="transparent",
                                    font=ctk.CTkFont(size=18, weight="normal"))

        closed2Label.place(x=180, y=135)

        self.closed2 = ctk.CTkEntry(self,
                                    width=100,
                                    height=30,
                                    font=ctk.CTkFont(size=18, weight="normal"))

        self.closed2.place(x=300, y=135)
        self.closed2.insert(0, "0")

        closed3Label = ctk.CTkLabel(self,
                                    text="Closed 3",
                                    width=100,
                                    height=30,
                                    fg_color="transparent",
                                    font=ctk.CTkFont(size=18, weight="normal"))

        closed3Label.place(x=180, y=170)

        self.closed3 = ctk.CTkEntry(self,
                                    width=100,
                                    height=30,
                                    font=ctk.CTkFont(size=18, weight="normal"))

        self.closed3.place(x=300, y=170)
        self.closed3.insert(0, "0")

        open4Label = ctk.CTkLabel(self,
                                  text="Open 4",
                                  width=100,
                                  height=30,
                                  fg_color="transparent",
                                  font=ctk.CTkFont(size=18, weight="normal"))

        open4Label.place(x=180, y=205)

        self.open4 = ctk.CTkEntry(self,
                                  width=100,
                                  height=30,
                                  font=ctk.CTkFont(size=18, weight="normal"))

        self.open4.place(x=300, y=205)
        self.open4.insert(0, "0")

        open6Label = ctk.CTkLabel(self,
                                  text="Open 6",
                                  width=100,
                                  height=30,
                                  fg_color="transparent",
                                  font=ctk.CTkFont(size=18, weight="normal"))

        open6Label.place(x=180, y=240)

        self.open6 = ctk.CTkEntry(self,
                                  width=100,
                                  height=30,
                                  font=ctk.CTkFont(size=18, weight="normal"))

        self.open6.place(x=300, y=240)
        self.open6.insert(0, "0")

        open8Label = ctk.CTkLabel(self,
                                  text="Open 8",
                                  width=100,
                                  height=30,
                                  fg_color="transparent",
                                  font=ctk.CTkFont(size=18, weight="normal"))

        open8Label.place(x=180, y=275)

        self.open8 = ctk.CTkEntry(self,
                                  width=100,
                                  height=30,
                                  font=ctk.CTkFont(size=18, weight="normal"))

        self.open8.place(x=300, y=275)
        self.open8.insert(0, "0")

        bonusAmountLabel = ctk.CTkLabel(self,
                                        text="Bonus",
                                        width=100,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        bonusAmountLabel.place(x=180, y=310)

        self.bonusAmount = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.bonusAmount.place(x=300, y=310)
        self.bonusAmount.insert(0, "0")

        otherAmountLabel = ctk.CTkLabel(self,
                                        text="Other Amount",
                                        width=100,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        otherAmountLabel.place(x=180, y=345)

        self.otherAmount = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.otherAmount.place(x=300, y=345)
        self.otherAmount.insert(0, "0")

        jobTotalPayLabel = ctk.CTkLabel(self,
                                        text="Job Total Pay",
                                        width=100,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        jobTotalPayLabel.place(x=180, y=400)

        self.jobTotalPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=25,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.jobTotalPay.place(x=300, y=400)
        self.jobTotalPay.configure(state='disabled')

        numWorkersLabel = ctk.CTkLabel(self,
                                       text="Total Workers",
                                       width=100,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        numWorkersLabel.place(x=180, y=435)

        self.numWorkers = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.numWorkers.place(x=300, y=435)
        self.numWorkers.configure(state='disabled')

        totalPayPerLabel = ctk.CTkLabel(self,
                                        text="Total Pay Per Person",
                                        width=150,
                                        height=25,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        totalPayPerLabel.place(x=420, y=400)

        self.totalPayPer = ctk.CTkEntry(self,
                                        width=100,
                                        height=25,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.totalPayPer.place(x=595, y=400)
        self.totalPayPer.configure(state='disabled')

        calculateButton = ctk.CTkButton(self,
                                        text="Calculate",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.calculateButton_command)

        calculateButton.place(x=5, y=430)

        resetFormButton = ctk.CTkButton(self,
                                        text="Reset Form",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.resetFormButton_command)

        resetFormButton.place(x=5, y=465)

        saveFileButton = ctk.CTkButton(self,
                                       text="Save File",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.saveFileButton_command)

        saveFileButton.place(x=5, y=500)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=535)

    def calculateButton_command(self):
        print("")
        print('Calculating Foam Pay Sheet...')
        self.numWorkers.configure(state='normal')
        self.jobTotalPay.configure(state='normal')
        self.totalPayPer.configure(state='normal')
        self.numWorkers.insert(0, str(self.count))
        nameFile = str(self.jobName.get())
        foamPayGet0 = self.closed34.get()
        foamPayGet1 = self.closed1.get()
        foamPayGet2 = self.closed2.get()
        foamPayGet3 = self.closed3.get()
        foamPayGet4 = self.open4.get()
        foamPayGet5 = self.open6.get()
        foamPayGet6 = self.open8.get()
        foamPayGet7 = self.bonusAmount.get()
        foamPayGet8 = self.otherAmount.get()

        foamPayCalc0 = float(foamPayGet0) * closed_under34.rate
        foamPayCalc1 = float(foamPayGet1) * closed_under1.rate
        foamPayCalc2 = float(foamPayGet2) * closed_under2.rate
        foamPayCalc3 = float(foamPayGet3) * closed_3.rate
        foamPayCalc4 = float(foamPayGet4) * open_cell4.rate
        foamPayCalc5 = float(foamPayGet5) * open_cell6.rate
        foamPayCalc6 = float(foamPayGet6) * open_cell8.rate
        foamPayCalc7 = float(foamPayGet7)
        foamPayCalc8 = float(foamPayGet8)

        foamPayTotalCalc = round(foamPayCalc0 + foamPayCalc1 + foamPayCalc2 + foamPayCalc3 + foamPayCalc4 +
                                 foamPayCalc5 + foamPayCalc6 + foamPayCalc7 + foamPayCalc8, 2)

        foamPaySplit = round(float(foamPayTotalCalc) / float(self.numWorkers.get()), 2)

        self.jobTotalPay.insert(0, '$' + str(round(foamPayTotalCalc, 2)))
        self.totalPayPer.insert(0, '$' + str(round(foamPaySplit, 2)))

        print("")
        print("Job Name: " + nameFile)
        print("")
        print("Closed Cell >= 3/4 Inch Installed: " + str(foamPayGet0) + " sqft")
        print(str(foamPayGet0) + " x " + str(closed_under34.rate) + " = $" + str(foamPayCalc0))
        print("")
        print("Closed Cell >1 Inch Installed: " + str(foamPayGet1) + " sqft")
        print(str(foamPayGet1) + " x " + str(closed_under1.rate) + " = $" + str(foamPayCalc1))
        print("")
        print("Closed Cell >= 2 Inch Installed: " + str(foamPayGet2) + " sqft")
        print(str(foamPayGet2) + " x " + str(closed_under2.rate) + " = $" + str(foamPayCalc2))
        print("")
        print("Closed Cell <= 3 Inch Installed: " + str(foamPayGet3) + " sqft")
        print(str(foamPayGet3) + " x " + str(closed_3.rate) + " = $" + str(foamPayCalc3))
        print("")
        print("Open Cell >= 4 Inch Installed: " + str(foamPayGet4) + " sqft")
        print(str(foamPayGet4) + " x " + str(open_cell4.rate) + " = $" + str(foamPayCalc4))
        print("")
        print("Open Cell <= 6 Inch Installed: " + str(foamPayGet5) + " sqft")
        print(str(foamPayGet5) + " x " + str(open_cell6.rate) + " = $" + str(foamPayCalc5))
        print("")
        print("Open Cell <= 8 Inch Installed: " + str(foamPayGet6) + " sqft")
        print(str(foamPayGet6) + " x " + str(open_cell8.rate) + " = $" + str(foamPayCalc6))
        print("")
        print("Total Bonus Amount = $" + str(foamPayGet7))
        print("")
        print("Total Other Amount = $" + str(foamPayGet8))
        print("")
        print("Total Job Pay: $" + str(foamPayTotalCalc))
        print("")
        print("Total Number Of Employee's: " + str(self.numWorkers.get()))
        print("")
        print("Total Pay Per Person: $" + str(foamPaySplit))
        print("")

    def resetFormButton_command(self):
        print('Resetting Form')
        self.count = 0
        self.jobName.delete(0, 'end')
        self.closed34.delete(0, 'end')
        self.numWorkers.delete(0, 'end')
        self.closed1.delete(0, 'end')
        self.closed2.delete(0, 'end')
        self.closed3.delete(0, 'end')
        self.open4.delete(0, 'end')
        self.open6.delete(0, 'end')
        self.open8.delete(0, 'end')
        self.bonusAmount.delete(0, 'end')
        self.otherAmount.delete(0, 'end')
        self.jobTotalPay.delete(0, 'end')
        self.totalPayPer.delete(0, 'end')
        self.closed34.insert(0, '0')
        self.closed1.insert(0, '0')
        self.closed2.insert(0, '0')
        self.closed3.insert(0, '0')
        self.open4.insert(0, '0')
        self.open6.insert(0, '0')
        self.open8.insert(0, '0')
        self.bonusAmount.insert(0, '0')
        self.otherAmount.insert(0, '0')

        if self.emp1Checkbox_var.get() == 1:
            self.emp1Checkbox_var.set(0)

        if self.emp2Checkbox_var.get() == 1:
            self.emp2Checkbox_var.set(0)

        if self.emp3Checkbox_var.get() == 1:
            self.emp3Checkbox_var.set(0)

        if self.emp4Checkbox_var.get() == 1:
            self.emp4Checkbox_var.set(0)

        if self.emp5Checkbox_var.get() == 1:
            self.emp5Checkbox_var.set(0)

        if self.emp6Checkbox_var.get() == 1:
            self.emp6Checkbox_var.set(0)

        if self.emp7Checkbox_var.get() == 1:
            self.emp7Checkbox_var.set(0)

        if self.emp8Checkbox_var.get() == 1:
            self.emp8Checkbox_var.set(0)

        if self.emp9Checkbox_var.get() == 1:
            self.emp9Checkbox_var.set(0)

        if self.emp10Checkbox_var.get() == 1:
            self.emp10Checkbox_var.set(0)

        self.numWorkers.configure(state='disabled')
        self.jobTotalPay.configure(state='disabled')
        self.totalPayPer.configure(state='disabled')

    def saveFileButton_command(self):
        namefile = self.jobName.get()
        filename = str(namefile + ".xlsx")

        if not os.path.exists(
                "D:/Insulpro/temp/paysheets/foam pay sheet/" + filename):
            print('Saving File...')
            self.savePaySheet()
            self.saveWorkRecord()
            self.saveEmployeeRecord()
            self.savePaytreeFoam()
            self.savePayrollSpreadsheet()

            self.resetFormButton_command()
            self.closeWindowButton_command()

            messagebox.showinfo(title='Save', message='File Save Successful')

        else:
            messagebox.showerror(title='File Name Already Exist',
                                 message='Please change job name to something else!')

    def closeWindowButton_command(self):
        self.destroy()

    def savePaySheet(self):
        print('Saving Pay Sheet')
        foamPayGet0 = self.closed34.get()
        foamPayGet1 = self.closed1.get()
        foamPayGet2 = self.closed2.get()
        foamPayGet3 = self.closed3.get()
        foamPayGet4 = self.open4.get()
        foamPayGet5 = self.open6.get()
        foamPayGet6 = self.open8.get()
        foamPayGet7 = self.bonusAmount.get()
        foamPayGet8 = self.otherAmount.get()

        foamPayCalc0 = float(foamPayGet0) * closed_under34.rate
        foamPayCalc1 = float(foamPayGet1) * closed_under1.rate
        foamPayCalc2 = float(foamPayGet2) * closed_under2.rate
        foamPayCalc3 = float(foamPayGet3) * closed_3.rate
        foamPayCalc4 = float(foamPayGet4) * open_cell4.rate
        foamPayCalc5 = float(foamPayGet5) * open_cell6.rate
        foamPayCalc6 = float(foamPayGet6) * open_cell8.rate
        foamPayCalc7 = float(foamPayGet7)
        foamPayCalc8 = float(foamPayGet8)

        splitPay0 = round(float(foamPayCalc0) / float(self.numWorkers.get()), 2)
        splitPay1 = round(float(foamPayCalc1) / float(self.numWorkers.get()), 2)
        splitPay2 = round(float(foamPayCalc2 / float(self.numWorkers.get())), 2)
        splitPay3 = round(float(foamPayCalc3) / float(self.numWorkers.get()), 2)
        splitPay4 = round(float(foamPayCalc4) / float(self.numWorkers.get()), 2)
        splitPay5 = round(float(foamPayCalc5) / float(self.numWorkers.get()), 2)
        splitPay6 = round(float(foamPayCalc6) / float(self.numWorkers.get()), 2)
        splitPay7 = round(float(foamPayCalc7) / float(self.numWorkers.get()), 2)
        splitPay8 = round(float(foamPayCalc8) / float(self.numWorkers.get()), 2)

        source = "D:/Insulpro/system/spreadsheet/Foam Pay Sheet.xlsx"

        destination = ("D:/Insulpro/temp/paysheets/foam pay sheet/Foam Pay "
                       "Sheet.xlsx")

        shutil.copy(source, destination)

        nameFile = self.jobName.get()

        filename = str(nameFile + '.xlsx')

        os.rename("D:/Insulpro/temp/paysheets/foam pay sheet/Foam Pay Sheet.xlsx",
                  "D:/Insulpro/temp/paysheets/foam pay sheet/" + filename)

        workbook = load_workbook(
            "D:/Insulpro/temp/paysheets/foam pay sheet/" + filename)

        sheet = workbook.active

        jobNameSheet = str(nameFile)

        sheet["A2"] = jobNameSheet
        sheet['A5'] = employee1.fullname
        sheet['A6'] = employee2.fullname
        sheet['A7'] = employee3.fullname
        sheet['A8'] = employee4.fullname
        sheet['A9'] = employee5.fullname
        sheet['A10'] = employee6.fullname
        sheet['A11'] = employee7.fullname
        sheet['A12'] = employee8.fullname
        sheet['A13'] = employee9.fullname
        sheet['A14'] = employee10.fullname

        if self.emp1Checkbox_var.get() == 1:
            sheet['B5'] = splitPay0
            sheet['C5'] = splitPay1
            sheet['D5'] = splitPay2
            sheet['E5'] = splitPay3
            sheet['F5'] = splitPay4
            sheet['G5'] = splitPay5
            sheet['H5'] = splitPay6
            sheet['I5'] = splitPay7
            sheet['J5'] = splitPay8

        if self.emp2Checkbox_var.get() == 1:
            sheet['B6'] = splitPay0
            sheet['C6'] = splitPay1
            sheet['D6'] = splitPay2
            sheet['E6'] = splitPay3
            sheet['F6'] = splitPay4
            sheet['G6'] = splitPay5
            sheet['H6'] = splitPay6
            sheet['I6'] = splitPay7
            sheet['J6'] = splitPay8

        if self.emp3Checkbox_var.get() == 1:
            sheet['B7'] = splitPay0
            sheet['C7'] = splitPay1
            sheet['D7'] = splitPay2
            sheet['E7'] = splitPay3
            sheet['F7'] = splitPay4
            sheet['G7'] = splitPay5
            sheet['H7'] = splitPay6
            sheet['I7'] = splitPay7
            sheet['J7'] = splitPay8

        if self.emp4Checkbox_var.get() == 1:
            sheet['B8'] = splitPay0
            sheet['C8'] = splitPay1
            sheet['D8'] = splitPay2
            sheet['E8'] = splitPay3
            sheet['F8'] = splitPay4
            sheet['G8'] = splitPay5
            sheet['H8'] = splitPay6
            sheet['I8'] = splitPay7
            sheet['J8'] = splitPay8

        if self.emp5Checkbox_var.get() == 1:
            sheet['B9'] = splitPay0
            sheet['C9'] = splitPay1
            sheet['D9'] = splitPay2
            sheet['E9'] = splitPay3
            sheet['F9'] = splitPay4
            sheet['G9'] = splitPay5
            sheet['H9'] = splitPay6
            sheet['I9'] = splitPay7
            sheet['J9'] = splitPay8

        if self.emp6Checkbox_var.get() == 1:
            sheet['B10'] = splitPay0
            sheet['C10'] = splitPay1
            sheet['D10'] = splitPay2
            sheet['E10'] = splitPay3
            sheet['F10'] = splitPay4
            sheet['G10'] = splitPay5
            sheet['H10'] = splitPay6
            sheet['I10'] = splitPay7
            sheet['J10'] = splitPay8

        if self.emp7Checkbox_var.get() == 1:
            sheet['B11'] = splitPay0
            sheet['C11'] = splitPay1
            sheet['D11'] = splitPay2
            sheet['E11'] = splitPay3
            sheet['F11'] = splitPay4
            sheet['G11'] = splitPay5
            sheet['H11'] = splitPay6
            sheet['I11'] = splitPay7
            sheet['J11'] = splitPay8

        if self.emp8Checkbox_var.get() == 1:
            sheet['B12'] = splitPay0
            sheet['C12'] = splitPay1
            sheet['D12'] = splitPay2
            sheet['E12'] = splitPay3
            sheet['F12'] = splitPay4
            sheet['G12'] = splitPay5
            sheet['H12'] = splitPay6
            sheet['I12'] = splitPay7
            sheet['J12'] = splitPay8

        if self.emp9Checkbox_var.get() == 1:
            sheet['B13'] = splitPay0
            sheet['C13'] = splitPay1
            sheet['D13'] = splitPay2
            sheet['E13'] = splitPay3
            sheet['F13'] = splitPay4
            sheet['G13'] = splitPay5
            sheet['H13'] = splitPay6
            sheet['I13'] = splitPay7
            sheet['J13'] = splitPay8

        if self.emp10Checkbox_var.get() == 1:
            sheet['B14'] = splitPay0
            sheet['C14'] = splitPay1
            sheet['D14'] = splitPay2
            sheet['E14'] = splitPay3
            sheet['F14'] = splitPay4
            sheet['G14'] = splitPay5
            sheet['H14'] = splitPay6
            sheet['I14'] = splitPay7
            sheet['J14'] = splitPay8

        workbook.save("D:/Insulpro/temp/paysheets/foam pay sheet/" + filename)

    def saveWorkRecord(self):
        print('Saving Work Record')
        nameFile = str(self.jobName.get())
        foamPayGet0 = self.closed34.get()
        foamPayGet1 = self.closed1.get()
        foamPayGet2 = self.closed2.get()
        foamPayGet3 = self.closed3.get()
        foamPayGet4 = self.open4.get()
        foamPayGet5 = self.open6.get()
        foamPayGet6 = self.open8.get()
        foamPayGet7 = self.bonusAmount.get()
        foamPayGet8 = self.otherAmount.get()

        workbook = load_workbook(
            "D:/Insulpro/payroll records/Work Record.xlsx")

        sheet = workbook['Foam Work Record']

        sheet.append([nameFile, foamPayGet0, foamPayGet1, foamPayGet2, foamPayGet3, foamPayGet4, foamPayGet5,
                      foamPayGet6, foamPayGet7, foamPayGet8])

        workbook.save("D:/Insulpro/payroll records/Work Record.xlsx")

    def savePaytreeFoam(self):
        empFile1 = str(employee1.fullname) + "/foam record.xlsx"
        empFile2 = str(employee2.fullname) + "/foam record.xlsx"
        empFile3 = str(employee3.fullname) + "/foam record.xlsx"
        empFile4 = str(employee4.fullname) + "/foam record.xlsx"
        empFile5 = str(employee5.fullname) + "/foam record.xlsx"
        empFile6 = str(employee6.fullname) + "/foam record.xlsx"
        empFile7 = str(employee7.fullname) + "/foam record.xlsx"
        empFile8 = str(employee8.fullname) + "/foam record.xlsx"
        empFile9 = str(employee9.fullname) + "/foam record.xlsx"
        empFile10 = str(employee10.fullname) + "/foam record.xlsx"

        namefile = self.jobName.get()
        foamPayGet0 = self.closed34.get()
        foamPayGet1 = self.closed1.get()
        foamPayGet2 = self.closed2.get()
        foamPayGet3 = self.closed3.get()
        foamPayGet4 = self.open4.get()
        foamPayGet5 = self.open6.get()
        foamPayGet6 = self.open8.get()
        foamPayGet7 = self.bonusAmount.get()
        foamPayGet8 = self.otherAmount.get()

        foamPayCalc0 = float(foamPayGet0) * closed_under34.rate
        foamPayCalc1 = float(foamPayGet1) * closed_under1.rate
        foamPayCalc2 = float(foamPayGet2) * closed_under2.rate
        foamPayCalc3 = float(foamPayGet3) * closed_3.rate
        foamPayCalc4 = float(foamPayGet4) * open_cell4.rate
        foamPayCalc5 = float(foamPayGet5) * open_cell6.rate
        foamPayCalc6 = float(foamPayGet6) * open_cell8.rate
        foamPayCalc7 = float(foamPayGet7)
        foamPayCalc8 = float(foamPayGet8)

        splitPay0 = round(float(foamPayCalc0) / float(self.numWorkers.get()), 2)
        splitPay1 = round(float(foamPayCalc1) / float(self.numWorkers.get()), 2)
        splitPay2 = round(float(foamPayCalc2 / float(self.numWorkers.get())), 2)
        splitPay3 = round(float(foamPayCalc3) / float(self.numWorkers.get()), 2)
        splitPay4 = round(float(foamPayCalc4) / float(self.numWorkers.get()), 2)
        splitPay5 = round(float(foamPayCalc5) / float(self.numWorkers.get()), 2)
        splitPay6 = round(float(foamPayCalc6) / float(self.numWorkers.get()), 2)
        splitPay7 = round(float(foamPayCalc7) / float(self.numWorkers.get()), 2)
        splitPay8 = round(float(foamPayCalc8) / float(self.numWorkers.get()), 2)

        foamPaySplit = (splitPay0 + splitPay1 + splitPay2 + splitPay3 + splitPay4 + splitPay5 + splitPay6 +
                        splitPay7 + splitPay8)

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile1)

            sheet = workbook['Foam Work Record']

            sheet.append([namefile, foamPaySplit, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile1)

            workbook.close()

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile2)

            sheet = workbook['Foam Work Record']

            sheet.append([namefile, foamPaySplit, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile2)

            workbook.close()

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile3)

            sheet = workbook['Foam Work Record']

            sheet.append([namefile, foamPaySplit, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile3)

            workbook.close()

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile4)

            sheet = workbook['Foam Work Record']

            sheet.append([namefile, foamPaySplit, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile4)

            workbook.close()

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile5)

            sheet = workbook['Foam Work Record']

            sheet.append([namefile, foamPaySplit, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile5)

            workbook.close()

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile6)

            sheet = workbook['Foam Work Record']

            sheet.append([namefile, foamPaySplit, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile6)

            workbook.close()

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile7)

            sheet = workbook['Foam Work Record']

            sheet.append([namefile, foamPaySplit, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile7)

            workbook.close()

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile8)

            sheet = workbook['Foam Work Record']

            sheet.append([namefile, foamPaySplit, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile8)

            workbook.close()

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile9)

            sheet = workbook['Foam Work Record']

            sheet.append([namefile, foamPaySplit, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile9)

            workbook.close()

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile10)

            sheet = workbook['Foam Work Record']

            sheet.append([namefile, foamPaySplit, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee paytree/" + empFile10)

            workbook.close()

    def saveEmployeeRecord(self):
        print('Saving Employee Records')
        empFile1 = str(employee1.fullname + '.xlsx')
        empFile2 = str(employee2.fullname + '.xlsx')
        empFile3 = str(employee3.fullname + '.xlsx')
        empFile4 = str(employee4.fullname + '.xlsx')
        empFile5 = str(employee5.fullname + '.xlsx')
        empFile6 = str(employee6.fullname + '.xlsx')
        empFile7 = str(employee7.fullname + '.xlsx')
        empFile8 = str(employee8.fullname + '.xlsx')
        empFile9 = str(employee9.fullname + '.xlsx')
        empFile10 = str(employee10.fullname + '.xlsx')

        nameFile = str(self.jobName.get())
        foamPayGet0 = self.closed34.get()
        foamPayGet1 = self.closed1.get()
        foamPayGet2 = self.closed2.get()
        foamPayGet3 = self.closed3.get()
        foamPayGet4 = self.open4.get()
        foamPayGet5 = self.open6.get()
        foamPayGet6 = self.open8.get()
        foamPayGet7 = self.bonusAmount.get()
        foamPayGet8 = self.otherAmount.get()

        foamPayCalc0 = float(foamPayGet0) * closed_under34.rate
        foamPayCalc1 = float(foamPayGet1) * closed_under1.rate
        foamPayCalc2 = float(foamPayGet2) * closed_under2.rate
        foamPayCalc3 = float(foamPayGet3) * closed_3.rate
        foamPayCalc4 = float(foamPayGet4) * open_cell4.rate
        foamPayCalc5 = float(foamPayGet5) * open_cell6.rate
        foamPayCalc6 = float(foamPayGet6) * open_cell8.rate
        foamPayCalc7 = float(foamPayGet7)
        foamPayCalc8 = float(foamPayGet8)

        splitPay0 = round(float(foamPayCalc0) / float(self.numWorkers.get()), 2)
        splitPay1 = round(float(foamPayCalc1) / float(self.numWorkers.get()), 2)
        splitPay2 = round(float(foamPayCalc2 / float(self.numWorkers.get())), 2)
        splitPay3 = round(float(foamPayCalc3) / float(self.numWorkers.get()), 2)
        splitPay4 = round(float(foamPayCalc4) / float(self.numWorkers.get()), 2)
        splitPay5 = round(float(foamPayCalc5) / float(self.numWorkers.get()), 2)
        splitPay6 = round(float(foamPayCalc6) / float(self.numWorkers.get()), 2)
        splitPay7 = round(float(foamPayCalc7) / float(self.numWorkers.get()), 2)
        splitPay8 = round(float(foamPayCalc8) / float(self.numWorkers.get()), 2)

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile1)

            sheet = workbook['Foam Work Record']

            sheet.append([nameFile, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile1)

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile2)

            sheet = workbook['Foam Work Record']

            sheet.append([nameFile, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])
            workbook.save("D:/Insulpro/employee work records/" + empFile2)

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile3)

            sheet = workbook['Foam Work Record']

            sheet.append([nameFile, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])
            workbook.save("D:/Insulpro/employee work records/" + empFile3)

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile4)

            sheet = workbook['Foam Work Record']

            sheet.append([nameFile, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile4)

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile5)

            sheet = workbook['Foam Work Record']

            sheet.append([nameFile, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile5)

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile6)

            sheet = workbook['Foam Work Record']

            sheet.append([nameFile, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile6)

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile7)

            sheet = workbook['Foam Work Record']

            sheet.append([nameFile, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile7)

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile8)

            sheet = workbook['Foam Work Record']

            sheet.append([nameFile, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile8)

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile9)

            sheet = workbook['Foam Work Record']

            sheet.append([nameFile, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile9)

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile10)

            sheet = workbook['Foam Work Record']

            sheet.append([nameFile, splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8])

            workbook.save("D:/Insulpro/employee work records/" + empFile10)

    def savePayrollSpreadsheet(self):
        print('Saving Payroll Spreadsheet')
        namefile = self.jobName.get()
        foamPayGet0 = self.closed34.get()
        foamPayGet1 = self.closed1.get()
        foamPayGet2 = self.closed2.get()
        foamPayGet3 = self.closed3.get()
        foamPayGet4 = self.open4.get()
        foamPayGet5 = self.open6.get()
        foamPayGet6 = self.open8.get()
        foamPayGet7 = self.bonusAmount.get()
        foamPayGet8 = self.otherAmount.get()

        foamPayCalc0 = float(foamPayGet0) * closed_under34.rate
        foamPayCalc1 = float(foamPayGet1) * closed_under1.rate
        foamPayCalc2 = float(foamPayGet2) * closed_under2.rate
        foamPayCalc3 = float(foamPayGet3) * closed_3.rate
        foamPayCalc4 = float(foamPayGet4) * open_cell4.rate
        foamPayCalc5 = float(foamPayGet5) * open_cell6.rate
        foamPayCalc6 = float(foamPayGet6) * open_cell8.rate
        foamPayCalc7 = float(foamPayGet7)
        foamPayCalc8 = float(foamPayGet8)

        splitPay0 = round(float(foamPayCalc0) / float(self.numWorkers.get()), 2)
        splitPay1 = round(float(foamPayCalc1) / float(self.numWorkers.get()), 2)
        splitPay2 = round(float(foamPayCalc2 / float(self.numWorkers.get())), 2)
        splitPay3 = round(float(foamPayCalc3) / float(self.numWorkers.get()), 2)
        splitPay4 = round(float(foamPayCalc4) / float(self.numWorkers.get()), 2)
        splitPay5 = round(float(foamPayCalc5) / float(self.numWorkers.get()), 2)
        splitPay6 = round(float(foamPayCalc6) / float(self.numWorkers.get()), 2)
        splitPay7 = round(float(foamPayCalc7) / float(self.numWorkers.get()), 2)
        splitPay8 = round(float(foamPayCalc8) / float(self.numWorkers.get()), 2)

        workbook = load_workbook("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        if self.emp1Checkbox_var.get() == 1:
            sheet = workbook['Emp1Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, 0, 0])

        if self.emp2Checkbox_var.get() == 1:
            sheet = workbook['Emp2Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, 0, 0])

        if self.emp3Checkbox_var.get() == 1:
            sheet = workbook['Emp3Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, 0, 0])

        if self.emp4Checkbox_var.get() == 1:
            sheet = workbook['Emp4Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, 0, 0])

        if self.emp5Checkbox_var.get() == 1:
            sheet = workbook['Emp5Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, 0, 0])

        if self.emp6Checkbox_var.get() == 1:
            sheet = workbook['Emp6Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, 0, 0])

        if self.emp7Checkbox_var.get() == 1:
            sheet = workbook['Emp7Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, 0, 0])

        if self.emp8Checkbox_var.get() == 1:
            sheet = workbook['Emp8Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, 0, 0])

        if self.emp9Checkbox_var.get() == 1:
            sheet = workbook['Emp9Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, 0, 0])

        if self.emp10Checkbox_var.get() == 1:
            sheet = workbook['Emp10Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          splitPay0, splitPay1, splitPay2, splitPay3, splitPay4, splitPay5,
                          splitPay6, splitPay7, splitPay8, 0, 0])

        workbook.save("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        workbook.close()

    def emp1Checkbox_command(self):
        print(employee1.fullname + " selected")
        self.emp1Checkbox_var.set(1)
        self.count += 1

    def emp2Checkbox_command(self):
        print(employee2.fullname + " selected")
        self.emp2Checkbox_var.set(1)
        self.count += 1

    def emp3Checkbox_command(self):
        print(employee3.fullname + " selected")
        self.emp3Checkbox_var.set(1)
        self.count += 1

    def emp4Checkbox_command(self):
        print(employee4.fullname + " selected")
        self.emp4Checkbox_var.set(1)
        self.count += 1

    def emp5Checkbox_command(self):
        print(employee5.fullname + " selected")
        self.emp5Checkbox_var.set(1)
        self.count += 1

    def emp6Checkbox_command(self):
        print(employee6.fullname + " selected")
        self.emp6Checkbox_var.set(1)
        self.count += 1

    def emp7Checkbox_command(self):
        print(employee7.fullname + " selected")
        self.emp7Checkbox_var.set(1)
        self.count += 1

    def emp8Checkbox_command(self):
        print(employee8.fullname + " selected")
        self.emp8Checkbox_var.set(1)
        self.count += 1

    def emp9Checkbox_command(self):
        print(employee9.fullname + " selected")
        self.emp9Checkbox_var.set(1)
        self.count += 1

    def emp10Checkbox_command(self):
        print(employee10.fullname + " selected")
        self.emp10Checkbox_var.set(1)
        self.count += 1


class MultipleDayPaySheet(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x300")
        self.title("Multiple Day Pay Sheet")
        # setting window size
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)

        self.emp1Checkbox_var = tk.IntVar(value=0)
        self.emp2Checkbox_var = tk.IntVar(value=0)
        self.emp3Checkbox_var = tk.IntVar(value=0)
        self.emp4Checkbox_var = tk.IntVar(value=0)
        self.emp5Checkbox_var = tk.IntVar(value=0)
        self.emp6Checkbox_var = tk.IntVar(value=0)
        self.emp7Checkbox_var = tk.IntVar(value=0)
        self.emp8Checkbox_var = tk.IntVar(value=0)
        self.emp9Checkbox_var = tk.IntVar(value=0)
        self.emp10Checkbox_var = tk.IntVar(value=0)

        emp1Checkbox = ctk.CTkCheckBox(self,
                                       text=employee1.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp1Checkbox_command,
                                       variable=self.emp1Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp1Checkbox.place(x=5, y=30)

        emp2Checkbox = ctk.CTkCheckBox(self,
                                       text=employee2.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp2Checkbox_command,
                                       variable=self.emp2Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp2Checkbox.place(x=5, y=65)

        emp3Checkbox = ctk.CTkCheckBox(self,
                                       text=employee3.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp3Checkbox_command,
                                       variable=self.emp3Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp3Checkbox.place(x=5, y=100)

        emp4Checkbox = ctk.CTkCheckBox(self,
                                       text=employee4.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp4Checkbox_command,
                                       variable=self.emp4Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp4Checkbox.place(x=5, y=135)

        emp5Checkbox = ctk.CTkCheckBox(self,
                                       text=employee5.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp5Checkbox_command,
                                       variable=self.emp5Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp5Checkbox.place(x=5, y=170)

        emp6Checkbox = ctk.CTkCheckBox(self,
                                       text=employee6.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp6Checkbox_command,
                                       variable=self.emp6Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp6Checkbox.place(x=5, y=205)

        emp7Checkbox = ctk.CTkCheckBox(self,
                                       text=employee7.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp7Checkbox_command,
                                       variable=self.emp7Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp7Checkbox.place(x=5, y=240)

        emp8Checkbox = ctk.CTkCheckBox(self,
                                       text=employee8.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp8Checkbox_command,
                                       variable=self.emp8Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp8Checkbox.place(x=5, y=275)

        emp9Checkbox = ctk.CTkCheckBox(self,
                                       text=employee9.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp9Checkbox_command,
                                       variable=self.emp9Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp9Checkbox.place(x=5, y=310)

        emp10Checkbox = ctk.CTkCheckBox(self,
                                        text=employee10.fullname,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.emp10Checkbox_command,
                                        variable=self.emp10Checkbox_var,
                                        onvalue=1,
                                        offvalue=0)
        emp10Checkbox.place(x=5, y=345)

        amountDaysLabel = ctk.CTkLabel(self,
                                       text="Amount of Days",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        amountDaysLabel.place(x=180, y=0)

        self.emp1Days = ctk.CTkEntry(self,
                                     width=100,
                                     height=30,
                                     font=ctk.CTkFont(size=18, weight="normal"))

        self.emp1Days.place(x=180, y=30)
        self.emp1Days.insert(0, '0')

        self.emp2Days = ctk.CTkEntry(self,
                                     width=100,
                                     height=30,
                                     font=ctk.CTkFont(size=18, weight="normal"))

        self.emp2Days.place(x=180, y=65)
        self.emp2Days.insert(0, '0')

        self.emp3Days = ctk.CTkEntry(self,
                                     width=100,
                                     height=30,
                                     font=ctk.CTkFont(size=18, weight="normal"))

        self.emp3Days.place(x=180, y=100)
        self.emp3Days.insert(0, '0')

        self.emp4Days = ctk.CTkEntry(self,
                                     width=100,
                                     height=30,
                                     font=ctk.CTkFont(size=18, weight="normal"))

        self.emp4Days.place(x=180, y=135)
        self.emp4Days.insert(0, '0')

        self.emp5Days = ctk.CTkEntry(self,
                                     width=100,
                                     height=30,
                                     font=ctk.CTkFont(size=18, weight="normal"))

        self.emp5Days.place(x=180, y=170)
        self.emp5Days.insert(0, '0')

        self.emp6Days = ctk.CTkEntry(self,
                                     width=100,
                                     height=30,
                                     font=ctk.CTkFont(size=18, weight="normal"))

        self.emp6Days.place(x=180, y=205)
        self.emp6Days.insert(0, '0')

        self.emp7Days = ctk.CTkEntry(self,
                                     width=100,
                                     height=30,
                                     font=ctk.CTkFont(size=18, weight="normal"))

        self.emp7Days.place(x=180, y=240)
        self.emp7Days.insert(0, '0')

        self.emp8Days = ctk.CTkEntry(self,
                                     width=100,
                                     height=30,
                                     font=ctk.CTkFont(size=18, weight="normal"))

        self.emp8Days.place(x=180, y=275)
        self.emp8Days.insert(0, '0')

        self.emp9Days = ctk.CTkEntry(self,
                                     width=100,
                                     height=30,
                                     font=ctk.CTkFont(size=18, weight="normal"))

        self.emp9Days.place(x=180, y=310)
        self.emp9Days.insert(0, '0')

        self.emp10Days = ctk.CTkEntry(self,
                                      width=100,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"))

        self.emp10Days.place(x=180, y=345)
        self.emp10Days.insert(0, '0')

        payPerDayLabel = ctk.CTkLabel(self,
                                      text="Pay Per Day",
                                      width=150,
                                      height=30,
                                      fg_color="transparent",
                                      font=ctk.CTkFont(size=18, weight="normal"))

        payPerDayLabel.place(x=330, y=0)

        self.emp1PayPer = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.emp1PayPer.place(x=330, y=30)
        self.emp1PayPer.configure(state='disabled')

        self.emp2PayPer = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.emp2PayPer.place(x=330, y=65)
        self.emp2PayPer.configure(state='disabled')

        self.emp3PayPer = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.emp3PayPer.place(x=330, y=100)
        self.emp3PayPer.configure(state='disabled')

        self.emp4PayPer = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.emp4PayPer.place(x=330, y=135)
        self.emp4PayPer.configure(state='disabled')

        self.emp5PayPer = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.emp5PayPer.place(x=330, y=170)
        self.emp5PayPer.configure(state='disabled')

        self.emp6PayPer = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.emp6PayPer.place(x=330, y=205)
        self.emp6PayPer.configure(state='disabled')

        self.emp7PayPer = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.emp7PayPer.place(x=330, y=240)
        self.emp7PayPer.configure(state='disabled')

        self.emp8PayPer = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.emp8PayPer.place(x=330, y=275)
        self.emp8PayPer.configure(state='disabled')

        self.emp9PayPer = ctk.CTkEntry(self,
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"))

        self.emp9PayPer.place(x=330, y=310)
        self.emp9PayPer.configure(state='disabled')

        self.emp10PayPer = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp10PayPer.place(x=330, y=345)
        self.emp10PayPer.configure(state='disabled')

        jobNameLabel = ctk.CTkLabel(self,
                                    text="Job Name",
                                    width=100,
                                    height=30,
                                    fg_color="transparent",
                                    font=ctk.CTkFont(size=18, weight="normal"))

        jobNameLabel.place(x=460, y=30)

        self.jobName = ctk.CTkEntry(self,
                                    width=100,
                                    height=30,
                                    font=ctk.CTkFont(size=18, weight="normal"))

        self.jobName.place(x=575, y=30)

        jobTotalPayLabel = ctk.CTkLabel(self,
                                        text="Job Total Pay",
                                        width=100,
                                        height=30,
                                        fg_color="transparent",
                                        font=ctk.CTkFont(size=18, weight="normal"))

        jobTotalPayLabel.place(x=460, y=65)

        self.jobTotalPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.jobTotalPay.place(x=575, y=65)

        PayPerDayLabel = ctk.CTkLabel(self,
                                      text="Pay Per Day",
                                      width=100,
                                      height=30,
                                      fg_color="transparent",
                                      font=ctk.CTkFont(size=18, weight="normal"))

        PayPerDayLabel.place(x=460, y=100)

        self.PayPerDay = ctk.CTkEntry(self,
                                      width=100,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"))

        self.PayPerDay.place(x=575, y=100)
        self.PayPerDay.configure(state='disabled')

        calculateButton = ctk.CTkButton(self,
                                        text="Calculate",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.calculateButton_command)

        calculateButton.place(x=10, y=420)

        resetFormButton = ctk.CTkButton(self,
                                        text="Reset Form",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.resetFormButton_command)

        resetFormButton.place(x=10, y=455)

        saveFileButton = ctk.CTkButton(self,
                                       text="Save File",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.saveFileButton_command)

        saveFileButton.place(x=10, y=490)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=10, y=525)

    def calculateButton_command(self):
        print("calculating Split Work Order")
        self.PayPerDay.configure(state='normal')

        nameFile = str(self.jobName.get())
        getPayTotal = float(self.jobTotalPay.get())

        emp1DayValue = float(self.emp1Days.get())
        emp2DayValue = float(self.emp2Days.get())
        emp3DayValue = float(self.emp3Days.get())
        emp4DayValue = float(self.emp4Days.get())
        emp5DayValue = float(self.emp5Days.get())
        emp6DayValue = float(self.emp6Days.get())
        emp7DayValue = float(self.emp7Days.get())
        emp8DayValue = float(self.emp8Days.get())
        emp9DayValue = float(self.emp9Days.get())
        emp10DayValue = float(self.emp10Days.get())

        totalManDays = emp1DayValue + emp2DayValue + emp3DayValue + emp4DayValue + emp5DayValue + emp6DayValue + \
                       emp7DayValue + emp8DayValue + emp9DayValue + emp10DayValue

        payPerDay = round(getPayTotal / totalManDays, 2)

        self.PayPerDay.insert(0, "$" + str(payPerDay))

        if self.emp1Checkbox_var.get() == 1:
            self.emp1PayPer.configure(state='normal')
            emp1Pay = round(payPerDay * emp1DayValue, 2)
            self.emp1PayPer.insert(0, '$' + str(emp1Pay))

        if self.emp2Checkbox_var.get() == 1:
            self.emp2PayPer.configure(state='normal')
            emp2Pay = round(payPerDay * emp2DayValue, 2)
            self.emp2PayPer.insert(0, '$' + str(emp2Pay))

        if self.emp3Checkbox_var.get() == 1:
            self.emp3PayPer.configure(state='normal')
            emp3Pay = round(payPerDay * emp3DayValue, 2)
            self.emp3PayPer.insert(0, '$' + str(emp3Pay))

        if self.emp4Checkbox_var.get() == 1:
            self.emp4PayPer.configure(state='normal')
            emp4Pay = round(payPerDay * emp4DayValue, 2)
            self.emp4PayPer.insert(0, '$' + str(emp4Pay))

        if self.emp5Checkbox_var.get() == 1:
            self.emp5PayPer.configure(state='normal')
            emp5Pay = round(payPerDay * emp5DayValue, 2)
            self.emp5PayPer.insert(0, '$' + str(emp5Pay))

        if self.emp6Checkbox_var.get() == 1:
            self.emp6PayPer.configure(state='normal')
            emp6Pay = round(payPerDay * emp6DayValue, 2)
            self.emp6PayPer.insert(0, '$' + str(emp6Pay))

        if self.emp7Checkbox_var.get() == 1:
            self.emp7PayPer.configure(state='normal')
            emp7Pay = round(payPerDay * emp7DayValue, 2)
            self.emp7PayPer.insert(0, '$' + str(emp7Pay))

        if self.emp8Checkbox_var.get() == 1:
            self.emp8PayPer.configure(state='normal')
            emp8Pay = round(payPerDay * emp8DayValue, 2)
            self.emp8PayPer.insert(0, '$' + str(emp8Pay))

        if self.emp9Checkbox_var.get() == 1:
            self.emp9PayPer.configure(state='normal')
            emp9Pay = round(payPerDay * emp9DayValue, 2)
            self.emp9PayPer.insert(0, '$' + str(emp9Pay))

        if self.emp10Checkbox_var.get() == 1:
            self.emp10PayPer.configure(state='normal')
            emp10Pay = round(payPerDay * emp10DayValue, 2)
            self.emp10PayPer.insert(0, '$' + str(emp10Pay))

        print("")
        print("Job Name: " + nameFile)
        print("")
        print("Total Job Pay: $" + str(getPayTotal))
        print("")
        print('Employee Days On Job:')
        print("")
        print(employee1.fullname + " = " + str(emp1DayValue) + " day(s)")
        print(employee2.fullname + " = " + str(emp2DayValue) + " day(s)")
        print(employee3.fullname + " = " + str(emp3DayValue) + " day(s)")
        print(employee4.fullname + " = " + str(emp4DayValue) + " day(s)")
        print(employee5.fullname + " = " + str(emp5DayValue) + " day(s)")
        print(employee6.fullname + " = " + str(emp6DayValue) + " day(s)")
        print(employee7.fullname + " = " + str(emp7DayValue) + " day(s)")
        print(employee8.fullname + " = " + str(emp8DayValue) + " day(s)")
        print(employee9.fullname + " = " + str(emp9DayValue) + " day(s)")
        print(employee10.fullname + " = " + str(emp10DayValue) + " day(s)")
        print("")
        print("Total Man Days To Complete Job: " + str(totalManDays))
        print("")
        print("Total Pay Per Day: " + str(payPerDay))
        print("")
        print("Total Pay Per Employee:")
        print(employee1.fullname + " = " + str(self.emp1PayPer.get()))
        print(employee2.fullname + " = " + str(self.emp2PayPer.get()))
        print(employee3.fullname + " = " + str(self.emp3PayPer.get()))
        print(employee4.fullname + " = " + str(self.emp4PayPer.get()))
        print(employee5.fullname + " = " + str(self.emp5PayPer.get()))
        print(employee6.fullname + " = " + str(self.emp6PayPer.get()))
        print(employee7.fullname + " = " + str(self.emp7PayPer.get()))
        print(employee8.fullname + " = " + str(self.emp8PayPer.get()))
        print(employee9.fullname + " = " + str(self.emp9PayPer.get()))
        print(employee10.fullname + " = " + str(self.emp10PayPer.get()))
        print("")

    def resetFormButton_command(self):
        print('Clearing Form')

        self.emp1Days.delete(0, "end")
        self.emp2Days.delete(0, "end")
        self.emp3Days.delete(0, "end")
        self.emp4Days.delete(0, "end")
        self.emp5Days.delete(0, "end")
        self.emp6Days.delete(0, "end")
        self.emp7Days.delete(0, "end")
        self.emp8Days.delete(0, "end")
        self.emp9Days.delete(0, "end")
        self.emp10Days.delete(0, "end")
        self.PayPerDay.delete(0, 'end')
        self.jobTotalPay.delete(0, 'end')
        self.jobName.delete(0, 'end')

        if self.emp1Checkbox_var.get() == 1:
            self.emp1Checkbox_var.set(0)

        if self.emp2Checkbox_var.get() == 1:
            self.emp2Checkbox_var.set(0)

        if self.emp3Checkbox_var.get() == 1:
            self.emp3Checkbox_var.set(0)

        if self.emp4Checkbox_var.get() == 1:
            self.emp4Checkbox_var.set(0)

        if self.emp5Checkbox_var.get() == 1:
            self.emp5Checkbox_var.set(0)

        if self.emp6Checkbox_var.get() == 1:
            self.emp6Checkbox_var.set(0)

        if self.emp7Checkbox_var.get() == 1:
            self.emp7Checkbox_var.set(0)

        if self.emp8Checkbox_var.get() == 1:
            self.emp8Checkbox_var.set(0)

        if self.emp9Checkbox_var.get() == 1:
            self.emp9Checkbox_var.set(0)

        if self.emp10Checkbox_var.get() == 1:
            self.emp10Checkbox_var.set(0)

        self.emp1Days.insert(0, "0")
        self.emp2Days.insert(0, "0")
        self.emp3Days.insert(0, "0")
        self.emp4Days.insert(0, "0")
        self.emp5Days.insert(0, "0")
        self.emp6Days.insert(0, "0")
        self.emp7Days.insert(0, "0")
        self.emp8Days.insert(0, "0")
        self.emp9Days.insert(0, "0")
        self.emp10Days.insert(0, "0")

        self.PayPerDay.configure(state='disabled')

        self.emp1PayPer.delete(0, 'end')
        self.emp1PayPer.configure(state='disabled')
        self.emp2PayPer.delete(0, 'end')
        self.emp2PayPer.configure(state='disabled')
        self.emp3PayPer.delete(0, 'end')
        self.emp3PayPer.configure(state='disabled')
        self.emp4PayPer.delete(0, 'end')
        self.emp4PayPer.configure(state='disabled')
        self.emp5PayPer.delete(0, 'end')
        self.emp5PayPer.configure(state='disabled')
        self.emp6PayPer.delete(0, 'end')
        self.emp6PayPer.configure(state='disabled')
        self.emp7PayPer.delete(0, 'end')
        self.emp7PayPer.configure(state='disabled')
        self.emp8PayPer.delete(0, 'end')
        self.emp8PayPer.configure(state='disabled')
        self.emp9PayPer.delete(0, 'end')
        self.emp9PayPer.configure(state='disabled')
        self.emp10PayPer.delete(0, 'end')
        self.emp10PayPer.configure(state='disabled')

    def saveFileButton_command(self):
        self.savePaySheet()
        self.saveWorkRecord()
        self.saveEmployeeRecord()
        self.savePaytreeMulti()
        self.savePayrollSpreadSheet()

        self.resetFormButton_command()

        self.closeWindowButton_command()

        messagebox.showinfo(title='Save', message='Save File Is Successful')

    def closeWindowButton_command(self):
        self.destroy()

    def savePaySheet(self):
        print('Saving Paysheet')
        getPayTotal = float(self.jobTotalPay.get())

        emp1DayValue = float(self.emp1Days.get())
        emp2DayValue = float(self.emp2Days.get())
        emp3DayValue = float(self.emp3Days.get())
        emp4DayValue = float(self.emp4Days.get())
        emp5DayValue = float(self.emp5Days.get())
        emp6DayValue = float(self.emp6Days.get())
        emp7DayValue = float(self.emp7Days.get())
        emp8DayValue = float(self.emp8Days.get())
        emp9DayValue = float(self.emp9Days.get())
        emp10DayValue = float(self.emp10Days.get())

        totalManDays = emp1DayValue + emp2DayValue + emp3DayValue + emp4DayValue + emp5DayValue + emp6DayValue + \
                       emp7DayValue + emp8DayValue + emp9DayValue + emp10DayValue

        payPerDay = round(getPayTotal / totalManDays, 2)

        emp1Pay = round(payPerDay * emp1DayValue, 2)
        emp2Pay = round(payPerDay * emp2DayValue, 2)
        emp3Pay = round(payPerDay * emp3DayValue, 2)
        emp4Pay = round(payPerDay * emp4DayValue, 2)
        emp5Pay = round(payPerDay * emp5DayValue, 2)
        emp6Pay = round(payPerDay * emp6DayValue, 2)
        emp7Pay = round(payPerDay * emp7DayValue, 2)
        emp8Pay = round(payPerDay * emp8DayValue, 2)
        emp9Pay = round(payPerDay * emp9DayValue, 2)
        emp10Pay = round(payPerDay * emp10DayValue, 2)

        source = "D:/Insulpro/system/spreadsheet/Batt Pay Sheet.xlsx"

        destination = ("D:/Insulpro/temp/paysheets/batt pay sheet/Batt Pay "
                       "Sheet.xlsx")

        shutil.copy(source, destination)

        nameFile = self.jobName.get()

        filename = str(nameFile + "-MDS" + ".xlsx")

        os.rename("D:/Insulpro/temp/paysheets/batt pay sheet/Batt Pay Sheet.xlsx",
                  "D:/Insulpro/temp/paysheets/batt pay sheet/" + filename)

        workbook = load_workbook(
            "D:/Insulpro/temp/paysheets/batt pay sheet/" + filename)

        sheet = workbook.active

        jobNameSheet = str(nameFile)

        sheet["A2"] = jobNameSheet
        sheet['A5'] = employee1.fullname
        sheet['A6'] = employee2.fullname
        sheet['A7'] = employee3.fullname
        sheet['A8'] = employee4.fullname
        sheet['A9'] = employee5.fullname
        sheet['A10'] = employee6.fullname
        sheet['A11'] = employee7.fullname
        sheet['A12'] = employee8.fullname
        sheet['A13'] = employee9.fullname
        sheet['A14'] = employee10.fullname

        if self.emp1Checkbox_var.get() == 1:
            sheet['H5'] = emp1Pay

        if self.emp2Checkbox_var.get() == 1:
            sheet['H6'] = emp2Pay

        if self.emp3Checkbox_var.get() == 1:
            sheet['H7'] = emp3Pay

        if self.emp4Checkbox_var.get() == 1:
            sheet['H8'] = emp4Pay

        if self.emp5Checkbox_var.get() == 1:
            sheet['H9'] = emp5Pay

        if self.emp6Checkbox_var.get() == 1:
            sheet['H10'] = emp6Pay

        if self.emp7Checkbox_var.get() == 1:
            sheet['H11'] = emp7Pay

        if self.emp8Checkbox_var.get() == 1:
            sheet['H12'] = emp8Pay

        if self.emp9Checkbox_var.get() == 1:
            sheet['H13'] = emp9Pay

        if self.emp10Checkbox_var.get() == 1:
            sheet['H14'] = emp10Pay

        workbook.save("D:/Insulpro/temp/paysheets/batt pay sheet/" + filename)

    def saveWorkRecord(self):
        print('Saving Work Record')
        nameFile = self.jobName.get()
        getPayTotal = float(self.jobTotalPay.get())

        emp1DayValue = float(self.emp1Days.get())
        emp2DayValue = float(self.emp2Days.get())
        emp3DayValue = float(self.emp3Days.get())
        emp4DayValue = float(self.emp4Days.get())
        emp5DayValue = float(self.emp5Days.get())
        emp6DayValue = float(self.emp6Days.get())
        emp7DayValue = float(self.emp7Days.get())
        emp8DayValue = float(self.emp8Days.get())
        emp9DayValue = float(self.emp9Days.get())
        emp10DayValue = float(self.emp10Days.get())

        totalManDays = emp1DayValue + emp2DayValue + emp3DayValue + emp4DayValue + emp5DayValue + emp6DayValue + \
                       emp7DayValue + emp8DayValue + emp9DayValue + emp10DayValue

        workbook = load_workbook("D:/Insulpro/payroll records/Work Record.xlsx")

        sheet = workbook['MultipleDay']

        sheet.append([nameFile, getPayTotal, totalManDays])

        workbook.save("D:/Insulpro/payroll records/Work Record.xlsx")

    def savePaytreeMulti(self):
        empFile1 = str(employee1.fullname) + "/multi record.xlsx"
        empFile2 = str(employee2.fullname) + "/multi record.xlsx"
        empFile3 = str(employee3.fullname) + "/multi record.xlsx"
        empFile4 = str(employee4.fullname) + "/multi record.xlsx"
        empFile5 = str(employee5.fullname) + "/multi record.xlsx"
        empFile6 = str(employee6.fullname) + "/multi record.xlsx"
        empFile7 = str(employee7.fullname) + "/multi record.xlsx"
        empFile8 = str(employee8.fullname) + "/multi record.xlsx"
        empFile9 = str(employee9.fullname) + "/multi record.xlsx"
        empFile10 = str(employee10.fullname) + "/multi record.xlsx"

        namefile = self.jobName.get()
        getPayTotal = float(self.jobTotalPay.get())

        emp1DayValue = float(self.emp1Days.get())
        emp2DayValue = float(self.emp2Days.get())
        emp3DayValue = float(self.emp3Days.get())
        emp4DayValue = float(self.emp4Days.get())
        emp5DayValue = float(self.emp5Days.get())
        emp6DayValue = float(self.emp6Days.get())
        emp7DayValue = float(self.emp7Days.get())
        emp8DayValue = float(self.emp8Days.get())
        emp9DayValue = float(self.emp9Days.get())
        emp10DayValue = float(self.emp10Days.get())

        totalManDays = emp1DayValue + emp2DayValue + emp3DayValue + emp4DayValue + emp5DayValue + emp6DayValue + \
                       emp7DayValue + emp8DayValue + emp9DayValue + emp10DayValue

        payPerDay = round(getPayTotal / totalManDays, 2)

        emp1Pay = round(payPerDay * emp1DayValue, 2)
        emp2Pay = round(payPerDay * emp2DayValue, 2)
        emp3Pay = round(payPerDay * emp3DayValue, 2)
        emp4Pay = round(payPerDay * emp4DayValue, 2)
        emp5Pay = round(payPerDay * emp5DayValue, 2)
        emp6Pay = round(payPerDay * emp6DayValue, 2)
        emp7Pay = round(payPerDay * emp7DayValue, 2)
        emp8Pay = round(payPerDay * emp8DayValue, 2)
        emp9Pay = round(payPerDay * emp9DayValue, 2)
        emp10Pay = round(payPerDay * emp10DayValue, 2)

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile1)

            sheet = workbook['Multi Work Record']

            sheet.append([namefile, emp1Pay, emp1DayValue, payPerDay, totalManDays, getPayTotal])

            workbook.save("D:/Insulpro/employee paytree/" + empFile1)

            workbook.close()

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile2)

            sheet = workbook['Multi Work Record']

            sheet.append([namefile, emp2Pay, emp2DayValue, payPerDay, totalManDays, getPayTotal])

            workbook.save("D:/Insulpro/employee paytree/" + empFile2)

            workbook.close()

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile3)

            sheet = workbook['Multi Work Record']

            sheet.append([namefile, emp3Pay, emp3DayValue, payPerDay, totalManDays, getPayTotal])

            workbook.save("D:/Insulpro/employee paytree/" + empFile3)

            workbook.close()

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile4)

            sheet = workbook['Multi Work Record']

            sheet.append([namefile, emp4Pay, emp4DayValue, payPerDay, totalManDays, getPayTotal])

            workbook.save("D:/Insulpro/employee paytree/" + empFile4)

            workbook.close()

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile5)

            sheet = workbook['Multi Work Record']

            sheet.append([namefile, emp5Pay, emp5DayValue, payPerDay, totalManDays, getPayTotal])

            workbook.save("D:/Insulpro/employee paytree/" + empFile5)

            workbook.close()

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile6)

            sheet = workbook['Multi Work Record']

            sheet.append([namefile, emp6Pay, emp6DayValue, payPerDay, totalManDays, getPayTotal])

            workbook.save("D:/Insulpro/employee paytree/" + empFile6)

            workbook.close()

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile7)

            sheet = workbook['Multi Work Record']

            sheet.append([namefile, emp7Pay, emp7DayValue, payPerDay, totalManDays, getPayTotal])

            workbook.save("D:/Insulpro/employee paytree/" + empFile7)

            workbook.close()

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile8)

            sheet = workbook['Multi Work Record']

            sheet.append([namefile, emp8Pay, emp8DayValue, payPerDay, totalManDays, getPayTotal])

            workbook.save("D:/Insulpro/employee paytree/" + empFile8)

            workbook.close()

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile9)

            sheet = workbook['Multi Work Record']

            sheet.append([namefile, emp9Pay, emp9DayValue, payPerDay, totalManDays, getPayTotal])

            workbook.save("D:/Insulpro/employee paytree/" + empFile9)

            workbook.close()

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile10)

            sheet = workbook['Multi Work Record']

            sheet.append([namefile, emp10Pay, emp10DayValue, payPerDay, totalManDays, getPayTotal])

            workbook.save("D:/Insulpro/employee paytree/" + empFile10)

            workbook.close()

    def saveEmployeeRecord(self):
        print('Saving Employee Records')
        empFile1 = str(employee1.fullname + '.xlsx')
        empFile2 = str(employee2.fullname + '.xlsx')
        empFile3 = str(employee3.fullname + '.xlsx')
        empFile4 = str(employee4.fullname + '.xlsx')
        empFile5 = str(employee5.fullname + '.xlsx')
        empFile6 = str(employee6.fullname + '.xlsx')
        empFile7 = str(employee7.fullname + '.xlsx')
        empFile8 = str(employee8.fullname + '.xlsx')
        empFile9 = str(employee9.fullname + '.xlsx')
        empFile10 = str(employee10.fullname + '.xlsx')

        namefile = str(self.jobName.get() + '-MDS')
        getPayTotal = float(self.jobTotalPay.get())

        emp1DayValue = float(self.emp1Days.get())
        emp2DayValue = float(self.emp2Days.get())
        emp3DayValue = float(self.emp3Days.get())
        emp4DayValue = float(self.emp4Days.get())
        emp5DayValue = float(self.emp5Days.get())
        emp6DayValue = float(self.emp6Days.get())
        emp7DayValue = float(self.emp7Days.get())
        emp8DayValue = float(self.emp8Days.get())
        emp9DayValue = float(self.emp9Days.get())
        emp10DayValue = float(self.emp10Days.get())

        totalManDays = emp1DayValue + emp2DayValue + emp3DayValue + emp4DayValue + emp5DayValue + emp6DayValue + \
                       emp7DayValue + emp8DayValue + emp9DayValue + emp10DayValue

        payPerDay = round(getPayTotal / totalManDays, 2)

        emp1Pay = round(payPerDay * emp1DayValue, 2)
        emp2Pay = round(payPerDay * emp2DayValue, 2)
        emp3Pay = round(payPerDay * emp3DayValue, 2)
        emp4Pay = round(payPerDay * emp4DayValue, 2)
        emp5Pay = round(payPerDay * emp5DayValue, 2)
        emp6Pay = round(payPerDay * emp6DayValue, 2)
        emp7Pay = round(payPerDay * emp7DayValue, 2)
        emp8Pay = round(payPerDay * emp8DayValue, 2)
        emp9Pay = round(payPerDay * emp9DayValue, 2)
        emp10Pay = round(payPerDay * emp10DayValue, 2)

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile1)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp1Pay, 0])

            workbook.save("D:/Insulpro/employee work records/" + empFile1)

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile2)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp2Pay, 0])

            workbook.save("D:/Insulpro/employee work records/" + empFile2)

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile3)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp3Pay, 0])

            workbook.save("D:/Insulpro/employee work records/" + empFile3)

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile4)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp4Pay, 0])

            workbook.save("D:/Insulpro/employee work records/" + empFile4)

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile5)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp5Pay, 0])

            workbook.save("D:/Insulpro/employee work records/" + empFile5)

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile6)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp6Pay, 0])

            workbook.save("D:/Insulpro/employee work records/" + empFile6)

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile7)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp7Pay, 0])

            workbook.save("D:/Insulpro/employee work records/" + empFile7)

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile8)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp8Pay, 0])

            workbook.save("D:/Insulpro/employee work records/" + empFile8)

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile9)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp9Pay, 0])

            workbook.save("D:/Insulpro/employee work records/" + empFile9)

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile10)

            sheet = workbook['Batt Work Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp10Pay, 0])

            workbook.save("D:/Insulpro/employee work records/" + empFile10)

    def savePayrollSpreadSheet(self):
        print('Saving Payroll Spreadsheet')
        name = self.jobName.get()
        namefile = name + '-MDS'
        getPayTotal = float(self.jobTotalPay.get())

        emp1DayValue = float(self.emp1Days.get())
        emp2DayValue = float(self.emp2Days.get())
        emp3DayValue = float(self.emp3Days.get())
        emp4DayValue = float(self.emp4Days.get())
        emp5DayValue = float(self.emp5Days.get())
        emp6DayValue = float(self.emp6Days.get())
        emp7DayValue = float(self.emp7Days.get())
        emp8DayValue = float(self.emp8Days.get())
        emp9DayValue = float(self.emp9Days.get())
        emp10DayValue = float(self.emp10Days.get())

        totalManDays = emp1DayValue + emp2DayValue + emp3DayValue + emp4DayValue + emp5DayValue + emp6DayValue + \
                       emp7DayValue + emp8DayValue + emp9DayValue + emp10DayValue

        payPerDay = round(getPayTotal / totalManDays, 2)

        emp1Pay = round(payPerDay * emp1DayValue, 2)
        emp2Pay = round(payPerDay * emp2DayValue, 2)
        emp3Pay = round(payPerDay * emp3DayValue, 2)
        emp4Pay = round(payPerDay * emp4DayValue, 2)
        emp5Pay = round(payPerDay * emp5DayValue, 2)
        emp6Pay = round(payPerDay * emp6DayValue, 2)
        emp7Pay = round(payPerDay * emp7DayValue, 2)
        emp8Pay = round(payPerDay * emp8DayValue, 2)
        emp9Pay = round(payPerDay * emp9DayValue, 2)
        emp10Pay = round(payPerDay * emp10DayValue, 2)

        workbook = load_workbook("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        if self.emp1Checkbox_var.get() == 1:
            sheet = workbook['Emp1Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp1Pay, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp2Checkbox_var.get() == 1:
            sheet = workbook['Emp2Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp2Pay, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp3Checkbox_var.get() == 1:
            sheet = workbook['Emp3Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp3Pay, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp4Checkbox_var.get() == 1:
            sheet = workbook['Emp4Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp4Pay, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp5Checkbox_var.get() == 1:
            sheet = workbook['Emp5Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp5Pay, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp6Checkbox_var.get() == 1:
            sheet = workbook['Emp6Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp6Pay, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp7Checkbox_var.get() == 1:
            sheet = workbook['Emp7Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp7Pay, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp8Checkbox_var.get() == 1:
            sheet = workbook['Emp8Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp8Pay, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp9Checkbox_var.get() == 1:
            sheet = workbook['Emp9Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp9Pay, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        if self.emp10Checkbox_var.get() == 1:
            sheet = workbook['Emp10Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, emp10Pay, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

        workbook.save("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        workbook.close()

    def emp1Checkbox_command(self):
        print(employee1.fullname + " selected")
        self.emp1Checkbox_var.set(1)

    def emp2Checkbox_command(self):
        print(employee2.fullname + " selected")
        self.emp2Checkbox_var.set(1)

    def emp3Checkbox_command(self):
        print(employee3.fullname + " selected")
        self.emp3Checkbox_var.set(1)

    def emp4Checkbox_command(self):
        print(employee4.fullname + " selected")
        self.emp4Checkbox_var.set(1)

    def emp5Checkbox_command(self):
        print(employee5.fullname + " selected")
        self.emp5Checkbox_var.set(1)

    def emp6Checkbox_command(self):
        print(employee6.fullname + " selected")
        self.emp6Checkbox_var.set(1)

    def emp7Checkbox_command(self):
        print(employee7.fullname + " selected")
        self.emp7Checkbox_var.set(1)

    def emp8Checkbox_command(self):
        print(employee8.fullname + " selected")
        self.emp8Checkbox_var.set(1)

    def emp9Checkbox_command(self):
        print(employee9.fullname + " selected")
        self.emp9Checkbox_var.set(1)

    def emp10Checkbox_command(self):
        print(employee10.fullname + " selected")
        self.emp10Checkbox_var.set(1)


class VacationRequest(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x300")
        self.title("Vacation Request")
        # setting window size
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)

        self.count = 0

        self.emp1Checkbox_var = tk.IntVar(value=0)
        self.emp2Checkbox_var = tk.IntVar(value=0)
        self.emp3Checkbox_var = tk.IntVar(value=0)
        self.emp4Checkbox_var = tk.IntVar(value=0)
        self.emp5Checkbox_var = tk.IntVar(value=0)
        self.emp6Checkbox_var = tk.IntVar(value=0)
        self.emp7Checkbox_var = tk.IntVar(value=0)
        self.emp8Checkbox_var = tk.IntVar(value=0)
        self.emp9Checkbox_var = tk.IntVar(value=0)
        self.emp10Checkbox_var = tk.IntVar(value=0)

        emp1Checkbox = ctk.CTkCheckBox(self,
                                       text=employee1.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp1Checkbox_command,
                                       variable=self.emp1Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp1Checkbox.place(x=5, y=30)

        emp2Checkbox = ctk.CTkCheckBox(self,
                                       text=employee2.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp2Checkbox_command,
                                       variable=self.emp2Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp2Checkbox.place(x=5, y=65)

        emp3Checkbox = ctk.CTkCheckBox(self,
                                       text=employee3.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp3Checkbox_command,
                                       variable=self.emp3Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp3Checkbox.place(x=5, y=100)

        emp4Checkbox = ctk.CTkCheckBox(self,
                                       text=employee4.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp4Checkbox_command,
                                       variable=self.emp4Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp4Checkbox.place(x=5, y=135)

        emp5Checkbox = ctk.CTkCheckBox(self,
                                       text=employee5.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp5Checkbox_command,
                                       variable=self.emp5Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp5Checkbox.place(x=5, y=170)

        emp6Checkbox = ctk.CTkCheckBox(self,
                                       text=employee6.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp6Checkbox_command,
                                       variable=self.emp6Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp6Checkbox.place(x=5, y=205)

        emp7Checkbox = ctk.CTkCheckBox(self,
                                       text=employee7.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp7Checkbox_command,
                                       variable=self.emp7Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp7Checkbox.place(x=5, y=240)

        emp8Checkbox = ctk.CTkCheckBox(self,
                                       text=employee8.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp8Checkbox_command,
                                       variable=self.emp8Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp8Checkbox.place(x=5, y=275)

        emp9Checkbox = ctk.CTkCheckBox(self,
                                       text=employee9.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp9Checkbox_command,
                                       variable=self.emp9Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp9Checkbox.place(x=5, y=310)

        emp10Checkbox = ctk.CTkCheckBox(self,
                                        text=employee10.fullname,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.emp10Checkbox_command,
                                        variable=self.emp10Checkbox_var,
                                        onvalue=1,
                                        offvalue=0)
        emp10Checkbox.place(x=5, y=345)

        daysRequestedLabel = ctk.CTkLabel(self,
                                          text="Amount of Requested Days",
                                          width=200,
                                          height=30,
                                          fg_color="transparent",
                                          font=ctk.CTkFont(size=18, weight="normal"))

        daysRequestedLabel.place(x=210, y=30)

        self.daysRequested = ctk.CTkEntry(self,
                                          width=100,
                                          height=25,
                                          font=ctk.CTkFont(size=18, weight="normal"))

        self.daysRequested.place(x=435, y=30)

        saveFileButton = ctk.CTkButton(self,
                                       text="Save File",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.saveFileButton_command)

        saveFileButton.place(x=210, y=80)

        resetFormButton = ctk.CTkButton(self,
                                        text="Reset Form",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.resetFormButton_command)

        resetFormButton.place(x=210, y=115)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=210, y=150)

    def saveFileButton_command(self):
        print('Requesting Vacation')
        self.savePaytreeVacation()
        self.saveEmployeeRecord()
        self.savePayrollSpreadSheet()
        date = datetime.datetime.now()

        source = "D:/Insulpro/system/spreadsheet/Vacation Request.xlsx"

        destination = "D:/Insulpro/temp/vacation/Vacation Request.xlsx"

        shutil.copy(source, destination)

        if self.emp1Checkbox_var.get() == 1:
            filename = employee1.fullname + '.xlsx'

            os.rename("D:/Insulpro/temp/vacation/Vacation Request.xlsx",
                      "D:/Insulpro/temp/vacation/" + filename)

            workbook = load_workbook("D:/Insulpro/temp/vacation/" + filename)

            sheet = workbook.active

            sheet['B9'] = int(self.daysRequested.get())
            sheet['C7'] = employee1.fullname

            workbook.save("D:/Insulpro/temp/vacation/" + filename)

            workbook1 = load_workbook(
                'D:/Insulpro/payroll records/vacation record.xlsx')

            sheet1 = workbook1['Emp1']

            sheet1.append([date, int(self.daysRequested.get())])

            workbook1.save('D:/Insulpro/payroll records/vacation record.xlsx')

            messagebox.showinfo(title='Save', message='Save File Is Successful')

        if self.emp2Checkbox_var.get() == 1:
            filename = employee2.fullname + '.xlsx'

            os.rename("D:/Insulpro/temp/vacation/Vacation Request.xlsx",
                      "D:/Insulpro/temp/vacation/" + filename)

            workbook = load_workbook("D:/Insulpro/temp/vacation/" + filename)

            sheet = workbook.active

            sheet['B9'] = int(self.daysRequested.get())
            sheet['C7'] = employee2.fullname

            workbook.save("D:/Insulpro/temp/vacation/" + filename)

            workbook1 = load_workbook(
                'D:/Insulpro/payroll records/vacation record.xlsx')

            sheet1 = workbook1['Emp2']

            sheet1.append([date, int(self.daysRequested.get())])

            workbook1.save('D:/Insulpro/payroll records/vacation record.xlsx')

            messagebox.showinfo(title='Save', message='Save File Is Successful')

        if self.emp3Checkbox_var.get() == 1:
            filename = employee3.fullname + '.xlsx'

            os.rename("D:/Insulpro/temp/vacation/Vacation Request.xlsx",
                      "D:/Insulpro/temp/vacation/" + filename)

            workbook = load_workbook("D:/Insulpro/temp/vacation/" + filename)

            sheet = workbook.active

            sheet['B9'] = int(self.daysRequested.get())
            sheet['C7'] = employee3.fullname

            workbook.save("D:/Insulpro/temp/vacation/" + filename)

            workbook1 = load_workbook(
                'D:/Insulpro/payroll records/vacation record.xlsx')

            sheet1 = workbook1['Emp3']

            sheet1.append([date, int(self.daysRequested.get())])

            workbook1.save('D:/Insulpro/payroll records/vacation record.xlsx')

            messagebox.showinfo(title='Save', message='Save File Is Successful')

        if self.emp4Checkbox_var.get() == 1:
            filename = employee4.fullname + '.xlsx'

            os.rename("D:/Insulpro/temp/vacation/Vacation Request.xlsx",
                      "D:/Insulpro/temp/vacation/" + filename)

            workbook = load_workbook("D:/Insulpro/temp/vacation/" + filename)

            sheet = workbook.active

            sheet['B9'] = int(self.daysRequested.get())
            sheet['C7'] = employee4.fullname

            workbook.save("D:/Insulpro/temp/vacation/" + filename)

            workbook1 = load_workbook(
                'D:/Insulpro/payroll records/vacation record.xlsx')

            sheet1 = workbook1['Emp4']

            sheet1.append([date, int(self.daysRequested.get())])

            workbook1.save('D:/Insulpro/payroll records/vacation record.xlsx')

            messagebox.showinfo(title='Save', message='Save File Is Successful')

        if self.emp5Checkbox_var.get() == 1:
            filename = employee5.fullname + '.xlsx'

            os.rename("D:/Insulpro/temp/vacation/Vacation Request.xlsx",
                      "D:/Insulpro/temp/vacation/" + filename)

            workbook = load_workbook("D:/Insulpro/temp/vacation/" + filename)

            sheet = workbook.active

            sheet['B9'] = int(self.daysRequested.get())
            sheet['C7'] = employee5.fullname

            workbook.save("D:/Insulpro/temp/vacation/" + filename)

            workbook1 = load_workbook(
                'D:/Insulpro/payroll records/vacation record.xlsx')

            sheet1 = workbook1['Emp5']

            sheet1.append([date, int(self.daysRequested.get())])

            workbook1.save('D:/Insulpro/payroll records/vacation record.xlsx')

            messagebox.showinfo(title='Save', message='Save File Is Successful')

        if self.emp6Checkbox_var.get() == 1:
            filename = employee6.fullname + '.xlsx'

            os.rename("D:/Insulpro/temp/vacation/Vacation Request.xlsx",
                      "D:/Insulpro/temp/vacation/" + filename)

            workbook = load_workbook("D:/Insulpro/temp/vacation/" + filename)

            sheet = workbook.active

            sheet['B9'] = int(self.daysRequested.get())
            sheet['C7'] = employee6.fullname

            workbook.save("D:/Insulpro/temp/vacation/" + filename)

            workbook1 = load_workbook(
                'D:/Insulpro/payroll records/vacation record.xlsx')

            sheet1 = workbook1['Emp6']

            sheet1.append([date, int(self.daysRequested.get())])

            workbook1.save('D:/Insulpro/payroll records/vacation record.xlsx')

            messagebox.showinfo(title='Save', message='Save File Is Successful')

        if self.emp7Checkbox_var.get() == 1:
            filename = employee7.fullname + '.xlsx'

            os.rename("D:/Insulpro/temp/vacation/Vacation Request.xlsx",
                      "D:/Insulpro/temp/vacation/" + filename)

            workbook = load_workbook("D:/Insulpro/temp/vacation/" + filename)

            sheet = workbook.active

            sheet['B9'] = int(self.daysRequested.get())
            sheet['C7'] = employee7.fullname

            workbook.save("D:/Insulpro/temp/vacation/" + filename)

            workbook1 = load_workbook(
                'D:/Insulpro/payroll records/vacation record.xlsx')

            sheet1 = workbook1['Emp7']

            sheet1.append([date, int(self.daysRequested.get())])

            workbook1.save('D:/Insulpro/payroll records/vacation record.xlsx')

            messagebox.showinfo(title='Save', message='Save File Is Successful')

        if self.emp8Checkbox_var.get() == 1:
            filename = employee8.fullname + '.xlsx'

            os.rename("D:/Insulpro/temp/vacation/Vacation Request.xlsx",
                      "D:/Insulpro/temp/vacation/" + filename)

            workbook = load_workbook("D:/Insulpro/temp/vacation/" + filename)

            sheet = workbook.active

            sheet['B9'] = int(self.daysRequested.get())
            sheet['C7'] = employee8.fullname

            workbook.save("D:/Insulpro/temp/vacation/" + filename)

            workbook1 = load_workbook(
                'D:/Insulpro/payroll records/vacation record.xlsx')

            sheet1 = workbook1['Emp8']

            sheet1.append([date, int(self.daysRequested.get())])

            workbook1.save('D:/Insulpro/payroll records/vacation record.xlsx')

            messagebox.showinfo(title='Save', message='Save File Is Successful')

        if self.emp9Checkbox_var.get() == 1:
            filename = employee9.fullname + '.xlsx'

            os.rename("D:/Insulpro/temp/vacation/Vacation Request.xlsx",
                      "D:/Insulpro/temp/vacation/" + filename)

            workbook = load_workbook("D:/Insulpro/temp/vacation/" + filename)

            sheet = workbook.active

            sheet['B9'] = int(self.daysRequested.get())
            sheet['C7'] = employee9.fullname

            workbook.save("D:/Insulpro/temp/vacation/" + filename)

            workbook1 = load_workbook(
                'D:/Insulpro/payroll records/vacation record.xlsx')

            sheet1 = workbook1['Emp9']

            sheet1.append([date, int(self.daysRequested.get())])

            workbook1.save('D:/Insulpro/payroll records/vacation record.xlsx')

            messagebox.showinfo(title='Save', message='Save File Is Successful')

        if self.emp10Checkbox_var.get() == 1:
            filename = employee10.fullname + '.xlsx'

            os.rename("D:/Insulpro/temp/vacation/Vacation Request.xlsx",
                      "D:/Insulpro/temp/vacation/" + filename)

            workbook = load_workbook("D:/Insulpro/temp/vacation/" + filename)

            sheet = workbook.active

            sheet['B9'] = int(self.daysRequested.get())
            sheet['C7'] = employee10.fullname

            workbook.save("D:/Insulpro/temp/vacation/" + filename)
            workbook1 = load_workbook(
                'D:/Insulpro/payroll records/vacation record.xlsx')

            sheet1 = workbook1['Emp10']

            sheet1.append([date, int(self.daysRequested.get())])

            workbook1.save('D:/Insulpro/payroll records/vacation record.xlsx')

            messagebox.showinfo(title='Save', message='Save File Is Successful')

        self.resetFormButton_command()

        self.closeWindowButton_command()

    def resetFormButton_command(self):
        print('Clearing Form')
        if self.emp1Checkbox_var.get() == 1:
            self.emp1Checkbox_var.set(0)

        if self.emp2Checkbox_var.get() == 1:
            self.emp2Checkbox_var.set(0)

        if self.emp3Checkbox_var.get() == 1:
            self.emp3Checkbox_var.set(0)

        if self.emp4Checkbox_var.get() == 1:
            self.emp4Checkbox_var.set(0)

        if self.emp5Checkbox_var.get() == 1:
            self.emp5Checkbox_var.set(0)

        if self.emp6Checkbox_var.get() == 1:
            self.emp6Checkbox_var.set(0)

        if self.emp7Checkbox_var.get() == 1:
            self.emp7Checkbox_var.set(0)

        if self.emp8Checkbox_var.get() == 1:
            self.emp8Checkbox_var.set(0)

        if self.emp9Checkbox_var.get() == 1:
            self.emp9Checkbox_var.set(0)

        if self.emp10Checkbox_var.get() == 1:
            self.emp10Checkbox_var.set(0)

        self.daysRequested.delete(0, 'end')

    def closeWindowButton_command(self):
        self.destroy()

    def savePaytreeVacation(self):
        empFile1 = str(employee1.fullname) + "/vacation record.xlsx"
        empFile2 = str(employee2.fullname) + "/vacation record.xlsx"
        empFile3 = str(employee3.fullname) + "/vacation record.xlsx"
        empFile4 = str(employee4.fullname) + "/vacation record.xlsx"
        empFile5 = str(employee5.fullname) + "/vacation record.xlsx"
        empFile6 = str(employee6.fullname) + "/vacation record.xlsx"
        empFile7 = str(employee7.fullname) + "/vacation record.xlsx"
        empFile8 = str(employee8.fullname) + "/vacation record.xlsx"
        empFile9 = str(employee9.fullname) + "/vacation record.xlsx"
        empFile10 = str(employee10.fullname) + "/vacation record.xlsx"

        date = datetime.datetime.now()
        vacationDaysTaken = int(self.daysRequested.get())
        vacationPay = vacationDaysTaken * vacation_day.rate

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile1)

            sheet = workbook['Vacation Work Record']

            sheet.append([date, vacationDaysTaken, vacationPay])

            workbook.save("D:/Insulpro/employee paytree/" + empFile1)

            workbook.close()

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile2)

            sheet = workbook['Vacation Work Record']

            sheet.append([date, vacationDaysTaken, vacationPay])

            workbook.save("D:/Insulpro/employee paytree/" + empFile2)

            workbook.close()

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile3)

            sheet = workbook['Vacation Work Record']

            sheet.append([date, vacationDaysTaken, vacationPay])

            workbook.save("D:/Insulpro/employee paytree/" + empFile3)

            workbook.close()

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile4)

            sheet = workbook['Vacation Work Record']

            sheet.append([date, vacationDaysTaken, vacationPay])

            workbook.save("D:/Insulpro/employee paytree/" + empFile4)

            workbook.close()

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile5)

            sheet = workbook['Vacation Work Record']

            sheet.append([date, vacationDaysTaken, vacationPay])

            workbook.save("D:/Insulpro/employee paytree/" + empFile5)

            workbook.close()

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile6)

            sheet = workbook['Vacation Work Record']

            sheet.append([date, vacationDaysTaken, vacationPay])

            workbook.save("D:/Insulpro/employee paytree/" + empFile6)

            workbook.close()

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile7)

            sheet = workbook['Vacation Work Record']

            sheet.append([date, vacationDaysTaken, vacationPay])

            workbook.save("D:/Insulpro/employee paytree/" + empFile7)

            workbook.close()

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile8)

            sheet = workbook['Vacation Work Record']

            sheet.append([date, vacationDaysTaken, vacationPay])

            workbook.save("D:/Insulpro/employee paytree/" + empFile8)

            workbook.close()

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile9)

            sheet = workbook['Vacation Work Record']

            sheet.append([date, vacationDaysTaken, vacationPay])

            workbook.save("D:/Insulpro/employee paytree/" + empFile9)

            workbook.close()

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile10)

            sheet = workbook['Vacation Work Record']

            sheet.append([date, vacationDaysTaken, vacationPay])

            workbook.save("D:/Insulpro/employee paytree/" + empFile10)

            workbook.close()

    def saveEmployeeRecord(self):
        print('Saving Employee Record')
        empFile1 = str(employee1.fullname + '.xlsx')
        empFile2 = str(employee2.fullname + '.xlsx')
        empFile3 = str(employee3.fullname + '.xlsx')
        empFile4 = str(employee4.fullname + '.xlsx')
        empFile5 = str(employee5.fullname + '.xlsx')
        empFile6 = str(employee6.fullname + '.xlsx')
        empFile7 = str(employee7.fullname + '.xlsx')
        empFile8 = str(employee8.fullname + '.xlsx')
        empFile9 = str(employee9.fullname + '.xlsx')
        empFile10 = str(employee10.fullname + '.xlsx')

        date = datetime.datetime.now()
        vacationDaysTaken = int(self.daysRequested.get())

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile1)

            sheet = workbook['Vacation Record']

            sheet.append([date, vacationDaysTaken])

            workbook.save("D:/Insulpro/employee work records/" + empFile1)

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile2)

            sheet = workbook['Vacation Record']

            sheet.append([date, vacationDaysTaken])

            workbook.save("D:/Insulpro/employee work records/" + empFile2)

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile3)

            sheet = workbook['Vacation Record']

            sheet.append([date, vacationDaysTaken])

            workbook.save("D:/Insulpro/employee work records/" + empFile3)

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile4)

            sheet = workbook['Vacation Record']

            sheet.append([date, vacationDaysTaken])

            workbook.save("D:/Insulpro/employee work records/" + empFile4)

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile5)

            sheet = workbook['Vacation Record']

            sheet.append([date, vacationDaysTaken])

            workbook.save("D:/Insulpro/employee work records/" + empFile5)

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile6)

            sheet = workbook['Vacation Record']

            sheet.append([date, vacationDaysTaken])

            workbook.save("D:/Insulpro/employee work records/" + empFile6)

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile7)

            sheet = workbook['Vacation Record']

            sheet.append([date, vacationDaysTaken])

            workbook.save("D:/Insulpro/employee work records/" + empFile7)

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile8)

            sheet = workbook['Vacation Record']

            sheet.append([date, vacationDaysTaken])

            workbook.save("D:/Insulpro/employee work records/" + empFile8)

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile9)

            sheet = workbook['Vacation Record']

            sheet.append([date, vacationDaysTaken])

            workbook.save("D:/Insulpro/employee work records/" + empFile9)

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile10)

            sheet = workbook['Vacation Record']

            sheet.append([date, vacationDaysTaken])
            workbook.save("D:/Insulpro/employee work records/" + empFile10)

    def savePayrollSpreadSheet(self):
        print('Saving Payroll Spreadsheet')
        namefile = 'Vacation Pay'
        vacationDaysTaken = int(self.daysRequested.get())
        vacationPay = vacationDaysTaken * vacation_day.rate

        workbook = load_workbook("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        if self.emp1Checkbox_var.get() == 1:
            sheet = workbook['Emp1Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, vacationPay, 0])

        if self.emp2Checkbox_var.get() == 1:
            sheet = workbook['Emp2Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, vacationPay, 0])

        if self.emp3Checkbox_var.get() == 1:
            sheet = workbook['Emp3Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, vacationPay, 0])

        if self.emp4Checkbox_var.get() == 1:
            sheet = workbook['Emp4Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, vacationPay, 0])

        if self.emp5Checkbox_var.get() == 1:
            sheet = workbook['Emp5Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, vacationPay, 0])

        if self.emp6Checkbox_var.get() == 1:
            sheet = workbook['Emp6Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, vacationPay, 0])

        if self.emp7Checkbox_var.get() == 1:
            sheet = workbook['Emp7Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, vacationPay, 0])

        if self.emp8Checkbox_var.get() == 1:
            sheet = workbook['Emp8Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, vacationPay, 0])

        if self.emp9Checkbox_var.get() == 1:
            sheet = workbook['Emp9Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, vacationPay, 0])

        if self.emp10Checkbox_var.get() == 1:
            sheet = workbook['Emp10Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, vacationPay, 0])

        workbook.save("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        workbook.close()

    def emp1Checkbox_command(self):
        print(employee1.fullname + " selected")
        self.emp1Checkbox_var.set(1)

    def emp2Checkbox_command(self):
        print(employee2.fullname + " selected")
        self.emp2Checkbox_var.set(1)

    def emp3Checkbox_command(self):
        print(employee3.fullname + " selected")
        self.emp3Checkbox_var.set(1)

    def emp4Checkbox_command(self):
        print(employee4.fullname + " selected")
        self.emp4Checkbox_var.set(1)

    def emp5Checkbox_command(self):
        print(employee5.fullname + " selected")
        self.emp5Checkbox_var.set(1)

    def emp6Checkbox_command(self):
        print(employee6.fullname + " selected")
        self.emp6Checkbox_var.set(1)

    def emp7Checkbox_command(self):
        print(employee7.fullname + " selected")
        self.emp7Checkbox_var.set(1)

    def emp8Checkbox_command(self):
        print(employee8.fullname + " selected")
        self.emp8Checkbox_var.set(1)

    def emp9Checkbox_command(self):
        print(employee9.fullname + " selected")
        self.emp9Checkbox_var.set(1)

    def emp10Checkbox_command(self):
        print(employee10.fullname + " selected")
        self.emp10Checkbox_var.set(1)


class ShopPaySheet(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Shop Pay Sheet")
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)

        self.emp1Checkbox_var = tk.IntVar(value=0)
        self.emp2Checkbox_var = tk.IntVar(value=0)
        self.emp3Checkbox_var = tk.IntVar(value=0)
        self.emp4Checkbox_var = tk.IntVar(value=0)
        self.emp5Checkbox_var = tk.IntVar(value=0)
        self.emp6Checkbox_var = tk.IntVar(value=0)
        self.emp7Checkbox_var = tk.IntVar(value=0)
        self.emp8Checkbox_var = tk.IntVar(value=0)
        self.emp9Checkbox_var = tk.IntVar(value=0)
        self.emp10Checkbox_var = tk.IntVar(value=0)

        emp1Checkbox = ctk.CTkCheckBox(self,
                                       text=employee1.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp1Checkbox_command,
                                       variable=self.emp1Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp1Checkbox.place(x=5, y=30)

        emp2Checkbox = ctk.CTkCheckBox(self,
                                       text=employee2.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp2Checkbox_command,
                                       variable=self.emp2Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp2Checkbox.place(x=5, y=65)

        emp3Checkbox = ctk.CTkCheckBox(self,
                                       text=employee3.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp3Checkbox_command,
                                       variable=self.emp3Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp3Checkbox.place(x=5, y=100)

        emp4Checkbox = ctk.CTkCheckBox(self,
                                       text=employee4.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp4Checkbox_command,
                                       variable=self.emp4Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp4Checkbox.place(x=5, y=135)

        emp5Checkbox = ctk.CTkCheckBox(self,
                                       text=employee5.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp5Checkbox_command,
                                       variable=self.emp5Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp5Checkbox.place(x=5, y=170)

        emp6Checkbox = ctk.CTkCheckBox(self,
                                       text=employee6.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp6Checkbox_command,
                                       variable=self.emp6Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp6Checkbox.place(x=5, y=205)

        emp7Checkbox = ctk.CTkCheckBox(self,
                                       text=employee7.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp7Checkbox_command,
                                       variable=self.emp7Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp7Checkbox.place(x=5, y=240)

        emp8Checkbox = ctk.CTkCheckBox(self,
                                       text=employee8.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp8Checkbox_command,
                                       variable=self.emp8Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp8Checkbox.place(x=5, y=275)

        emp9Checkbox = ctk.CTkCheckBox(self,
                                       text=employee9.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp9Checkbox_command,
                                       variable=self.emp9Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp9Checkbox.place(x=5, y=310)

        emp10Checkbox = ctk.CTkCheckBox(self,
                                        text=employee10.fullname,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.emp10Checkbox_command,
                                        variable=self.emp10Checkbox_var,
                                        onvalue=1,
                                        offvalue=0)
        emp10Checkbox.place(x=5, y=345)

        shopPayAmountLabel = ctk.CTkLabel(self,
                                          text="Shop Pay Amount",
                                          width=100,
                                          height=30,
                                          fg_color="transparent",
                                          font=ctk.CTkFont(size=18, weight="normal"))

        shopPayAmountLabel.place(x=180, y=0)

        self.emp1ShopPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp1ShopPay.place(x=180, y=30)
        self.emp1ShopPay.insert(0, '0')

        self.emp2ShopPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp2ShopPay.place(x=180, y=65)
        self.emp2ShopPay.insert(0, '0')

        self.emp3ShopPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp3ShopPay.place(x=180, y=100)
        self.emp3ShopPay.insert(0, '0')

        self.emp4ShopPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp4ShopPay.place(x=180, y=135)
        self.emp4ShopPay.insert(0, '0')

        self.emp5ShopPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp5ShopPay.place(x=180, y=170)
        self.emp5ShopPay.insert(0, '0')

        self.emp6ShopPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp6ShopPay.place(x=180, y=205)
        self.emp6ShopPay.insert(0, '0')

        self.emp7ShopPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp7ShopPay.place(x=180, y=240)
        self.emp7ShopPay.insert(0, '0')

        self.emp8ShopPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp8ShopPay.place(x=180, y=275)
        self.emp8ShopPay.insert(0, '0')

        self.emp9ShopPay = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp9ShopPay.place(x=180, y=310)
        self.emp9ShopPay.insert(0, '0')

        self.emp10ShopPay = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp10ShopPay.place(x=180, y=345)
        self.emp10ShopPay.insert(0, '0')

        dayWeekLabel = ctk.CTkLabel(self,
                                    text="Day Of Week",
                                    width=100,
                                    height=30,
                                    fg_color="transparent",
                                    font=ctk.CTkFont(size=18, weight="normal"))

        dayWeekLabel.place(x=325, y=30)

        self.optionmenu_var = ctk.StringVar(value="Monday")
        self.optionmenu = ctk.CTkOptionMenu(self, values=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
                                                          "Saturday", "Sunday"],
                                            command=self.optionmenu_callback,
                                            variable=self.optionmenu_var)
        self.optionmenu.place(x=435, y=30)

        saveFileButton = ctk.CTkButton(self,
                                       text="Save File",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.saveFileButton_command)

        saveFileButton.place(x=20, y=425)

        resetFormButton = ctk.CTkButton(self,
                                        text="Reset Form",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.resetFormButton_command)

        resetFormButton.place(x=20, y=460)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=20, y=495)

    def saveFileButton_command(self):
        date = datetime.datetime.now()
        dateStamp = date.strftime('%m%d%y%H%M%S')
        nameFile = self.optionmenu.get()
        filename = str(nameFile) + str(dateStamp) + ".xlsx"

        if not os.path.exists(
                "D:/Insulpro/temp/paysheets/shop pay sheet/" + filename):
            print('Saving File...')

            self.savePaySheet()
            self.saveWorkRecord()
            self.savePaytreeShop()
            self.saveEmployeeRecord()
            self.savePayrollSpreadsheet()

            self.resetFormButton_command()
            self.closeWindowButton_command()

            messagebox.showinfo(title='Save', message='File Save Successful')
            print("File was saved successfully")

        else:
            messagebox.showerror(title='File Name Already Exist',
                                 message='Please change job name to something else!')

    def resetFormButton_command(self):
        print('Clearing Form')
        if self.emp1Checkbox_var.get() == 1:
            self.emp1Checkbox_var.set(0)

        if self.emp2Checkbox_var.get() == 1:
            self.emp2Checkbox_var.set(0)

        if self.emp3Checkbox_var.get() == 1:
            self.emp3Checkbox_var.set(0)

        if self.emp4Checkbox_var.get() == 1:
            self.emp4Checkbox_var.set(0)

        if self.emp5Checkbox_var.get() == 1:
            self.emp5Checkbox_var.set(0)

        if self.emp6Checkbox_var.get() == 1:
            self.emp6Checkbox_var.set(0)

        if self.emp7Checkbox_var.get() == 1:
            self.emp7Checkbox_var.set(0)

        if self.emp8Checkbox_var.get() == 1:
            self.emp8Checkbox_var.set(0)

        if self.emp9Checkbox_var.get() == 1:
            self.emp9Checkbox_var.set(0)

        if self.emp10Checkbox_var.get() == 1:
            self.emp10Checkbox_var.set(0)

        self.emp1ShopPay.delete(0, "end")
        self.emp2ShopPay.delete(0, "end")
        self.emp3ShopPay.delete(0, "end")
        self.emp4ShopPay.delete(0, "end")
        self.emp5ShopPay.delete(0, "end")
        self.emp6ShopPay.delete(0, "end")
        self.emp7ShopPay.delete(0, "end")
        self.emp8ShopPay.delete(0, "end")
        self.emp9ShopPay.delete(0, "end")
        self.emp10ShopPay.delete(0, "end")

        self.emp1ShopPay.insert(0, "0")
        self.emp2ShopPay.insert(0, "0")
        self.emp3ShopPay.insert(0, "0")
        self.emp4ShopPay.insert(0, "0")
        self.emp5ShopPay.insert(0, "0")
        self.emp6ShopPay.insert(0, "0")
        self.emp7ShopPay.insert(0, "0")
        self.emp8ShopPay.insert(0, "0")
        self.emp9ShopPay.insert(0, "0")
        self.emp10ShopPay.insert(0, "0")

    def closeWindowButton_command(self):
        self.destroy()

    @staticmethod
    def optionmenu_callback(choice):
        print("optionmenu dropdown clicked:", choice)

    def savePaySheet(self):
        print('Saving Shop Paysheet')
        shopPayGet1 = self.emp1ShopPay.get()
        shopPayGet2 = self.emp2ShopPay.get()
        shopPayGet3 = self.emp3ShopPay.get()
        shopPayGet4 = self.emp4ShopPay.get()
        shopPayGet5 = self.emp5ShopPay.get()
        shopPayGet6 = self.emp6ShopPay.get()
        shopPayGet7 = self.emp7ShopPay.get()
        shopPayGet8 = self.emp8ShopPay.get()
        shopPayGet9 = self.emp9ShopPay.get()
        shopPayGet10 = self.emp10ShopPay.get()

        source = "D:/Insulpro/system/spreadsheet/Shop Pay Sheet.xlsx"

        destination = ("D:/Insulpro/temp/paysheets/shop pay sheet/Shop Pay "
                       "Sheet.xlsx")

        shutil.copy(source, destination)

        date = datetime.datetime.now()
        dateStamp = date.strftime('%m%d%y%H%M%S')
        nameFile = self.optionmenu.get()
        filename = str(nameFile) + str(dateStamp) + ".xlsx"

        os.rename("D:/Insulpro/temp/paysheets/shop pay sheet/Shop Pay Sheet.xlsx",
                  "D:/Insulpro/temp/paysheets/shop pay sheet/" + filename)

        workbook = load_workbook(
            "D:/Insulpro/temp/paysheets/shop pay sheet/" + filename)

        sheet = workbook.active

        sheet["A2"] = str(nameFile)
        sheet['A5'] = employee1.fullname
        sheet['A6'] = employee2.fullname
        sheet['A7'] = employee3.fullname
        sheet['A8'] = employee4.fullname
        sheet['A9'] = employee5.fullname
        sheet['A10'] = employee6.fullname
        sheet['A11'] = employee7.fullname
        sheet['A12'] = employee8.fullname
        sheet['A13'] = employee9.fullname
        sheet['A14'] = employee10.fullname

        if self.emp1Checkbox_var.get() == 1:
            sheet['B5'] = shopPayGet1

        if self.emp2Checkbox_var.get() == 1:
            sheet['B6'] = shopPayGet2

        if self.emp3Checkbox_var.get() == 1:
            sheet['B7'] = shopPayGet3

        if self.emp4Checkbox_var.get() == 1:
            sheet['B8'] = shopPayGet4

        if self.emp5Checkbox_var.get() == 1:
            sheet['B9'] = shopPayGet5

        if self.emp6Checkbox_var.get() == 1:
            sheet['B10'] = shopPayGet6

        if self.emp7Checkbox_var.get() == 1:
            sheet['B11'] = shopPayGet7

        if self.emp8Checkbox_var.get() == 1:
            sheet['B12'] = shopPayGet8

        if self.emp9Checkbox_var.get() == 1:
            sheet['B13'] = shopPayGet9

        if self.emp10Checkbox_var.get() == 1:
            sheet['B14'] = shopPayGet10

        workbook.save("D:/Insulpro/temp/paysheets/shop pay sheet/" + filename)

        workbook.close()

    def saveWorkRecord(self):
        print('Saving Work Record')
        shopPayGet1 = self.emp1ShopPay.get()
        shopPayGet2 = self.emp2ShopPay.get()
        shopPayGet3 = self.emp3ShopPay.get()
        shopPayGet4 = self.emp4ShopPay.get()
        shopPayGet5 = self.emp5ShopPay.get()
        shopPayGet6 = self.emp6ShopPay.get()
        shopPayGet7 = self.emp7ShopPay.get()
        shopPayGet8 = self.emp8ShopPay.get()
        shopPayGet9 = self.emp9ShopPay.get()
        shopPayGet10 = self.emp10ShopPay.get()

        shopPayTotal = (float(shopPayGet1) + float(shopPayGet2) + float(shopPayGet3) + float(shopPayGet4) +
                        float(shopPayGet5) + float(shopPayGet6) + float(shopPayGet7) + float(shopPayGet8) +
                        float(shopPayGet9) + float(shopPayGet10))

        date = datetime.datetime.now()
        dateStamp = date.strftime('%m%d%y%H%M%S')
        nameFile = self.optionmenu.get()
        filename = str(nameFile) + str(dateStamp)

        workbook = load_workbook("D:/Insulpro/payroll records/Work Record.xlsx")

        sheet = workbook['Shop Record']

        sheet.append([filename, shopPayTotal])

        workbook.save("D:/Insulpro/payroll records/Work Record.xlsx")

    def savePaytreeShop(self):
        empFile1 = str(employee1.fullname) + "/shop record.xlsx"
        empFile2 = str(employee2.fullname) + "/shop record.xlsx"
        empFile3 = str(employee3.fullname) + "/shop record.xlsx"
        empFile4 = str(employee4.fullname) + "/shop record.xlsx"
        empFile5 = str(employee5.fullname) + "/shop record.xlsx"
        empFile6 = str(employee6.fullname) + "/shop record.xlsx"
        empFile7 = str(employee7.fullname) + "/shop record.xlsx"
        empFile8 = str(employee8.fullname) + "/shop record.xlsx"
        empFile9 = str(employee9.fullname) + "/shop record.xlsx"
        empFile10 = str(employee10.fullname) + "/shop record.xlsx"

        shopPayGet1 = self.emp1ShopPay.get()
        shopPayGet2 = self.emp2ShopPay.get()
        shopPayGet3 = self.emp3ShopPay.get()
        shopPayGet4 = self.emp4ShopPay.get()
        shopPayGet5 = self.emp5ShopPay.get()
        shopPayGet6 = self.emp6ShopPay.get()
        shopPayGet7 = self.emp7ShopPay.get()
        shopPayGet8 = self.emp8ShopPay.get()
        shopPayGet9 = self.emp9ShopPay.get()
        shopPayGet10 = self.emp10ShopPay.get()

        date = datetime.datetime.now()
        dateStamp = date.strftime('%m%d%y%H%M%S')
        nameFile = self.optionmenu.get()
        filename = str(nameFile) + str(dateStamp)

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile1)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet1)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile1)

            workbook.close()

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile2)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet2)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile2)

            workbook.close()

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile3)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet3)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile3)

            workbook.close()

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile4)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet4)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile4)

            workbook.close()

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile5)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet5)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile5)

            workbook.close()

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile6)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet6)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile6)

            workbook.close()

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile7)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet7)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile7)

            workbook.close()

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile8)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet8)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile8)

            workbook.close()

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile9)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet9)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile9)

            workbook.close()

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile10)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet10)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile10)

            workbook.close()

    def saveEmployeeRecord(self):
        print('Saving Employee Records')
        empFile1 = str(employee1.fullname + '.xlsx')
        empFile2 = str(employee2.fullname + '.xlsx')
        empFile3 = str(employee3.fullname + '.xlsx')
        empFile4 = str(employee4.fullname + '.xlsx')
        empFile5 = str(employee5.fullname + '.xlsx')
        empFile6 = str(employee6.fullname + '.xlsx')
        empFile7 = str(employee7.fullname + '.xlsx')
        empFile8 = str(employee8.fullname + '.xlsx')
        empFile9 = str(employee9.fullname + '.xlsx')
        empFile10 = str(employee10.fullname + '.xlsx')

        shopPayGet1 = self.emp1ShopPay.get()
        shopPayGet2 = self.emp2ShopPay.get()
        shopPayGet3 = self.emp3ShopPay.get()
        shopPayGet4 = self.emp4ShopPay.get()
        shopPayGet5 = self.emp5ShopPay.get()
        shopPayGet6 = self.emp6ShopPay.get()
        shopPayGet7 = self.emp7ShopPay.get()
        shopPayGet8 = self.emp8ShopPay.get()
        shopPayGet9 = self.emp9ShopPay.get()
        shopPayGet10 = self.emp10ShopPay.get()

        date = datetime.datetime.now()
        dateStamp = date.strftime('%m%d%y%H%M%S')
        nameFile = self.optionmenu.get()
        filename = str(nameFile) + str(dateStamp)

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile1)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet1)])

            workbook.save("D:/Insulpro/employee work records/" + empFile1)

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile2)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet2)])

            workbook.save("D:/Insulpro/employee work records/" + empFile2)

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile3)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet3)])

            workbook.save("D:/Insulpro/employee work records/" + empFile3)

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile4)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet4)])

            workbook.save("D:/Insulpro/employee work records/" + empFile4)

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile5)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet5)])

            workbook.save("D:/Insulpro/employee work records/" + empFile5)

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile6)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet6)])

            workbook.save("D:/Insulpro/employee work records/" + empFile6)

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile7)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet7)])

            workbook.save("D:/Insulpro/employee work records/" + empFile7)

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile8)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet8)])

            workbook.save("D:/Insulpro/employee work records/" + empFile8)

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile9)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet9)])

            workbook.save("D:/Insulpro/employee work records/" + empFile9)

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook(
                "D:/Insulpro/employee work records/" + empFile10)

            sheet = workbook['Shop Work Record']

            sheet.append([filename, float(shopPayGet10)])

            workbook.save("D:/Insulpro/employee work records/" + empFile10)

    def savePayrollSpreadsheet(self):
        print('Saving Payroll Spreadsheet')

        shopPayGet1 = float(self.emp1ShopPay.get())
        shopPayGet2 = float(self.emp2ShopPay.get())
        shopPayGet3 = float(self.emp3ShopPay.get())
        shopPayGet4 = float(self.emp4ShopPay.get())
        shopPayGet5 = float(self.emp5ShopPay.get())
        shopPayGet6 = float(self.emp6ShopPay.get())
        shopPayGet7 = float(self.emp7ShopPay.get())
        shopPayGet8 = float(self.emp8ShopPay.get())
        shopPayGet9 = float(self.emp9ShopPay.get())
        shopPayGet10 = float(self.emp10ShopPay.get())

        date = datetime.datetime.now()
        dateStamp = date.strftime('%m%d%y%H%M%S')
        fileName = self.optionmenu.get()
        namefile = str(fileName) + str(dateStamp)

        workbook = load_workbook("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        if self.emp1Checkbox_var.get() == 1:
            sheet = workbook['Emp1Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, shopPayGet1])

        if self.emp2Checkbox_var.get() == 1:
            sheet = workbook['Emp2Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, shopPayGet2])

        if self.emp3Checkbox_var.get() == 1:
            sheet = workbook['Emp3Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, shopPayGet3])

        if self.emp4Checkbox_var.get() == 1:
            sheet = workbook['Emp4Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, shopPayGet4])

        if self.emp5Checkbox_var.get() == 1:
            sheet = workbook['Emp5Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, shopPayGet5])

        if self.emp6Checkbox_var.get() == 1:
            sheet = workbook['Emp6Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, shopPayGet6])

        if self.emp7Checkbox_var.get() == 1:
            sheet = workbook['Emp7Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, shopPayGet7])

        if self.emp8Checkbox_var.get() == 1:
            sheet = workbook['Emp8Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, shopPayGet8])

        if self.emp9Checkbox_var.get() == 1:
            sheet = workbook['Emp9Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, shopPayGet9])

        if self.emp10Checkbox_var.get() == 1:
            sheet = workbook['Emp10Record']

            sheet.append([namefile, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0, shopPayGet10])

        workbook.save("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

        workbook.close()

    def emp1Checkbox_command(self):
        print(employee1.fullname + " selected")
        self.emp1Checkbox_var.set(1)

    def emp2Checkbox_command(self):
        print(employee2.fullname + " selected")
        self.emp2Checkbox_var.set(1)

    def emp3Checkbox_command(self):
        print(employee3.fullname + " selected")
        self.emp3Checkbox_var.set(1)

    def emp4Checkbox_command(self):
        print(employee4.fullname + " selected")
        self.emp4Checkbox_var.set(1)

    def emp5Checkbox_command(self):
        print(employee5.fullname + " selected")
        self.emp5Checkbox_var.set(1)

    def emp6Checkbox_command(self):
        print(employee6.fullname + " selected")
        self.emp6Checkbox_var.set(1)

    def emp7Checkbox_command(self):
        print(employee7.fullname + " selected")
        self.emp7Checkbox_var.set(1)

    def emp8Checkbox_command(self):
        print(employee8.fullname + " selected")
        self.emp8Checkbox_var.set(1)

    def emp9Checkbox_command(self):
        print(employee9.fullname + " selected")
        self.emp9Checkbox_var.set(1)

    def emp10Checkbox_command(self):
        print(employee10.fullname + " selected")
        self.emp10Checkbox_var.set(1)


# Attendance Section
class DailyTimeSheet(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x300")
        self.title("Daily Time Sheet")
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)

        self.emp1Checkbox_var = tk.IntVar(value=0)
        self.emp2Checkbox_var = tk.IntVar(value=0)
        self.emp3Checkbox_var = tk.IntVar(value=0)
        self.emp4Checkbox_var = tk.IntVar(value=0)
        self.emp5Checkbox_var = tk.IntVar(value=0)
        self.emp6Checkbox_var = tk.IntVar(value=0)
        self.emp7Checkbox_var = tk.IntVar(value=0)
        self.emp8Checkbox_var = tk.IntVar(value=0)
        self.emp9Checkbox_var = tk.IntVar(value=0)
        self.emp10Checkbox_var = tk.IntVar(value=0)

        emp1Checkbox = ctk.CTkCheckBox(self,
                                       text=employee1.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp1Checkbox_command,
                                       variable=self.emp1Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp1Checkbox.place(x=5, y=30)

        emp2Checkbox = ctk.CTkCheckBox(self,
                                       text=employee2.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp2Checkbox_command,
                                       variable=self.emp2Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp2Checkbox.place(x=5, y=65)

        emp3Checkbox = ctk.CTkCheckBox(self,
                                       text=employee3.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp3Checkbox_command,
                                       variable=self.emp3Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp3Checkbox.place(x=5, y=100)

        emp4Checkbox = ctk.CTkCheckBox(self,
                                       text=employee4.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp4Checkbox_command,
                                       variable=self.emp4Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp4Checkbox.place(x=5, y=135)

        emp5Checkbox = ctk.CTkCheckBox(self,
                                       text=employee5.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp5Checkbox_command,
                                       variable=self.emp5Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp5Checkbox.place(x=5, y=170)

        emp6Checkbox = ctk.CTkCheckBox(self,
                                       text=employee6.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp6Checkbox_command,
                                       variable=self.emp6Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp6Checkbox.place(x=5, y=205)

        emp7Checkbox = ctk.CTkCheckBox(self,
                                       text=employee7.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp7Checkbox_command,
                                       variable=self.emp7Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp7Checkbox.place(x=5, y=240)

        emp8Checkbox = ctk.CTkCheckBox(self,
                                       text=employee8.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp8Checkbox_command,
                                       variable=self.emp8Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp8Checkbox.place(x=5, y=275)

        emp9Checkbox = ctk.CTkCheckBox(self,
                                       text=employee9.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp9Checkbox_command,
                                       variable=self.emp9Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp9Checkbox.place(x=5, y=310)

        emp10Checkbox = ctk.CTkCheckBox(self,
                                        text=employee10.fullname,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.emp10Checkbox_command,
                                        variable=self.emp10Checkbox_var,
                                        onvalue=1,
                                        offvalue=0)
        emp10Checkbox.place(x=5, y=345)

        clockedInLabel = ctk.CTkLabel(self,
                                      text="Clocked  In",
                                      width=100,
                                      height=30,
                                      fg_color="transparent",
                                      font=ctk.CTkFont(size=18, weight="normal"))

        clockedInLabel.place(x=180, y=0)

        self.emp1ClockIn = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp1ClockIn.place(x=180, y=30)

        self.emp2ClockIn = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp2ClockIn.place(x=180, y=65)

        self.emp3ClockIn = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp3ClockIn.place(x=180, y=100)

        self.emp4ClockIn = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp4ClockIn.place(x=180, y=135)

        self.emp5ClockIn = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp5ClockIn.place(x=180, y=170)

        self.emp6ClockIn = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp6ClockIn.place(x=180, y=205)

        self.emp7ClockIn = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp7ClockIn.place(x=180, y=240)

        self.emp8ClockIn = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp8ClockIn.place(x=180, y=275)

        self.emp9ClockIn = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp9ClockIn.place(x=180, y=310)

        self.emp10ClockIn = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp10ClockIn.place(x=180, y=345)

        clockedOutLabel = ctk.CTkLabel(self,
                                       text="Clocked Out",
                                       width=100,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        clockedOutLabel.place(x=300, y=0)

        self.emp1ClockOut = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp1ClockOut.place(x=300, y=30)

        self.emp2ClockOut = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp2ClockOut.place(x=300, y=65)

        self.emp3ClockOut = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp3ClockOut.place(x=300, y=100)

        self.emp4ClockOut = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp4ClockOut.place(x=300, y=135)

        self.emp5ClockOut = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp5ClockOut.place(x=300, y=170)

        self.emp6ClockOut = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp6ClockOut.place(x=300, y=205)

        self.emp7ClockOut = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp7ClockOut.place(x=300, y=240)

        self.emp8ClockOut = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp8ClockOut.place(x=300, y=275)

        self.emp9ClockOut = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp9ClockOut.place(x=300, y=310)

        self.emp10ClockOut = ctk.CTkEntry(self,
                                          width=100,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"))

        self.emp10ClockOut.place(x=300, y=345)

        clockInButton = ctk.CTkButton(self,
                                      text="Clock In",
                                      width=100,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.clockInButton_command)

        clockInButton.place(x=180, y=380)

        clockOutButton = ctk.CTkButton(self,
                                       text="Clock Out",
                                       width=100,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.clockOutButton_command)

        clockOutButton.place(x=300, y=380)

        viewTimeSheetButton = ctk.CTkButton(self,
                                            text="View Time Sheet",
                                            width=150,
                                            height=30,
                                            font=ctk.CTkFont(size=18, weight="normal"),
                                            command=self.viewTimeSheetButton_command)

        viewTimeSheetButton.place(x=20, y=425)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=20, y=460)

    def clockInButton_command(self):
        print("command")
        dateToday = datetime.datetime.today()
        namefile = dateToday.strftime("%A%m%d%y")
        filename = str(namefile) + ".xlsx"

        clockin = datetime.datetime.now()
        clockinTime = clockin.strftime("%H:%M:%S")

        if os.path.exists("D:/Insulpro/payroll records/time sheets/" + filename):
            if self.emp1Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['B4'] = clockinTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['B4'].value
                self.emp1ClockIn.insert(0, x1)
                self.emp1Checkbox_var.set(0)

            if self.emp2Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['B5'] = clockinTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['B5'].value
                self.emp2ClockIn.insert(0, x1)
                self.emp2Checkbox_var.set(0)

            if self.emp3Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['B6'] = clockinTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['B6'].value
                self.emp3ClockIn.insert(0, x1)
                self.emp3Checkbox_var.set(0)

            if self.emp4Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['B7'] = clockinTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['B7'].value
                self.emp4ClockIn.insert(0, x1)
                self.emp4Checkbox_var.set(0)

            if self.emp5Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['B8'] = clockinTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['B8'].value
                self.emp5ClockIn.insert(0, x1)
                self.emp5Checkbox_var.set(0)

            if self.emp6Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['B9'] = clockinTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['B9'].value
                self.emp6ClockIn.insert(0, x1)
                self.emp6Checkbox_var.set(0)

            if self.emp7Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['B10'] = clockinTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['B10'].value
                self.emp7ClockIn.insert(0, x1)
                self.emp7Checkbox_var.set(0)

            if self.emp8Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['B11'] = clockinTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['B11'].value
                self.emp8ClockIn.insert(0, x1)
                self.emp8Checkbox_var.set(0)

            if self.emp9Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['B12'] = clockinTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['B12'].value
                self.emp9ClockIn.insert(0, x1)
                self.emp9Checkbox_var.set(0)

            if self.emp10Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['B13'] = clockinTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['B13'].value
                self.emp10ClockIn.insert(0, x1)
                self.emp10Checkbox_var.set(0)

    def clockOutButton_command(self):
        print("command")
        dateToday = datetime.datetime.today()
        namefile = dateToday.strftime("%A%m%d%y")
        filename = str(namefile) + ".xlsx"

        clockOut = datetime.datetime.now()
        clockOutTime = clockOut.strftime("%H:%M:%S")

        if os.path.exists("D:/Insulpro/payroll records/time sheets/" + filename):
            if self.emp1Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['C4'] = clockOutTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['C4'].value
                self.emp1ClockOut.insert(0, x1)
                self.emp1Checkbox_var.set(0)

            if self.emp2Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['C5'] = clockOutTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['C5'].value
                self.emp2ClockOut.insert(0, x1)
                self.emp2Checkbox_var.set(0)

            if self.emp3Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['C6'] = clockOutTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['C6'].value
                self.emp3ClockOut.insert(0, x1)
                self.emp3Checkbox_var.set(0)

            if self.emp4Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['C7'] = clockOutTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['C7'].value
                self.emp4ClockOut.insert(0, x1)
                self.emp4Checkbox_var.set(0)

            if self.emp5Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['C8'] = clockOutTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['C8'].value
                self.emp5ClockOut.insert(0, x1)
                self.emp5Checkbox_var.set(0)

            if self.emp6Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['C9'] = clockOutTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['C9'].value
                self.emp6ClockOut.insert(0, x1)
                self.emp6Checkbox_var.set(0)

            if self.emp7Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['C10'] = clockOutTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['C10'].value
                self.emp7ClockOut.insert(0, x1)
                self.emp7Checkbox_var.set(0)

            if self.emp8Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['C11'] = clockOutTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['C11'].value
                self.emp8ClockOut.insert(0, x1)
                self.emp8Checkbox_var.set(0)

            if self.emp9Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['C12'] = clockOutTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['C12'].value
                self.emp9ClockOut.insert(0, x1)
                self.emp9Checkbox_var.set(0)

            if self.emp10Checkbox_var.get() == 1:
                wb = load_workbook(
                    "D:/Insulpro/payroll records/time sheets/" + filename)

                sheet = wb['DailyTimeSheet']

                sheet['C13'] = clockOutTime
                wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

                wb.close()
                x1 = sheet['C13'].value
                self.emp10ClockOut.insert(0, x1)
                self.emp10Checkbox_var.set(0)

    def viewTimeSheetButton_command(self):
        print("command")
        dateToday = datetime.datetime.today()
        namefile = dateToday.strftime("%A%m%d%y")
        filename = str(namefile) + ".xlsx"

        if os.path.exists("D:/Insulpro/payroll records/time sheets/" + filename):
            wb = load_workbook("D:/Insulpro/payroll records/time sheets/" + filename)

            sheet = wb['DailyTimeSheet']

            x1 = sheet['B4'].value
            y1 = sheet['C4'].value
            x2 = sheet['B5'].value
            y2 = sheet['C5'].value
            x3 = sheet['B6'].value
            y3 = sheet['C6'].value
            x4 = sheet['B7'].value
            y4 = sheet['C7'].value
            x5 = sheet['B8'].value
            y5 = sheet['C8'].value
            x6 = sheet['B9'].value
            y6 = sheet['C9'].value
            x7 = sheet['B10'].value
            y7 = sheet['C10'].value
            x8 = sheet['B11'].value
            y8 = sheet['C11'].value
            x9 = sheet['B12'].value
            y9 = sheet['C12'].value
            x10 = sheet['B13'].value
            y10 = sheet['C13'].value

            self.emp1ClockIn.insert(0, str(x1))
            self.emp1ClockOut.insert(0, str(y1))
            self.emp2ClockIn.insert(0, str(x2))
            self.emp2ClockOut.insert(0, str(y2))
            self.emp3ClockIn.insert(0, str(x3))
            self.emp3ClockOut.insert(0, str(y3))
            self.emp4ClockIn.insert(0, str(x4))
            self.emp4ClockOut.insert(0, str(y4))
            self.emp5ClockIn.insert(0, str(x5))
            self.emp5ClockOut.insert(0, str(y5))
            self.emp6ClockIn.insert(0, str(x6))
            self.emp6ClockOut.insert(0, str(y6))
            self.emp7ClockIn.insert(0, str(x7))
            self.emp7ClockOut.insert(0, str(y7))
            self.emp8ClockIn.insert(0, str(x8))
            self.emp8ClockOut.insert(0, str(y8))
            self.emp9ClockIn.insert(0, str(x9))
            self.emp9ClockOut.insert(0, str(y9))
            self.emp10ClockIn.insert(0, str(x10))
            self.emp10ClockOut.insert(0, str(y10))

            wb.close()

    def closeWindowButton_command(self):
        self.destroy()

    def emp1Checkbox_command(self):
        print(employee1.fullname + " selected")
        self.emp1Checkbox_var.set(1)

    def emp2Checkbox_command(self):
        print(employee2.fullname + " selected")
        self.emp2Checkbox_var.set(1)

    def emp3Checkbox_command(self):
        print(employee3.fullname + " selected")
        self.emp3Checkbox_var.set(1)

    def emp4Checkbox_command(self):
        print(employee4.fullname + " selected")
        self.emp4Checkbox_var.set(1)

    def emp5Checkbox_command(self):
        print(employee5.fullname + " selected")
        self.emp5Checkbox_var.set(1)

    def emp6Checkbox_command(self):
        print(employee6.fullname + " selected")
        self.emp6Checkbox_var.set(1)

    def emp7Checkbox_command(self):
        print(employee7.fullname + " selected")
        self.emp7Checkbox_var.set(1)

    def emp8Checkbox_command(self):
        print(employee8.fullname + " selected")
        self.emp8Checkbox_var.set(1)

    def emp9Checkbox_command(self):
        print(employee9.fullname + " selected")
        self.emp9Checkbox_var.set(1)

    def emp10Checkbox_command(self):
        print(employee10.fullname + " selected")
        self.emp10Checkbox_var.set(1)


# todo create work order sections
# Work Order Sections
class NewConstructionWorkOrder(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x300")
        self.title("New Construction Work Order")


class ExistConstructionWorkOrder(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x300")
        self.title("Existing Construction Work Order")


class AtticWorkOrder(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x300")
        self.title("Attic Work Order")


class FoamWorkOrder(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x300")
        self.title("Foam Work Order")


class CelluloseWorkOrder(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x300")
        self.title("Cellulose Work Order")


class ShopWorkOrder(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x300")
        self.title("Shop Work Order")


class PayrollPaytreeEntry(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x300")
        self.title("Enter Employee Payroll Record")
        # setting window size
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)

        self.emp1Checkbox_var = tk.IntVar(value=0)
        self.emp2Checkbox_var = tk.IntVar(value=0)
        self.emp3Checkbox_var = tk.IntVar(value=0)
        self.emp4Checkbox_var = tk.IntVar(value=0)
        self.emp5Checkbox_var = tk.IntVar(value=0)
        self.emp6Checkbox_var = tk.IntVar(value=0)
        self.emp7Checkbox_var = tk.IntVar(value=0)
        self.emp8Checkbox_var = tk.IntVar(value=0)
        self.emp9Checkbox_var = tk.IntVar(value=0)
        self.emp10Checkbox_var = tk.IntVar(value=0)

        emp1Checkbox = ctk.CTkCheckBox(self,
                                       text=employee1.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp1Checkbox_command,
                                       variable=self.emp1Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp1Checkbox.place(x=5, y=30)

        emp2Checkbox = ctk.CTkCheckBox(self,
                                       text=employee2.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp2Checkbox_command,
                                       variable=self.emp2Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp2Checkbox.place(x=5, y=65)

        emp3Checkbox = ctk.CTkCheckBox(self,
                                       text=employee3.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp3Checkbox_command,
                                       variable=self.emp3Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp3Checkbox.place(x=5, y=100)

        emp4Checkbox = ctk.CTkCheckBox(self,
                                       text=employee4.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp4Checkbox_command,
                                       variable=self.emp4Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp4Checkbox.place(x=5, y=135)

        emp5Checkbox = ctk.CTkCheckBox(self,
                                       text=employee5.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp5Checkbox_command,
                                       variable=self.emp5Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp5Checkbox.place(x=5, y=170)

        emp6Checkbox = ctk.CTkCheckBox(self,
                                       text=employee6.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp6Checkbox_command,
                                       variable=self.emp6Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp6Checkbox.place(x=5, y=205)

        emp7Checkbox = ctk.CTkCheckBox(self,
                                       text=employee7.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp7Checkbox_command,
                                       variable=self.emp7Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp7Checkbox.place(x=5, y=240)

        emp8Checkbox = ctk.CTkCheckBox(self,
                                       text=employee8.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp8Checkbox_command,
                                       variable=self.emp8Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp8Checkbox.place(x=5, y=275)

        emp9Checkbox = ctk.CTkCheckBox(self,
                                       text=employee9.fullname,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.emp9Checkbox_command,
                                       variable=self.emp9Checkbox_var,
                                       onvalue=1,
                                       offvalue=0)
        emp9Checkbox.place(x=5, y=310)

        emp10Checkbox = ctk.CTkCheckBox(self,
                                        text=employee10.fullname,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.emp10Checkbox_command,
                                        variable=self.emp10Checkbox_var,
                                        onvalue=1,
                                        offvalue=0)
        emp10Checkbox.place(x=5, y=345)

        payrollAmountLabel = ctk.CTkLabel(self,
                                          text="Payroll Amount",
                                          width=100,
                                          height=30,
                                          fg_color="transparent",
                                          font=ctk.CTkFont(size=18, weight="normal"))

        payrollAmountLabel.place(x=180, y=0)

        self.emp1payroll = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp1payroll.place(x=180, y=30)
        self.emp1payroll.insert(0, '0')

        self.emp2payroll = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp2payroll.place(x=180, y=65)
        self.emp2payroll.insert(0, '0')

        self.emp3payroll = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp3payroll.place(x=180, y=100)
        self.emp3payroll.insert(0, '0')

        self.emp4payroll = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp4payroll.place(x=180, y=135)
        self.emp4payroll.insert(0, '0')

        self.emp5payroll = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp5payroll.place(x=180, y=170)
        self.emp5payroll.insert(0, '0')

        self.emp6payroll = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp6payroll.place(x=180, y=205)
        self.emp6payroll.insert(0, '0')

        self.emp7payroll = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp7payroll.place(x=180, y=240)
        self.emp7payroll.insert(0, '0')

        self.emp8payroll = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp8payroll.place(x=180, y=275)
        self.emp8payroll.insert(0, '0')

        self.emp9payroll = ctk.CTkEntry(self,
                                        width=100,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"))

        self.emp9payroll.place(x=180, y=310)
        self.emp9payroll.insert(0, '0')

        self.emp10payroll = ctk.CTkEntry(self,
                                         width=100,
                                         height=30,
                                         font=ctk.CTkFont(size=18, weight="normal"))

        self.emp10payroll.place(x=180, y=345)
        self.emp10payroll.insert(0, '0')

        saveFileButton = ctk.CTkButton(self,
                                       text="Save File",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.saveFileButton_command)

        saveFileButton.place(x=20, y=425)

        resetFormButton = ctk.CTkButton(self,
                                        text="Reset Form",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.resetFormButton_command)

        resetFormButton.place(x=20, y=460)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=20, y=495)

    def saveFileButton_command(self):
        empFile1 = str(employee1.fullname) + "/Payroll record.xlsx"
        empFile2 = str(employee2.fullname) + "/payroll record.xlsx"
        empFile3 = str(employee3.fullname) + "/payroll record.xlsx"
        empFile4 = str(employee4.fullname) + "/payroll record.xlsx"
        empFile5 = str(employee5.fullname) + "/payroll record.xlsx"
        empFile6 = str(employee6.fullname) + "/payroll record.xlsx"
        empFile7 = str(employee7.fullname) + "/payroll record.xlsx"
        empFile8 = str(employee8.fullname) + "/payroll record.xlsx"
        empFile9 = str(employee9.fullname) + "/payroll record.xlsx"
        empFile10 = str(employee10.fullname) + "/payroll record.xlsx"

        filedate = datetime.datetime.now()
        filename = filedate.strftime('%m%d%y')

        payrollGet1 = self.emp1payroll.get()
        payrollGet2 = self.emp2payroll.get()
        payrollGet3 = self.emp3payroll.get()
        payrollGet4 = self.emp4payroll.get()
        payrollGet5 = self.emp5payroll.get()
        payrollGet6 = self.emp6payroll.get()
        payrollGet7 = self.emp7payroll.get()
        payrollGet8 = self.emp8payroll.get()
        payrollGet9 = self.emp9payroll.get()
        payrollGet10 = self.emp10payroll.get()

        if self.emp1Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile1)

            sheet = workbook['Payroll Record']

            sheet.append([filename, float(payrollGet1)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile1)

            workbook.close()

        if self.emp2Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile2)

            sheet = workbook['Payroll Record']

            sheet.append([filename, float(payrollGet2)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile2)

            workbook.close()

        if self.emp3Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile3)

            sheet = workbook['Payroll Record']

            sheet.append([filename, float(payrollGet3)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile3)

            workbook.close()

        if self.emp4Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile4)

            sheet = workbook['Payroll Record']

            sheet.append([filename, float(payrollGet4)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile4)

            workbook.close()

        if self.emp5Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile5)

            sheet = workbook['Payroll Record']

            sheet.append([filename, float(payrollGet5)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile5)

            workbook.close()

        if self.emp6Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile6)

            sheet = workbook['Payroll Record']

            sheet.append([filename, float(payrollGet6)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile6)

            workbook.close()

        if self.emp7Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile7)

            sheet = workbook['Payroll Record']

            sheet.append([filename, float(payrollGet7)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile7)

            workbook.close()

        if self.emp8Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile8)

            sheet = workbook['Payroll Record']

            sheet.append([filename, float(payrollGet8)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile8)

            workbook.close()

        if self.emp9Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile9)

            sheet = workbook['Payroll Record']

            sheet.append([filename, float(payrollGet9)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile9)

            workbook.close()

        if self.emp10Checkbox_var.get() == 1:
            workbook = load_workbook("D:/Insulpro/employee paytree/" + empFile10)

            sheet = workbook['Payroll Record']

            sheet.append([filename, float(payrollGet10)])

            workbook.save("D:/Insulpro/employee paytree/" + empFile10)

            workbook.close()

        self.resetFormButton_command()
        self.closeWindowButton_command()

        messagebox.showinfo(title='Save', message='File Save Successful')
        print("File was saved successfully")

    def resetFormButton_command(self):
        print('Clearing Form')
        if self.emp1Checkbox_var.get() == 1:
            self.emp1Checkbox_var.set(0)

        if self.emp2Checkbox_var.get() == 1:
            self.emp2Checkbox_var.set(0)

        if self.emp3Checkbox_var.get() == 1:
            self.emp3Checkbox_var.set(0)

        if self.emp4Checkbox_var.get() == 1:
            self.emp4Checkbox_var.set(0)

        if self.emp5Checkbox_var.get() == 1:
            self.emp5Checkbox_var.set(0)

        if self.emp6Checkbox_var.get() == 1:
            self.emp6Checkbox_var.set(0)

        if self.emp7Checkbox_var.get() == 1:
            self.emp7Checkbox_var.set(0)

        if self.emp8Checkbox_var.get() == 1:
            self.emp8Checkbox_var.set(0)

        if self.emp9Checkbox_var.get() == 1:
            self.emp9Checkbox_var.set(0)

        if self.emp10Checkbox_var.get() == 1:
            self.emp10Checkbox_var.set(0)

        self.emp1payroll.delete(0, "end")
        self.emp2payroll.delete(0, "end")
        self.emp3payroll.delete(0, "end")
        self.emp4payroll.delete(0, "end")
        self.emp5payroll.delete(0, "end")
        self.emp6payroll.delete(0, "end")
        self.emp7payroll.delete(0, "end")
        self.emp8payroll.delete(0, "end")
        self.emp9payroll.delete(0, "end")
        self.emp10payroll.delete(0, "end")

        self.emp1payroll.insert(0, "0")
        self.emp2payroll.insert(0, "0")
        self.emp3payroll.insert(0, "0")
        self.emp4payroll.insert(0, "0")
        self.emp5payroll.insert(0, "0")
        self.emp6payroll.insert(0, "0")
        self.emp7payroll.insert(0, "0")
        self.emp8payroll.insert(0, "0")
        self.emp9payroll.insert(0, "0")
        self.emp10payroll.insert(0, "0")

    def closeWindowButton_command(self):
        self.destroy()

    def emp1Checkbox_command(self):
        print(employee1.fullname + " selected")
        self.emp1Checkbox_var.set(1)

    def emp2Checkbox_command(self):
        print(employee2.fullname + " selected")
        self.emp2Checkbox_var.set(1)

    def emp3Checkbox_command(self):
        print(employee3.fullname + " selected")
        self.emp3Checkbox_var.set(1)

    def emp4Checkbox_command(self):
        print(employee4.fullname + " selected")
        self.emp4Checkbox_var.set(1)

    def emp5Checkbox_command(self):
        print(employee5.fullname + " selected")
        self.emp5Checkbox_var.set(1)

    def emp6Checkbox_command(self):
        print(employee6.fullname + " selected")
        self.emp6Checkbox_var.set(1)

    def emp7Checkbox_command(self):
        print(employee7.fullname + " selected")
        self.emp7Checkbox_var.set(1)

    def emp8Checkbox_command(self):
        print(employee8.fullname + " selected")
        self.emp8Checkbox_var.set(1)

    def emp9Checkbox_command(self):
        print(employee9.fullname + " selected")
        self.emp9Checkbox_var.set(1)

    def emp10Checkbox_command(self):
        print(employee10.fullname + " selected")
        self.emp10Checkbox_var.set(1)


# Employee Payroll Sections

# Employee 1
class Employee1Record(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x400")
        self.title(employee1.fullname)

        self.toplevel_window = None

        workRecordLabel = ctk.CTkLabel(self,
                                       text="Work Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        workRecordLabel.place(x=5, y=30)

        battWorkButton = ctk.CTkButton(self,
                                       text="Batt Work",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.battWorkButton_command)

        battWorkButton.place(x=5, y=65)

        atticWorkButton = ctk.CTkButton(self,
                                        text="Attic Work",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.atticWorkButton_command)

        atticWorkButton.place(x=5, y=100)

        sprayFoamButton = ctk.CTkButton(self,
                                        text="Spray Foam",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.sprayFoamButton_command)

        sprayFoamButton.place(x=5, y=135)

        multiDayButton = ctk.CTkButton(self,
                                       text="Multiple Day",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.multiDayButton_command)

        multiDayButton.place(x=5, y=170)

        shopPayButton = ctk.CTkButton(self,
                                      text="Shop Pay",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.shopPayButton_command)

        shopPayButton.place(x=5, y=205)

        payRecordsLabel = ctk.CTkLabel(self,
                                       text="Pay Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        payRecordsLabel.place(x=210, y=30)

        payrollButton = ctk.CTkButton(self,
                                      text="Payroll",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.payrollButton_command)

        payrollButton.place(x=210, y=65)

        vacationButton = ctk.CTkButton(self,
                                       text="Vacation Record",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.vacationButton_command)

        vacationButton.place(x=210, y=100)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=300)

    def battWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee1PayBatt(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def atticWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee1PayAttic(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def sprayFoamButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee1PayFoam(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def multiDayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee1PayMulti(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopPayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee1PayShop(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def payrollButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee1Payroll(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def vacationButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee1PayVacation(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def closeWindowButton_command(self):
        self.destroy()


class Employee1PayBatt(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee1.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile1 = employee1.fullname + '/batt record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile1
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee1PayAttic(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee1.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile1 = employee1.fullname + '/attic record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile1
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee1PayFoam(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee1.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile1 = employee1.fullname + '/foam record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile1
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee1PayMulti(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee1.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile1 = employee1.fullname + '/multi record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile1
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee1PayVacation(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee1.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile1 = employee1.fullname + '/vacation record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile1
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee1PayShop(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee1.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile1 = employee1.fullname + '/Shop record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile1
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee1Payroll(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee1.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile1 = employee1.fullname + '/Payroll record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile1
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


# Employee 2
class Employee2Record(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x400")
        self.title(employee2.fullname)

        self.toplevel_window = None

        workRecordLabel = ctk.CTkLabel(self,
                                       text="Work Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        workRecordLabel.place(x=5, y=30)

        battWorkButton = ctk.CTkButton(self,
                                       text="Batt Work",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.battWorkButton_command)

        battWorkButton.place(x=5, y=65)

        atticWorkButton = ctk.CTkButton(self,
                                        text="Attic Work",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.atticWorkButton_command)

        atticWorkButton.place(x=5, y=100)

        sprayFoamButton = ctk.CTkButton(self,
                                        text="Spray Foam",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.sprayFoamButton_command)

        sprayFoamButton.place(x=5, y=135)

        multiDayButton = ctk.CTkButton(self,
                                       text="Multiple Day",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.multiDayButton_command)

        multiDayButton.place(x=5, y=170)

        shopPayButton = ctk.CTkButton(self,
                                      text="Shop Pay",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.shopPayButton_command)

        shopPayButton.place(x=5, y=205)

        payRecordsLabel = ctk.CTkLabel(self,
                                       text="Pay Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        payRecordsLabel.place(x=210, y=30)

        payrollButton = ctk.CTkButton(self,
                                      text="Payroll",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.payrollButton_command)

        payrollButton.place(x=210, y=65)

        vacationButton = ctk.CTkButton(self,
                                       text="Vacation Record",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.vacationButton_command)

        vacationButton.place(x=210, y=100)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=300)

    def battWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee2PayBatt(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def atticWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee2PayAttic(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def sprayFoamButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee2PayFoam(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def multiDayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee2PayMulti(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopPayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee2PayShop(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def payrollButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee2Payroll(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def vacationButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee2PayVacation(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def closeWindowButton_command(self):
        self.destroy()


class Employee2PayBatt(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee2.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile2 = employee2.fullname + '/batt record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile2
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee2PayAttic(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee2.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile2 = employee2.fullname + '/attic record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile2
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee2PayFoam(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee2.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile2 = employee2.fullname + '/foam record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile2
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee2PayMulti(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee2.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile2 = employee2.fullname + '/multi record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile2
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee2PayVacation(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee2.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile2 = employee2.fullname + '/vacation record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile2
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee2PayShop(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee2.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile2 = employee2.fullname + '/Shop record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile2
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee2Payroll(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee2.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile2 = employee2.fullname + '/Payroll record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile2
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


# Employee 3
class Employee3Record(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x400")
        self.title(employee3.fullname)

        self.toplevel_window = None

        workRecordLabel = ctk.CTkLabel(self,
                                       text="Work Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        workRecordLabel.place(x=5, y=30)

        battWorkButton = ctk.CTkButton(self,
                                       text="Batt Work",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.battWorkButton_command)

        battWorkButton.place(x=5, y=65)

        atticWorkButton = ctk.CTkButton(self,
                                        text="Attic Work",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.atticWorkButton_command)

        atticWorkButton.place(x=5, y=100)

        sprayFoamButton = ctk.CTkButton(self,
                                        text="Spray Foam",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.sprayFoamButton_command)

        sprayFoamButton.place(x=5, y=135)

        multiDayButton = ctk.CTkButton(self,
                                       text="Multiple Day",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.multiDayButton_command)

        multiDayButton.place(x=5, y=170)

        shopPayButton = ctk.CTkButton(self,
                                      text="Shop Pay",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.shopPayButton_command)

        shopPayButton.place(x=5, y=205)

        payRecordsLabel = ctk.CTkLabel(self,
                                       text="Pay Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        payRecordsLabel.place(x=210, y=30)

        payrollButton = ctk.CTkButton(self,
                                      text="Payroll",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.payrollButton_command)

        payrollButton.place(x=210, y=65)

        vacationButton = ctk.CTkButton(self,
                                       text="Vacation Record",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.vacationButton_command)

        vacationButton.place(x=210, y=100)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=300)

    def battWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee3PayBatt(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def atticWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee3PayAttic(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def sprayFoamButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee3PayFoam(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def multiDayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee3PayMulti(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopPayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee3PayShop(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def payrollButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee3Payroll(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def vacationButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee3PayVacation(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def closeWindowButton_command(self):
        self.destroy()


class Employee3PayBatt(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee3.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile3 = employee3.fullname + '/batt record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile3
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee3PayAttic(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee3.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile3 = employee3.fullname + '/attic record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile3
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee3PayFoam(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee3.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile3 = employee3.fullname + '/foam record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile3
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee3PayMulti(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee3.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile3 = employee3.fullname + '/multi record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile3
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee3PayVacation(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee3.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile3 = employee3.fullname + '/vacation record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile3
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee3PayShop(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee3.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile3 = employee3.fullname + '/Shop record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile3
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee3Payroll(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee3.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile3 = employee3.fullname + '/Payroll record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile3
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


# Employee 4
class Employee4Record(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x400")
        self.title(employee4.fullname)

        self.toplevel_window = None

        workRecordLabel = ctk.CTkLabel(self,
                                       text="Work Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        workRecordLabel.place(x=5, y=30)

        battWorkButton = ctk.CTkButton(self,
                                       text="Batt Work",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.battWorkButton_command)

        battWorkButton.place(x=5, y=65)

        atticWorkButton = ctk.CTkButton(self,
                                        text="Attic Work",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.atticWorkButton_command)

        atticWorkButton.place(x=5, y=100)

        sprayFoamButton = ctk.CTkButton(self,
                                        text="Spray Foam",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.sprayFoamButton_command)

        sprayFoamButton.place(x=5, y=135)

        multiDayButton = ctk.CTkButton(self,
                                       text="Multiple Day",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.multiDayButton_command)

        multiDayButton.place(x=5, y=170)

        shopPayButton = ctk.CTkButton(self,
                                      text="Shop Pay",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.shopPayButton_command)

        shopPayButton.place(x=5, y=205)

        payRecordsLabel = ctk.CTkLabel(self,
                                       text="Pay Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        payRecordsLabel.place(x=210, y=30)

        payrollButton = ctk.CTkButton(self,
                                      text="Payroll",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.payrollButton_command)

        payrollButton.place(x=210, y=65)

        vacationButton = ctk.CTkButton(self,
                                       text="Vacation Record",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.vacationButton_command)

        vacationButton.place(x=210, y=100)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=300)

    def battWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee4PayBatt(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def atticWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee4PayAttic(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def sprayFoamButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee4PayFoam(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def multiDayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee4PayMulti(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopPayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee4PayShop(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def payrollButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee4Payroll(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def vacationButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee4PayVacation(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def closeWindowButton_command(self):
        self.destroy()


class Employee4PayBatt(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee4.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile4 = employee4.fullname + '/batt record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile4
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee4PayAttic(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee4.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile4 = employee4.fullname + '/attic record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile4
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee4PayFoam(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee4.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile4 = employee4.fullname + '/foam record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile4
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee4PayMulti(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee4.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile4 = employee4.fullname + '/multi record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile4
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee4PayVacation(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee4.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile4 = employee4.fullname + '/vacation record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile4
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee4PayShop(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee4.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile4 = employee4.fullname + '/Shop record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile4
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee4Payroll(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee4.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile4 = employee4.fullname + '/Payroll record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile4
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


# Employee 5
class Employee5Record(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x400")
        self.title(employee5.fullname)

        self.toplevel_window = None

        workRecordLabel = ctk.CTkLabel(self,
                                       text="Work Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        workRecordLabel.place(x=5, y=30)

        battWorkButton = ctk.CTkButton(self,
                                       text="Batt Work",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.battWorkButton_command)

        battWorkButton.place(x=5, y=65)

        atticWorkButton = ctk.CTkButton(self,
                                        text="Attic Work",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.atticWorkButton_command)

        atticWorkButton.place(x=5, y=100)

        sprayFoamButton = ctk.CTkButton(self,
                                        text="Spray Foam",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.sprayFoamButton_command)

        sprayFoamButton.place(x=5, y=135)

        multiDayButton = ctk.CTkButton(self,
                                       text="Multiple Day",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.multiDayButton_command)

        multiDayButton.place(x=5, y=170)

        shopPayButton = ctk.CTkButton(self,
                                      text="Shop Pay",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.shopPayButton_command)

        shopPayButton.place(x=5, y=205)

        payRecordsLabel = ctk.CTkLabel(self,
                                       text="Pay Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        payRecordsLabel.place(x=210, y=30)

        payrollButton = ctk.CTkButton(self,
                                      text="Payroll",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.payrollButton_command)

        payrollButton.place(x=210, y=65)

        vacationButton = ctk.CTkButton(self,
                                       text="Vacation Record",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.vacationButton_command)

        vacationButton.place(x=210, y=100)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=300)

    def battWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee5PayBatt(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def atticWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee5PayAttic(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def sprayFoamButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee5PayFoam(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def multiDayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee5PayMulti(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopPayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee5PayShop(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def payrollButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee5Payroll(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def vacationButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee5PayVacation(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def closeWindowButton_command(self):
        self.destroy()


class Employee5PayBatt(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee5.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile5 = employee5.fullname + '/batt record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile5
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee5PayAttic(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee5.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile5 = employee5.fullname + '/attic record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile5
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee5PayFoam(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee5.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile5 = employee5.fullname + '/foam record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile5
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee5PayMulti(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee5.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile5 = employee5.fullname + '/multi record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile5
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee5PayVacation(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee5.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile5 = employee5.fullname + '/vacation record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile5
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee5PayShop(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee5.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile5 = employee5.fullname + '/Shop record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile5
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee5Payroll(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee5.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile5 = employee5.fullname + '/Payroll record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile5
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


# Employee 6
class Employee6Record(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x400")
        self.title(employee6.fullname)

        self.toplevel_window = None

        workRecordLabel = ctk.CTkLabel(self,
                                       text="Work Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        workRecordLabel.place(x=5, y=30)

        battWorkButton = ctk.CTkButton(self,
                                       text="Batt Work",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.battWorkButton_command)

        battWorkButton.place(x=5, y=65)

        atticWorkButton = ctk.CTkButton(self,
                                        text="Attic Work",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.atticWorkButton_command)

        atticWorkButton.place(x=5, y=100)

        sprayFoamButton = ctk.CTkButton(self,
                                        text="Spray Foam",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.sprayFoamButton_command)

        sprayFoamButton.place(x=5, y=135)

        multiDayButton = ctk.CTkButton(self,
                                       text="Multiple Day",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.multiDayButton_command)

        multiDayButton.place(x=5, y=170)

        shopPayButton = ctk.CTkButton(self,
                                      text="Shop Pay",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.shopPayButton_command)

        shopPayButton.place(x=5, y=205)

        payRecordsLabel = ctk.CTkLabel(self,
                                       text="Pay Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        payRecordsLabel.place(x=210, y=30)

        payrollButton = ctk.CTkButton(self,
                                      text="Payroll",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.payrollButton_command)

        payrollButton.place(x=210, y=65)

        vacationButton = ctk.CTkButton(self,
                                       text="Vacation Record",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.vacationButton_command)

        vacationButton.place(x=210, y=100)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=300)

    def battWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee6PayBatt(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def atticWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee6PayAttic(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def sprayFoamButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee6PayFoam(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def multiDayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee6PayMulti(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopPayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee6PayShop(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def payrollButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee6Payroll(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def vacationButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee6PayVacation(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def closeWindowButton_command(self):
        self.destroy()


class Employee6PayBatt(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee6.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile6 = employee6.fullname + '/batt record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile6
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee6PayAttic(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee6.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile6 = employee6.fullname + '/attic record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile6
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee6PayFoam(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee6.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile6 = employee6.fullname + '/foam record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile6
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee6PayMulti(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee6.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile6 = employee6.fullname + '/multi record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile6
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee6PayVacation(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee6.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile6 = employee6.fullname + '/vacation record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile6
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee6PayShop(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee6.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile6 = employee6.fullname + '/Shop record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile6
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee6Payroll(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee6.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile6 = employee6.fullname + '/Payroll record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile6
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


# Employee 7
class Employee7Record(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x400")
        self.title(employee7.fullname)

        self.toplevel_window = None

        workRecordLabel = ctk.CTkLabel(self,
                                       text="Work Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        workRecordLabel.place(x=5, y=30)

        battWorkButton = ctk.CTkButton(self,
                                       text="Batt Work",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.battWorkButton_command)

        battWorkButton.place(x=5, y=65)

        atticWorkButton = ctk.CTkButton(self,
                                        text="Attic Work",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.atticWorkButton_command)

        atticWorkButton.place(x=5, y=100)

        sprayFoamButton = ctk.CTkButton(self,
                                        text="Spray Foam",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.sprayFoamButton_command)

        sprayFoamButton.place(x=5, y=135)

        multiDayButton = ctk.CTkButton(self,
                                       text="Multiple Day",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.multiDayButton_command)

        multiDayButton.place(x=5, y=170)

        shopPayButton = ctk.CTkButton(self,
                                      text="Shop Pay",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.shopPayButton_command)

        shopPayButton.place(x=5, y=205)

        payRecordsLabel = ctk.CTkLabel(self,
                                       text="Pay Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        payRecordsLabel.place(x=210, y=30)

        payrollButton = ctk.CTkButton(self,
                                      text="Payroll",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.payrollButton_command)

        payrollButton.place(x=210, y=65)

        vacationButton = ctk.CTkButton(self,
                                       text="Vacation Record",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.vacationButton_command)

        vacationButton.place(x=210, y=100)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=300)

    def battWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee7PayBatt(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def atticWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee7PayAttic(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def sprayFoamButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee7PayFoam(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def multiDayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee7PayMulti(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopPayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee7PayShop(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def payrollButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee7Payroll(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def vacationButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee7PayVacation(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def closeWindowButton_command(self):
        self.destroy()


class Employee7PayBatt(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee7.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile7 = employee7.fullname + '/batt record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile7
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee7PayAttic(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee7.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile7 = employee7.fullname + '/attic record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile7
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee7PayFoam(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee7.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile7 = employee7.fullname + '/foam record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile7
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee7PayMulti(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee7.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile7 = employee7.fullname + '/multi record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile7
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee7PayVacation(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee7.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile7 = employee7.fullname + '/vacation record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile7
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee7PayShop(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee7.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile7 = employee7.fullname + '/Shop record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile7
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee7Payroll(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee7.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile7 = employee7.fullname + '/Payroll record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile7
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


# Employee 8
class Employee8Record(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x400")
        self.title(employee8.fullname)

        self.toplevel_window = None

        workRecordLabel = ctk.CTkLabel(self,
                                       text="Work Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        workRecordLabel.place(x=5, y=30)

        battWorkButton = ctk.CTkButton(self,
                                       text="Batt Work",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.battWorkButton_command)

        battWorkButton.place(x=5, y=65)

        atticWorkButton = ctk.CTkButton(self,
                                        text="Attic Work",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.atticWorkButton_command)

        atticWorkButton.place(x=5, y=100)

        sprayFoamButton = ctk.CTkButton(self,
                                        text="Spray Foam",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.sprayFoamButton_command)

        sprayFoamButton.place(x=5, y=135)

        multiDayButton = ctk.CTkButton(self,
                                       text="Multiple Day",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.multiDayButton_command)

        multiDayButton.place(x=5, y=170)

        shopPayButton = ctk.CTkButton(self,
                                      text="Shop Pay",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.shopPayButton_command)

        shopPayButton.place(x=5, y=205)

        payRecordsLabel = ctk.CTkLabel(self,
                                       text="Pay Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        payRecordsLabel.place(x=210, y=30)

        payrollButton = ctk.CTkButton(self,
                                      text="Payroll",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.payrollButton_command)

        payrollButton.place(x=210, y=65)

        vacationButton = ctk.CTkButton(self,
                                       text="Vacation Record",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.vacationButton_command)

        vacationButton.place(x=210, y=100)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=300)

    def battWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee8PayBatt(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def atticWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee8PayAttic(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def sprayFoamButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee8PayFoam(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def multiDayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee8PayMulti(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopPayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee8PayShop(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def payrollButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee8Payroll(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def vacationButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee8PayVacation(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def closeWindowButton_command(self):
        self.destroy()


class Employee8PayBatt(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee8.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile8 = employee8.fullname + '/batt record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile8
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee8PayAttic(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee8.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile8 = employee8.fullname + '/attic record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile8
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee8PayFoam(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee8.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile8 = employee8.fullname + '/foam record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile8
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee8PayMulti(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee8.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile8 = employee8.fullname + '/multi record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile8
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee8PayVacation(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee8.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile8 = employee8.fullname + '/vacation record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile8
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee8PayShop(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee8.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile8 = employee8.fullname + '/Shop record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile8
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee8Payroll(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee8.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile8 = employee8.fullname + '/Payroll record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile8
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


# Employee 9
class Employee9Record(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x400")
        self.title(employee9.fullname)

        self.toplevel_window = None

        workRecordLabel = ctk.CTkLabel(self,
                                       text="Work Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        workRecordLabel.place(x=5, y=30)

        battWorkButton = ctk.CTkButton(self,
                                       text="Batt Work",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.battWorkButton_command)

        battWorkButton.place(x=5, y=65)

        atticWorkButton = ctk.CTkButton(self,
                                        text="Attic Work",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.atticWorkButton_command)

        atticWorkButton.place(x=5, y=100)

        sprayFoamButton = ctk.CTkButton(self,
                                        text="Spray Foam",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.sprayFoamButton_command)

        sprayFoamButton.place(x=5, y=135)

        multiDayButton = ctk.CTkButton(self,
                                       text="Multiple Day",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.multiDayButton_command)

        multiDayButton.place(x=5, y=170)

        shopPayButton = ctk.CTkButton(self,
                                      text="Shop Pay",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.shopPayButton_command)

        shopPayButton.place(x=5, y=205)

        payRecordsLabel = ctk.CTkLabel(self,
                                       text="Pay Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        payRecordsLabel.place(x=210, y=30)

        payrollButton = ctk.CTkButton(self,
                                      text="Payroll",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.payrollButton_command)

        payrollButton.place(x=210, y=65)

        vacationButton = ctk.CTkButton(self,
                                       text="Vacation Record",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.vacationButton_command)

        vacationButton.place(x=210, y=100)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=300)

    def battWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee9PayBatt(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def atticWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee9PayAttic(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def sprayFoamButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee9PayFoam(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def multiDayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee9PayMulti(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopPayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee9PayShop(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def payrollButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee9Payroll(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def vacationButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee9PayVacation(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def closeWindowButton_command(self):
        self.destroy()


class Employee9PayBatt(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee9.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile9 = employee9.fullname + '/batt record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile9
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee9PayAttic(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee9.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile9 = employee9.fullname + '/attic record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile9
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee9PayFoam(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee9.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile9 = employee9.fullname + '/foam record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile9
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee9PayMulti(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee9.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile9 = employee9.fullname + '/multi record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile9
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee9PayVacation(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee9.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile9 = employee9.fullname + '/vacation record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile9
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee9PayShop(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee9.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile9 = employee9.fullname + '/Shop record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile9
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee9Payroll(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee9.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile9 = employee9.fullname + '/Payroll record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile9
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


# Employee 10
class Employee10Record(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("400x400")
        self.title(employee10.fullname)

        self.toplevel_window = None

        workRecordLabel = ctk.CTkLabel(self,
                                       text="Work Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        workRecordLabel.place(x=5, y=30)

        battWorkButton = ctk.CTkButton(self,
                                       text="Batt Work",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.battWorkButton_command)

        battWorkButton.place(x=5, y=65)

        atticWorkButton = ctk.CTkButton(self,
                                        text="Attic Work",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.atticWorkButton_command)

        atticWorkButton.place(x=5, y=100)

        sprayFoamButton = ctk.CTkButton(self,
                                        text="Spray Foam",
                                        width=150,
                                        height=30,
                                        font=ctk.CTkFont(size=18, weight="normal"),
                                        command=self.sprayFoamButton_command)

        sprayFoamButton.place(x=5, y=135)

        multiDayButton = ctk.CTkButton(self,
                                       text="Multiple Day",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.multiDayButton_command)

        multiDayButton.place(x=5, y=170)

        shopPayButton = ctk.CTkButton(self,
                                      text="Shop Pay",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.shopPayButton_command)

        shopPayButton.place(x=5, y=205)

        payRecordsLabel = ctk.CTkLabel(self,
                                       text="Pay Records",
                                       width=150,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=18, weight="normal"))

        payRecordsLabel.place(x=210, y=30)

        payrollButton = ctk.CTkButton(self,
                                      text="Payroll",
                                      width=150,
                                      height=30,
                                      font=ctk.CTkFont(size=18, weight="normal"),
                                      command=self.payrollButton_command)

        payrollButton.place(x=210, y=65)

        vacationButton = ctk.CTkButton(self,
                                       text="Vacation Record",
                                       width=150,
                                       height=30,
                                       font=ctk.CTkFont(size=18, weight="normal"),
                                       command=self.vacationButton_command)

        vacationButton.place(x=210, y=100)

        closeWindowButton = ctk.CTkButton(self,
                                          text="Close Window",
                                          width=150,
                                          height=30,
                                          font=ctk.CTkFont(size=18, weight="normal"),
                                          command=self.closeWindowButton_command)

        closeWindowButton.place(x=5, y=300)

    def battWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee10PayBatt(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def atticWorkButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee10PayAttic(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def sprayFoamButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee10PayFoam(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def multiDayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee10PayMulti(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopPayButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee1PayShop(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def payrollButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee10Payroll(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def vacationButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Employee10PayVacation(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def closeWindowButton_command(self):
        self.destroy()


class Employee10PayBatt(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee10.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile10 = employee10.fullname + '/batt record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile10
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee10PayAttic(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee10.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile10 = employee10.fullname + '/attic record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile10
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee10PayFoam(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee10.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile10 = employee10.fullname + '/foam record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile10
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee10PayMulti(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee10.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile10 = employee10.fullname + '/multi record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile10
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee10PayVacation(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee10.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile10 = employee10.fullname + '/vacation record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile10
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee10PayShop(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee10.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile10 = employee10.fullname + '/Shop record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile10
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


class Employee10Payroll(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title(employee10.fullname)
        width = 800
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)
        self.attributes('-fullscreen', False)
        self.config(background='white')
        ft = ctk.CTkFont(size=18, weight="normal")

        style = ttk.Style()
        style.theme_use('clam')

        self.tv1 = ttk.Treeview(self)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(
            self, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(
            self, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        self.Load_excel_data()

        closeWindowBtn = tk.Button(self)
        closeWindowBtn["bg"] = "#f0f0f0"
        closeWindowBtn["font"] = ft
        closeWindowBtn["fg"] = "#000000"
        closeWindowBtn["justify"] = "center"
        closeWindowBtn["text"] = "Close"
        closeWindowBtn["relief"] = "solid"
        closeWindowBtn.pack(side='bottom')
        closeWindowBtn["command"] = self.closeWindowBtn_command

    def closeWindowBtn_command(self):
        self.destroy()

    # noinspection PyTypeChecker
    def Load_excel_data(self):
        empFile10 = employee10.fullname + '/Payroll record.xlsx'
        file_path = "D:/Insulpro/employee paytree/" + empFile10
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            self.tv1.insert("", "end", values=row)
        return None


# Main Window Section
class MainWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title('InsulPro Insulation')
        width = 600
        height = 600
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)

        # Menu Bar-Main Window
        self.menubar = tk.Menu(self)
        self.config(menu=self.menubar)

        self.fileMenu = tk.Menu(self.menubar, tearoff=0, font=("Comic San", 14))
        self.menubar.add_cascade(label="File", menu=self.fileMenu)
        self.fileMenu.add_command(label='Open File', command=self.open_command)
        self.fileMenu.add_command(label="Exit", command=self.quit)

        self.toolMenu = tk.Menu(self.menubar, tearoff=0, font=("Comic San MS", 14))
        self.menubar.add_cascade(label="Tools", menu=self.toolMenu)
        self.toolMenu.add_command(label='Calculator', command=self.calculator)

        # Create Pay Sheet Widgets
        paySheetLabel = ctk.CTkLabel(self,
                                     text="Create Pay Sheet",
                                     width=175,
                                     height=30,
                                     fg_color="transparent",
                                     font=ctk.CTkFont(size=20, weight="normal"))
        paySheetLabel.place(x=10, y=0)

        battPaySheetButton = ctk.CTkButton(self,
                                           text="Batt Pay Sheet",
                                           width=175,
                                           height=30,
                                           font=ctk.CTkFont(size=17),
                                           command=self.battPaySheetButton_command)
        battPaySheetButton.place(x=5, y=30)

        atticPaySheetButton = ctk.CTkButton(self,
                                            text="Attic Pay Sheet",
                                            width=175,
                                            height=30,
                                            font=ctk.CTkFont(size=17),
                                            command=self.atticPaySheetButton_command)
        atticPaySheetButton.place(x=5, y=70)

        foamPaySheetButton = ctk.CTkButton(self,
                                           text="Foam Pay Sheet",
                                           width=175,
                                           height=30,
                                           font=ctk.CTkFont(size=17),
                                           command=self.sprayFoamPaySheetButton_command)
        foamPaySheetButton.place(x=5, y=110)

        multiPaySheetButton = ctk.CTkButton(self,
                                            text='Multiple Day Split',
                                            width=175,
                                            height=30,
                                            font=ctk.CTkFont(size=17),
                                            command=self.multiPaySheetButton_command)
        multiPaySheetButton.place(x=5, y=150)

        vacationRequestButton = ctk.CTkButton(self,
                                              text="Vacation Request",
                                              width=175,
                                              height=30,
                                              font=ctk.CTkFont(size=17),
                                              command=self.vacationRequestButton_command)
        vacationRequestButton.place(x=5, y=190)

        # Create Work Order Widgets
        workOrderLabel = ctk.CTkLabel(self,
                                      text="Create Work Order",
                                      width=150,
                                      height=30,
                                      fg_color="transparent",
                                      font=ctk.CTkFont(size=20, weight="normal"))
        workOrderLabel.place(x=200, y=0)

        newConstructionButton = ctk.CTkButton(self,
                                              text="New Construction",
                                              width=175,
                                              height=30,
                                              font=ctk.CTkFont(size=17),
                                              command=self.newConstructionButton_command)
        newConstructionButton.place(x=205, y=30)

        existConstructionButton = ctk.CTkButton(self,
                                                text="Existing Construction",
                                                width=175,
                                                height=30,
                                                font=ctk.CTkFont(size=17),
                                                command=self.existConstructionButton_command)
        existConstructionButton.place(x=205, y=70)

        atticWorkOrderButton = ctk.CTkButton(self,
                                             text="Attic Work Order",
                                             width=175,
                                             height=30,
                                             font=ctk.CTkFont(size=17),
                                             command=self.atticWorkOrderButton_command)
        atticWorkOrderButton.place(x=205, y=110)

        foamWorkOrderButton = ctk.CTkButton(self,
                                            text="Spray Foam",
                                            width=175,
                                            height=30,
                                            font=ctk.CTkFont(size=17),
                                            command=self.foamWorkOrderButton_command)
        foamWorkOrderButton.place(x=205, y=150)

        cellWorkOrderButton = ctk.CTkButton(self,
                                            text="Cellulose",
                                            width=175,
                                            height=30,
                                            font=ctk.CTkFont(size=17),
                                            command=self.cellWorkOrderButton_command)
        cellWorkOrderButton.place(x=205, y=190)

        # Shop work order and pay widgets
        shopTicketLabel = ctk.CTkLabel(self,
                                       text="Shop Tickets",
                                       width=175,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=20, weight="normal"))
        shopTicketLabel.place(x=5, y=230)

        shopWorkOrderButton = ctk.CTkButton(self,
                                            text="Shop Work Order",
                                            width=175,
                                            height=30,
                                            font=ctk.CTkFont(size=17),
                                            command=self.shopWorkOrderButton_command)
        shopWorkOrderButton.place(x=5, y=270)

        shopPaySheetButton = ctk.CTkButton(self,
                                           text="Shop Pay Sheet",
                                           width=175, height=30,
                                           font=ctk.CTkFont(size=17),
                                           command=self.shopPaySheetButton_command)
        shopPaySheetButton.place(x=5, y=310)

        # Attendance widgets
        attendanceLabel = ctk.CTkLabel(self,
                                       text="Attendance",
                                       width=175,
                                       height=30,
                                       fg_color="transparent",
                                       font=ctk.CTkFont(size=20, weight="normal"))
        attendanceLabel.place(x=205, y=230)

        timeSheetButton = ctk.CTkButton(self,
                                        text="Daily Time Sheet",
                                        width=175,
                                        height=30,
                                        font=ctk.CTkFont(size=17),
                                        command=self.timeSheetButton_command)
        timeSheetButton.place(x=205, y=270)

        # admin tool section
        adminLabel = ctk.CTkLabel(self,
                                  text="Admin Tools",
                                  width=150,
                                  height=30,
                                  fg_color="transparent",
                                  font=ctk.CTkFont(size=20, weight="normal"))
        adminLabel.place(x=400, y=0)

        createPayrollButton = ctk.CTkButton(self,
                                            text="Create Payroll File",
                                            width=175, height=30,
                                            font=ctk.CTkFont(size=17),
                                            command=self.savePayrollRecordButton_command)
        createPayrollButton.place(x=400, y=30)

        createBackupButton = ctk.CTkButton(self,
                                           text="Backup All Files",
                                           width=175, height=30,
                                           font=ctk.CTkFont(size=17),
                                           command=self.saveBackupButton_command)
        createBackupButton.place(x=400, y=65)

        createEmpPayrollButton = ctk.CTkButton(self,
                                               text="Employee Payroll",
                                               width=175, height=30,
                                               font=ctk.CTkFont(size=17),
                                               command=self.payrollPaytreeEntryButton_command)
        createEmpPayrollButton.place(x=400, y=100)

        self.appearance_mode_label = ctk.CTkLabel(self,
                                                  text="Appearance Mode:",
                                                  width=175,
                                                  height=30,
                                                  fg_color="transparent",
                                                  font=ctk.CTkFont(size=20, weight="normal"))
        self.appearance_mode_label.place(x=5, y=490)

        self.appearance_mode_optionmenu = ctk.CTkOptionMenu(self,
                                                            values=["Light", "Dark", "System"],
                                                            command=self.change_appearance_mode_event)

        self.appearance_mode_optionmenu.place(x=5, y=520)

        self.toplevel_window = None
        self.createStarterFiles()

    @staticmethod
    def change_appearance_mode_event(new_appearance_mode: str):
        ctk.set_appearance_mode(new_appearance_mode)

    @staticmethod
    def open_command():
        filepath_main = filedialog.askopenfilename(
            initialdir="D:/Insulpro",
            title="Open Files",
            filetypes=(("excel files", "*.xlsx"), ("all files", "*.*")))

        subprocess.Popen([filepath_main], shell=True)

    def calculator(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = Calculator(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def battPaySheetButton_command(self):
        window = BattPaySheet(self)
        window.grab_set()

    def atticPaySheetButton_command(self):
        window = AtticPaySheet(self)
        window.grab_set()

    def sprayFoamPaySheetButton_command(self):
        window = SprayFoamPaySheet(self)
        window.grab_set()

    def multiPaySheetButton_command(self):
        window = MultipleDayPaySheet(self)
        window.grab_set()

    def vacationRequestButton_command(self):
        window = VacationRequest(self)
        window.grab_set()

    def newConstructionButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = NewConstructionWorkOrder(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def existConstructionButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = ExistConstructionWorkOrder(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def atticWorkOrderButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = AtticWorkOrder(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def foamWorkOrderButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = FoamWorkOrder(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def cellWorkOrderButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = CelluloseWorkOrder(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopWorkOrderButton_command(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = ShopWorkOrder(self)  # create window if its None or destroyed
        else:
            self.toplevel_window.focus()  # if window exists focus it

    def shopPaySheetButton_command(self):
        window = ShopPaySheet(self)
        window.grab_set()

    def timeSheetButton_command(self):
        window = DailyTimeSheet(self)
        window.grab_set()

    def payrollPaytreeEntryButton_command(self):
        window = PayrollPaytreeEntry(self)
        window.grab_set()

    def createStarterFiles(self):
        print('Starting Program Sequence.....')
        self.createDirectory_command()
        self.createPaytreeDirectory()
        self.createEmployeeRecords()
        self.createEmployeePaytreeBatts()
        self.createEmployeePaytreeAttic()
        self.createEmployeePaytreeFoam()
        self.createEmployeePaytreeMulti()
        self.createEmployeePaytreeVacation()
        self.createEmployeePaytreeShop()
        self.createEmployeePaytreePayroll()
        self.createJobRecords()
        self.createVacationRecord()
        self.createPayrollFile()
        self.createTimeSheetRecord()

        print('Program Running>>>>>')
        print("")

    @staticmethod
    def createDirectory_command():
        print("Creating A Working Directory")
        dir_pathtemp = "D:/Insulpro/temp"
        dir_path00 = "D:/Insulpro/temp/work orders/"
        dir_path0 = "D:/Insulpro/temp/paysheets/"
        dir_path1 = "D:/Insulpro/temp/paysheets/attic pay sheet"
        dir_path2 = "D:/Insulpro/temp/work orders/attic work orders"
        dir_path3 = "D:/Insulpro/temp/paysheets/batt pay sheet"
        dir_path4 = "D:/Insulpro/temp/work orders/new work orders"
        dir_path5 = "D:/Insulpro/temp/work orders/existing work orders"
        dir_path6 = "D:/Insulpro/temp/paysheets/foam pay sheet"
        dir_path7 = "D:/Insulpro/temp/work orders/foam work orders"
        dir_path8 = "D:/Insulpro/temp/vacation"
        dir_path9 = "D:/Insulpro/employee paytree"
        dir_path10 = "D:/Insulpro/employee work records"
        dir_path11 = "D:/Insulpro/payroll records"
        dir_path12 = "D:/Insulpro/temp/weekly payroll"
        dir_path13 = "D:/Insulpro/temp/paysheets/shop pay sheet"
        dir_path14 = "D:/Insulpro/payroll records/time sheets"
        dir_path15 = "D:/Insulpro/temp/work orders/shop work orders"

        if not os.path.exists(dir_pathtemp):
            os.mkdir(dir_pathtemp)

        if not os.path.exists(dir_path00):
            os.mkdir(dir_path00)

        if not os.path.exists(dir_path0):
            os.mkdir(dir_path0)

        if not os.path.exists(dir_path1):
            os.mkdir(dir_path1)

        if not os.path.exists(dir_path2):
            os.mkdir(dir_path2)

        if not os.path.exists(dir_path3):
            os.mkdir(dir_path3)

        if not os.path.exists(dir_path4):
            os.mkdir(dir_path4)

        if not os.path.exists(dir_path5):
            os.mkdir(dir_path5)

        if not os.path.exists(dir_path6):
            os.mkdir(dir_path6)

        if not os.path.exists(dir_path7):
            os.mkdir(dir_path7)

        if not os.path.exists(dir_path8):
            os.mkdir(dir_path8)

        if not os.path.exists(dir_path9):
            os.mkdir(dir_path9)

        if not os.path.exists(dir_path10):
            os.mkdir(dir_path10)

        if not os.path.exists(dir_path11):
            os.mkdir(dir_path11)

        if not os.path.exists(dir_path12):
            os.mkdir(dir_path12)

        if not os.path.exists(dir_path13):
            os.mkdir(dir_path13)

        if not os.path.exists(dir_path14):
            os.mkdir(dir_path14)

        if not os.path.exists(dir_path15):
            os.mkdir(dir_path15)
        else:
            print('Working Directory Structure Exists...')

    @staticmethod
    def createPaytreeDirectory():
        print("Creating Paytree Directory")
        dir_Path1 = "D:/Insulpro/employee paytree/" + employee1.fullname
        dir_Path2 = "D:/Insulpro/employee paytree/" + employee2.fullname
        dir_Path3 = "D:/Insulpro/employee paytree/" + employee3.fullname
        dir_Path4 = "D:/Insulpro/employee paytree/" + employee4.fullname
        dir_Path5 = "D:/Insulpro/employee paytree/" + employee5.fullname
        dir_Path6 = "D:/Insulpro/employee paytree/" + employee6.fullname
        dir_Path7 = "D:/Insulpro/employee paytree/" + employee7.fullname
        dir_Path8 = "D:/Insulpro/employee paytree/" + employee8.fullname
        dir_Path9 = "D:/Insulpro/employee paytree/" + employee9.fullname
        dir_Path10 = "D:/Insulpro/employee paytree/" + employee10.fullname

        if not os.path.exists(dir_Path1):
            os.mkdir(dir_Path1)

        if not os.path.exists(dir_Path2):
            os.mkdir(dir_Path2)

        if not os.path.exists(dir_Path3):
            os.mkdir(dir_Path3)

        if not os.path.exists(dir_Path4):
            os.mkdir(dir_Path4)

        if not os.path.exists(dir_Path5):
            os.mkdir(dir_Path5)

        if not os.path.exists(dir_Path6):
            os.mkdir(dir_Path6)

        if not os.path.exists(dir_Path7):
            os.mkdir(dir_Path7)

        if not os.path.exists(dir_Path8):
            os.mkdir(dir_Path8)

        if not os.path.exists(dir_Path9):
            os.mkdir(dir_Path9)

        if not os.path.exists(dir_Path10):
            os.mkdir(dir_Path10)

        else:
            print('Employee Paytree Directory Exists')

    @staticmethod
    def createEmployeeRecords():
        print("Creating Employee Records")
        empFile1 = str(employee1.fullname) + ".xlsx"
        empFile2 = str(employee2.fullname) + ".xlsx"
        empFile3 = str(employee3.fullname) + ".xlsx"
        empFile4 = str(employee4.fullname) + ".xlsx"
        empFile5 = str(employee5.fullname) + ".xlsx"
        empFile6 = str(employee6.fullname) + ".xlsx"
        empFile7 = str(employee7.fullname) + ".xlsx"
        empFile8 = str(employee8.fullname) + ".xlsx"
        empFile9 = str(employee9.fullname) + ".xlsx"
        empFile10 = str(employee10.fullname) + ".xlsx"

        if not os.path.exists("D:/Insulpro/employee work records/" + empFile1):
            source = "D:/Insulpro/system/spreadsheet/Employee Work Record.xlsx"

            destination = "D:/Insulpro/employee work records/" + empFile1

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee work records/" + empFile2):
            source = "D:/Insulpro/system/spreadsheet/Employee Work Record.xlsx"

            destination = "D:/Insulpro/employee work records/" + empFile2

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee work records/" + empFile3):
            source = "D:/Insulpro/system/spreadsheet/Employee Work Record.xlsx"

            destination = "D:/Insulpro/employee work records/" + empFile3

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee work records/" + empFile4):
            source = "D:/Insulpro/system/spreadsheet/Employee Work Record.xlsx"

            destination = "D:/Insulpro/employee work records/" + empFile4

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee work records/" + empFile5):
            source = "D:/Insulpro/system/spreadsheet/Employee Work Record.xlsx"

            destination = "D:/Insulpro/employee work records/" + empFile5

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee work records/" + empFile6):
            source = "D:/Insulpro/system/spreadsheet/Employee Work Record.xlsx"

            destination = "D:/Insulpro/employee work records/" + empFile6

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee work records/" + empFile7):
            source = "D:/Insulpro/system/spreadsheet/Employee Work Record.xlsx"

            destination = "D:/Insulpro/employee work records/" + empFile7

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee work records/" + empFile8):
            source = "D:/Insulpro/system/spreadsheet/Employee Work Record.xlsx"

            destination = "D:/Insulpro/employee work records/" + empFile8

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee work records/" + empFile9):
            source = "D:/Insulpro/system/spreadsheet/Employee Work Record.xlsx"

            destination = "D:/Insulpro/employee work records/" + empFile9

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee work records/" + empFile10):
            source = "D:/Insulpro/system/spreadsheet/Employee Work Record.xlsx"

            destination = "D:/Insulpro/employee work records/" + empFile10

            shutil.copy(source, destination)

            print('Creating Employee Work Records')

        else:
            print('Employee Work Records Directories Exists....')

    @staticmethod
    def createJobRecords():
        print("Creating Work Record")
        if not os.path.exists("D:/Insulpro/payroll records/Work Record.xlsx"):
            source = "D:/Insulpro/system/spreadsheet/Work Record.xlsx"

            destination = "D:/Insulpro/payroll records/Work Record.xlsx"

            shutil.copy(source, destination)

        else:
            print('Work Record Exists....')

    @staticmethod
    def createVacationRecord():
        print("Creating Vacation Record")
        if not os.path.exists("D:/Insulpro/payroll records/Vacation Record.xlsx"):

            source = "D:/Insulpro/system/spreadsheet/Vacation Record.xlsx"

            destination = "D:/Insulpro/payroll records/Vacation Record.xlsx"

            shutil.copy(source, destination)

            workbook = load_workbook(
                'D:/Insulpro/payroll records/Vacation Record.xlsx')

            sheet = workbook['Vacation Tracker']

            sheet['A4'] = employee1.fullname
            sheet['B4'] = employee1.vacationdays

            sheet['A5'] = employee2.fullname
            sheet['B5'] = employee2.vacationdays

            sheet['A6'] = employee3.fullname
            sheet['B6'] = employee3.vacationdays

            sheet['A7'] = employee4.fullname
            sheet['B7'] = employee4.vacationdays

            sheet['A8'] = employee5.fullname
            sheet['B8'] = employee5.vacationdays

            sheet['A9'] = employee6.fullname
            sheet['B9'] = employee6.vacationdays

            sheet['A10'] = employee7.fullname
            sheet['B10'] = employee7.vacationdays

            sheet['A11'] = employee8.fullname
            sheet['B11'] = employee8.vacationdays

            sheet['A12'] = employee9.fullname
            sheet['B12'] = employee9.vacationdays

            sheet['A13'] = employee10.fullname
            sheet['B13'] = employee10.vacationdays

            sheet1 = workbook['Emp1']
            sheet2 = workbook['Emp2']
            sheet3 = workbook['Emp3']
            sheet4 = workbook['Emp4']
            sheet5 = workbook['Emp5']
            sheet6 = workbook['Emp6']
            sheet7 = workbook['Emp7']
            sheet8 = workbook['Emp8']
            sheet9 = workbook['Emp9']
            sheet10 = workbook['Emp10']

            sheet1['B3'] = employee1.fullname
            sheet2['B3'] = employee2.fullname
            sheet3['B3'] = employee3.fullname
            sheet4['B3'] = employee4.fullname
            sheet5['B3'] = employee5.fullname
            sheet6['B3'] = employee6.fullname
            sheet7['B3'] = employee7.fullname
            sheet8['B3'] = employee8.fullname
            sheet9['B3'] = employee9.fullname
            sheet10['B3'] = employee10.fullname

            workbook.save('D:/Insulpro/payroll records/Vacation Record.xlsx')

            workbook.close()

        else:
            print('Vacation Record Exists....')

    @staticmethod
    def createPayrollFile():
        print("Creating Payroll File")
        if not os.path.exists("D:/Insulpro/temp/weekly payroll/payroll.xlsx"):
            source = "D:/Insulpro/system/spreadsheet/payroll spreadsheet.xlsx"

            destination = "D:/Insulpro/temp/weekly payroll/payroll.xlsx"

            shutil.copy(source, destination)

            workbook = load_workbook("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

            sheet = workbook['Emp1&Emp2']
            sheet_1 = workbook['Emp3&Emp4']
            sheet_2 = workbook['Emp5&Emp6']
            sheet_3 = workbook['Emp7&Emp8']
            sheet_4 = workbook['Emp9&Emp10']
            sheet_8 = workbook['Pay Rates']

            sheet['B1'] = employee1.fullname
            sheet['B19'] = employee2.fullname
            sheet_1['B1'] = employee3.fullname
            sheet_1['B19'] = employee4.fullname
            sheet_2['B1'] = employee5.fullname
            sheet_2['B19'] = employee6.fullname
            sheet_3['B1'] = employee7.fullname
            sheet_3['B19'] = employee8.fullname
            sheet_4['B1'] = employee9.fullname
            sheet_4['B19'] = employee10.fullname

            sheet_8['A3'] = batt1.name
            sheet_8['B3'] = batt1.rate
            sheet_8['A4'] = batt2.name
            sheet_8['B4'] = batt2.rate
            sheet_8['A5'] = existing_soffit.name
            sheet_8['B5'] = existing_soffit.rate
            sheet_8['A6'] = bibs_full.name
            sheet_8['B6'] = bibs_full.rate
            sheet_8['A7'] = bibs_hung.name
            sheet_8['B7'] = bibs_hung.rate
            sheet_8['A8'] = bibs_tacked.name
            sheet_8['B8'] = bibs_tacked.rate
            sheet_8['A9'] = bibs_blown.name
            sheet_8['B9'] = bibs_blown.rate
            sheet_8['A10'] = r19_blown.name
            sheet_8['B10'] = r19_blown.rate
            sheet_8['A11'] = r30_blown.name
            sheet_8['B11'] = r30_blown.rate
            sheet_8['A12'] = r38_blown.name
            sheet_8['B12'] = r38_blown.rate
            sheet_8['A13'] = r49_blown.name
            sheet_8['B13'] = r49_blown.rate
            sheet_8['A14'] = cell_blown.name
            sheet_8['B14'] = cell_blown.rate
            sheet_8['A15'] = closed_under1.name
            sheet_8['B15'] = closed_under1.rate
            sheet_8['A16'] = closed_under2.name
            sheet_8['B16'] = closed_under2.rate
            sheet_8['A17'] = closed_3.name
            sheet_8['B17'] = closed_3.rate
            sheet_8['A18'] = open_cell6.name
            sheet_8['B18'] = open_cell6.rate
            sheet_8['A19'] = open_cell8.name
            sheet_8['B19'] = open_cell8.rate
            sheet_8['A20'] = existing_soffit.name
            sheet_8['B20'] = existing_soffit.rate
            sheet_8['A21'] = vacation_day.name
            sheet_8['B21'] = vacation_day.rate
            sheet_8['A22'] = closed_under34.name
            sheet_8['B22'] = closed_under34.rate
            sheet_8['A23'] = open_cell4.name
            sheet_8['B23'] = open_cell4.rate

            workbook.save("D:/Insulpro/temp/weekly payroll/payroll.xlsx")

            workbook.close()

        else:
            print('Payroll Spreadsheet Exists....')

    @staticmethod
    def createEmployeePaytreeBatts():
        print("Creating Employee Paytree Batts")
        empFile1 = str(employee1.fullname) + "/batt record.xlsx"
        empFile2 = str(employee2.fullname) + "/batt record.xlsx"
        empFile3 = str(employee3.fullname) + "/batt record.xlsx"
        empFile4 = str(employee4.fullname) + "/batt record.xlsx"
        empFile5 = str(employee5.fullname) + "/batt record.xlsx"
        empFile6 = str(employee6.fullname) + "/batt record.xlsx"
        empFile7 = str(employee7.fullname) + "/batt record.xlsx"
        empFile8 = str(employee8.fullname) + "/batt record.xlsx"
        empFile9 = str(employee9.fullname) + "/batt record.xlsx"
        empFile10 = str(employee10.fullname) + "/batt record.xlsx"

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile1):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Batt.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile1

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile2):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Batt.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile2

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile3):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Batt.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile3

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile4):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Batt.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile4

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile5):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Batt.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile5

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile6):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Batt.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile6

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile7):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Batt.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile7

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile8):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Batt.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile8

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile9):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Batt.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile9

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile10):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Batt.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile10

            shutil.copy(source, destination)
        else:
            print('Employee Paytree Batts Exists......')

    @staticmethod
    def createEmployeePaytreeAttic():
        print("Creating Employee Paytree Attic")
        empFile1 = str(employee1.fullname) + "/attic record.xlsx"
        empFile2 = str(employee2.fullname) + "/attic record.xlsx"
        empFile3 = str(employee3.fullname) + "/attic record.xlsx"
        empFile4 = str(employee4.fullname) + "/attic record.xlsx"
        empFile5 = str(employee5.fullname) + "/attic record.xlsx"
        empFile6 = str(employee6.fullname) + "/attic record.xlsx"
        empFile7 = str(employee7.fullname) + "/attic record.xlsx"
        empFile8 = str(employee8.fullname) + "/attic record.xlsx"
        empFile9 = str(employee9.fullname) + "/attic record.xlsx"
        empFile10 = str(employee10.fullname) + "/attic record.xlsx"

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile1):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Attic.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile1

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile2):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Attic.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile2

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile3):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Attic.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile3

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile4):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Attic.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile4

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile5):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Attic.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile5

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile6):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Attic.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile6

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile7):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Attic.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile7

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile8):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Attic.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile8

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile9):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Attic.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile9

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile10):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Attic.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile10

            shutil.copy(source, destination)

        else:
            print('Employee Paytree Attics Exists....')

    @staticmethod
    def createEmployeePaytreeFoam():
        print("Creating Employee Paytree Foam")
        empFile1 = str(employee1.fullname) + "/foam record.xlsx"
        empFile2 = str(employee2.fullname) + "/foam record.xlsx"
        empFile3 = str(employee3.fullname) + "/foam record.xlsx"
        empFile4 = str(employee4.fullname) + "/foam record.xlsx"
        empFile5 = str(employee5.fullname) + "/foam record.xlsx"
        empFile6 = str(employee6.fullname) + "/foam record.xlsx"
        empFile7 = str(employee7.fullname) + "/foam record.xlsx"
        empFile8 = str(employee8.fullname) + "/foam record.xlsx"
        empFile9 = str(employee9.fullname) + "/foam record.xlsx"
        empFile10 = str(employee10.fullname) + "/foam record.xlsx"

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile1):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Foam.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile1

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile2):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Foam.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile2

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile3):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Foam.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile3

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile4):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Foam.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile4

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile5):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Foam.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile5

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile6):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Foam.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile6

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile7):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Foam.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile7

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile8):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Foam.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile8

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile9):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Foam.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile9

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile10):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Foam.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile10

            shutil.copy(source, destination)

        else:
            print('Employee Paytree Foam Exists.....')

    @staticmethod
    def createEmployeePaytreeMulti():
        print("Creating Employee Paytree Multi")
        empFile1 = str(employee1.fullname) + "/multi record.xlsx"
        empFile2 = str(employee2.fullname) + "/multi record.xlsx"
        empFile3 = str(employee3.fullname) + "/multi record.xlsx"
        empFile4 = str(employee4.fullname) + "/multi record.xlsx"
        empFile5 = str(employee5.fullname) + "/multi record.xlsx"
        empFile6 = str(employee6.fullname) + "/multi record.xlsx"
        empFile7 = str(employee7.fullname) + "/multi record.xlsx"
        empFile8 = str(employee8.fullname) + "/multi record.xlsx"
        empFile9 = str(employee9.fullname) + "/multi record.xlsx"
        empFile10 = str(employee10.fullname) + "/multi record.xlsx"

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile1):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Multi.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile1

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile2):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Multi.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile2

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile3):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Multi.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile3

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile4):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Multi.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile4

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile5):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Multi.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile5

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile6):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Multi.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile6

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile7):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Multi.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile7

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile8):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Multi.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile8

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile9):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Multi.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile9

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile10):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Multi.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile10

            shutil.copy(source, destination)

        else:
            print('Employee Paytree Multi Exists.....')

    @staticmethod
    def createEmployeePaytreeVacation():
        print("Creating Employee Paytree Vacation")
        empFile1 = str(employee1.fullname) + "/vacation record.xlsx"
        empFile2 = str(employee2.fullname) + "/vacation record.xlsx"
        empFile3 = str(employee3.fullname) + "/vacation record.xlsx"
        empFile4 = str(employee4.fullname) + "/vacation record.xlsx"
        empFile5 = str(employee5.fullname) + "/vacation record.xlsx"
        empFile6 = str(employee6.fullname) + "/vacation record.xlsx"
        empFile7 = str(employee7.fullname) + "/vacation record.xlsx"
        empFile8 = str(employee8.fullname) + "/vacation record.xlsx"
        empFile9 = str(employee9.fullname) + "/vacation record.xlsx"
        empFile10 = str(employee10.fullname) + "/vacation record.xlsx"

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile1):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Vacation.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile1

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile2):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Vacation.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile2

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile3):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Vacation.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile3

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile4):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Vacation.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile4

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile5):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Vacation.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile5

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile6):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Vacation.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile6

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile7):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Vacation.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile7

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile8):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Vacation.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile8

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile9):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Vacation.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile9

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile10):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Vacation.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile10

            shutil.copy(source, destination)

        else:
            print('Employee Paytree Vacation Exists....')

    @staticmethod
    def createEmployeePaytreeShop():
        print("Creating Employee Paytree Shop")
        empFile1 = str(employee1.fullname) + "/Shop record.xlsx"
        empFile2 = str(employee2.fullname) + "/Shop record.xlsx"
        empFile3 = str(employee3.fullname) + "/Shop record.xlsx"
        empFile4 = str(employee4.fullname) + "/Shop record.xlsx"
        empFile5 = str(employee5.fullname) + "/Shop record.xlsx"
        empFile6 = str(employee6.fullname) + "/Shop record.xlsx"
        empFile7 = str(employee7.fullname) + "/Shop record.xlsx"
        empFile8 = str(employee8.fullname) + "/Shop record.xlsx"
        empFile9 = str(employee9.fullname) + "/Shop record.xlsx"
        empFile10 = str(employee10.fullname) + "/Shop record.xlsx"

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile1):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Shop.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile1

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile2):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Shop.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile2

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile3):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Shop.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile3

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile4):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Shop.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile4

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile5):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Shop.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile5

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile6):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Shop.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile6

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile7):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Shop.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile7

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile8):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Shop.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile8

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile9):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Shop.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile9

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile10):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Shop.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile10

            shutil.copy(source, destination)

        else:
            print('Employee Paytree Shop Exists....')

    @staticmethod
    def createEmployeePaytreePayroll():
        print("Creating Employee Paytree Payroll ")
        empFile1 = str(employee1.fullname) + "/Payroll record.xlsx"
        empFile2 = str(employee2.fullname) + "/Payroll record.xlsx"
        empFile3 = str(employee3.fullname) + "/Payroll record.xlsx"
        empFile4 = str(employee4.fullname) + "/Payroll record.xlsx"
        empFile5 = str(employee5.fullname) + "/Payroll record.xlsx"
        empFile6 = str(employee6.fullname) + "/Payroll record.xlsx"
        empFile7 = str(employee7.fullname) + "/Payroll record.xlsx"
        empFile8 = str(employee8.fullname) + "/Payroll record.xlsx"
        empFile9 = str(employee9.fullname) + "/Payroll record.xlsx"
        empFile10 = str(employee10.fullname) + "/Payroll record.xlsx"

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile1):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Payroll.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile1

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile2):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Payroll.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile2

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile3):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Payroll.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile3

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile4):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Payroll.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile4

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile5):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Payroll.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile5

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile6):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Payroll.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile6

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile7):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Payroll.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile7

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile8):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Payroll.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile8

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile9):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Payroll.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile9

            shutil.copy(source, destination)

        if not os.path.exists("D:/Insulpro/employee paytree/" + empFile10):
            source = "D:/Insulpro/system/spreadsheet/Employee Paytree Payroll.xlsx"

            destination = "D:/Insulpro/employee paytree/" + empFile10

            shutil.copy(source, destination)

        else:
            print('Employee Paytree Payroll  Exists....')

    @staticmethod
    def createTimeSheetRecord():
        print("Creating Time Sheet")
        dateToday = datetime.datetime.today()
        namefile = dateToday.strftime("%A%m%d%y")
        filename = str(namefile) + ".xlsx"

        if not os.path.exists(
                "D:/Insulpro/payroll records/time sheets/" + filename):
            source = "D:/Insulpro/system/spreadsheet/Daily Time Sheet.xlsx"
            destination = ("D:/Insulpro/payroll records/time sheets/Daily Time "
                           "Sheet.xlsx")

            shutil.copy(source, destination)

            os.rename(
                "D:/Insulpro/payroll records/time sheets/Daily Time Sheet.xlsx",
                "D:/Insulpro/payroll records/time sheets/" + filename)

            wb = load_workbook(
                "D:/Insulpro/payroll records/time sheets/" + filename)

            sheet = wb['DailyTimeSheet']

            sheet['D1'] = str(namefile)

            sheet['A4'] = employee1.fullname
            sheet['A5'] = employee2.fullname
            sheet['A6'] = employee3.fullname
            sheet['A7'] = employee4.fullname
            sheet['A8'] = employee5.fullname
            sheet['A9'] = employee6.fullname
            sheet['A10'] = employee7.fullname
            sheet['A11'] = employee8.fullname
            sheet['A12'] = employee9.fullname
            sheet['A13'] = employee10.fullname

            wb.save("D:/Insulpro/payroll records/time sheets/" + filename)

            wb.close()
        else:
            print('Daily Time Sheet Already Exists')

    def savePayrollRecordButton_command(self):
        print('Saving Weekly Payroll')
        answer = messagebox.askyesno(title='Save Payroll Record',
                                     message="Would You Really Like To Save Payroll Record?")
        if answer:
            print('Creating Payroll Folder')
            self.copyBattPaySheet()
            self.copyAtticPaySheet()
            self.copyFoamPaySheet()
            self.copyVacation()
            self.copyShopPaySheet()
            self.copyNewWorkOrders()
            self.copyExistWorkOrders()
            self.copyAtticWorkOrders()
            self.copyFoamWorkOrders()
            self.copyShopWorkOrders()
            self.copyPayrollFolder()

            self.clearPayrollDirectory()

            messagebox.showinfo(title='Success', message='Payroll Record Created Successfully')

    def saveBackupButton_command(self):
        print('Backing Up All Files')
        answer = messagebox.askyesno(title='Backup Files',
                                     message="Would You Really Like To Backup All Files?")
        if answer:
            print('Creating Backup Folder')
            self.copyPayrollRecords()
            self.copyEmployeeRecords()
            self.copyEmployeePaytree()
            self.clearWorkRecordDirectory()

            messagebox.showinfo(title='Success', message='All Files Backed Up Successfully')

    @staticmethod
    def copyBattPaySheet():
        source = "D:/Insulpro/temp/paysheets/batt pay sheet/"

        destination = "D:/Insulpro/temp/weekly payroll/paysheets/batt pay sheet/"

        shutil.copytree(source, destination)

    @staticmethod
    def copyAtticPaySheet():
        source = "D:/Insulpro/temp/paysheets/attic pay sheet/"

        destination = "D:/Insulpro/temp/weekly payroll/paysheets/attic pay sheet/"

        shutil.copytree(source, destination)

    @staticmethod
    def copyFoamPaySheet():
        source = "D:/Insulpro/temp/paysheets/foam pay sheet/"

        destination = "D:/Insulpro/temp/weekly payroll/paysheets/foam pay sheet/"

        shutil.copytree(source, destination)

    @staticmethod
    def copyVacation():
        source = "D:/Insulpro/temp/vacation/"

        destination = "D:/Insulpro/temp/weekly payroll/vacation/"

        shutil.copytree(source, destination)

    @staticmethod
    def copyShopPaySheet():
        source = "D:/Insulpro/temp/paysheets/shop pay sheet/"

        destination = "D:/Insulpro/temp/weekly payroll/paysheets/shop pay sheet/"

        shutil.copytree(source, destination)

    @staticmethod
    def copyNewWorkOrders():
        source = "D:/Insulpro/temp/work orders/new work orders/"

        destination = "D:/Insulpro/temp/weekly payroll/work orders/new work orders/"

        shutil.copytree(source, destination)

    @staticmethod
    def copyExistWorkOrders():
        source = "D:/Insulpro/temp/work orders/existing work orders/"

        destination = "D:/Insulpro/temp/weekly payroll/work orders/existing work orders/"

        shutil.copytree(source, destination)

    @staticmethod
    def copyAtticWorkOrders():
        source = "D:/Insulpro/temp/work orders/attic work orders"

        destination = "D:/Insulpro/temp/weekly payroll/work orders/attic work orders"

        shutil.copytree(source, destination)

    @staticmethod
    def copyFoamWorkOrders():
        source = "D:/Insulpro/temp/work orders/foam work orders"

        destination = "D:/Insulpro/temp/weekly payroll/work orders/foam work orders"

        shutil.copytree(source, destination)

    @staticmethod
    def copyShopWorkOrders():
        source = "D:/Insulpro/temp/work orders/shop work orders"

        destination = "D:/Insulpro/temp/weekly payroll/work orders/shop work orders"

        shutil.copytree(source, destination)

    @staticmethod
    def copyPayrollFolder():
        filedate = datetime.datetime.now()
        filedate_1 = filedate.strftime('%m%d%y')

        source = 'D:/Insulpro/temp/weekly payroll/'

        destination = 'D:/Insulpro/payroll records/weekly payroll/payroll' + str(filedate_1)

        shutil.copytree(source, destination)

    @staticmethod
    def clearPayrollDirectory():
        dir_path = "D:/Insulpro/temp/paysheets/batt pay sheet"
        dir_path1 = "D:/Insulpro/temp/paysheets/attic pay sheet"
        dir_path2 = "D:/Insulpro/temp/paysheets/foam pay sheet"
        dir_path3 = "D:/Insulpro/temp/paysheets/shop pay sheet"
        dir_path4 = "D:/Insulpro/temp/vacation"
        dir_path5 = "D:/Insulpro/temp/weekly payroll"
        dir_path6 = "D:/Insulpro/temp/work orders/new work orders"
        dir_path7 = "D:/Insulpro/temp/work orders/existing work orders"
        dir_path8 = "D:/Insulpro/temp/work orders/attic work orders"
        dir_path9 = "D:/Insulpro/temp/work orders/foam work orders"
        dir_path10 = "D:/Insulpro/temp/work orders/shop work orders"

        shutil.rmtree(dir_path)
        shutil.rmtree(dir_path1)
        shutil.rmtree(dir_path2)
        shutil.rmtree(dir_path3)
        shutil.rmtree(dir_path4)
        shutil.rmtree(dir_path5)
        shutil.rmtree(dir_path6)
        shutil.rmtree(dir_path7)
        shutil.rmtree(dir_path8)
        shutil.rmtree(dir_path9)
        shutil.rmtree(dir_path10)

        os.mkdir(dir_path)
        os.mkdir(dir_path1)
        os.mkdir(dir_path2)
        os.mkdir(dir_path3)
        os.mkdir(dir_path4)
        os.mkdir(dir_path5)
        os.mkdir(dir_path6)
        os.mkdir(dir_path7)
        os.mkdir(dir_path8)
        os.mkdir(dir_path9)
        os.mkdir(dir_path10)

    @staticmethod
    def copyPayrollRecords():
        source = "D:/Insulpro/payroll records/"

        destination = "D:/Insulpro/backup/payroll records"

        shutil.copytree(source, destination)

    @staticmethod
    def copyEmployeeRecords():
        source = "D:/Insulpro/employee work records/"

        destination = "D:/Insulpro/backup/employee work records"

        shutil.copytree(source, destination)

    @staticmethod
    def copyEmployeePaytree():
        source = "D:/Insulpro/employee paytree/"

        destination = "D:/Insulpro/backup/employee paytree/"

        shutil.copytree(source, destination)

    @staticmethod
    def clearWorkRecordDirectory():
        dir_path = "D:/Insulpro/employee work records"
        dir_path1 = "D:/Insulpro/payroll records"
        dir_path2 = "D:/Insulpro/employee paytree"

        shutil.rmtree(dir_path)
        shutil.rmtree(dir_path1)
        shutil.rmtree(dir_path2)

        os.mkdir(dir_path)
        os.mkdir(dir_path1)
        os.mkdir(dir_path2)


# Login Screen
class LoginScreen(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title('InsulPro Insulation-Login Screen')
        # setting window size
        width = 400
        height = 200
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(alignstr)
        self.resizable(width=False, height=False)

        usernameLabel = ctk.CTkLabel(self, text="Username", width=150, height=30,
                                     font=ctk.CTkFont(size=20, weight="normal"))
        usernameLabel.place(x=10, y=10)

        self.username = ctk.CTkEntry(self, width=175, height=30)
        self.username.place(x=170, y=10)

        passwordLabel = ctk.CTkLabel(self, text="Password", width=150, height=30,
                                     font=ctk.CTkFont(size=20, weight="normal"))
        passwordLabel.place(x=10, y=50)

        self.password = ctk.CTkEntry(self, width=175, height=30)
        self.password.place(x=170, y=50)

        loginButton = ctk.CTkButton(self, text='Login', width=175, height=30,
                                    font=ctk.CTkFont(size=17), command=self.loginButton_command)
        loginButton.place(x=10, y=100)

        closeWindowButton = ctk.CTkButton(self, text="Close Window", width=175, height=30,
                                          font=ctk.CTkFont(size=17), command=self.closeWindowButton_command)
        closeWindowButton.place(x=190, y=100)

        self.toplevel_window = None

    def loginButton_command(self):
        if self.username.get() == employee0.username and self.password.get() == employee0.pwd:
            window = MainWindow(self)
            window.grab_set()
            self.username.delete(0, 'end')
            self.password.delete(0, 'end')

        elif self.username.get() == employee1.username and self.password.get() == employee1.pwd:
            if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
                self.toplevel_window = Employee1Record(self)
                self.username.delete(0, 'end')
                self.password.delete(0, 'end')
            else:
                self.toplevel_window.focus()

        elif self.username.get() == employee2.username and self.password.get() == employee2.pwd:
            if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
                self.toplevel_window = Employee2Record(self)
                self.username.delete(0, 'end')
                self.password.delete(0, 'end')
            else:
                self.toplevel_window.focus()

        elif self.username.get() == employee3.username and self.password.get() == employee3.pwd:
            if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
                self.toplevel_window = Employee3Record(self)
                self.username.delete(0, 'end')
                self.password.delete(0, 'end')
            else:
                self.toplevel_window.focus()

        elif self.username.get() == employee4.username and self.password.get() == employee4.pwd:
            if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
                self.toplevel_window = Employee4Record(self)
                self.username.delete(0, 'end')
                self.password.delete(0, 'end')
            else:
                self.toplevel_window.focus()

        elif self.username.get() == employee5.username and self.password.get() == employee5.pwd:
            if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
                self.toplevel_window = Employee5Record(self)
                self.username.delete(0, 'end')
                self.password.delete(0, 'end')
            else:
                self.toplevel_window.focus()

        elif self.username.get() == employee6.username and self.password.get() == employee6.pwd:
            if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
                self.toplevel_window = Employee6Record(self)
                self.username.delete(0, 'end')
                self.password.delete(0, 'end')
            else:
                self.toplevel_window.focus()

        elif self.username.get() == employee7.username and self.password.get() == employee7.pwd:
            if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
                self.toplevel_window = Employee7Record(self)
                self.username.delete(0, 'end')
                self.password.delete(0, 'end')
            else:
                self.toplevel_window.focus()

        elif self.username.get() == employee8.username and self.password.get() == employee8.pwd:
            if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
                self.toplevel_window = Employee8Record(self)
                self.username.delete(0, 'end')
                self.password.delete(0, 'end')
            else:
                self.toplevel_window.focus()

        elif self.username.get() == employee9.username and self.password.get() == employee9.pwd:
            if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
                self.toplevel_window = Employee9Record(self)
                self.username.delete(0, 'end')
                self.password.delete(0, 'end')
            else:
                self.toplevel_window.focus()

        elif self.username.get() == employee10.username and self.password.get() == employee10.pwd:
            if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
                self.toplevel_window = Employee10Record(self)
                self.username.delete(0, 'end')
                self.password.delete(0, 'end')
            else:
                self.toplevel_window.focus()

        else:
            self.username.delete(0, 'end')
            self.password.delete(0, 'end')
            messagebox.showerror(title="Invalid Credentials", message="Please Enter A Valid Username and/or Password")

    def closeWindowButton_command(self):
        self.destroy()


# Run Program
app = LoginScreen()
app.mainloop()
