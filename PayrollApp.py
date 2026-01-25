# =========================================================
# Standard Library Imports
# =========================================================
import datetime
import os
import shutil
import tkinter as tk
from tkinter import messagebox

# =========================================================
# Third‑Party Libraries
# =========================================================
import customtkinter as ctk
import pandas as pd
from openpyxl import load_workbook

# =========================================================
# Startup Log
# =========================================================
print("Importing Standard Library")
print("  - datetime")
print("  - os")
print("  - shutil")
print("  - tkinter (tk)")
print("  - tkinter.messagebox")
print("----------------------------------")
print("Importing Third‑Party Libraries")
print("  - customtkinter (ctk)")
print("  - pandas (pd)")
print("  - openpyxl.load_workbook")
print("----------------------------------")
print("Opening Login Manager...")
print("----------------------------------")

# =========================================================
# Network Root Path
# (Update this if the network drive or folder changes)
# =========================================================
NETWORK_ROOT = r"\\Kenny-pc\d\Insulpro"


def netpath(*parts):
    """Build a full path inside the network root."""
    return os.path.join(NETWORK_ROOT, *parts)


# =========================================================
# Directory Initialization
# =========================================================
MASTER_PAYROLL_DIR = netpath("temp", "master_payroll")

# Ensure master payroll directory exists
os.makedirs(MASTER_PAYROLL_DIR, exist_ok=True)

# =========================================================
# File Paths
# =========================================================
EMPLOYEE_FILE = netpath("system", "employees.xlsx")


# ---------------------------------------------------------
#  Employee class
# ---------------------------------------------------------
class Employee:
    def __init__(self, first, last, vacationdays=0, passwd=""):
        self.first = first
        self.last = last
        self.vacationdays = vacationdays
        self.fullname = f"{first} {last}"
        self.username = f"{first[0]}{last}".lower()
        self.pwd = passwd


# ---------------------------------------------------------
#  Load employees from Excel Function
# ---------------------------------------------------------
def load_employees_from_excel(path):
    df = pd.read_excel(path)
    loaded_employees = []

    for _, row in df.iterrows():
        emp = Employee(
            first=row["first"],
            last=row["last"],
            vacationdays=row.get("vacationdays", 0),
            passwd=row.get("password", "")
        )
        loaded_employees.append(emp)

    return loaded_employees

# ---------------------------------------------------------
#  Load Employees from Excel
# ---------------------------------------------------------
employees = load_employees_from_excel(
    netpath("system", "employees.xlsx")
)


# ---------------------------------------------------------
# Payrate Class
# ---------------------------------------------------------
class Payrate:
    def __init__(self, name, rate):
        self.name = name
        self.rate = float(rate)


# ---------------------------------------------------------
# Load Payrates from Excel Function
# ---------------------------------------------------------
def load_payrates_from_excel(path):
    df = pd.read_excel(path)
    loaded_payrates = {}

    for _, row in df.iterrows():
        name = str(row["name"]).strip()
        rate = float(row["rate"])

        payrate_obj = Payrate(name, rate)

        key = (
            name.lower()
            .replace(" ", "_")
            .replace("/", "_")
            .replace('"', "")
            .replace(">", "")
            .replace("<", "")
        )

        loaded_payrates[key] = payrate_obj

    return loaded_payrates


# ---------------------------------------------------------
# Load Payrates from Excel
# ---------------------------------------------------------
payrates = load_payrates_from_excel(
    netpath("system", "payrates.xlsx")
)


# ---------------------------------------------------------
# Function to append payroll file
# ---------------------------------------------------------
def append_to_master_payroll(emp: Employee, rows: list):
    """
    rows = list of tuples:
    (date, job_name, pay_item, qty, rate, total, split)
    """

    filename = f"{emp.fullname}.xlsx"
    path = os.path.join(MASTER_PAYROLL_DIR, filename)

    # If file doesn't exist, create it with headers
    if not os.path.exists(path):
        workbook = load_workbook(netpath("system", "spreadsheet", "MasterPayrollTemplate.xlsx"))
        sheet = workbook.active
        _ = sheet  # prevents false unused-variable warnings
        workbook.save(path)
        workbook.close()

    # Open existing file
    workbook = load_workbook(path)
    sheet = workbook.active

    # Find next empty row
    row = sheet.max_row + 1
    while all(sheet[f"{col}{row}"].value is None for col in "ABCDEFG"):
        row -= 1
    row += 1

    # Append rows
    for r in rows:
        sheet[f"A{row}"] = r[0]  # Date
        sheet[f"B{row}"] = r[1]  # Job Name
        sheet[f"C{row}"] = r[2]  # Pay Item
        sheet[f"D{row}"] = r[3]  # Quantity
        sheet[f"E{row}"] = r[4]  # Rate
        sheet[f"F{row}"] = r[5]  # Total
        sheet[f"G{row}"] = r[6]  # Split
        row += 1

    workbook.save(path)
    workbook.close()


# ---------------------------------------------------------
# App Controller to manage frame switching
# ---------------------------------------------------------
class AppController(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("InsulPro - INSULPAY")
        self.geometry("600x600")
        self.resizable(False, False)

        # centers the screen
        self.update_idletasks()
        w, h = 800, 600
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = int((sw - w) / 2)
        y = int((sh - h) / 2)
        self.geometry(f"{w}x{h}+{x}+{y}")

        # Store frames
        self.frames = {}

        # Frame Class Attribute
        self.FRAME_CLASSES = {
            "login": LoginFrame,
            "main": MainWindowFrame,
            "paysheet": PaySheetFrame,
            "weekly": WeeklyPayrollFrame,
            "view_weekly": ViewWeeklyPayrollFrame,
            "payroll_tools": PayrollToolsFrame,
            "ytd": YTDPayrollFrame,

            # EMPLOYEE MODULE
            "employee_records": EmployeeRecordsFrame,
            "employee_detail": EmployeeDetailFrame,
            "add_employee": AddEmployeeFrame,

        }

        # Create and place frames
        for name, FrameClass in self.FRAME_CLASSES.items():
            frame = FrameClass(self)
            frame.place(relwidth=1, relheight=1)
            self.frames[name] = frame

        # Show login first
        self.show_frame("login")

    def show_frame(self, name):
        frame = self.frames[name]

        if name == "login":
            frame.reset_fields()

        frame.tkraise()


# ---------------------------------------------------------
# Login System Frame
# ---------------------------------------------------------
class LoginFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # Username
        ctk.CTkLabel(self, text="Username").grid(row=0, column=0, padx=10, pady=10)
        self.username_entry = ctk.CTkEntry(self, width=200)
        self.username_entry.grid(row=0, column=1, padx=10, pady=10)

        # Password
        ctk.CTkLabel(self, text="Password").grid(row=1, column=0, padx=10, pady=10)
        self.password_entry = ctk.CTkEntry(self, width=200, show="*")
        self.password_entry.grid(row=1, column=1, padx=10, pady=10)

        # Show Password Toggle
        self.show_password_var = tk.BooleanVar(value=False)
        show_pw_checkbox = ctk.CTkCheckBox(
            self,
            text="Show Password",
            variable=self.show_password_var,
            command=self.toggle_password_visibility
        )
        show_pw_checkbox.grid(row=2, column=1, sticky="w", padx=10)

        # Login Button
        login_btn = ctk.CTkButton(
            self,
            text="Login",
            width=150,
            command=self.login_button_command
        )
        login_btn.grid(row=3, column=0, columnspan=2, pady=20)

        # Close Button
        close_btn = ctk.CTkButton(
            self,
            text="Close Window",
            width=150,
            command=self.master.destroy
        )
        close_btn.grid(row=4, column=0, columnspan=2, pady=10)

        # Enter key triggers login
        self.master.bind("<Return>", lambda event: self.login_button_command())

    # Toggle password visibility
    def toggle_password_visibility(self):
        if self.show_password_var.get():
            self.password_entry.configure(show="")
        else:
            self.password_entry.configure(show="*")

    # Login logic
    def login_button_command(self):
        username = self.username_entry.get().strip().lower()
        password = self.password_entry.get().strip()



        matched = next(
            (emp for emp in employees if emp.username == username and emp.pwd == password),
            None
        )

        if not matched:
            messagebox.showerror("Invalid Credentials", "Please enter a valid username and/or password")
            self.username_entry.delete(0, 'end')
            self.password_entry.delete(0, 'end')
            return

        # Admin → Main Menu
        if matched.username in ("admin", "mfrank"):
            self.master.show_frame("main")
            return

        # Employee windows will be added later
        messagebox.showinfo("Login Successful", f"Welcome, {matched.fullname}")

        self.username_entry.focus_set()

    def reset_fields(self):
        self.username_entry.delete(0, 'end')
        self.password_entry.delete(0, 'end')
        self.show_password_var.set(False)
        self.password_entry.configure(show="*")


# ---------------------------------------------------------
# Admin Main Frame
# ---------------------------------------------------------
class MainWindowFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        title_label = ctk.CTkLabel(
            self,
            text="Main Menu",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(pady=20)

        btn_frame = ctk.CTkFrame(self)
        btn_frame.pack(pady=10)

        ctk.CTkButton(btn_frame, text="Create Pay Sheet", width=200,
                      command=self.open_pay_sheet).pack(pady=10)

        ctk.CTkButton(btn_frame, text="Employee Records", width=200,
                      command=self.open_employee_records).pack(pady=10)

        ctk.CTkButton(btn_frame, text="Payroll Tools", width=200,
                      command=self.open_payroll_tools).pack(pady=10)

        ctk.CTkButton(btn_frame, text="Backup / Reset Tools", width=200,
                      command=self.open_backup_tools).pack(pady=10)

        appearance_label = ctk.CTkLabel(
            self,
            text="Appearance Mode:",
            font=ctk.CTkFont(size=16)
        )
        appearance_label.pack(pady=(40, 5))

        appearance_menu = ctk.CTkOptionMenu(
            self,
            values=["Light", "Dark", "System"],
            command=self.change_appearance
        )
        appearance_menu.pack()

        ctk.CTkButton(
            self,
            text="Logout",
            width=200,
            command=lambda: self.master.show_frame("login")
        ).pack(pady=20)

    @staticmethod
    def change_appearance(mode):
        ctk.set_appearance_mode(mode)

    def open_pay_sheet(self):
        self.master.show_frame("paysheet")

    def open_employee_records(self):
        self.master.show_frame("employee_records")
    def open_payroll_tools(self):
        self.master.show_frame("payroll_tools")

    def open_backup_tools(self):
        messagebox.showinfo("Coming Soon", "Backup Tools module will be added.")


# ---------------------------------------------------------
# Pay Sheet Frame
# ---------------------------------------------------------
class PaySheetFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # Title
        title = ctk.CTkLabel(
            self,
            text="Create Pay Sheet",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title.pack(pady=20)

        # =====================================================
        # JOB NAME + DATE FIELDS
        # =====================================================
        header_frame = ctk.CTkFrame(self)
        header_frame.pack(pady=10)

        # Job Name
        ctk.CTkLabel(header_frame, text="Job Name:").grid(row=0, column=0, padx=10, pady=5)
        self.jobname_entry = ctk.CTkEntry(header_frame, width=200)
        self.jobname_entry.grid(row=0, column=1, padx=10, pady=5)

        # Date
        ctk.CTkLabel(header_frame, text="Date:").grid(row=1, column=0, padx=10, pady=5)
        self.date_entry = ctk.CTkEntry(header_frame, width=200)
        self.date_entry.grid(row=1, column=1, padx=10, pady=5)

        # Auto-fill today's date
        today = datetime.date.today().strftime("%m-%d-%Y")
        self.date_entry.insert(0, today)

        # =====================================================
        # EMPLOYEE CHECKBOX LIST
        # =====================================================
        emp_frame = ctk.CTkScrollableFrame(self, width=250, height=300)
        emp_frame.pack(side="left", padx=20, pady=10)

        ctk.CTkLabel(emp_frame, text="Select Employees:",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(pady=5)

        self.employee_vars = {}  # username → BooleanVar

        for emp in employees:
            var = tk.BooleanVar()
            chk = ctk.CTkCheckBox(
                emp_frame,
                text=emp.fullname,
                variable=var
            )
            chk.pack(anchor="w", pady=2)
            self.employee_vars[emp.username] = var

        # =====================================================
        # PAYRATE INPUT GRID
        # =====================================================
        rate_frame = ctk.CTkScrollableFrame(self, width=300, height=300)
        rate_frame.pack(side="right", padx=20, pady=10)

        ctk.CTkLabel(rate_frame, text="Enter Quantities:",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(pady=5)

        self.payrate_entries = {}  # key → entry widget

        for key, pr in payrates.items():
            row = ctk.CTkFrame(rate_frame)
            row.pack(fill="x", pady=3)

            label = ctk.CTkLabel(row, text=pr.name, width=180, anchor="w")
            label.pack(side="left", padx=5)

            entry = ctk.CTkEntry(row, width=80)
            entry.pack(side="right", padx=5)

            self.payrate_entries[key] = entry

        # =====================================================
        # ACTION BUTTONS
        # =====================================================
        btn_frame = ctk.CTkFrame(self)
        btn_frame.pack(pady=20)

        ctk.CTkButton(
            btn_frame,
            text="Calculate Split",
            width=200,
            command=self.calculate_split
        ).pack(pady=10)

        ctk.CTkButton(
            btn_frame,
            text="Save Pay Sheet",
            width=200,
            command=self.save_paysheet
        ).pack(pady=10)

        ctk.CTkButton(
            btn_frame,
            text="Reset Form",
            width=200,
            command=self.reset_form
        ).pack(pady=10)

        ctk.CTkButton(
            btn_frame,
            text="Back to Main Menu",
            width=200,
            command=lambda: self.master.show_frame("main")
        ).pack(pady=10)

    # ---------------------------------------------------------
    # CALCULATE SPLIT LOGIC
    # ---------------------------------------------------------
    def calculate_split(self):
        # Get Job name and date
        job_name = self.jobname_entry.get().strip()
        date_value = self.date_entry.get().strip()

        if not job_name:
            messagebox.showerror("Missing Job Name", "Please enter a job name.")
            return

        if not date_value:
            messagebox.showerror("Missing Date", "Please enter a date.")
            return

        # 1. Get selected employees
        selected = [
            username for username, var in self.employee_vars.items()
            if var.get()
        ]

        if not selected:
            messagebox.showerror("No Employees Selected",
                                 "Please select at least one employee.")
            return

        num_workers = len(selected)

        # 2. Loop through payrate entries
        results = []

        for key, entry in self.payrate_entries.items():
            qty_text = entry.get().strip()
            if not qty_text:
                continue  # skip empty fields

            try:
                qty = float(qty_text)
            except ValueError:
                messagebox.showerror("Invalid Quantity",
                                     f"Invalid number for {payrates[key].name}")
                return

            rate = payrates[key].rate
            total = qty * rate
            split = total / num_workers

            results.append((payrates[key].name, qty, rate, total, split))

        # 3. Show results
        if not results:
            messagebox.showinfo("No Data", "No quantities entered.")
            return

        msg = f"Job: {job_name}\nDate: {date_value}\n\nPay Split Results:\n\n"

        for name, qty, rate, total, split in results:
            msg += (
                f"{name}\n"
                f"  Qty: {qty}\n"
                f"  Rate: {rate}\n"
                f"  Total: ${total:.2f}\n"
                f"  Split per worker: ${split:.2f}\n\n"
            )

        messagebox.showinfo("Pay Split", msg)

    # ---------------------------------------------------------
    # SAVE PAY SHEET TO EXCEL
    # ---------------------------------------------------------
    def save_paysheet(self):
        job_name = self.jobname_entry.get().strip()
        date_value = self.date_entry.get().strip()

        if not job_name:
            messagebox.showerror("Missing Job Name", "Please enter a job name.")
            return

        if not date_value:
            messagebox.showerror("Missing Date", "Please enter a date.")
            return

        # 1. Collect selected employees
        selected_emps = [
            emp for emp in employees
            if self.employee_vars[emp.username].get()
        ]

        if not selected_emps:
            messagebox.showerror("No Employees Selected", "Select at least one employee.")
            return

        num_workers = len(selected_emps)

        # 2. Collect payrate entries
        pay_items = []
        for key, entry in self.payrate_entries.items():
            qty_text = entry.get().strip()
            if not qty_text:
                continue

            try:
                qty = float(qty_text)
            except ValueError:
                messagebox.showerror("Invalid Quantity", f"Invalid number for {payrates[key].name}")
                return

            rate = payrates[key].rate
            total = qty * rate
            split = total / num_workers

            pay_items.append((payrates[key].name, qty, rate, total, split))

        if not pay_items:
            messagebox.showerror("No Pay Items", "Enter at least one quantity.")
            return

        # 3. Copy template → create new file
        template = netpath("system", "spreadsheet", "PaySheetTemplate.xlsx")
        save_dir = netpath("temp", "paysheets")

        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        filename = f"{job_name}.xlsx"
        new_path = os.path.join(save_dir, filename)

        # ---------------------------------------------------------
        # CHECK IF FILE ALREADY EXISTS
        # ---------------------------------------------------------
        if os.path.exists(new_path):
            messagebox.showerror(
                "File Already Exists",
                f"A pay sheet named '{filename}' already exists.\n"
                "Please choose a different job name."
            )
            return

        # Safe to create the file
        shutil.copy(template, new_path)

        # 4. Write data into Excel
        workbook = load_workbook(new_path)
        sheet = workbook.active

        # Header
        sheet["B1"] = job_name
        sheet["B2"] = date_value

        # Employees
        row = 5
        for emp in selected_emps:
            sheet[f"A{row}"] = emp.fullname
            row += 1

        # Pay Items
        row = 5
        for name, qty, rate, total, split in pay_items:
            sheet[f"C{row}"] = name
            sheet[f"D{row}"] = qty
            sheet[f"E{row}"] = rate
            sheet[f"F{row}"] = total
            sheet[f"G{row}"] = split
            row += 1

        workbook.save(new_path)
        workbook.close()

        # ---------------------------------------------------------
        # UPDATE MASTER PAYROLL FILES
        # ---------------------------------------------------------
        for emp in selected_emps:
            rows = []
            for name, qty, rate, total, split in pay_items:
                rows.append((
                    date_value,
                    job_name,
                    name,
                    qty,
                    rate,
                    total,
                    split
                ))
            append_to_master_payroll(emp, rows)

        messagebox.showinfo("Saved", f"Pay sheet saved:\n{new_path}")

        self.reset_form()

    # ---------------------------------------------------------
    # RESET FORM
    # ---------------------------------------------------------
    def reset_form(self):
        # Reset job name
        self.jobname_entry.delete(0, "end")

        # Reset date to today
        today = datetime.date.today().strftime("%m-%d-%Y")
        self.date_entry.delete(0, "end")
        self.date_entry.insert(0, today)

        # Uncheck all employees
        for var in self.employee_vars.values():
            var.set(False)

        # Clear all payrate entries
        for entry in self.payrate_entries.values():
            entry.delete(0, "end")


# ---------------------------------------------------------
# Weekly Payroll Generator Frame
# ---------------------------------------------------------
# noinspection PyBroadException
class WeeklyPayrollFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        title = ctk.CTkLabel(
            self,
            text="Weekly Payroll Generator",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title.pack(pady=20)

        # Date range inputs
        frame = ctk.CTkFrame(self)
        frame.pack(pady=10)

        ctk.CTkLabel(frame, text="Start Date (MM-DD-YYYY):").grid(row=0, column=0, padx=10, pady=5)
        self.start_entry = ctk.CTkEntry(frame, width=150)
        self.start_entry.grid(row=0, column=1, padx=10, pady=5)

        ctk.CTkLabel(frame, text="End Date (MM-DD-YYYY):").grid(row=1, column=0, padx=10, pady=5)
        self.end_entry = ctk.CTkEntry(frame, width=150)
        self.end_entry.grid(row=1, column=1, padx=10, pady=5)

        # Buttons
        ctk.CTkButton(
            self,
            text="Generate Weekly Payroll",
            width=200,
            command=self.generate_weekly_payroll
        ).pack(pady=20)

        ctk.CTkButton(
            self,
            text="Reset",
            width=200,
            command=self.reset_fields
        ).pack(pady=10)

        ctk.CTkButton(
            self,
            text="Back to Main Menu",
            width=200,
            command=lambda: self.master.show_frame("main")
        ).pack(pady=10)

    def reset_fields(self):
            self.start_entry.delete(0, 'end')
            self.end_entry.delete(0, 'end')

    # ---------------------------------------------------------
    # GENERATE WEEKLY PAYROLL
    # ---------------------------------------------------------
    def generate_weekly_payroll(self):
        start_text = self.start_entry.get().strip()
        end_text = self.end_entry.get().strip()

        if not start_text or not end_text:
            messagebox.showerror("Missing Dates", "Please enter both start and end dates.")
            return

        try:
            start_date = datetime.datetime.strptime(start_text, "%m-%d-%Y").date()
            end_date = datetime.datetime.strptime(end_text, "%m-%d-%Y").date()
        except ValueError:
            messagebox.showerror("Invalid Date", "Please enter valid dates in MM-DD-YYYY format.")
            return

        if end_date < start_date:
            messagebox.showerror("Invalid Range", "End date must be after start date.")
            return

        save_dir = netpath("weekly payroll")
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        output_path = os.path.join(save_dir, f"WeeklyPayroll_{start_text}_to_{end_text}.xlsx")

        if os.path.exists(output_path):
            messagebox.showerror(
                "File Already Exists",
                f"A weekly payroll file for this date range already exists:\n{output_path}"
            )
            return

        workbook = load_workbook(netpath("system", "spreadsheet", "WeeklyPayrollTemplate.xlsx"))
        sheet = workbook.active

        sheet["B2"] = start_text
        sheet["B3"] = end_text

        emp_row = item_row = job_row = day_row = 5

        global_pay_item_totals = {}
        global_job_totals = {}
        global_day_totals = {}

        for emp in employees:
            filename = f"{emp.fullname}.xlsx"
            path = os.path.join(MASTER_PAYROLL_DIR, filename)

            if not os.path.exists(path):
                continue

            emp_book = load_workbook(path, data_only=True)
            emp_sheet = emp_book.active

            total_pay = 0

            for r in range(2, emp_sheet.max_row + 1):
                date_val = emp_sheet[f"A{r}"].value

                if isinstance(date_val, datetime.datetime):
                    date_val = date_val.date()

                if isinstance(date_val, str):
                    for fmt in ("%m-%d-%Y", "%Y-%m-%d"):
                        try:
                            date_val = datetime.datetime.strptime(date_val, fmt).date()
                            break
                        except:
                            pass

                if not isinstance(date_val, datetime.date):
                    continue

                if not (start_date <= date_val <= end_date):
                    continue

                split = emp_sheet[f"G{r}"].value
                if split is None:
                    continue

                try:
                    split = float(split)
                except:
                    continue

                total_pay += split

                pay_item = emp_sheet[f"C{r}"].value
                job_name = emp_sheet[f"B{r}"].value

                if pay_item:
                    global_pay_item_totals[pay_item] = global_pay_item_totals.get(pay_item, 0) + split

                if job_name:
                    global_job_totals[job_name] = global_job_totals.get(job_name, 0) + split

                global_day_totals[date_val] = global_day_totals.get(date_val, 0) + split

            sheet[f"A{emp_row}"] = emp.fullname
            sheet[f"B{emp_row}"] = total_pay
            emp_row += 1

            emp_book.close()

        sheet["D4"] = "Totals Per Pay Item"
        for item, total in sorted(global_pay_item_totals.items()):
            sheet[f"D{item_row}"] = item
            sheet[f"E{item_row}"] = total
            item_row += 1

        sheet["G4"] = "Totals Per Job"
        for job, total in sorted(global_job_totals.items()):
            sheet[f"G{job_row}"] = job
            sheet[f"H{job_row}"] = total
            job_row += 1

        sheet["J4"] = "Totals Per Day"
        for day, total in sorted(global_day_totals.items()):
            sheet[f"J{day_row}"] = day.strftime("%m-%d-%Y")
            sheet[f"K{day_row}"] = total
            day_row += 1

        weekly_total = sum(global_day_totals.values())
        sheet["D2"] = "Weekly Total:"
        sheet["E2"] = weekly_total

        workbook.save(output_path)
        workbook.close()

        messagebox.showinfo("Weekly Payroll Generated", f"Saved to:\n{output_path}")


# ---------------------------------------------------------
# View Weekly Payroll Frame
# ---------------------------------------------------------
class ViewWeeklyPayrollFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        title = ctk.CTkLabel(
            self,
            text="View Weekly Payroll",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title.pack(pady=20)

        # File selection
        file_frame = ctk.CTkFrame(self)
        file_frame.pack(pady=10)

        ctk.CTkLabel(file_frame, text="Select Weekly Payroll File:").grid(row=0, column=0, padx=10)
        self.file_entry = ctk.CTkEntry(file_frame, width=350)
        self.file_entry.grid(row=0, column=1, padx=10)

        ctk.CTkButton(
            file_frame,
            text="Browse",
            command=self.browse_file
        ).grid(row=0, column=2, padx=10)

        ctk.CTkButton(
            self,
            text="Load Payroll",
            width=200,
            command=self.load_payroll
        ).pack(pady=10)

        ctk.CTkButton(
            self,
            text="Back to Main Menu",
            width=200,
            command=self.go_back
        ).pack(pady=10)

        # Scrollable display area
        self.scroll = ctk.CTkScrollableFrame(self, width=900, height=500)
        self.scroll.pack(pady=20)

    def browse_file(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            initialdir=netpath("weekly payroll"),
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if path:
            self.file_entry.delete(0, "end")
            self.file_entry.insert(0, path)

    def load_payroll(self):
        path = self.file_entry.get().strip()
        if not os.path.exists(path):
            messagebox.showerror("Error", "File not found.")
            return

        # Clear previous content
        for widget in self.scroll.winfo_children():
            widget.destroy()

        book = load_workbook(path, data_only=True)
        sheet = book.active

        # Weekly summary
        summary = ctk.CTkLabel(
            self.scroll,
            text=f"Weekly Total: {sheet['E2'].value}",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        summary.pack(pady=10)

        sections = [
            ("Employee Totals", "A", "B"),
            ("Totals Per Pay Item", "D", "E"),
            ("Totals Per Job", "G", "H"),
            ("Totals Per Day", "J", "K")
        ]

        for title, col1, col2 in sections:
            ctk.CTkLabel(
                self.scroll,
                text=title,
                font=ctk.CTkFont(size=16, weight="bold")
            ).pack(pady=10)

            table = ctk.CTkFrame(self.scroll)
            table.pack(pady=5)

            # Header row
            ctk.CTkLabel(table, text="Name", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10)
            ctk.CTkLabel(table, text="Amount", font=ctk.CTkFont(weight="bold")).grid(row=0, column=1, padx=10)

            # Read rows
            rows = []
            row = 5
            while True:
                v1 = sheet[f"{col1}{row}"].value
                v2 = sheet[f"{col2}{row}"].value
                if v1 is None and v2 is None:
                    break
                rows.append((v1, v2))
                row += 1

            if not rows:
                ctk.CTkLabel(table, text="No data available").grid(row=1, column=0, columnspan=2)
                continue

            for i, (v1, v2) in enumerate(sorted(rows), start=1):
                ctk.CTkLabel(table, text=str(v1), width=200, anchor="w").grid(row=i, column=0, padx=10)
                ctk.CTkLabel(table, text=str(v2), width=200, anchor="w").grid(row=i, column=1, padx=10)

        book.close()

    def reset_view(self):
        self.file_entry.delete(0, "end")
        for widget in self.scroll.winfo_children():
            widget.destroy()

    def go_back(self):
        self.reset_view()
        self.master.show_frame("main")


# ---------------------------------------------------------
# Payroll Tools Frame
# ---------------------------------------------------------
class PayrollToolsFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        title = ctk.CTkLabel(
            self,
            text="Payroll Tools",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title.pack(pady=20)

        btn_frame = ctk.CTkFrame(self)
        btn_frame.pack(pady=20)

        ctk.CTkButton(
            btn_frame,
            text="Weekly Payroll Generator",
            width=220,
            command=lambda: self.master.show_frame("weekly")
        ).pack(pady=10)

        ctk.CTkButton(
            btn_frame,
            text="View Weekly Payroll",
            width=220,
            command=lambda: self.master.show_frame("view_weekly")
        ).pack(pady=10)

        ctk.CTkButton(
            btn_frame,
            text="YTD Payroll Summary",
            width=220,
            command=lambda: self.master.show_frame("ytd")
        ).pack(pady=10)

        ctk.CTkButton(
            btn_frame,
            text="Back",
            width=220,
            command=lambda: self.master.show_frame("main")
        ).pack(pady=20)


# ---------------------------------------------------------
# YTD Payroll Frame
# ---------------------------------------------------------
# noinspection PyBroadException
class YTDPayrollFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        title = ctk.CTkLabel(
            self,
            text="Year-To-Date Payroll Summary",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title.pack(pady=20)

        form = ctk.CTkFrame(self)
        form.pack(pady=20)

        # YEAR
        ctk.CTkLabel(form, text="Year:").grid(row=0, column=0, padx=10, pady=5)
        self.year_entry = ctk.CTkEntry(form, width=150)
        self.year_entry.grid(row=0, column=1, padx=10, pady=5)

        # EMPLOYEE FILTER
        ctk.CTkLabel(form, text="Employee (optional):").grid(row=1, column=0, padx=10, pady=5)
        self.employee_option = ctk.CTkComboBox(
            form,
            values=["All"] + [emp.fullname for emp in employees],
            width=200
        )
        self.employee_option.grid(row=1, column=1, padx=10, pady=5)
        self.employee_option.set("All")

        jobs, payitems = self.collect_job_and_payitems()

        # JOB FILTER
        ctk.CTkLabel(form, text="Job (optional):").grid(row=2, column=0, padx=10, pady=5)
        self.job_option = ctk.CTkComboBox(
            form,
            values=["All"] + jobs,
            width=200
        )
        self.job_option.grid(row=2, column=1, padx=10, pady=5)
        self.job_option.set("All")

        # GENERATE BUTTON
        ctk.CTkButton(
            self,
            text="Generate YTD Payroll Summary",
            width=250,
            command=self.generate_ytd_summary
        ).pack(pady=20)

        # BACK BUTTON
        ctk.CTkButton(
            self,
            text="Back to Payroll Tools",
            width=200,
            command=lambda: self.master.show_frame("payroll_tools")
        ).pack(pady=10)

    @staticmethod
    def collect_job_and_payitems():
        jobs = set()
        payitems = set()

        for emp in employees:
            filename = f"{emp.fullname}.xlsx"
            path = os.path.join(MASTER_PAYROLL_DIR, filename)

            if not os.path.exists(path):
                continue

            book = load_workbook(path)
            sheet = book.active

            for r in range(2, sheet.max_row + 1):
                job = sheet[f"B{r}"].value
                payitem = sheet[f"C{r}"].value

                if job:
                    jobs.add(str(job).strip())

                if payitem:
                    payitems.add(str(payitem).strip())

            book.close()

        return sorted(jobs), sorted(payitems)

    def generate_ytd_summary(self):
        year_text = self.year_entry.get().strip()
        employee_filter = self.employee_option.get()
        job_filter = self.job_option.get()

        if job_filter == "All":
            job_filter = ""

        if not year_text.isdigit():
            messagebox.showerror("Invalid Year", "Please enter a valid year (e.g., 2025).")
            return

        year = int(year_text)

        # Prepare output directory
        save_dir = netpath("ytd payroll")
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        output_path = os.path.join(save_dir, f"YTD_{year}.xlsx")

        # Load template
        workbook = load_workbook(netpath("system", "spreadsheet", "YTDPayrollTemplate.xlsx"))
        sheet = workbook.active

        sheet["A3"] = "Year:"
        sheet["B3"] = year

        # Summary dictionaries
        employee_totals = {}
        pay_item_totals = {}
        job_totals = {}
        month_totals = {}

        # Loop through employees
        for emp in employees:

            # Employee filter
            if employee_filter != "All" and emp.fullname != employee_filter:
                continue

            filename = f"{emp.fullname}.xlsx"
            path = os.path.join(MASTER_PAYROLL_DIR, filename)

            if not os.path.exists(path):
                continue

            emp_book = load_workbook(path)
            emp_sheet = emp_book.active

            for r in range(2, emp_sheet.max_row + 1):

                date_val = emp_sheet[f"A{r}"].value

                if isinstance(date_val, datetime.datetime):
                    date_val = date_val.date()

                if isinstance(date_val, str):
                    try:
                        date_val = datetime.datetime.strptime(date_val, "%m-%d-%Y").date()
                    except:
                        continue

                if not isinstance(date_val, datetime.date):
                    continue

                if date_val.year != year:
                    continue

                # Filters
                job_name = str(emp_sheet[f"B{r}"].value).lower()
                pay_item = str(emp_sheet[f"C{r}"].value).lower()

                if job_filter and job_filter not in job_name:
                    continue

                split = emp_sheet[f"G{r}"].value
                if not split:
                    continue

                try:
                    split = float(split)
                except:
                    continue

                # Employee totals
                employee_totals[emp.fullname] = employee_totals.get(emp.fullname, 0) + split

                # Pay item totals
                if pay_item:
                    pay_item_totals[pay_item] = pay_item_totals.get(pay_item, 0) + split

                # Job totals
                if job_name:
                    job_totals[job_name] = job_totals.get(job_name, 0) + split

                # Month totals
                month = date_val.strftime("%B")
                month_totals[month] = month_totals.get(month, 0) + split

            emp_book.close()

        # Write results to Excel
        row = 5

        # Employee totals
        sheet["A4"] = "Employee Totals"
        for name, total in employee_totals.items():
            sheet[f"A{row}"] = name
            sheet[f"B{row}"] = total
            row += 1

        # Pay item totals
        row = 5
        sheet["D4"] = "Pay Item Totals"
        for item, total in pay_item_totals.items():
            sheet[f"D{row}"] = item
            sheet[f"E{row}"] = total
            row += 1

        # Job totals
        row = 5
        sheet["G4"] = "Job Totals"
        for job, total in job_totals.items():
            sheet[f"G{row}"] = job
            sheet[f"H{row}"] = total
            row += 1

        # Month totals
        row = 5
        sheet["J4"] = "Month Totals"
        for month, total in month_totals.items():
            sheet[f"J{row}"] = month
            sheet[f"K{row}"] = total
            row += 1

        # Grand total
        grand_total = sum(employee_totals.values())
        sheet["A2"] = "Grand Total:"
        sheet["B2"] = grand_total

        workbook.save(output_path)
        workbook.close()

        messagebox.showinfo("YTD Payroll Summary Generated", f"Saved to:\n{output_path}")




if __name__ == "__main__":
    app = AppController()
    app.mainloop()
