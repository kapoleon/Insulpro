# =========================================================
# Standard Library Imports
# =========================================================

import datetime
import os
import shutil
import tkinter as tk
from tkinter import messagebox

import customtkinter as ctk
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# =========================================================
# Network Root Path
# (Update this if the network drive or folder changes)
# =========================================================

USE_NETWORK = True

if USE_NETWORK:
    NETWORK_ROOT = r"\\Kenny-pc\d\Insulpro"
else:
    NETWORK_ROOT = r"D:\Insulpro"


def netpath(*parts):
    """Build a full path inside the network root."""
    return os.path.join(NETWORK_ROOT, *parts)




# =========================================================
# Startup Log
# =========================================================

def startup_log():
    print("==========================================")
    print(" InsulPay — Application Startup")
    print("==========================================")
    print(f"Startup Time: {datetime.datetime.now()}")
    print("------------------------------------------")
    print("Loading Standard Libraries:")
    print("  • datetime")
    print("  • os")
    print("  • shutil")
    print("  • tkinter (tk)")
    print("  • tkinter.messagebox")
    print("------------------------------------------")
    print("Loading Third‑Party Libraries:")
    print("  • customtkinter (ctk)")
    print("  • pandas (pd)")
    print("  • openpyxl")
    print("  • openpyxl.load_workbook")
    print("  • openpyxl.styles.Font")
    print("  • openpyxl.utils.get_column_letter")
    print("------------------------------------------")
    print("Initializing Login Manager...")
    print("==========================================")
