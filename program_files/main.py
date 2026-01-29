# =========================================================
# Standard Library Imports
# =========================================================

import datetime
import os
import shutil
import tkinter as tk
from tkinter import messagebox

# =========================================================
# Third‑Party Libraries
# =========================================================

import customtkinter as ctk
import pandas as pd
from openpyxl import load_workbook

# =========================================================
# Network Root Path
# (Update this if the network drive or folder changes)
# =========================================================

USE_NETWORK = True

if USE_NETWORK:
    NETWORK_ROOT = r"\\Kenny-pc\d\Insulpro"
else:
    NETWORK_ROOT = r"D:\Insulpro"

def netpath(*parts):
    """Build a full path inside the network root."""
    return os.path.join(NETWORK_ROOT, *parts)

# =========================================================
# Startup Log
# =========================================================

def startup_log():
    print("Importing Standard Library")
    print("  - datetime")
    print("  - os")
    print("  - shutil")
    print("  - tkinter (tk)")
    print("  - tkinter.messagebox")
    print("----------------------------------")
    print("Importing Third‑Party Libraries")
    print("  - customtkinter (ctk)")
    print("  - pandas (pd)")
    print("  - openpyxl.load_workbook")
    print("----------------------------------")
    print("Opening Login Manager...")
    print("----------------------------------")

startup_log()

# =========================================================
# File Paths
# =========================================================

EMPLOYEE_FILE        = netpath("data", "employees.xlsx")
PAYRATE_FILE         = netpath("data", "payrates.xlsx")

# Vacation Log
VACATION_TEMPLATE    = netpath("data", "spreadsheet", "VacationLogTemplate.xlsx")
VACATION_LOG         = netpath("payroll_records", "vacation", "vacation_log.xlsx")

# Vacation Requests
REQUEST_TEMPLATE     = netpath("data", "spreadsheet", "VacationRequestTemplate.xlsx")
REQUEST_FILE         = netpath("payroll_records", "vacation", "vacation_requests.xlsx")

# Master Payroll Directory
MASTER_PAYROLL_DIR   = netpath("payroll_records", "master_payroll")

# Ensure master payroll directory exists
os.makedirs(MASTER_PAYROLL_DIR, exist_ok=True)

# =========================================================
# Helpers
# =========================================================

# =========================================================
# Copy Template Helper
# =========================================================

def ensure_file_exists(target, template):
    os.makedirs(os.path.dirname(target), exist_ok=True)

    if not os.path.exists(target):
        if not os.path.exists(template):
            raise FileNotFoundError(f"Template not found: {template}")
        shutil.copy(template, target)
        print(f"Created: {os.path.basename(target)} from template.")

ensure_file_exists(VACATION_LOG, VACATION_TEMPLATE)
ensure_file_exists(REQUEST_FILE, REQUEST_TEMPLATE)

# ---------------------------------------------------------
# Helper to append master_payroll file
# ---------------------------------------------------------
def append_to_master_payroll(emp: Employee, rows: list):
    """
    rows = list of tuples:
    (date, job_name, pay_item, qty, rate, total, split)
    """

    filename = f"{emp.fullname}.xlsx"
    path = os.path.join(MASTER_PAYROLL_DIR, filename)

    # If file doesn't exist, create it with headers
    if not os.path.exists(path):
        workbook = load_workbook(netpath("data", "spreadsheet", "MasterPayrollTemplate.xlsx"))
        sheet = workbook.active
        _ = sheet  # prevents false unused-variable warnings
        workbook.save(path)
        workbook.close()

    # Open existing file
    workbook = load_workbook(path)
    sheet = workbook.active

    # Find next empty row
    row = sheet.max_row + 1
    while all(sheet[f"{col}{row}"].value is None for col in "ABCDEFG"):
        row -= 1
    row += 1

    # Append rows
    for r in rows:
        sheet[f"A{row}"] = r[0]  # Date
        sheet[f"B{row}"] = r[1]  # Job Name
        sheet[f"C{row}"] = r[2]  # Pay Item
        sheet[f"D{row}"] = r[3]  # Quantity
        sheet[f"E{row}"] = r[4]  # Rate
        sheet[f"F{row}"] = r[5]  # Total
        sheet[f"G{row}"] = r[6]  # Split
        row += 1

    workbook.save(path)
    workbook.close()

# =========================================================
# Authentication Helper
# =========================================================
def authenticate(username, password):
    """Return the matching Employee object if credentials are valid."""
    username = username.strip().lower()

    for emp in employees:
        if emp.username.lower() == username and emp.pwd == password:
            return emp

    return None


# =========================================================
# Employee Excel Column Constants
# =========================================================
COL_FIRST             = "first"
COL_LAST              = "last"
COL_VACATION_MAX      = "vacation_max"
COL_VACATION_REMAIN   = "vacation_remaining"
COL_ROLE              = "role"
COL_USERNAME          = "username"
COL_PASSWORD          = "password"


# =========================================================
# Employee Class
# =========================================================
class Employee:
    """Data model representing a single employee record."""

    def __init__(self, first, last, vacation_max=0, vacation_remaining=0,
                 passwd="", username=None, role="employee"):

        self.first = first
        self.last = last
        self.vacation_max = int(vacation_max)
        self.vacation_remaining = int(vacation_remaining)
        self.pwd = passwd
        self.role = role.lower().strip()

        # Derived fields
        self.fullname = f"{first} {last}"

        # Username is generated ONCE unless explicitly provided
        self.username = username if username else f"{first[0]}{last}".lower()

    def to_dict(self):
        return {
            COL_FIRST: self.first,
            COL_LAST: self.last,
            COL_VACATION_MAX: self.vacation_max,
            COL_VACATION_REMAIN: self.vacation_remaining,
            COL_ROLE: self.role,
            COL_USERNAME: self.username,
            COL_PASSWORD: self.pwd
        }

    def __repr__(self):
        return (
            f"Employee(fullname='{self.fullname}', "
            f"username='{self.username}', role='{self.role}', "
            f"vacation_remaining={self.vacation_remaining})"
        )

# =========================================================
# Load Employees from Excel
# =========================================================
def load_employees_from_excel(path):
    """Load employee records from an Excel file into Employee objects."""

    if not os.path.exists(path):
        print(f"WARNING: Employee file not found at: {path}")
        print("Employee list initialized as empty.")
        return []

    df = pd.read_excel(path)
    loaded_employees = []

    for _, row in df.iterrows():
        emp = Employee(
            first=row[COL_FIRST],
            last=row[COL_LAST],
            vacation_max=row.get(COL_VACATION_MAX, 0),
            vacation_remaining=row.get(COL_VACATION_REMAIN, 0),
            passwd=row.get(COL_PASSWORD, ""),
            username=row.get(COL_USERNAME),
            role=row.get(COL_ROLE, "employee")
        )

        loaded_employees.append(emp)

    return loaded_employees

# =========================================================
# Initialize Global Employee List
# =========================================================
employees = load_employees_from_excel(EMPLOYEE_FILE)
print(f"Loaded {len(employees)} employees from Excel.")


# =========================================================
# Save Employees to Excel
# =========================================================
def save_employees_to_excel():
    """Write all employee records back to the Excel file."""

    # Ensure directory exists
    os.makedirs(os.path.dirname(EMPLOYEE_FILE), exist_ok=True)

    df = pd.DataFrame(
        [emp.to_dict() for emp in employees],
        columns=[
            COL_FIRST,
            COL_LAST,
            COL_VACATION_MAX,
            COL_VACATION_REMAIN,
            COL_ROLE,
            COL_USERNAME,
            COL_PASSWORD
        ]
    )

    df.to_excel(EMPLOYEE_FILE, index=False)
    print("Employees saved to Excel.")


# =========================================================
# Payrate Excel Column Constants
# =========================================================
COL_PAYRATE_NAME = "name"
COL_PAYRATE_RATE = "rate"


# =========================================================
# Helper: Clean dictionary keys
# =========================================================
def clean_key(text):
    return (
        text.lower()
        .replace(" ", "_")
        .replace("/", "_")
        .replace('"', "")
        .replace(">", "")
        .replace("<", "")
    )


# =========================================================
# Payrate Class
# =========================================================
class Payrate:
    """Represents a payrate entry with a name and pricing rate."""

    def __init__(self, name, rate):
        self.name = name
        self.rate = float(rate)

    def __repr__(self):
        return f"Payrate(name='{self.name}', rate={self.rate})"

    def to_dict(self):
        return {
            COL_PAYRATE_NAME: self.name,
            COL_PAYRATE_RATE: self.rate
        }


# =========================================================
# Load Payrates from Excel
# =========================================================
def load_payrates_from_excel(path):
    """Load payrate records from an Excel file into a dictionary."""

    if not os.path.exists(path):
        print(f"WARNING: Payrate file not found at: {path}")
        print("Payrate list initialized as empty.")
        return {}

    df = pd.read_excel(path)

    # Normalize column names
    df.columns = [str(c).strip().lower() for c in df.columns]

    loaded_payrates = {}

    for _, row in df.iterrows():
        name = str(row[COL_PAYRATE_NAME]).strip()
        rate = float(row[COL_PAYRATE_RATE])

        payrate_obj = Payrate(name, rate)
        key = clean_key(name)

        loaded_payrates[key] = payrate_obj

    return loaded_payrates


# =========================================================
# Save Payrates to Excel
# =========================================================
def save_payrates_to_excel(path, payrate_dict):
    """Write payrate records back to Excel."""

    os.makedirs(os.path.dirname(path), exist_ok=True)

    data = [pr.to_dict() for pr in payrate_dict.values()]
    df = pd.DataFrame(data, columns=[COL_PAYRATE_NAME, COL_PAYRATE_RATE])

    df.to_excel(path, index=False)
    print("Payrates saved to Excel.")


# =========================================================
# Initialize Global Payrate Dictionary
# =========================================================
payrates = load_payrates_from_excel(PAYRATE_FILE)
print(f"Loaded {len(payrates)} payrates from Excel.")

# =========================================================
# Application Controller
# =========================================================
class AppController(ctk.CTk):
    """Main application controller that manages all frames, navigation, and role-based menus."""

    def __init__(self):
        super().__init__()

        # -----------------------------
        # Window Configuration
        # -----------------------------
        self.title("Insulpro - InsulPay")
        self.geometry("900x600")
        self.resizable(True, True)
        self.state("zoomed")
        self.bind_shortcuts()
        self.bind("<Control-Shift-P>", lambda e: self.open_command_palette())

        # Logged-in user reference
        self.current_user = None

        # Frame registry (lazy loading)
        self.frames = {}
        self.frame_classes = {
            # Authentication
            "login": LoginFrame,
            "change_password": ChangePasswordFrame,

            # Admin Screens
            "main_admin": MainAdminFrame,
            "employee_management": EmployeeManagementFrame,
            "add_employee": AddEmployeeFrame,
            "employee_detail": EmployeeDetailFrame,

            # Payroll Screens
            "payroll_tools_menu": PayrollToolsMenuFrame,
            "paysheet": PaySheetFrame,
            "weekly_payroll": WeeklyPayrollFrame,
            "view_weekly": ViewWeeklyPayrollFrame,
            "ytd": YTDPayrollFrame,
            "view_ytd": ViewYTDSummaryFrame,

            # Vacation Screens
            "vacation_tool_menu": VacationToolsMenuFrame,
            "vacation_tool": VacationPayrollFrame,
            "view_vacation": VacationHistoryFrame,
            "vacation_approval": VacationRequestApprovalFrame,

            # Employee Screens
            "main_employee": MainEmployeeFrame,
            "employee_info": EmployeeInfoFrame,
            "request_vacation": RequestVacationFrame,
            "view_employee_payroll": EmployeeWeeklyPayFrame,
        }

        # Show login screen first
        self.show_frame("login")

    # =========================================================
    # Bind Menu Shortcuts
    # =========================================================

    def bind_shortcuts(self):
        # Global shortcuts
        self.bind("<Control-q>", lambda e: self.quit())
        self.bind("<Control-l>", lambda e: self.logout_from_menu())
        self.bind("<F1>", lambda e: self.show_about())

        # Admin-only shortcuts (checked at runtime)
        self.bind("<Control-e>", lambda e: self._shortcut_to("employee_management"))
        self.bind("<Control-p>", lambda e: self._shortcut_to("payroll_tools_menu"))
        self.bind("<Control-v>", lambda e: self._shortcut_to("vacation_tool_menu"))

    # =========================================================
    # Lazy Frame Loader
    # =========================================================
    def get_frame(self, name):
        """Create the frame only when needed (lazy loading)."""
        if name not in self.frames:
            cls = self.frame_classes[name]
            self.frames[name] = cls(self)
        return self.frames[name]

    # =========================================================
    # Role-Based Menu Builder
    # =========================================================
    def build_menu_for_role(self, role):
        menu_bar = tk.Menu(self)

        # -----------------------------
        # File Menu
        # -----------------------------
        file_menu = tk.Menu(menu_bar, tearoff=0)
        file_menu.add_command(label="Logout", command=self.logout_from_menu)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.quit)
        menu_bar.add_cascade(label="File", menu=file_menu)

        # -----------------------------
        # Tools Menu (ADMIN ONLY)
        # -----------------------------
        if role == "admin":
            tools_menu = tk.Menu(menu_bar, tearoff=0)
            tools_menu.add_command(label="Employee Management",
                                   command=lambda: self.show_frame("employee_management"))
            tools_menu.add_command(label="Payroll Tools",
                                   command=lambda: self.show_frame("payroll_tools_menu"))
            tools_menu.add_command(label="Vacation Tools",
                                   command=lambda: self.show_frame("vacation_tool_menu"))
            menu_bar.add_cascade(label="Tools", menu=tools_menu)

        # -----------------------------
        # Help Menu
        # -----------------------------
        help_menu = tk.Menu(menu_bar, tearoff=0)
        help_menu.add_command(label="About", command=self.show_about)
        menu_bar.add_cascade(label="Help", menu=help_menu)

        # Attach menu to window
        self.config(menu=menu_bar)

    # =========================================================
    # Frame Navigation
    # =========================================================
    def show_frame(self, name):
        """Show a frame by name and hide all others."""

        # Hide all existing frames
        for frame in self.frames.values():
            frame.place_forget()

        # Create or retrieve the frame
        frame = self.get_frame(name)

        # Show it
        frame.place(relwidth=1, relheight=1)

        # Update menu bar based on role
        if self.current_user:
            self.build_menu_for_role(self.current_user.role)
        else:
            self.build_menu_for_role("none")

        # Call on_show() if the frame defines it
        if hasattr(frame, "on_show"):
            frame.on_show()

    def _shortcut_to(self, frame_name):
        if self.current_user and self.current_user.role == "admin":
            self.show_frame(frame_name)

    # =========================================================
    # Menu Actions
    # =========================================================
    def logout_from_menu(self):
        self.current_user = None
        self.show_frame("login")

    @staticmethod
    def show_about():
        messagebox.showinfo(
            "About InsulPay",
            "InsulPay Payroll System\nCreated by Kenneth Ray"
        )

    def open_command_palette(self):
        if not self.current_user:
            return

        role = self.current_user.role

        # Base commands for all users
        commands = {
            "Logout": lambda: self.logout_from_menu(),
            "Change Password": lambda: self.show_frame("change_password"),
            "About": lambda: self.show_about(),
        }

        # Employee commands
        if role == "employee":
            commands.update({
                "My Info": lambda: self.show_frame("employee_info"),
                "Request Vacation": lambda: self.show_frame("request_vacation"),
                "View My Weekly Payroll": lambda: self.show_frame("view_employee_payroll"),
            })

        # Admin commands
        if role == "admin":
            commands.update({
                "Admin Dashboard": lambda: self.show_frame("main_admin"),
                "Employee Management": lambda: self.show_frame("employee_management"),
                "Add Employee": lambda: self.show_frame("add_employee"),
                "Payroll Tools": lambda: self.show_frame("payroll_tools_menu"),
                "Weekly Payroll": lambda: self.show_frame("weekly_payroll"),
                "YTD Payroll": lambda: self.show_frame("ytd"),
                "Vacation Tools": lambda: self.show_frame("vacation_tool_menu"),
                "Vacation Approvals": lambda: self.show_frame("vacation_approval"),
            })

        # Open the palette
        CommandPalette(self, commands)


# =========================================================
# Login Screen Frame
# =========================================================
class LoginFrame(ctk.CTkFrame):
    """Login screen for authenticating employees."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Login",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title.pack(pady=40)

        # -----------------------------
        # Form Container
        # -----------------------------
        form = ctk.CTkFrame(self)
        form.pack(pady=20)

        # Username
        ctk.CTkLabel(form, text="Username:").grid(
            row=0, column=0, padx=10, pady=10
        )
        self.username_entry = ctk.CTkEntry(form, width=250)
        self.username_entry.grid(row=0, column=1, padx=10, pady=10)

        # Password
        ctk.CTkLabel(form, text="Password:").grid(
            row=1, column=0, padx=10, pady=10
        )
        self.password_entry = ctk.CTkEntry(
            form, width=250, show="*"
        )
        self.password_entry.grid(row=1, column=1, padx=10, pady=10)

        # Press Enter to login
        self.username_entry.bind("<Return>", lambda e: self.handle_login())
        self.password_entry.bind("<Return>", lambda e: self.handle_login())

        # -----------------------------
        # Login Button
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Login",
            width=200,
            command=self.handle_login
        ).pack(pady=30)

        # Error Label
        self.error_label = ctk.CTkLabel(
            self,
            text="",
            text_color="red"
        )
        self.error_label.pack()

        # -----------------------------
        # Footer Branding
        # -----------------------------
        footer = ctk.CTkLabel(
            self,
            text="InsulPay Payroll System - Login System",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        footer.pack(side="bottom", pady=10)

    # ---------------------------------------------------------
    # Login Handler
    # ---------------------------------------------------------
    def handle_login(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()

        user = authenticate(username, password)

        if not user:
            self.error_label.configure(text="Invalid username or password.")
            self.password_entry.delete(0, "end")
            return

        # Clear error message
        self.error_label.configure(text="")

        # Store logged-in user
        self.master.current_user = user

        # Route based on role
        if user.role == "admin":
            self.master.show_frame("main_admin")
        else:
            self.master.show_frame("main_employee")


# =========================================================
# Change Password Frame
# =========================================================
class ChangePasswordFrame(ctk.CTkFrame):
    """Allows the employee to change their password."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Back Button (Top Left)
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Back",
            width=120,
            command=lambda: self.master.show_frame("employee_info")
        ).place(x=20, y=20)

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Change Password",
            font=ctk.CTkFont(size=26, weight="bold")
        )
        title.pack(pady=60)

        # -----------------------------
        # Form Container
        # -----------------------------
        form = ctk.CTkFrame(self)
        form.pack(pady=20)

        # Current Password
        ctk.CTkLabel(form, text="Current Password:").grid(
            row=0, column=0, padx=10, pady=10
        )
        self.current_entry = ctk.CTkEntry(form, width=250, show="*")
        self.current_entry.grid(row=0, column=1, padx=10, pady=10)

        # New Password
        ctk.CTkLabel(form, text="New Password:").grid(
            row=1, column=0, padx=10, pady=10
        )
        self.new_entry = ctk.CTkEntry(form, width=250, show="*")
        self.new_entry.grid(row=1, column=1, padx=10, pady=10)

        # Confirm New Password
        ctk.CTkLabel(form, text="Confirm New Password:").grid(
            row=2, column=0, padx=10, pady=10
        )
        self.confirm_entry = ctk.CTkEntry(form, width=250, show="*")
        self.confirm_entry.grid(row=2, column=1, padx=10, pady=10)

        # Bind Enter to submit
        self.current_entry.bind("<Return>", lambda e: self.save_password())
        self.new_entry.bind("<Return>", lambda e: self.save_password())
        self.confirm_entry.bind("<Return>", lambda e: self.save_password())

        # -----------------------------
        # Save Button
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Save Password",
            width=200,
            command=self.save_password
        ).pack(pady=20)

    # ---------------------------------------------------------
    # When frame is shown
    # ---------------------------------------------------------
    def on_show(self):
        self.current_entry.delete(0, "end")
        self.new_entry.delete(0, "end")
        self.confirm_entry.delete(0, "end")
        self.current_entry.focus()

    # ---------------------------------------------------------
    # Save New Password
    # ---------------------------------------------------------
    def save_password(self):
        user = self.master.current_user

        current = self.current_entry.get().strip()
        new = self.new_entry.get().strip()
        confirm = self.confirm_entry.get().strip()

        # Validate current password
        if current != user.pwd:
            messagebox.showerror("Error", "Current password is incorrect.")
            return

        # Validate match
        if new != confirm:
            messagebox.showerror("Error", "New passwords do not match.")
            return

        # Validate length
        if len(new) < 3:
            messagebox.showerror("Error", "Password must be at least 3 characters.")
            return

        # Save new password
        user.pwd = new
        save_employees_to_excel()

        messagebox.showinfo("Success", "Password updated successfully.")
        self.master.show_frame("employee_info")


# =========================================================
# Main Admin Window
# =========================================================
class MainAdminFrame(ctk.CTkFrame):
    """Main dashboard for admin users."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Logout Button (Top Left)
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Logout",
            width=120,
            fg_color="red",
            hover_color="#8b0000",
            command=self.logout
        ).place(x=20, y=20)

        # -----------------------------
        # Title / Welcome
        # -----------------------------
        self.title_label = ctk.CTkLabel(
            self,
            text="InsulPay Dashboard",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        self.title_label.pack(pady=(60, 10))

        self.welcome_label = ctk.CTkLabel(
            self,
            text="",
            font=ctk.CTkFont(size=18)
        )
        self.welcome_label.pack(pady=10)

        # -----------------------------
        # Button Container
        # -----------------------------
        button_frame = ctk.CTkFrame(self)
        button_frame.pack(pady=30)

        # Manage Employee Records
        ctk.CTkButton(
            button_frame,
            text="Manage Employee Data",
            width=250,
            command=lambda: self.master.show_frame("employee_management")
        ).pack(pady=10)

        # Payroll Tools
        ctk.CTkButton(
            button_frame,
            text="Payroll Tools",
            width=250,
            command=lambda: self.master.show_frame("payroll_tools_menu")
        ).pack(pady=10)

        # Vacation Tools
        ctk.CTkButton(
            button_frame,
            text="Vacation Tools",
            width=250,
            command=lambda: self.master.show_frame("vacation_tool_menu")
        ).pack(pady=10)

        # -----------------------------
        # Footer
        # -----------------------------
        footer = ctk.CTkLabel(
            self,
            text="InsulPay Payroll System — Admin Panel",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        footer.pack(side="bottom", pady=10)

    # ---------------------------------------------------------
    # Update welcome message when frame is shown
    # ---------------------------------------------------------
    def on_show(self):
        user = self.master.current_user
        self.welcome_label.configure(text=f"Welcome, {user.fullname}")

    # ---------------------------------------------------------
    # Logout Handler
    # ---------------------------------------------------------
    def logout(self):
        self.master.current_user = None
        self.master.show_frame("login")

# =========================================================
# Employee Records (Admin Module)
# =========================================================
class EmployeeManagementFrame(ctk.CTkFrame):
    """Admin view for listing, adding, editing, and deleting employees."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Employee Management",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title.pack(pady=30)

        # -----------------------------
        # Table Container (Scrollable)
        # -----------------------------
        self.table_frame = ctk.CTkScrollableFrame(
            self,
            width=700,
            height=350
        )
        self.table_frame.pack(pady=10)

        # -----------------------------
        # Buttons
        # -----------------------------
        button_frame = ctk.CTkFrame(self)
        button_frame.pack(pady=20)

        ctk.CTkButton(
            button_frame,
            text="Add Employee",
            width=200,
            command=lambda: self.master.show_frame("add_employee")
        ).grid(row=0, column=0, padx=10)

        ctk.CTkButton(
            button_frame,
            text="Back",
            width=200,
            command=lambda: self.master.show_frame("main_admin")
        ).grid(row=0, column=1, padx=10)

    # ---------------------------------------------------------
    # Refresh table when frame is shown
    # ---------------------------------------------------------
    def on_show(self):
        # Clear old rows
        for widget in self.table_frame.winfo_children():
            widget.destroy()

        # Table headers
        headers = ["Username", "Full Name", "Role", "Vacation Days", "Edit", "Delete"]
        for col, text in enumerate(headers):
            ctk.CTkLabel(
                self.table_frame,
                text=text,
                font=ctk.CTkFont(weight="bold")
            ).grid(row=0, column=col, padx=10, pady=5)

        # Table rows
        for row_index, emp in enumerate(employees, start=1):

            ctk.CTkLabel(self.table_frame, text=emp.username).grid(row=row_index, column=0, padx=10, pady=5)
            ctk.CTkLabel(self.table_frame, text=emp.fullname).grid(row=row_index, column=1, padx=10, pady=5)
            ctk.CTkLabel(self.table_frame, text=emp.role).grid(row=row_index, column=2, padx=10, pady=5)

            vac_text = f"{emp.vacation_remaining} / {emp.vacation_max}"
            ctk.CTkLabel(self.table_frame, text=vac_text).grid(row=row_index, column=3, padx=10, pady=5)

            ctk.CTkButton(
                self.table_frame,
                text="Edit",
                width=80,
                command=lambda e=emp: self.open_edit(e)
            ).grid(row=row_index, column=4, padx=10)

            ctk.CTkButton(
                self.table_frame,
                text="Delete",
                width=80,
                fg_color="red",
                hover_color="#8b0000",
                command=lambda e=emp: self.delete_employee(e)
            ).grid(row=row_index, column=5, padx=10)

    # ---------------------------------------------------------
    # Open Edit Screen (Lazy-loading safe)
    # ---------------------------------------------------------
    def open_edit(self, employee):
        frame = self.master.get_frame("employee_detail")
        frame.set_employee(employee)
        self.master.show_frame("employee_detail")

    # ---------------------------------------------------------
    # Delete Employee
    # ---------------------------------------------------------
    def delete_employee(self, employee):
        confirm = messagebox.askyesno(
            "Confirm Delete",
            f"Delete {employee.fullname}?"
        )
        if not confirm:
            return

        employees.remove(employee)
        save_employees_to_excel()
        self.on_show()

# =========================================================
# Employee Detail / Edit Screen
# =========================================================
class EmployeeDetailFrame(ctk.CTkFrame):
    """Admin screen for editing an existing employee."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.employee = None

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Edit Employee",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title.pack(pady=30)

        # -----------------------------
        # Form Container
        # -----------------------------
        form = ctk.CTkFrame(self)
        form.pack(pady=20)

        # First Name
        ctk.CTkLabel(form, text="First Name:").grid(
            row=0, column=0, padx=10, pady=10, sticky="e"
        )
        self.first_entry = ctk.CTkEntry(form, width=250)
        self.first_entry.grid(row=0, column=1)

        # Last Name
        ctk.CTkLabel(form, text="Last Name:").grid(
            row=1, column=0, padx=10, pady=10, sticky="e"
        )
        self.last_entry = ctk.CTkEntry(form, width=250)
        self.last_entry.grid(row=1, column=1)

        # Role
        ctk.CTkLabel(form, text="Role:").grid(
            row=2, column=0, padx=10, pady=10, sticky="e"
        )
        self.role_option = ctk.CTkOptionMenu(form, values=["employee", "admin"])
        self.role_option.grid(row=2, column=1)

        # Vacation Max
        ctk.CTkLabel(form, text="Vacation Max Days:").grid(
            row=3, column=0, padx=10, pady=10, sticky="e"
        )
        self.vac_max_entry = ctk.CTkEntry(form, width=250)
        self.vac_max_entry.grid(row=3, column=1)

        # Vacation Remaining
        ctk.CTkLabel(form, text="Vacation Remaining:").grid(
            row=4, column=0, padx=10, pady=10, sticky="e"
        )
        self.vac_remaining_entry = ctk.CTkEntry(form, width=250)
        self.vac_remaining_entry.grid(row=4, column=1)

        # Password
        ctk.CTkLabel(form, text="Password:").grid(
            row=5, column=0, padx=10, pady=10, sticky="e"
        )
        self.pass_entry = ctk.CTkEntry(form, width=250)
        self.pass_entry.grid(row=5, column=1)

        # -----------------------------
        # Buttons
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Save Changes",
            width=200,
            command=self.save_changes
        ).pack(pady=20)

        ctk.CTkButton(
            self,
            text="Back",
            width=200,
            command=lambda: self.master.show_frame("employee_management")
        ).pack()

    # ---------------------------------------------------------
    # Load employee data into fields
    # ---------------------------------------------------------
    def set_employee(self, employee):
        self.employee = employee

        self.first_entry.delete(0, "end")
        self.first_entry.insert(0, employee.first)

        self.last_entry.delete(0, "end")
        self.last_entry.insert(0, employee.last)

        self.role_option.set(employee.role)

        self.vac_max_entry.delete(0, "end")
        self.vac_max_entry.insert(0, employee.vacation_max)

        self.vac_remaining_entry.delete(0, "end")
        self.vac_remaining_entry.insert(0, employee.vacation_remaining)

        self.pass_entry.delete(0, "end")
        self.pass_entry.insert(0, employee.pwd)

    # ---------------------------------------------------------
    # Save updated employee data
    # ---------------------------------------------------------
    def save_changes(self):
        if not self.employee:
            return

        self.employee.first = self.first_entry.get().strip()
        self.employee.last = self.last_entry.get().strip()
        self.employee.fullname = f"{self.employee.first} {self.employee.last}"

        self.employee.role = self.role_option.get().strip().lower()

        self.employee.vacation_max = int(self.vac_max_entry.get().strip())
        self.employee.vacation_remaining = int(self.vac_remaining_entry.get().strip())

        self.employee.pwd = self.pass_entry.get().strip()

        save_employees_to_excel()

        messagebox.showinfo("Success", "Employee updated successfully.")
        self.master.show_frame("employee_management")

# =========================================================
# Add Employee Screen
# =========================================================
class AddEmployeeFrame(ctk.CTkFrame):
    """Screen for creating a new employee."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Add New Employee",
            font=ctk.CTkFont(size=26, weight="bold")
        )
        title.pack(pady=30)

        # -----------------------------
        # Form Container
        # -----------------------------
        form = ctk.CTkFrame(self)
        form.pack(pady=20)

        # First Name
        ctk.CTkLabel(form, text="First Name:").grid(
            row=0, column=0, padx=10, pady=10
        )
        self.first_entry = ctk.CTkEntry(form, width=250)
        self.first_entry.grid(row=0, column=1)

        # Last Name
        ctk.CTkLabel(form, text="Last Name:").grid(
            row=1, column=0, padx=10, pady=10
        )
        self.last_entry = ctk.CTkEntry(form, width=250)
        self.last_entry.grid(row=1, column=1)

        # Vacation Days
        ctk.CTkLabel(form, text="Vacation Days:").grid(
            row=2, column=0, padx=10, pady=10
        )
        self.vac_entry = ctk.CTkEntry(form, width=250)
        self.vac_entry.grid(row=2, column=1)

        # Role
        ctk.CTkLabel(form, text="Role:").grid(
            row=3, column=0, padx=10, pady=10
        )
        self.role_option = ctk.CTkOptionMenu(form, values=["employee", "admin"])
        self.role_option.grid(row=3, column=1)

        # Password
        ctk.CTkLabel(form, text="Password:").grid(
            row=4, column=0, padx=10, pady=10
        )
        self.pass_entry = ctk.CTkEntry(form, width=250, show="*")
        self.pass_entry.grid(row=4, column=1)

        # -----------------------------
        # Save Button
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Create Employee",
            width=200,
            command=self.create_employee
        ).pack(pady=20)

        # -----------------------------
        # Back Button
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Back",
            width=200,
            command=lambda: self.master.show_frame("employee_management")
        ).pack()

    # ---------------------------------------------------------
    # Reset fields when frame is shown
    # ---------------------------------------------------------
    def on_show(self):
        self.first_entry.delete(0, "end")
        self.last_entry.delete(0, "end")
        self.vac_entry.delete(0, "end")
        self.pass_entry.delete(0, "end")
        self.role_option.set("employee")
        self.first_entry.focus()

    # ---------------------------------------------------------
    # Create new employee
    # ---------------------------------------------------------
    def create_employee(self):
        first = self.first_entry.get().strip()
        last = self.last_entry.get().strip()
        vac = int(self.vac_entry.get().strip())
        role = self.role_option.get().strip().lower()
        pwd = self.pass_entry.get().strip()

        username = f"{first[0]}{last}".lower()

        new_emp = Employee(
            first=first,
            last=last,
            vacation_max=vac,
            vacation_remaining=vac,
            passwd=pwd,
            username=username,
            role=role
        )

        employees.append(new_emp)
        save_employees_to_excel()

        messagebox.showinfo("Success", "Employee created successfully.")
        self.master.show_frame("employee_management")

# =========================================================
# Payroll Tools Menu
# =========================================================
class PayrollToolsMenuFrame(ctk.CTkFrame):
    """Menu screen for all payroll-related tools."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Payroll Tools",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title.pack(pady=40)

        # -----------------------------
        # Button Container
        # -----------------------------
        button_frame = ctk.CTkFrame(self)
        button_frame.pack(pady=20)

        # Generate Paysheet Button
        ctk.CTkButton(
            button_frame,
            text="Generate Paysheet",
            width=260,
            command=lambda: self.master.show_frame("paysheet")
        ).pack(pady=10)

        # Generate Weekly Payroll Button
        ctk.CTkButton(
            button_frame,
            text="Generate Weekly Payroll",
            width=260,
            command=lambda: self.master.show_frame("weekly_payroll")
        ).pack(pady=10)

        # View Weekly Payroll Button
        ctk.CTkButton(
            button_frame,
            text="View Weekly Payroll",
            width=260,
            command=lambda: self.master.show_frame("view_weekly")
        ).pack(pady=10)

        # Generate YTD Payroll Button
        ctk.CTkButton(
            button_frame,
            text="Generate YTD Payroll",
            width=260,
            command=lambda: self.master.show_frame("ytd")
        ).pack(pady=10)

        # View Yearly Payroll Button
        ctk.CTkButton(
            button_frame,
            text="View YTD Payroll",
            width=260,
            command=lambda: self.master.show_frame("view_ytd")
        ).pack(pady=10)

        # Back button
        ctk.CTkButton(
            self,
            text="Back",
            width=220,
            command=lambda: self.master.show_frame("main_admin")
        ).pack(pady=40)

    def on_show(self):
        pass  # Nothing dynamic yet, but ready for future updates

# =========================================================
# Pay Sheet Frame
# =========================================================
class PaySheetFrame(ctk.CTkFrame):
    """Screen for creating a job paysheet with employee selection and payrate inputs."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Create Pay Sheet",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title.pack(pady=20)

        # =====================================================
        # JOB NAME + DATE FIELDS
        # =====================================================
        header_frame = ctk.CTkFrame(self)
        header_frame.pack(pady=10)

        ctk.CTkLabel(header_frame, text="Job Name:").grid(
            row=0, column=0, padx=10, pady=5
        )
        self.jobname_entry = ctk.CTkEntry(header_frame, width=200)
        self.jobname_entry.grid(row=0, column=1, padx=10, pady=5)

        ctk.CTkLabel(header_frame, text="Date:").grid(
            row=1, column=0, padx=10, pady=5
        )
        self.date_entry = ctk.CTkEntry(header_frame, width=200)
        self.date_entry.grid(row=1, column=1, padx=10, pady=5)

        # =====================================================
        # EMPLOYEE CHECKBOX LIST
        # =====================================================
        self.emp_frame = ctk.CTkScrollableFrame(self, width=250, height=300)
        self.emp_frame.pack(side="left", padx=20, pady=10)

        ctk.CTkLabel(
            self.emp_frame,
            text="Select Employees:",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(pady=5)

        self.employee_vars = {}

        # =====================================================
        # PAYRATE INPUT GRID
        # =====================================================
        self.rate_frame = ctk.CTkScrollableFrame(self, width=300, height=300)
        self.rate_frame.pack(side="right", padx=20, pady=10)

        ctk.CTkLabel(
            self.rate_frame,
            text="Enter Quantities:",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(pady=5)

        self.payrate_entries = {}

        # =====================================================
        # ACTION BUTTONS
        # =====================================================
        btn_frame = ctk.CTkFrame(self)
        btn_frame.pack(pady=20)

        ctk.CTkButton(
            btn_frame,
            text="Calculate Split",
            width=200,
            command=self.calculate_split
        ).pack(pady=10)

        ctk.CTkButton(
            btn_frame,
            text="Save Pay Sheet",
            width=200,
            command=self.save_paysheet
        ).pack(pady=10)

        ctk.CTkButton(
            btn_frame,
            text="Reset Form",
            width=200,
            command=self.reset_form
        ).pack(pady=10)

        ctk.CTkButton(
            btn_frame,
            text="Back",
            width=200,
            command=lambda: self.master.show_frame("payroll_tools_menu")
        ).pack(pady=10)

    def on_show(self):
        # Reset job name
        self.jobname_entry.delete(0, "end")

        # Reset date
        self.date_entry.delete(0, "end")
        today = datetime.date.today().strftime("%m-%d-%Y")
        self.date_entry.insert(0, today)

        # -----------------------------
        # Rebuild employee list
        # -----------------------------
        for widget in self.emp_frame.winfo_children():
            if isinstance(widget, ctk.CTkCheckBox):
                widget.destroy()

        self.employee_vars.clear()

        for emp in employees:
            var = tk.BooleanVar()
            chk = ctk.CTkCheckBox(
                self.emp_frame,
                text=emp.fullname,
                variable=var
            )
            chk.pack(anchor="w", pady=2)
            self.employee_vars[emp.username] = var

        # -----------------------------
        # Rebuild payrate list
        # -----------------------------
        # Clear old widgets
        for widget in self.rate_frame.winfo_children():
            widget.destroy()

        # Reset dictionary
        self.payrate_entries.clear()

        # Re-add header
        ctk.CTkLabel(
            self.rate_frame,
            text="Enter Quantities:",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(pady=5)

        # Build fresh list
        for key, pr in payrates.items():
            row = ctk.CTkFrame(self.rate_frame)
            row.pack(fill="x", pady=3)

            ctk.CTkLabel(row, text=pr.name, width=180, anchor="w").pack(side="left", padx=5)

            entry = ctk.CTkEntry(row, width=80)
            entry.pack(side="right", padx=5)

            self.payrate_entries[key] = entry

    def calculate_split(self):
        # -----------------------------
        # Validate job name and date
        # -----------------------------
        job_name = self.jobname_entry.get().strip()
        date_value = self.date_entry.get().strip()

        if not job_name:
            messagebox.showerror("Missing Job Name", "Please enter a job name.")
            return

        if not date_value:
            messagebox.showerror("Missing Date", "Please enter a date.")
            return

        # -----------------------------
        # Collect selected employees
        # -----------------------------
        selected = [
            username for username, var in self.employee_vars.items()
            if var.get()
        ]

        if not selected:
            messagebox.showerror(
                "No Employees Selected",
                "Please select at least one employee."
            )
            return

        num_workers = len(selected)

        # -----------------------------
        # Collect payrate quantities
        # -----------------------------
        results = []

        for key, entry in self.payrate_entries.items():
            qty_text = entry.get().strip()
            if not qty_text:
                continue

            try:
                qty = float(qty_text)
            except ValueError:
                messagebox.showerror(
                    "Invalid Quantity",
                    f"Invalid number for {payrates[key].name}"
                )
                return

            rate = payrates[key].rate
            total = qty * rate
            split = total / num_workers

            results.append((payrates[key].name, qty, rate, total, split))

        # -----------------------------
        # No quantities entered
        # -----------------------------
        if not results:
            messagebox.showinfo("No Data", "No quantities entered.")
            return

        # -----------------------------
        # Open the results window
        # -----------------------------
        PaySplitResultsWindow(
            self,
            job_name,
            date_value,
            num_workers,
            results
        )

    # =========================================================
    # SAVE PAY SHEET TO EXCEL
    # =========================================================
    def save_paysheet(self):
        # -----------------------------
        # Validate job name and date
        # -----------------------------
        job_name = self.jobname_entry.get().strip()
        date_value = self.date_entry.get().strip()

        if not job_name:
            messagebox.showerror("Missing Job Name", "Please enter a job name.")
            return

        if not date_value:
            messagebox.showerror("Missing Date", "Please enter a date.")
            return

        # -----------------------------
        # Collect selected employees
        # -----------------------------
        selected_emps = [
            emp for emp in employees
            if self.employee_vars[emp.username].get()
        ]

        if not selected_emps:
            messagebox.showerror("No Employees Selected", "Select at least one employee.")
            return

        num_workers = len(selected_emps)

        # -----------------------------
        # Collect payrate entries
        # -----------------------------
        pay_items = []

        for key, entry in self.payrate_entries.items():
            qty_text = entry.get().strip()
            if not qty_text:
                continue

            try:
                qty = float(qty_text)
            except ValueError:
                messagebox.showerror(
                    "Invalid Quantity",
                    f"Invalid number for {payrates[key].name}"
                )
                return

            rate = payrates[key].rate
            total = qty * rate
            split = total / num_workers

            pay_items.append((payrates[key].name, qty, rate, total, split))

        if not pay_items:
            messagebox.showerror("No Pay Items", "Enter at least one quantity.")
            return

        # -----------------------------
        # Create paysheet file
        # -----------------------------
        template = netpath("data", "spreadsheet", "PaySheetTemplate.xlsx")
        save_dir = netpath("payroll_records", "paysheets")

        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        filename = f"{job_name}.xlsx"
        new_path = os.path.join(save_dir, filename)

        # Prevent overwriting existing file
        if os.path.exists(new_path):
            messagebox.showerror(
                "File Already Exists",
                f"A pay sheet named '{filename}' already exists.\n"
                "Please choose a different job name."
            )
            return

        shutil.copy(template, new_path)

        # -----------------------------
        # Write data into Excel
        # -----------------------------
        workbook = load_workbook(new_path)
        sheet = workbook.active

        # Header
        sheet["B1"] = job_name
        sheet["B2"] = date_value

        # Employees
        row = 5
        for emp in selected_emps:
            sheet[f"A{row}"] = emp.fullname
            row += 1

        # Pay Items
        row = 5
        for name, qty, rate, total, split in pay_items:
            sheet[f"C{row}"] = name
            sheet[f"D{row}"] = qty
            sheet[f"E{row}"] = rate
            sheet[f"F{row}"] = total
            sheet[f"G{row}"] = split
            row += 1

        workbook.save(new_path)
        workbook.close()

        # -----------------------------
        # Update master payroll files
        # -----------------------------
        for emp in selected_emps:
            rows = []
            for name, qty, rate, total, split in pay_items:
                rows.append((
                    date_value,
                    job_name,
                    name,
                    qty,
                    rate,
                    total,
                    split
                ))
            append_to_master_payroll(emp, rows)

        messagebox.showinfo("Saved", f"Pay sheet saved:\n{new_path}")

        self.reset_form()

    # =========================================================
    # RESET FORM
    # =========================================================
    def reset_form(self):
        # Reset job name
        self.jobname_entry.delete(0, "end")

        # Reset date to today
        today = datetime.date.today().strftime("%m-%d-%Y")
        self.date_entry.delete(0, "end")
        self.date_entry.insert(0, today)

        # Uncheck all employees
        for var in self.employee_vars.values():
            var.set(False)

        # Clear all payrate entries
        for entry in self.payrate_entries.values():
            entry.delete(0, "end")

# =========================================================
# Pay Split Calculation Window
# =========================================================
class PaySplitResultsWindow(ctk.CTkToplevel):
    """Standalone window showing detailed pay split results."""

    def __init__(self, master, job_name, date_value, num_workers, results):
        super().__init__(master)

        self.title("Pay Split Results")
        self.geometry("700x600")
        self.resizable(True, True)
        self.grab_set()

        self.job_name = job_name
        self.date_value = date_value
        self.num_workers = num_workers
        self.results = results

        # Compute totals
        self.grand_total = sum(total for _, _, _, total, _ in results)
        self.per_worker = self.grand_total / num_workers

        # =====================================================
        # HEADER
        # =====================================================
        ctk.CTkLabel(
            self,
            text="Pay Split Results",
            font=ctk.CTkFont(size=26, weight="bold")
        ).pack(pady=20)

        # =====================================================
        # JOB SUMMARY
        # =====================================================
        summary_frame = ctk.CTkFrame(self)
        summary_frame.pack(fill="x", padx=20, pady=10)

        summary_items = [
            f"Job Name: {job_name}",
            f"Date: {date_value}",
            f"Employees Selected: {num_workers}",
            f"Total Job Pay: ${self.grand_total:.2f}",
            f"Pay Per Worker: ${self.per_worker:.2f}"
        ]

        for item in summary_items:
            ctk.CTkLabel(
                summary_frame,
                text=item,
                font=ctk.CTkFont(size=16)
            ).pack(anchor="w", pady=2)

        # =====================================================
        # SCROLLABLE BREAKDOWN
        # =====================================================
        detail_frame = ctk.CTkScrollableFrame(self, width=650, height=300)
        detail_frame.pack(padx=20, pady=20, fill="both", expand=True)

        ctk.CTkLabel(
            detail_frame,
            text="Detailed Pay Breakdown",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(pady=10)

        for name, qty, rate, total, split in results:
            row = ctk.CTkFrame(detail_frame)
            row.pack(fill="x", pady=5)

            ctk.CTkLabel(row, text=name, width=180, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(row, text=f"Qty: {qty}").pack(side="left", padx=5)
            ctk.CTkLabel(row, text=f"Rate: ${rate:.2f}").pack(side="left", padx=5)
            ctk.CTkLabel(row, text=f"Total: ${total:.2f}").pack(side="left", padx=5)
            ctk.CTkLabel(row, text=f"Split: ${split:.2f}").pack(side="left", padx=5)

        # =====================================================
        # BUTTONS
        # =====================================================
        btn_frame = ctk.CTkFrame(self)
        btn_frame.pack(pady=20)

        ctk.CTkButton(btn_frame, text="Save to Excel", width=150, command=self.save_to_excel).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="Print Preview", width=150, command=self.print_preview).pack(side="left", padx=10)

    # =========================================================
    # SAVE TO EXCEL
    # =========================================================
    def save_to_excel(self):
        import pandas as pd
        from tkinter import filedialog

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not save_path:
            return

        data = []
        for name, qty, rate, total, split in self.results:
            data.append({
                "Pay Item": name,
                "Quantity": qty,
                "Rate": rate,
                "Total": total,
                "Split Per Worker": split
            })

        df = pd.DataFrame(data)
        df.to_excel(save_path, index=False)

        messagebox.showinfo("Saved", "Pay sheet saved to Excel.")

    # =========================================================
    # PRINT PREVIEW
    # =========================================================
    def print_preview(self):
        preview = ctk.CTkToplevel(self)
        preview.title("Print Preview")
        preview.geometry("600x700")
        preview.grab_set()

        ctk.CTkLabel(
            preview,
            text=f"Pay Sheet - {self.job_name}",
            font=ctk.CTkFont(size=24, weight="bold")
        ).pack(pady=20)

        text = tk.Text(preview, wrap="word", font=("Arial", 12))
        text.pack(fill="both", expand=True, padx=20, pady=20)

        # Build printable text
        printable = f"Job Name: {self.job_name}\n"
        printable += f"Date: {self.date_value}\n"
        printable += f"Employees: {self.num_workers}\n"
        printable += f"Total Job Pay: ${self.grand_total:.2f}\n"
        printable += f"Pay Per Worker: ${self.per_worker:.2f}\n\n"
        printable += "Detailed Breakdown:\n\n"

        for name, qty, rate, total, split in self.results:
            printable += (
                f"{name}\n"
                f"  Qty: {qty}\n"
                f"  Rate: ${rate:.2f}\n"
                f"  Total: ${total:.2f}\n"
                f"  Split: ${split:.2f}\n\n"
            )

        text.insert("1.0", printable)


# =========================================================
# Weekly Payroll Generator Frame
# =========================================================
class WeeklyPayrollFrame(ctk.CTkFrame):
    """Generates a weekly payroll summary from master payroll files."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Weekly Payroll Generator",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title.pack(pady=20)

        # -----------------------------
        # Date Range Inputs
        # -----------------------------
        frame = ctk.CTkFrame(self)
        frame.pack(pady=10)

        ctk.CTkLabel(frame, text="Start Date (MM-DD-YYYY):").grid(
            row=0, column=0, padx=10, pady=5
        )
        self.start_entry = ctk.CTkEntry(frame, width=150)
        self.start_entry.grid(row=0, column=1, padx=10, pady=5)

        ctk.CTkLabel(frame, text="End Date (MM-DD-YYYY):").grid(
            row=1, column=0, padx=10, pady=5
        )
        self.end_entry = ctk.CTkEntry(frame, width=150)
        self.end_entry.grid(row=1, column=1, padx=10, pady=5)

        # -----------------------------
        # Buttons
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Generate Weekly Payroll",
            width=200,
            command=self.generate_weekly_payroll
        ).pack(pady=20)

        ctk.CTkButton(
            self,
            text="Reset",
            width=200,
            command=self.reset_fields
        ).pack(pady=10)

        ctk.CTkButton(
            self,
            text="Back",
            width=200,
            command=lambda: self.master.show_frame("payroll_tools_menu")
        ).pack(pady=10)

    # ---------------------------------------------------------
    # Reset Fields
    # ---------------------------------------------------------
    def reset_fields(self):
        self.start_entry.delete(0, "end")
        self.end_entry.delete(0, "end")

    # =========================================================
    # GENERATE WEEKLY PAYROLL
    # =========================================================
    def generate_weekly_payroll(self):
        # -----------------------------
        # Validate date inputs
        # -----------------------------
        start_text = self.start_entry.get().strip()
        end_text = self.end_entry.get().strip()

        if not start_text or not end_text:
            messagebox.showerror("Missing Dates", "Please enter both start and end dates.")
            return

        try:
            start_date = datetime.datetime.strptime(start_text, "%m-%d-%Y").date()
            end_date = datetime.datetime.strptime(end_text, "%m-%d-%Y").date()
        except ValueError:
            messagebox.showerror("Invalid Date", "Please enter valid dates in MM-DD-YYYY format.")
            return

        if end_date < start_date:
            messagebox.showerror("Invalid Range", "End date must be after start date.")
            return

        # -----------------------------
        # Prepare output directory
        # -----------------------------
        save_dir = netpath("payroll_records", "weekly payroll")
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        output_path = os.path.join(
            save_dir,
            f"WeeklyPayroll_{start_text}_to_{end_text}.xlsx"
        )

        if os.path.exists(output_path):
            messagebox.showerror(
                "File Already Exists",
                f"A weekly payroll file for this date range already exists:\n{output_path}"
            )
            return

        # -----------------------------
        # Load template
        # -----------------------------
        workbook = load_workbook(netpath("data", "spreadsheet", "WeeklyPayrollTemplate.xlsx"))
        sheet = workbook.active

        sheet["B2"] = start_text
        sheet["B3"] = end_text

        # Starting rows for each section
        emp_row = item_row = job_row = day_row = 5

        global_pay_item_totals = {}
        global_job_totals = {}
        global_day_totals = {}

        # =====================================================
        # PROCESS EACH EMPLOYEE
        # =====================================================
        for emp in employees:
            filename = f"{emp.fullname}.xlsx"
            path = os.path.join(MASTER_PAYROLL_DIR, filename)

            if not os.path.exists(path):
                continue

            emp_book = load_workbook(path, data_only=True)
            emp_sheet = emp_book.active

            total_pay = 0

            # -----------------------------
            # Loop through employee rows
            # -----------------------------
            for r in range(2, emp_sheet.max_row + 1):
                date_val = emp_sheet[f"A{r}"].value

                # Normalize date formats
                if isinstance(date_val, datetime.datetime):
                    date_val = date_val.date()

                if isinstance(date_val, str):
                    for fmt in ("%m-%d-%Y", "%Y-%m-%d"):
                        try:
                            date_val = datetime.datetime.strptime(date_val, fmt).date()
                            break
                        except ValueError:
                            pass

                if not isinstance(date_val, datetime.date):
                    continue

                if not (start_date <= date_val <= end_date):
                    continue

                split = emp_sheet[f"G{r}"].value
                if split is None:
                    continue

                try:
                    split = float(split)
                except ValueError:
                    continue

                total_pay += split

                pay_item = emp_sheet[f"C{r}"].value
                job_name = emp_sheet[f"B{r}"].value

                # Totals per pay item
                if pay_item:
                    global_pay_item_totals[pay_item] = (
                        global_pay_item_totals.get(pay_item, 0) + split
                    )

                # Totals per job
                if job_name:
                    global_job_totals[job_name] = (
                        global_job_totals.get(job_name, 0) + split
                    )

                # Totals per day
                global_day_totals[date_val] = (
                    global_day_totals.get(date_val, 0) + split
                )

            # Write employee total
            sheet[f"A{emp_row}"] = emp.fullname
            sheet[f"B{emp_row}"] = total_pay
            emp_row += 1

            emp_book.close()

        # =====================================================
        # WRITE SUMMARY SECTIONS
        # =====================================================

        # Totals per pay item
        sheet["D4"] = "Totals Per Pay Item"
        for item, total in sorted(global_pay_item_totals.items()):
            sheet[f"D{item_row}"] = item
            sheet[f"E{item_row}"] = total
            item_row += 1

        # Totals per job
        sheet["G4"] = "Totals Per Job"
        for job, total in sorted(global_job_totals.items()):
            sheet[f"G{job_row}"] = job
            sheet[f"H{job_row}"] = total
            job_row += 1

        # Totals per day
        sheet["J4"] = "Totals Per Day"
        for day, total in sorted(global_day_totals.items()):
            sheet[f"J{day_row}"] = day.strftime("%m-%d-%Y")
            sheet[f"K{day_row}"] = total
            day_row += 1

        # Weekly total
        weekly_total = sum(global_day_totals.values())
        sheet["D2"] = "Weekly Total:"
        sheet["E2"] = weekly_total

        # Save file
        workbook.save(output_path)
        workbook.close()

        messagebox.showinfo("Weekly Payroll Generated", f"Saved to:\n{output_path}")

# =========================================================
# View Weekly Payroll Frame
# =========================================================
class ViewWeeklyPayrollFrame(ctk.CTkFrame):
    """Displays a selected weekly payroll Excel file in a clean, scrollable layout."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="View Weekly Payroll",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title.pack(pady=20)

        # -----------------------------
        # File Selection
        # -----------------------------
        file_frame = ctk.CTkFrame(self)
        file_frame.pack(pady=10)

        ctk.CTkLabel(
            file_frame,
            text="Select Weekly Payroll File:"
        ).grid(row=0, column=0, padx=10)

        self.file_entry = ctk.CTkEntry(file_frame, width=350)
        self.file_entry.grid(row=0, column=1, padx=10)

        ctk.CTkButton(
            file_frame,
            text="Browse",
            command=self.browse_file
        ).grid(row=0, column=2, padx=10)

        # Load + Back Buttons
        ctk.CTkButton(
            self,
            text="Load Payroll",
            width=200,
            command=self.load_payroll
        ).pack(pady=10)

        ctk.CTkButton(
            self,
            text="Back",
            width=200,
            command=self.go_back
        ).pack(pady=10)

        # -----------------------------
        # Scrollable Display Area
        # -----------------------------
        self.scroll = ctk.CTkScrollableFrame(self, width=900, height=500)
        self.scroll.pack(pady=20)

    # ---------------------------------------------------------
    # Browse for file
    # ---------------------------------------------------------
    def browse_file(self):
        from tkinter import filedialog

        path = filedialog.askopenfilename(
            initialdir=netpath("payroll_records", "weekly payroll"),
            filetypes=[("Excel Files", "*.xlsx")]
        )

        if path:
            self.file_entry.delete(0, "end")
            self.file_entry.insert(0, path)

    # ---------------------------------------------------------
    # Load and display payroll file
    # ---------------------------------------------------------
    def load_payroll(self):
        path = self.file_entry.get().strip()

        if not os.path.exists(path):
            messagebox.showerror("Error", "File not found.")
            return

        # Clear previous content
        for widget in self.scroll.winfo_children():
            widget.destroy()

        book = load_workbook(path, data_only=True)
        sheet = book.active

        # -----------------------------
        # Weekly Summary
        # -----------------------------
        summary = ctk.CTkLabel(
            self.scroll,
            text=f"Weekly Total: {sheet['E2'].value}",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        summary.pack(pady=10)

        # -----------------------------
        # Sections to Display
        # -----------------------------
        sections = [
            ("Employee Totals",      "A", "B"),
            ("Totals Per Pay Item",  "D", "E"),
            ("Totals Per Job",       "G", "H"),
            ("Totals Per Day",       "J", "K")
        ]

        for title, col1, col2 in sections:

            # Section Title
            ctk.CTkLabel(
                self.scroll,
                text=title,
                font=ctk.CTkFont(size=16, weight="bold")
            ).pack(pady=10)

            table = ctk.CTkFrame(self.scroll)
            table.pack(pady=5)

            # Header Row
            ctk.CTkLabel(
                table,
                text="Name",
                font=ctk.CTkFont(weight="bold")
            ).grid(row=0, column=0, padx=10)

            ctk.CTkLabel(
                table,
                text="Amount",
                font=ctk.CTkFont(weight="bold")
            ).grid(row=0, column=1, padx=10)

            # Read rows
            rows = []
            row = 5

            while True:
                v1 = sheet[f"{col1}{row}"].value
                v2 = sheet[f"{col2}{row}"].value

                if v1 is None and v2 is None:
                    break

                rows.append((v1, v2))
                row += 1

            # No data case
            if not rows:
                ctk.CTkLabel(
                    table,
                    text="No data available"
                ).grid(row=1, column=0, columnspan=2)
                continue

            # Display rows
            for i, (v1, v2) in enumerate(sorted(rows), start=1):
                ctk.CTkLabel(
                    table,
                    text=str(v1),
                    width=200,
                    anchor="w"
                ).grid(row=i, column=0, padx=10)

                ctk.CTkLabel(
                    table,
                    text=str(v2),
                    width=200,
                    anchor="w"
                ).grid(row=i, column=1, padx=10)

        book.close()

    # ---------------------------------------------------------
    # Reset view
    # ---------------------------------------------------------
    def reset_view(self):
        self.file_entry.delete(0, "end")
        for widget in self.scroll.winfo_children():
            widget.destroy()

    # ---------------------------------------------------------
    # Back navigation
    # ---------------------------------------------------------
    def go_back(self):
        self.reset_view()
        self.master.show_frame("payroll_tools_menu")

# =========================================================
# YTD Payroll Frame
# =========================================================
class YTDPayrollFrame(ctk.CTkFrame):
    """Generates a Year-To-Date payroll summary with optional filters."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Year-To-Date Payroll Summary",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title.pack(pady=20)

        # -----------------------------
        # Form Container
        # -----------------------------
        form = ctk.CTkFrame(self)
        form.pack(pady=20)

        # Year
        ctk.CTkLabel(form, text="Year:").grid(
            row=0, column=0, padx=10, pady=5
        )
        self.year_entry = ctk.CTkEntry(form, width=150)
        self.year_entry.grid(row=0, column=1, padx=10, pady=5)

        # Employee Filter
        ctk.CTkLabel(form, text="Employee (optional):").grid(
            row=1, column=0, padx=10, pady=5
        )
        self.employee_option = ctk.CTkComboBox(
            form,
            values=["All"] + [emp.fullname for emp in employees],
            width=200
        )
        self.employee_option.grid(row=1, column=1, padx=10, pady=5)
        self.employee_option.set("All")

        # Job Filter Options
        jobs, payitems = self.collect_job_and_payitems()

        ctk.CTkLabel(form, text="Job (optional):").grid(
            row=2, column=0, padx=10, pady=5
        )
        self.job_option = ctk.CTkComboBox(
            form,
            values=["All"] + jobs,
            width=200
        )
        self.job_option.grid(row=2, column=1, padx=10, pady=5)
        self.job_option.set("All")

        # -----------------------------
        # Generate Button
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Generate YTD Payroll Summary",
            width=250,
            command=self.generate_ytd_summary
        ).pack(pady=20)

        # Back Button
        ctk.CTkButton(
            self,
            text="Back",
            width=200,
            command=lambda: self.master.show_frame("payroll_tools_menu")
        ).pack(pady=10)

    # ---------------------------------------------------------
    # Collect all job names and pay items from master files
    # ---------------------------------------------------------
    @staticmethod
    def collect_job_and_payitems():
        jobs = set()
        payitems = set()

        for emp in employees:
            filename = f"{emp.fullname}.xlsx"
            path = os.path.join(MASTER_PAYROLL_DIR, filename)

            if not os.path.exists(path):
                continue

            book = load_workbook(path)
            sheet = book.active

            for r in range(2, sheet.max_row + 1):
                job = sheet[f"B{r}"].value
                payitem = sheet[f"C{r}"].value

                if job:
                    jobs.add(str(job).strip())

                if payitem:
                    payitems.add(str(payitem).strip())

            book.close()

        return sorted(jobs), sorted(payitems)

    # =========================================================
    # Generate YTD Summary
    # =========================================================
    def generate_ytd_summary(self):
        year_text = self.year_entry.get().strip()
        employee_filter = self.employee_option.get()
        job_filter = self.job_option.get()

        # Normalize job filter
        if job_filter == "All":
            job_filter = ""

        # Validate year
        if not year_text.isdigit():
            messagebox.showerror(
                "Invalid Year",
                "Please enter a valid year (e.g., 2025)."
            )
            return

        year = int(year_text)

        # -----------------------------
        # Prepare output directory
        # -----------------------------
        save_dir = netpath("payroll_records", "ytd payroll")
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        output_path = os.path.join(save_dir, f"YTD_{year}.xlsx")

        # -----------------------------
        # Load template
        # -----------------------------
        workbook = load_workbook(
            netpath("data", "spreadsheet", "YTDPayrollTemplate.xlsx")
        )
        sheet = workbook.active

        sheet["A3"] = "Year:"
        sheet["B3"] = year

        # Summary dictionaries
        employee_totals = {}
        pay_item_totals = {}
        job_totals = {}
        month_totals = {}

        # =====================================================
        # Loop through employees
        # =====================================================
        for emp in employees:

            # Employee filter
            if employee_filter != "All" and emp.fullname != employee_filter:
                continue

            filename = f"{emp.fullname}.xlsx"
            path = os.path.join(MASTER_PAYROLL_DIR, filename)

            if not os.path.exists(path):
                continue

            emp_book = load_workbook(path)
            emp_sheet = emp_book.active

            # -----------------------------
            # Loop through employee rows
            # -----------------------------
            for r in range(2, emp_sheet.max_row + 1):

                date_val = emp_sheet[f"A{r}"].value

                # Normalize date formats
                if isinstance(date_val, datetime.datetime):
                    date_val = date_val.date()

                if isinstance(date_val, str):
                    try:
                        date_val = datetime.datetime.strptime(
                            date_val, "%m-%d-%Y"
                        ).date()
                    except ValueError:
                        continue

                if not isinstance(date_val, datetime.date):
                    continue

                if date_val.year != year:
                    continue

                # Filters
                job_name = str(emp_sheet[f"B{r}"].value).lower()
                pay_item = str(emp_sheet[f"C{r}"].value).lower()

                if job_filter and job_filter not in job_name:
                    continue

                split = emp_sheet[f"G{r}"].value
                if not split:
                    continue

                try:
                    split = float(split)
                except ValueError, TypeError:
                    continue

                # Employee totals
                employee_totals[emp.fullname] = (
                    employee_totals.get(emp.fullname, 0) + split
                )

                # Pay item totals
                if pay_item:
                    pay_item_totals[pay_item] = (
                        pay_item_totals.get(pay_item, 0) + split
                    )

                # Job totals
                if job_name:
                    job_totals[job_name] = (
                        job_totals.get(job_name, 0) + split
                    )

                # Month totals
                month = date_val.strftime("%B")
                month_totals[month] = month_totals.get(month, 0) + split

            emp_book.close()

        # =====================================================
        # Write results to Excel
        # =====================================================
        row = 5

        # Employee totals
        sheet["A4"] = "Employee Totals"
        for name, total in employee_totals.items():
            sheet[f"A{row}"] = name
            sheet[f"B{row}"] = total
            row += 1

        # Pay item totals
        row = 5
        sheet["D4"] = "Pay Item Totals"
        for item, total in pay_item_totals.items():
            sheet[f"D{row}"] = item
            sheet[f"E{row}"] = total
            row += 1

        # Job totals
        row = 5
        sheet["G4"] = "Job Totals"
        for job, total in job_totals.items():
            sheet[f"G{row}"] = job
            sheet[f"H{row}"] = total
            row += 1

        # Month totals
        row = 5
        sheet["J4"] = "Month Totals"
        for month, total in month_totals.items():
            sheet[f"J{row}"] = month
            sheet[f"K{row}"] = total
            row += 1

        # Grand total
        grand_total = sum(employee_totals.values())
        sheet["A2"] = "Grand Total:"
        sheet["B2"] = grand_total

        # Save file
        workbook.save(output_path)
        workbook.close()

        messagebox.showinfo(
            "YTD Payroll Summary Generated",
            f"Saved to:\n{output_path}"
        )

# =========================================================
# View YTD Payroll Summary Frame
# =========================================================
class ViewYTDSummaryFrame(ctk.CTkFrame):
    """Displays YTD payroll summary with section selector."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Top-Left Back Button
        # -----------------------------
        ctk.CTkButton(
            self,
            text="← Back",
            width=120,
            command=lambda: self.master.show_frame("payroll_tools_menu")
        ).place(x=20, y=20)

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="YTD Payroll Summary Viewer",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title.pack(pady=60)

        # -----------------------------
        # Filter Controls
        # -----------------------------
        filter_frame = ctk.CTkFrame(self)
        filter_frame.pack(pady=10)

        # Year
        ctk.CTkLabel(filter_frame, text="Year:").grid(
            row=0, column=0, padx=10, pady=10
        )
        self.year_entry = ctk.CTkEntry(filter_frame, width=150)
        self.year_entry.grid(row=0, column=1, padx=10, pady=10)

        # Section Selector
        ctk.CTkLabel(filter_frame, text="Section:").grid(
            row=1, column=0, padx=10, pady=10
        )
        self.section_option = ctk.CTkOptionMenu(
            filter_frame,
            values=[
                "Employee Totals",
                "Pay Item Totals",
                "Job Totals",
                "Month Totals"
            ]
        )
        self.section_option.grid(row=1, column=1, padx=10, pady=10)

        # View Button
        ctk.CTkButton(
            filter_frame,
            text="Load Summary",
            width=200,
            command=self.load_ytd_summary
        ).grid(row=2, column=0, columnspan=2, pady=20)

        # -----------------------------
        # Scrollable Table
        # -----------------------------
        self.table_frame = ctk.CTkScrollableFrame(self, width=900, height=400)
        self.table_frame.pack(pady=20)

        # -----------------------------
        # Bottom Navigation
        # -----------------------------
        nav = ctk.CTkFrame(self)
        nav.pack(pady=10)

        ctk.CTkButton(
            nav,
            text="Back to Payroll Tools",
            width=200,
            command=lambda: self.master.show_frame("payroll_tools_menu")
        ).grid(row=0, column=0, padx=10)

        ctk.CTkButton(
            nav,
            text="Back to Main Menu",
            width=200,
            command=lambda: self.master.show_frame("main_admin")
        ).grid(row=0, column=1, padx=10)

    # ---------------------------------------------------------
    # Load YTD summary into table
    # ---------------------------------------------------------
    def load_ytd_summary(self):
        # Clear old rows
        for widget in self.table_frame.winfo_children():
            widget.destroy()

        year_text = self.year_entry.get().strip()

        # Validate year
        if not year_text.isdigit():
            ctk.CTkLabel(
                self.table_frame,
                text="Invalid year. Enter a number like 2025.",
                font=ctk.CTkFont(size=16)
            ).pack(pady=20)
            return

        year = int(year_text)

        # Load YTD file
        save_dir = netpath("payroll_records", "ytd payroll")
        ytd_path = os.path.join(save_dir, f"YTD_{year}.xlsx")

        if not os.path.exists(ytd_path):
            ctk.CTkLabel(
                self.table_frame,
                text="No YTD summary found for this year.",
                font=ctk.CTkFont(size=16)
            ).pack(pady=20)
            return

        df = pd.read_excel(ytd_path)

        # Determine which section to display
        section = self.section_option.get()

        if section == "Employee Totals":
            display_df = df.iloc[:, 0:2]
            display_df.columns = ["Employee", "Total"]

        elif section == "Pay Item Totals":
            display_df = df.iloc[:, 3:5]
            display_df.columns = ["Pay Item", "Total"]

        elif section == "Job Totals":
            display_df = df.iloc[:, 6:8]
            display_df.columns = ["Job", "Total"]

        elif section == "Month Totals":
            display_df = df.iloc[:, 9:11]
            display_df.columns = ["Month", "Total"]

        else:
            ctk.CTkLabel(
                self.table_frame,
                text="Unknown section.",
                font=ctk.CTkFont(size=16)
            ).pack(pady=20)
            return

        # -----------------------------
        # FIX: Remove empty rows (NaN rows)
        # -----------------------------
        display_df = display_df.dropna(how="all")

        if display_df.empty:
            ctk.CTkLabel(
                self.table_frame,
                text="No data found for this section.",
                font=ctk.CTkFont(size=16)
            ).pack(pady=20)
            return

        # -----------------------------
        # Table Headers
        # -----------------------------
        for col, text in enumerate(display_df.columns):
            ctk.CTkLabel(
                self.table_frame,
                text=text,
                font=ctk.CTkFont(weight="bold")
            ).grid(row=0, column=col, padx=10, pady=5)

        # -----------------------------
        # Table Rows
        # -----------------------------
        for row_index, row in enumerate(display_df.itertuples(index=False), start=1):
            for col_index, value in enumerate(row):
                ctk.CTkLabel(
                    self.table_frame,
                    text=str(value)
                ).grid(row=row_index, column=col_index, padx=10, pady=5)

# =========================================================
# Vacation Tools Menu Frame
# =========================================================
class VacationToolsMenuFrame(ctk.CTkFrame):
    """Menu screen for all vacation-related tools."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Vacation Tools",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title.pack(pady=40)

        # -----------------------------
        # Button Container
        # -----------------------------
        button_frame = ctk.CTkFrame(self)
        button_frame.pack(pady=20)

        # Add Vacation Record
        ctk.CTkButton(
            button_frame,
            text="Add Vacation Record",
            width=250,
            command=lambda: self.master.show_frame("vacation_tool")
        ).pack(pady=10)

        # View Vacation Record
        ctk.CTkButton(
            button_frame,
            text="View Vacation Record",
            width=250,
            command=lambda: self.master.show_frame("view_vacation")
        ).pack(pady=10)

        # Approve Vacation
        ctk.CTkButton(
            button_frame,
            text="Approve Vacation",
            width=250,
            command=lambda: self.master.show_frame("vacation_approval")
        ).pack(pady=10)

        # -----------------------------
        # Back Button
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Back",
            width=220,
            command=lambda: self.master.show_frame("main_admin")
        ).pack(pady=40)

    # ---------------------------------------------------------
    # Frame shown hook
    # ---------------------------------------------------------
    def on_show(self):
        pass  # Reserved for future dynamic updates

# =========================================================
# Vacation Payroll Frame
# =========================================================
class VacationPayrollFrame(ctk.CTkFrame):
    """Admin tool for processing vacation pay and deducting vacation days."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        self.calculated_pay = None
        self.calculated_days = None

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Vacation Payroll Tool",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title.pack(pady=30)

        # -----------------------------
        # Employee Selection
        # -----------------------------
        form = ctk.CTkFrame(self)
        form.pack(pady=20)

        ctk.CTkLabel(form, text="Select Employee:").grid(
            row=0, column=0, padx=10, pady=10, sticky="e"
        )

        self.emp_option = ctk.CTkOptionMenu(
            form,
            values=[f"{e.fullname} ({e.username})" for e in employees],
            command=self.update_employee_info
        )
        self.emp_option.grid(row=0, column=1, padx=10, pady=10)

        # Vacation Max
        ctk.CTkLabel(form, text="Vacation Max Days:").grid(
            row=1, column=0, padx=10, pady=10, sticky="e"
        )
        self.vac_max_label = ctk.CTkLabel(form, text="")
        self.vac_max_label.grid(row=1, column=1, padx=10, pady=10, sticky="w")

        # Vacation Remaining
        ctk.CTkLabel(form, text="Vacation Remaining:").grid(
            row=2, column=0, padx=10, pady=10, sticky="e"
        )
        self.vac_remaining_label = ctk.CTkLabel(form, text="")
        self.vac_remaining_label.grid(row=2, column=1, padx=10, pady=10, sticky="w")

        # -----------------------------
        # Vacation Input Fields
        # -----------------------------
        ctk.CTkLabel(form, text="Days Used:").grid(
            row=3, column=0, padx=10, pady=10, sticky="e"
        )
        self.days_entry = ctk.CTkEntry(form, width=200)
        self.days_entry.grid(row=3, column=1, padx=10, pady=10)

        ctk.CTkLabel(form, text="Vacation Pay Rate ($/day):").grid(
            row=4, column=0, padx=10, pady=10, sticky="e"
        )
        self.rate_entry = ctk.CTkEntry(form, width=200)
        self.rate_entry.grid(row=4, column=1, padx=10, pady=10)

        # -----------------------------
        # Buttons
        # -----------------------------
        btn_frame = ctk.CTkFrame(self)
        btn_frame.pack(pady=20)

        ctk.CTkButton(
            btn_frame,
            text="Calculate Vacation Pay",
            width=220,
            command=self.calculate_vacation_pay
        ).grid(row=0, column=0, padx=10)

        ctk.CTkButton(
            btn_frame,
            text="Save Vacation Record",
            width=220,
            command=self.save_vacation_record
        ).grid(row=0, column=1, padx=10)

        ctk.CTkButton(
            btn_frame,
            text="Back",
            width=220,
            command=lambda: self.master.show_frame("vacation_tool_menu")
        ).grid(row=0, column=2, padx=10)

        # -----------------------------
        # Result Label
        # -----------------------------
        self.result_label = ctk.CTkLabel(
            self,
            text="",
            font=ctk.CTkFont(size=18)
        )
        self.result_label.pack(pady=20)

        # Track selected employee
        self.selected_employee = None

    # ---------------------------------------------------------
    # Update employee info when selected
    # ---------------------------------------------------------
    def update_employee_info(self, selection):
        username = selection.split("(")[-1].replace(")", "").strip()
        self.selected_employee = next(
            e for e in employees if e.username == username
        )

        self.vac_max_label.configure(
            text=str(self.selected_employee.vacation_max)
        )
        self.vac_remaining_label.configure(
            text=str(self.selected_employee.vacation_remaining)
        )

    # ---------------------------------------------------------
    # Calculate vacation pay
    # ---------------------------------------------------------
    def calculate_vacation_pay(self):
        if not self.selected_employee:
            self.result_label.configure(
                text="Select an employee first.",
                text_color="red"
            )
            return

        try:
            days_used = float(self.days_entry.get().strip())
            rate = float(self.rate_entry.get().strip())
        except (ValueError, TypeError):
            self.result_label.configure(
                text="Invalid input.",
                text_color="red"
            )
            return

        if days_used > self.selected_employee.vacation_remaining:
            self.result_label.configure(
                text="Not enough vacation days remaining.",
                text_color="red"
            )
            return

        total_pay = days_used * rate

        self.result_label.configure(
            text=f"Vacation Pay: ${total_pay:,.2f}",
            text_color="white"
        )

        self.calculated_pay = total_pay
        self.calculated_days = days_used

    # ---------------------------------------------------------
    # Save vacation record + update employee balance
    # ---------------------------------------------------------
    def save_vacation_record(self):
        if not hasattr(self, "calculated_pay"):
            self.result_label.configure(
                text="Calculate pay first.",
                text_color="red"
            )
            return

        emp = self.selected_employee

        # Deduct days
        emp.vacation_remaining -= self.calculated_days

        # Save employees.xlsx
        save_employees_to_excel()

        # Append to vacation_log.xlsx
        self.append_vacation_log(emp)

        self.result_label.configure(
            text="Vacation record saved successfully.",
            text_color="green"
        )

    # ---------------------------------------------------------
    # Append to vacation_log.xlsx
    # ---------------------------------------------------------
    def append_vacation_log(self, emp):

        df = pd.read_excel(VACATION_LOG)

        new_row = {
            "date": datetime.date.today().strftime("%Y-%m-%d"),
            "username": emp.username,
            "fullname": emp.fullname,
            "days_used": self.calculated_days,
            "rate": float(self.rate_entry.get().strip()),
            "total_pay": self.calculated_pay,
            "remaining_days": emp.vacation_remaining
        }

        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        df.to_excel(VACATION_LOG, index=False)

# =========================================================
# Vacation History Viewer
# =========================================================
class VacationHistoryFrame(ctk.CTkFrame):
    """View all vacation log entries for a selected employee."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Top-Left Back Button
        # -----------------------------
        ctk.CTkButton(
            self,
            text="← Back",
            width=120,
            command=lambda: self.master.show_frame("vacation_tool_menu")
        ).place(x=20, y=20)

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Vacation History Viewer",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title.pack(pady=60)  # pushed down to avoid overlap with top-left button

        # -----------------------------
        # Employee Selector
        # -----------------------------
        selector_frame = ctk.CTkFrame(self)
        selector_frame.pack(pady=10)

        ctk.CTkLabel(
            selector_frame,
            text="Select Employee:"
        ).grid(row=0, column=0, padx=10, pady=10)

        self.emp_option = ctk.CTkOptionMenu(
            selector_frame,
            values=[f"{e.fullname} ({e.username})" for e in employees],
            command=self.load_history
        )
        self.emp_option.grid(row=0, column=1, padx=10, pady=10)

        # -----------------------------
        # Scrollable Table
        # -----------------------------
        self.table_frame = ctk.CTkScrollableFrame(self, width=900, height=400)
        self.table_frame.pack(pady=20)

        # -----------------------------
        # Bottom Navigation Buttons
        # -----------------------------
        nav_frame = ctk.CTkFrame(self)
        nav_frame.pack(pady=10)


    # ---------------------------------------------------------
    # Load vacation history for selected employee
    # ---------------------------------------------------------
    def load_history(self, selection):
        # Clear old table rows
        for widget in self.table_frame.winfo_children():
            widget.destroy()

        # Extract username from "(username)"
        username = selection.split("(")[-1].replace(")", "").strip()

        # Load vacation log
        try:
            df = pd.read_excel(VACATION_LOG)
        except Exception as e:
            ctk.CTkLabel(
                self.table_frame,
                text=f"Error loading log: {e}"
            ).pack()
            return

        # Filter rows for this employee
        emp_rows = df[df["username"] == username]

        # If no history
        if emp_rows.empty:
            ctk.CTkLabel(
                self.table_frame,
                text="No vacation history found.",
                font=ctk.CTkFont(size=16)
            ).pack(pady=20)
            return

        # -----------------------------
        # Table Headers
        # -----------------------------
        headers = ["Date", "Days Used", "Rate", "Total Pay", "Remaining Days"]

        for col, text in enumerate(headers):
            ctk.CTkLabel(
                self.table_frame,
                text=text,
                font=ctk.CTkFont(weight="bold")
            ).grid(row=0, column=col, padx=10, pady=5)

        # -----------------------------
        # Table Rows
        # -----------------------------
        for row_index, (_, row) in enumerate(emp_rows.iterrows(), start=1):
            ctk.CTkLabel(
                self.table_frame,
                text=row["date"]
            ).grid(row=row_index, column=0, padx=10, pady=5)

            ctk.CTkLabel(
                self.table_frame,
                text=row["days_used"]
            ).grid(row=row_index, column=1, padx=10, pady=5)

            ctk.CTkLabel(
                self.table_frame,
                text=row["rate"]
            ).grid(row=row_index, column=2, padx=10, pady=5)

            ctk.CTkLabel(
                self.table_frame,
                text=row["total_pay"]
            ).grid(row=row_index, column=3, padx=10, pady=5)

            ctk.CTkLabel(
                self.table_frame,
                text=row["remaining_days"]
            ).grid(row=row_index, column=4, padx=10, pady=5)

# =========================================================
# Vacation Request Approval Frame
# =========================================================
class VacationRequestApprovalFrame(ctk.CTkFrame):
    """Admin screen to approve or deny employee vacation requests."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Top-Left Back Button
        # -----------------------------
        ctk.CTkButton(
            self,
            text="← Back",
            width=120,
            command=lambda: self.master.show_frame("vacation_tool_menu")
        ).place(x=20, y=20)

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Vacation Request Approval",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title.pack(pady=60)

        # -----------------------------
        # Scrollable Table
        # -----------------------------
        self.table_frame = ctk.CTkScrollableFrame(self, width=900, height=450)
        self.table_frame.pack(pady=20)

        # -----------------------------
        # Bottom Navigation
        # -----------------------------
        nav = ctk.CTkFrame(self)
        nav.pack(pady=10)


    # ---------------------------------------------------------
    # Refresh table when shown
    # ---------------------------------------------------------
    def on_show(self):
        self.load_requests()

    # ---------------------------------------------------------
    # Load pending requests
    # ---------------------------------------------------------
    def load_requests(self):
        # Clear old rows
        for widget in self.table_frame.winfo_children():
            widget.destroy()

        req_file = netpath(REQUEST_FILE)

        if not os.path.exists(req_file):
            ctk.CTkLabel(self.table_frame, text="No pending requests.").pack(pady=20)
            return

        df = pd.read_excel(req_file)
        pending = df[df["status"] == "Pending"]

        if pending.empty:
            ctk.CTkLabel(self.table_frame, text="No pending requests.").pack(pady=20)
            return

        # -----------------------------
        # Table Headers
        # -----------------------------
        headers = ["Date", "Employee", "Days", "Approve", "Deny"]
        for col, text in enumerate(headers):
            ctk.CTkLabel(
                self.table_frame,
                text=text,
                font=ctk.CTkFont(weight="bold")
            ).grid(row=0, column=col, padx=10, pady=5)

        # -----------------------------
        # Table Rows
        # -----------------------------
        for row_index, (idx, row) in enumerate(pending.iterrows(), start=1):

            ctk.CTkLabel(
                self.table_frame,
                text=row["date_requested"]
            ).grid(row=row_index, column=0, padx=10, pady=5)

            ctk.CTkLabel(
                self.table_frame,
                text=row["fullname"]
            ).grid(row=row_index, column=1, padx=10, pady=5)

            ctk.CTkLabel(
                self.table_frame,
                text=row["days_requested"]
            ).grid(row=row_index, column=2, padx=10, pady=5)

            # Approve button
            ctk.CTkButton(
                self.table_frame,
                text="Approve",
                width=100,
                fg_color="green",
                command=lambda i=idx: self.approve_request(i)
            ).grid(row=row_index, column=3, padx=10)

            # Deny button
            ctk.CTkButton(
                self.table_frame,
                text="Deny",
                width=100,
                fg_color="red",
                hover_color="#8b0000",
                command=lambda i=idx: self.deny_request(i)
            ).grid(row=row_index, column=4, padx=10)

    # ---------------------------------------------------------
    # Approve request
    # ---------------------------------------------------------
    def approve_request(self, request_index):
        req_file = netpath(REQUEST_FILE)
        df = pd.read_excel(req_file)

        row = df.loc[request_index]
        username = row["username"]
        days = float(row["days_requested"])

        # Find employee
        emp = next(e for e in employees if e.username == username)

        # Deduct days
        emp.vacation_remaining -= days
        save_employees_to_excel()

        # Log to vacation_log.xlsx
        self.append_vacation_log(emp, days)

        # Update request status
        df.at[request_index, "status"] = "Approved"
        df.to_excel(req_file, index=False)

        self.load_requests()

    # ---------------------------------------------------------
    # Deny request
    # ---------------------------------------------------------
    def deny_request(self, request_index):
        req_file = netpath(REQUEST_FILE)
        df = pd.read_excel(req_file)

        # Ask for admin comment
        comment = simpledialog.askstring("Deny Request", "Enter reason for denial:")

        df.at[request_index, "status"] = "Denied"
        df.at[request_index, "admin_comment"] = comment if comment else "No comment"
        df.to_excel(req_file, index=False)

        self.load_requests()

    # ---------------------------------------------------------
    # Append approved vacation to log
    # ---------------------------------------------------------
    @staticmethod
    def append_vacation_log(emp, days_used):

        df = pd.read_excel(VACATION_LOG)

        new_row = {
            "date": datetime.date.today().strftime("%Y-%m-%d"),
            "username": emp.username,
            "fullname": emp.fullname,
            "days_used": days_used,
            "rate": 0,          # Admin approval does not include pay rate
            "total_pay": 0,
            "remaining_days": emp.vacation_remaining
        }

        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        df.to_excel(VACATION_LOG, index=False)

# =========================================================
# Main Employee Window
# =========================================================
class MainEmployeeFrame(ctk.CTkFrame):
    """Main dashboard for regular employee users."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Title / Welcome
        # -----------------------------
        self.title_label = ctk.CTkLabel(
            self,
            text="Employee Dashboard",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        self.title_label.pack(pady=40)

        # Dynamic welcome message
        self.welcome_label = ctk.CTkLabel(
            self,
            text="",
            font=ctk.CTkFont(size=18)
        )
        self.welcome_label.pack(pady=10)

        # -----------------------------
        # Button Container
        # -----------------------------
        button_frame = ctk.CTkFrame(self)
        button_frame.pack(pady=30)

        # My Information
        ctk.CTkButton(
            button_frame,
            text="My Information",
            width=250,
            command=lambda: self.master.show_frame("employee_info")
        ).pack(pady=10)

        # Request Vacation
        ctk.CTkButton(
            button_frame,
            text="Request Vacation",
            width=250,
            command=lambda: self.master.show_frame("request_vacation")
        ).pack(pady=10)

        # View Weekly Pay button
        ctk.CTkButton(
            button_frame,
            text="View Weekly Pay",
            width=250,
            command=lambda: self.master.show_frame("view_employee_payroll")
        ).pack(pady=10)

        # -----------------------------
        # Logout Button
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Logout",
            fg_color="red",
            hover_color="#8b0000",
            width=200,
            command=self.logout
        ).pack(pady=40)

    # ---------------------------------------------------------
    # Update welcome message when frame is shown
    # ---------------------------------------------------------
    def on_show(self):
        user = self.master.current_user
        self.welcome_label.configure(text=f"Welcome, {user.fullname}")

    # ---------------------------------------------------------
    # Logout Handler
    # ---------------------------------------------------------
    def logout(self):
        self.master.current_user = None
        self.master.show_frame("login")

# =========================================================
# Employee Info Screen (Employee Side)
# =========================================================
class EmployeeInfoFrame(ctk.CTkFrame):
    """Displays personal information for the logged-in employee."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # Track password visibility
        self.password_visible = False

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="My Information",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title.pack(pady=30)

        # -----------------------------
        # Info Container
        # -----------------------------
        self.info_frame = ctk.CTkFrame(self)
        self.info_frame.pack(pady=20)

        # Static labels
        ctk.CTkLabel(self.info_frame, text="Full Name:").grid(
            row=0, column=0, padx=10, pady=10, sticky="e"
        )
        ctk.CTkLabel(self.info_frame, text="Username:").grid(
            row=1, column=0, padx=10, pady=10, sticky="e"
        )
        ctk.CTkLabel(self.info_frame, text="Role:").grid(
            row=2, column=0, padx=10, pady=10, sticky="e"
        )
        ctk.CTkLabel(self.info_frame, text="Vacation Days Earned:").grid(
            row=3, column=0, padx=10, pady=10, sticky="e"
        )
        ctk.CTkLabel(self.info_frame, text="Password:").grid(
            row=4, column=0, padx=10, pady=10, sticky="e"
        )

        # Dynamic labels
        self.fullname_label = ctk.CTkLabel(self.info_frame, text="")
        self.fullname_label.grid(row=0, column=1, padx=10, pady=10, sticky="w")

        self.username_label = ctk.CTkLabel(self.info_frame, text="")
        self.username_label.grid(row=1, column=1, padx=10, pady=10, sticky="w")

        self.role_label = ctk.CTkLabel(self.info_frame, text="")
        self.role_label.grid(row=2, column=1, padx=10, pady=10, sticky="w")

        self.vacation_label = ctk.CTkLabel(self.info_frame, text="")
        self.vacation_label.grid(row=3, column=1, padx=10, pady=10, sticky="w")

        # Masked password label
        self.password_label = ctk.CTkLabel(self.info_frame, text="")
        self.password_label.grid(row=4, column=1, padx=10, pady=10, sticky="w")

        # Show/Hide toggle button
        self.toggle_button = ctk.CTkButton(
            self.info_frame,
            text="Show",
            width=80,
            command=self.toggle_password
        )
        self.toggle_button.grid(row=4, column=2, padx=10)

        # -----------------------------
        # Action Buttons
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Change Password",
            width=200,
            command=lambda: self.master.show_frame("change_password")
        ).pack(pady=10)

        ctk.CTkButton(
            self,
            text="Back",
            width=200,
            command=lambda: self.master.show_frame("main_employee")
        ).pack(pady=20)

    # ---------------------------------------------------------
    # Populate fields when shown
    # ---------------------------------------------------------
    def on_show(self):
        user = self.master.current_user

        self.fullname_label.configure(text=user.fullname)
        self.username_label.configure(text=user.username)
        self.role_label.configure(text=user.role)

        # Updated vacation field
        self.vacation_label.configure(text=str(user.vacation_max))

        # Mask password by default
        self.password_visible = False
        self.password_label.configure(text="*" * len(user.pwd))
        self.toggle_button.configure(text="Show")

    # ---------------------------------------------------------
    # Toggle password visibility
    # ---------------------------------------------------------
    def toggle_password(self):
        user = self.master.current_user

        if self.password_visible:
            self.password_label.configure(text="*" * len(user.pwd))
            self.toggle_button.configure(text="Show")
            self.password_visible = False
        else:
            self.password_label.configure(text=user.pwd)
            self.toggle_button.configure(text="Hide")
            self.password_visible = True

# =========================================================
# Request Vacation Frame
# =========================================================
class RequestVacationFrame(ctk.CTkFrame):
    """Employees submit vacation day requests for admin approval."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="Request Vacation",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title.pack(pady=30)

        # -----------------------------
        # Info Frame
        # -----------------------------
        info = ctk.CTkFrame(self)
        info.pack(pady=20)

        ctk.CTkLabel(info, text="Vacation Days Remaining:").grid(
            row=0, column=0, padx=10, pady=10
        )
        self.remaining_label = ctk.CTkLabel(info, text="")
        self.remaining_label.grid(row=0, column=1, padx=10, pady=10)

        ctk.CTkLabel(info, text="Days Requested:").grid(
            row=1, column=0, padx=10, pady=10
        )
        self.days_entry = ctk.CTkEntry(info, width=200)
        self.days_entry.grid(row=1, column=1, padx=10, pady=10)

        # -----------------------------
        # Submit Button
        # -----------------------------
        ctk.CTkButton(
            self,
            text="Submit Request",
            width=200,
            command=self.submit_request
        ).pack(pady=20)

        # Status message
        self.message_label = ctk.CTkLabel(
            self,
            text="",
            font=ctk.CTkFont(size=16)
        )
        self.message_label.pack(pady=10)

        # Back button
        ctk.CTkButton(
            self,
            text="Back",
            width=200,
            command=lambda: self.master.show_frame("main_employee")
        ).pack(pady=20)

    # ---------------------------------------------------------
    # Populate fields when shown
    # ---------------------------------------------------------
    def on_show(self):
        user = self.master.current_user
        self.remaining_label.configure(text=str(user.vacation_remaining))
        self.days_entry.delete(0, "end")
        self.message_label.configure(text="")

    # ---------------------------------------------------------
    # Submit vacation request
    # ---------------------------------------------------------
    def submit_request(self):
        user = self.master.current_user

        # Validate input
        try:
            days_requested = float(self.days_entry.get().strip())
        except ValueError, TypeError:
            self.message_label.configure(
                text="Invalid number of days.",
                text_color="red"
            )
            return

        if days_requested <= 0:
            self.message_label.configure(
                text="Enter a positive number.",
                text_color="red"
            )
            return

        if days_requested > user.vacation_remaining:
            self.message_label.configure(
                text="Not enough vacation days remaining.",
                text_color="red"
            )
            return

        # -----------------------------
        # Load or create request file
        # -----------------------------
        req_file = netpath(REQUEST_FILE)

        if not os.path.exists(req_file):
            df = pd.DataFrame(columns=[
                "date_requested", "username", "fullname",
                "days_requested", "status", "admin_comment"
            ])
            df.to_excel(req_file, index=False)

        df = pd.read_excel(req_file)

        new_row = {
            "date_requested": datetime.date.today().strftime("%Y-%m-%d"),
            "username": user.username,
            "fullname": user.fullname,
            "days_requested": days_requested,
            "status": "Pending",
            "admin_comment": ""
        }

        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        df.to_excel(req_file, index=False)

        self.message_label.configure(
            text="Your request has been submitted for approval.",
            text_color="green"
        )

# =========================================================
# Employee Weekly Summary Frame
# =========================================================
# =========================================================
# Employee Weekly Summary Frame
# =========================================================
class EmployeeWeeklyPayFrame(ctk.CTkFrame):
    """Displays a weekly payroll summary for the logged-in employee."""

    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.place(relwidth=1, relheight=1)

        # Back button (top-left)
        ctk.CTkButton(
            self,
            text="Back",
            width=120,
            command=lambda: self.master.show_frame("main_employee")
        ).place(x=20, y=20)

        # -----------------------------
        # Title
        # -----------------------------
        title = ctk.CTkLabel(
            self,
            text="My Weekly Payroll Summary",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title.pack(pady=20)

        # -----------------------------
        # Week Selector
        # -----------------------------
        selector_frame = ctk.CTkFrame(self)
        selector_frame.pack(pady=10)

        ctk.CTkLabel(selector_frame, text="Choose Week:").grid(
            row=0, column=0, padx=10
        )

        self.week_list = self.generate_week_list()
        self.week_option = ctk.CTkOptionMenu(
            selector_frame,
            values=self.week_list,
            width=250,
            command=lambda _: self.load_week()
        )
        self.week_option.grid(row=0, column=1, padx=10)

        self.week_option.set(self.week_list[0])

        # -----------------------------
        # Scrollable Display Area
        # -----------------------------
        self.scroll = ctk.CTkScrollableFrame(self, width=900, height=500)
        self.scroll.pack(pady=20)

    # ---------------------------------------------------------
    # Generate last 12 weeks
    # ---------------------------------------------------------
    @staticmethod
    def generate_week_list():
        today = datetime.date.today()
        weeks = []

        for i in range(12):
            start = today - datetime.timedelta(days=today.weekday() + (7 * i))
            end = start + datetime.timedelta(days=6)
            label = f"{start.strftime('%b %d')} - {end.strftime('%b %d, %Y')}"
            weeks.append(label)

        return weeks

    # ---------------------------------------------------------
    # Parse week label safely
    # ---------------------------------------------------------
    @staticmethod
    def parse_week(label):
        start_text, end_text = label.split(" - ")
        end = datetime.datetime.strptime(end_text, "%b %d, %Y").date()
        start = datetime.datetime.strptime(
            f"{start_text} {end.year}", "%b %d %Y"
        ).date()
        return start, end

    # ---------------------------------------------------------
    # Load weekly summary
    # ---------------------------------------------------------
    def load_week(self):
        # Clear old content
        for widget in self.scroll.winfo_children():
            widget.destroy()

        user = self.master.current_user
        fullname = user.fullname

        # Payroll file path
        path = os.path.join(MASTER_PAYROLL_DIR, f"{fullname}.xlsx")
        if not os.path.exists(path):
            ctk.CTkLabel(self.scroll, text="No payroll file found.").pack(pady=20)
            return

        df = pd.read_excel(path)

        # Normalize headers
        df.columns = [str(c).strip().lower() for c in df.columns]

        # Detect required columns
        col_map = {
            "date": ["date", "day", "work_date"],
            "job": ["job", "job name"],
            "pay_item": ["pay item", "payitem"],
            "quantity": ["quantity", "qty"],
            "rate": ["rate", "payrate"],
            "split": ["split", "pay", "amount", "total"]
        }

        resolved = {}
        for target, options in col_map.items():
            for opt in options:
                if opt in df.columns:
                    resolved[target] = opt
                    break

        required = ["date", "job", "pay_item", "quantity", "rate", "split"]
        for col in required:
            if col not in resolved:
                ctk.CTkLabel(self.scroll, text=f"Missing column: {col}").pack(pady=20)
                return

        # Convert date
        df[resolved["date"]] = pd.to_datetime(df[resolved["date"]], errors="coerce").dt.date

        # Filter by selected week
        week_label = self.week_option.get()
        start_week, end_week = self.parse_week(week_label)

        df = df[(df[resolved["date"]] >= start_week) & (df[resolved["date"]] <= end_week)]

        if df.empty:
            ctk.CTkLabel(self.scroll, text="No payroll data for this week.").pack(pady=20)
            return

        # -----------------------------
        # Weekly Total
        # -----------------------------
        total = df[resolved["split"]].sum()

        ctk.CTkLabel(
            self.scroll,
            text=f"Weekly Total: ${total:,.2f}",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(pady=10)

        # -----------------------------
        # Section Builder
        # -----------------------------
        def build_section(title, data_dict):
            ctk.CTkLabel(
                self.scroll,
                text=title,
                font=ctk.CTkFont(size=16, weight="bold")
            ).pack(pady=10)

            table = ctk.CTkFrame(self.scroll)
            table.pack(pady=5)

            # Headers
            ctk.CTkLabel(table, text="Name", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10)
            ctk.CTkLabel(table, text="Amount", font=ctk.CTkFont(weight="bold")).grid(row=0, column=1, padx=10)

            if not data_dict:
                ctk.CTkLabel(table, text="No data available").grid(row=1, column=0, columnspan=2)
                return

            for i, (name, amount) in enumerate(sorted(data_dict.items()), start=1):
                ctk.CTkLabel(table, text=str(name)).grid(row=i, column=0, padx=10)
                ctk.CTkLabel(table, text=f"${amount:,.2f}").grid(row=i, column=1, padx=10)

        # -----------------------------
        # Totals Per Job
        # -----------------------------
        job_totals = df.groupby(resolved["job"])[resolved["split"]].sum().to_dict()
        build_section("Totals Per Job", job_totals)

        # -----------------------------
        # Totals Per Pay Item
        # -----------------------------
        pay_item_totals = df.groupby(resolved["pay_item"])[resolved["split"]].sum().to_dict()
        build_section("Totals Per Pay Item", pay_item_totals)

        # -----------------------------
        # Totals Per Day
        # -----------------------------
        day_totals = df.groupby(resolved["date"])[resolved["split"]].sum().to_dict()
        build_section("Totals Per Day", day_totals)


class CommandPalette(ctk.CTkToplevel):
    """A searchable command launcher (Ctrl+Shift+P)."""

    def __init__(self, master, commands):
        super().__init__(master)
        self.master = master
        self.commands = commands

        # Window setup
        self.title("Command Palette")
        self.geometry("500x400")
        self.resizable(False, False)
        self.grab_set()  # Modal behavior

        # Center window
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - 250
        y = (self.winfo_screenheight() // 2) - 200
        self.geometry(f"+{x}+{y}")

        # Search box
        self.search_entry = ctk.CTkEntry(self, placeholder_text="Type a command...")
        self.search_entry.pack(fill="x", padx=20, pady=20)
        self.search_entry.focus()

        # Listbox for results
        self.listbox = tk.Listbox(self, height=12, font=("Segoe UI", 12))
        self.listbox.pack(fill="both", expand=True, padx=20, pady=10)

        # Bindings
        self.search_entry.bind("<KeyRelease>", self.update_list)
        self.listbox.bind("<Double-Button-1>", self.execute_selected)
        self.bind("<Return>", lambda e: self.execute_selected(None))
        self.bind("<Escape>", lambda e: self.destroy())

        # Initial population
        self.update_list(None)

    def update_list(self, event):
        query = self.search_entry.get().lower()
        self.listbox.delete(0, tk.END)

        for label, action in self.commands.items():
            if query in label.lower():
                self.listbox.insert(tk.END, label)

    def execute_selected(self, event):
        selection = self.listbox.curselection()
        if not selection:
            return

        label = self.listbox.get(selection[0])
        action = self.commands[label]

        # Execute the command
        action()

        # Close the palette
        self.destroy()





if __name__ == "__main__":
    app = AppController()
    app.mainloop()
